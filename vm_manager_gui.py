#!/usr/bin/env python3
"""
Simple Dashboard - Giao dien don gian
=====================================
- Bam Start de chay
- Xem tien do tuc thi
- Click vao project de xem chi tiet
- Nut HIEN/AN CMD de toggle CMD windows

Usage:
    pythonw vm_manager_gui.py   (an console)
    python vm_manager_gui.py    (co console)
"""

import sys
import os

# AN CONSOLE WINDOW KHI CHAY GUI
if sys.platform == "win32":
    try:
        import ctypes
        # Ẩn console window
        ctypes.windll.user32.ShowWindow(ctypes.windll.kernel32.GetConsoleWindow(), 0)
    except:
        pass

    if sys.stdout:
        try:
            sys.stdout.reconfigure(encoding='utf-8', errors='replace')
        except:
            pass
    os.environ['PYTHONIOENCODING'] = 'utf-8'

import tkinter as tk
from tkinter import ttk
import threading
import time
from pathlib import Path
from typing import Dict, Optional

# PIL for thumbnails
try:
    from PIL import Image, ImageTk
    PIL_AVAILABLE = True
except:
    PIL_AVAILABLE = False

TOOL_DIR = Path(__file__).parent

# Import VM Manager
try:
    from vm_manager import VMManager
    VM_AVAILABLE = True
except:
    VM_AVAILABLE = False

# Import Central Logger
try:
    from modules.central_logger import get_recent_logs, tail_log, LOG_FILE, add_callback, remove_callback
    LOGGER_AVAILABLE = True
except:
    LOGGER_AVAILABLE = False

# Import Excel Status for detailed status checking
try:
    from modules.excel_status import check_project_status as check_excel_status, EXCEL_STEPS
    EXCEL_STATUS_AVAILABLE = True
except:
    EXCEL_STATUS_AVAILABLE = False


# ================================================================================
# SETTINGS WINDOW - Cau hinh va kiem tra tai nguyen
# ================================================================================

class SettingsWindow(tk.Toplevel):
    """Cua so cau hinh - kiem tra va thiet lap tai nguyen."""

    def __init__(self, parent):
        super().__init__(parent)
        self.title("CAU HINH - Settings")
        self.geometry("750x700")
        self.configure(bg='#1a1a2e')
        self.resizable(True, True)
        self.minsize(700, 650)

        self._build()
        self._load_settings()
        self._check_resources()

    def _build(self):
        # Header
        header = tk.Frame(self, bg='#e94560', height=50)
        header.pack(fill="x")
        header.pack_propagate(False)
        tk.Label(header, text="CAU HINH HE THONG", bg='#e94560', fg='white',
                 font=("Arial", 14, "bold")).pack(pady=12)

        # Main content with scroll
        main_frame = tk.Frame(self, bg='#1a1a2e')
        main_frame.pack(fill="both", expand=True, padx=10, pady=10)

        # === KIEM TRA TAI NGUYEN ===
        check_frame = tk.LabelFrame(main_frame, text=" KIEM TRA TAI NGUYEN ", bg='#16213e', fg='#00ff88',
                                    font=("Arial", 10, "bold"), padx=10, pady=10)
        check_frame.pack(fill="x", pady=5)

        self.resource_labels = {}
        resources = [
            ("chrome", "Chrome Portable", "Trinh duyet de tao anh/video"),
            ("api_key", "API Key (Gemini/DeepSeek)", "De tao noi dung Excel"),
            ("proxy_token", "Proxy API Token", "De tao video qua API"),
            ("projects", "Thu muc PROJECTS", "Chua cac du an"),
        ]

        for res_id, name, desc in resources:
            row = tk.Frame(check_frame, bg='#16213e')
            row.pack(fill="x", pady=3)

            # Status icon
            status_lbl = tk.Label(row, text="?", width=3, bg='#16213e', fg='#ffd93d',
                                  font=("Arial", 12, "bold"))
            status_lbl.pack(side="left", padx=5)
            self.resource_labels[res_id] = status_lbl

            # Name and desc
            tk.Label(row, text=name, width=25, bg='#16213e', fg='white',
                     font=("Arial", 10), anchor="w").pack(side="left", padx=5)
            tk.Label(row, text=desc, bg='#16213e', fg='#888',
                     font=("Arial", 9), anchor="w").pack(side="left", padx=5)

        # Refresh button
        tk.Button(check_frame, text="Kiem tra lai", command=self._check_resources,
                  bg='#0984e3', fg='white', font=("Arial", 9), relief="flat", padx=10).pack(pady=5)

        # === API KEYS ===
        api_frame = tk.LabelFrame(main_frame, text=" API KEYS ", bg='#16213e', fg='#ffd93d',
                                  font=("Arial", 10, "bold"), padx=10, pady=10)
        api_frame.pack(fill="x", pady=5)

        # DeepSeek API Key
        tk.Label(api_frame, text="DeepSeek API Key:", bg='#16213e', fg='white',
                 font=("Arial", 10)).pack(anchor="w")
        self.deepseek_var = tk.StringVar()
        deepseek_entry = tk.Entry(api_frame, textvariable=self.deepseek_var, width=60,
                                  font=("Consolas", 10), bg='#0f3460', fg='white',
                                  insertbackground='white')
        deepseek_entry.pack(fill="x", pady=2)

        # Gemini API Key
        tk.Label(api_frame, text="Gemini API Key:", bg='#16213e', fg='white',
                 font=("Arial", 10)).pack(anchor="w", pady=(10, 0))
        self.gemini_var = tk.StringVar()
        gemini_entry = tk.Entry(api_frame, textvariable=self.gemini_var, width=60,
                                font=("Consolas", 10), bg='#0f3460', fg='white',
                                insertbackground='white')
        gemini_entry.pack(fill="x", pady=2)

        # Proxy API Token
        tk.Label(api_frame, text="Proxy API Token (Video):", bg='#16213e', fg='white',
                 font=("Arial", 10)).pack(anchor="w", pady=(10, 0))
        self.proxy_token_var = tk.StringVar()
        proxy_entry = tk.Entry(api_frame, textvariable=self.proxy_token_var, width=60,
                               font=("Consolas", 10), bg='#0f3460', fg='white',
                               insertbackground='white')
        proxy_entry.pack(fill="x", pady=2)

        # === CHROME ===
        chrome_frame = tk.LabelFrame(main_frame, text=" CHROME PORTABLE ", bg='#16213e', fg='#00d9ff',
                                     font=("Arial", 10, "bold"), padx=10, pady=10)
        chrome_frame.pack(fill="x", pady=5)

        tk.Label(chrome_frame, text="Duong dan Chrome 1:", bg='#16213e', fg='white',
                 font=("Arial", 10)).pack(anchor="w")
        self.chrome1_var = tk.StringVar()
        chrome1_row = tk.Frame(chrome_frame, bg='#16213e')
        chrome1_row.pack(fill="x", pady=2)
        tk.Entry(chrome1_row, textvariable=self.chrome1_var, width=50,
                 font=("Consolas", 9), bg='#0f3460', fg='white',
                 insertbackground='white').pack(side="left", fill="x", expand=True)
        tk.Button(chrome1_row, text="Chon...", command=lambda: self._browse_chrome(1),
                  bg='#6c5ce7', fg='white', font=("Arial", 8), relief="flat").pack(side="left", padx=5)

        tk.Label(chrome_frame, text="Duong dan Chrome 2:", bg='#16213e', fg='white',
                 font=("Arial", 10)).pack(anchor="w", pady=(10, 0))
        self.chrome2_var = tk.StringVar()
        chrome2_row = tk.Frame(chrome_frame, bg='#16213e')
        chrome2_row.pack(fill="x", pady=2)
        tk.Entry(chrome2_row, textvariable=self.chrome2_var, width=50,
                 font=("Consolas", 9), bg='#0f3460', fg='white',
                 insertbackground='white').pack(side="left", fill="x", expand=True)
        tk.Button(chrome2_row, text="Chon...", command=lambda: self._browse_chrome(2),
                  bg='#6c5ce7', fg='white', font=("Arial", 8), relief="flat").pack(side="left", padx=5)

        # === BUTTONS ===
        btn_frame = tk.Frame(self, bg='#1a1a2e')
        btn_frame.pack(fill="x", padx=10, pady=10)

        tk.Button(btn_frame, text="LUU CAU HINH", command=self._save_settings,
                  bg='#00ff88', fg='#1a1a2e', font=("Arial", 11, "bold"),
                  relief="flat", padx=20, pady=5).pack(side="left", padx=5)

        tk.Button(btn_frame, text="DONG", command=self.destroy,
                  bg='#e94560', fg='white', font=("Arial", 11, "bold"),
                  relief="flat", padx=20, pady=5).pack(side="right", padx=5)

    def _load_settings(self):
        """Load settings tu file."""
        try:
            import yaml
            settings_path = TOOL_DIR / "config" / "settings.yaml"
            if settings_path.exists():
                with open(settings_path, "r", encoding="utf-8") as f:
                    config = yaml.safe_load(f) or {}

                self.deepseek_var.set(config.get('deepseek_api_key', ''))
                gemini_keys = config.get('gemini_api_keys', [''])
                self.gemini_var.set(gemini_keys[0] if gemini_keys else '')
                self.proxy_token_var.set(config.get('proxy_api_token', ''))
                self.chrome1_var.set(config.get('chrome_portable', ''))
                self.chrome2_var.set(config.get('chrome_portable_2', ''))
        except Exception as e:
            print(f"Error loading settings: {e}")

    def _save_settings(self):
        """Luu settings vao file."""
        try:
            import yaml
            settings_path = TOOL_DIR / "config" / "settings.yaml"

            # Doc config hien tai
            config = {}
            if settings_path.exists():
                with open(settings_path, "r", encoding="utf-8") as f:
                    config = yaml.safe_load(f) or {}

            # Update
            config['deepseek_api_key'] = self.deepseek_var.get().strip()
            gemini_key = self.gemini_var.get().strip()
            config['gemini_api_keys'] = [gemini_key] if gemini_key else ['']
            config['proxy_api_token'] = self.proxy_token_var.get().strip()
            config['chrome_portable'] = self.chrome1_var.get().strip()
            config['chrome_portable_2'] = self.chrome2_var.get().strip()

            # Luu
            with open(settings_path, "w", encoding="utf-8") as f:
                yaml.safe_dump(config, f, default_flow_style=False, allow_unicode=True)

            # Thong bao
            from tkinter import messagebox
            messagebox.showinfo("Thanh cong", "Da luu cau hinh!")
            self._check_resources()

        except Exception as e:
            from tkinter import messagebox
            messagebox.showerror("Loi", f"Khong the luu: {e}")

    def _check_resources(self):
        """Kiem tra cac tai nguyen."""
        # Chrome
        chrome1 = self.chrome1_var.get()
        chrome2 = self.chrome2_var.get()
        chrome_ok = (chrome1 and Path(chrome1).exists()) or (chrome2 and Path(chrome2).exists())
        self._set_status("chrome", chrome_ok)

        # API Keys
        deepseek = self.deepseek_var.get().strip()
        gemini = self.gemini_var.get().strip()
        api_ok = bool(deepseek) or bool(gemini)
        self._set_status("api_key", api_ok)

        # Proxy Token
        proxy_token = self.proxy_token_var.get().strip()
        self._set_status("proxy_token", bool(proxy_token))

        # Projects folder
        projects_dir = TOOL_DIR / "PROJECTS"
        self._set_status("projects", projects_dir.exists())

    def _set_status(self, res_id: str, ok: bool):
        """Set status icon."""
        if res_id in self.resource_labels:
            if ok:
                self.resource_labels[res_id].config(text="OK", fg='#00ff88')
            else:
                self.resource_labels[res_id].config(text="X", fg='#e94560')

    def _browse_chrome(self, num: int):
        """Chon file Chrome."""
        from tkinter import filedialog
        path = filedialog.askopenfilename(
            title=f"Chon Chrome Portable {num}",
            filetypes=[("Executable", "*.exe"), ("All files", "*.*")]
        )
        if path:
            if num == 1:
                self.chrome1_var.set(path)
            else:
                self.chrome2_var.set(path)
            self._check_resources()


# ================================================================================
# PROJECT DETAIL - Hiển thị chi tiết từng scene
# ================================================================================

def find_scene_image(img_folder: Path, scene_id: int) -> Path | None:
    """
    Tìm ảnh scene với nhiều format khác nhau:
    - scene_001.png (format chuẩn)
    - 1.png (format đơn giản)
    - scene_1.png (không padding)
    """
    if not img_folder or not img_folder.exists():
        return None

    # Các format có thể có
    candidates = [
        img_folder / f"scene_{scene_id:03d}.png",  # scene_001.png
        img_folder / f"{scene_id}.png",             # 1.png
        img_folder / f"scene_{scene_id}.png",       # scene_1.png
        img_folder / f"scene_{scene_id:03d}.jpg",  # scene_001.jpg
        img_folder / f"{scene_id}.jpg",             # 1.jpg
    ]

    for path in candidates:
        if path.exists():
            return path
    return None


def find_scene_video(vid_folder: Path, scene_id: int) -> Path | None:
    """
    Tìm video scene với nhiều format khác nhau.
    """
    if not vid_folder or not vid_folder.exists():
        return None

    candidates = [
        vid_folder / f"scene_{scene_id:03d}.mp4",
        vid_folder / f"{scene_id}.mp4",
        vid_folder / f"scene_{scene_id}.mp4",
    ]

    for path in candidates:
        if path.exists():
            return path
    return None


class ProjectDetail(tk.Toplevel):
    """Xem chi tiết project - Excel steps + từng scene."""

    def __init__(self, parent, code: str):
        super().__init__(parent)
        self.code = code
        self.title(f"{code} - Chi tiết")
        self.geometry("900x700")
        self.configure(bg='#1a1a2e')
        self._auto_refresh = True

        self._build()
        self._load()
        self._start_auto_refresh()

    def _start_auto_refresh(self):
        """Auto refresh mỗi 3 giây."""
        if self._auto_refresh and self.winfo_exists():
            self._load()
            self.after(3000, self._start_auto_refresh)

    def destroy(self):
        """Stop auto refresh khi đóng."""
        self._auto_refresh = False
        super().destroy()

    def _build(self):
        # Header
        tk.Label(
            self, text=f"PROJECT: {self.code}",
            font=("Arial", 16, "bold"),
            bg='#0f3460', fg='white',
            pady=15
        ).pack(fill="x")

        # Buttons
        btn_row = tk.Frame(self, bg='#1a1a2e')
        btn_row.pack(fill="x", padx=10, pady=5)

        tk.Button(btn_row, text="Mở thư mục", command=self._open_folder,
                  bg='#e94560', fg='white', font=("Arial", 10), relief="flat", padx=15).pack(side="left", padx=5)
        tk.Button(btn_row, text="Làm mới", command=self._load,
                  bg='#0f3460', fg='white', font=("Arial", 10), relief="flat", padx=15).pack(side="left", padx=5)

        # === EXCEL STEPS SECTION ===
        excel_frame = tk.LabelFrame(self, text=" EXCEL - 7 BƯỚC ", bg='#16213e', fg='white',
                                    font=("Arial", 10, "bold"), padx=10, pady=5)
        excel_frame.pack(fill="x", padx=10, pady=5)

        # Steps header
        steps_header = tk.Frame(excel_frame, bg='#0f3460')
        steps_header.pack(fill="x")
        for txt, w in [("Bước", 20), ("Trạng thái", 12), ("Thời gian", 15), ("Ghi chú", 30)]:
            tk.Label(steps_header, text=txt, width=w, bg='#0f3460', fg='white',
                     font=("Arial", 9, "bold")).pack(side="left", padx=2, pady=3)

        # Steps container
        self.steps_frame = tk.Frame(excel_frame, bg='#16213e')
        self.steps_frame.pack(fill="x")

        # === SCENES SECTION ===
        # Summary
        self.summary_var = tk.StringVar(value="Đang tải...")
        tk.Label(self, textvariable=self.summary_var, bg='#1a1a2e', fg='#00d9ff',
                 font=("Consolas", 12, "bold")).pack(pady=5)

        # Scenes list
        list_frame = tk.LabelFrame(self, text=" SCENES - ẢNH & VIDEO ", bg='#16213e', fg='white',
                                   font=("Arial", 10, "bold"), padx=5, pady=5)
        list_frame.pack(fill="both", expand=True, padx=10, pady=(0, 10))

        # Header
        header = tk.Frame(list_frame, bg='#0f3460')
        header.pack(fill="x")
        for txt, w in [("Scene", 6), ("Prompt", 35), ("Ảnh", 6), ("Video", 6), ("Status", 10)]:
            tk.Label(header, text=txt, width=w, bg='#0f3460', fg='white',
                     font=("Arial", 9, "bold")).pack(side="left", padx=2, pady=5)

        # Scrollable
        canvas = tk.Canvas(list_frame, bg='#16213e', highlightthickness=0)
        scrollbar = ttk.Scrollbar(list_frame, orient="vertical", command=canvas.yview)
        self.scenes_frame = tk.Frame(canvas, bg='#16213e')

        canvas.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side="right", fill="y")
        canvas.pack(side="left", fill="both", expand=True)
        canvas.create_window((0, 0), window=self.scenes_frame, anchor="nw")
        self.scenes_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))

        # Mouse wheel
        canvas.bind_all("<MouseWheel>", lambda e: canvas.yview_scroll(int(-1*(e.delta/120)), "units"))

    def _load(self):
        """Load Excel steps và scenes."""
        # Clear
        for w in self.steps_frame.winfo_children():
            w.destroy()
        for w in self.scenes_frame.winfo_children():
            w.destroy()

        project_dir = self._find_dir()
        if not project_dir:
            self.summary_var.set("Không tìm thấy project!")
            return

        excel_path = project_dir / f"{self.code}_prompts.xlsx"
        if not excel_path.exists():
            self.summary_var.set("Chưa có Excel!")
            self._show_no_excel_steps()
            return

        try:
            from modules.excel_manager import PromptWorkbook
            wb = PromptWorkbook(str(excel_path))
            wb.load_or_create()  # PHẢI gọi load trước khi dùng

            # Load Excel steps status
            self._load_excel_steps(wb)

            scenes = wb.get_scenes()

            if not scenes:
                self.summary_var.set("Excel chưa có scenes!")
                return

            img_folder = project_dir / "img"
            vid_folder = project_dir / "vid"

            img_ok = vid_ok = vid_need = 0

            for i, scene in enumerate(scenes):
                # Handle both Scene objects and dicts
                if hasattr(scene, 'scene_id'):
                    sid = scene.scene_id
                    prompt = (scene.img_prompt or "")[:35]
                    vid_enabled = scene.video_enabled if hasattr(scene, 'video_enabled') else False
                else:
                    sid = scene.get('scene_id', i + 1)
                    prompt = (scene.get('img_prompt') or "")[:35]
                    vid_enabled = scene.get('video_enabled', False)

                img_path = find_scene_image(img_folder, sid)
                vid_path = find_scene_video(vid_folder, sid)
                has_img = img_path is not None
                has_vid = vid_path is not None

                if has_img:
                    img_ok += 1
                if vid_enabled:
                    vid_need += 1
                    if has_vid:
                        vid_ok += 1

                # Row
                bg = '#1a1a2e' if i % 2 == 0 else '#16213e'
                row = tk.Frame(self.scenes_frame, bg=bg)
                row.pack(fill="x", pady=1)

                tk.Label(row, text=str(sid), width=4, bg=bg, fg='white', font=("Consolas", 9)).pack(side="left", padx=2)

                # Thumbnail (40x40)
                thumb_label = tk.Label(row, width=5, height=2, bg='#333', cursor="hand2")
                thumb_label.pack(side="left", padx=2)

                if has_img and img_path and PIL_AVAILABLE:
                    try:
                        img = Image.open(str(img_path))
                        img.thumbnail((40, 40))
                        photo = ImageTk.PhotoImage(img)
                        thumb_label.configure(image=photo, width=40, height=40)
                        thumb_label.image = photo  # Keep reference
                        thumb_label.bind("<Button-1>", lambda e, s=sid: self._open_img(s))
                    except:
                        thumb_label.configure(text="IMG", fg='#00ff88')
                        thumb_label.bind("<Button-1>", lambda e, s=sid: self._open_img(s))
                elif has_img:
                    thumb_label.configure(text="IMG", fg='#00ff88')
                    thumb_label.bind("<Button-1>", lambda e, s=sid: self._open_img(s))
                else:
                    thumb_label.configure(text="--", fg='#666')

                tk.Label(row, text=prompt, width=30, bg=bg, fg='#aaa', font=("Consolas", 8), anchor="w").pack(side="left", padx=2)

                # Video status
                if not vid_enabled:
                    vid_txt, vid_clr = "-", '#444'
                elif has_vid:
                    vid_txt, vid_clr = "VID", '#00ff88'
                else:
                    vid_txt, vid_clr = "...", '#666'

                vid_lbl = tk.Label(row, text=vid_txt, width=5, bg=bg, fg=vid_clr, font=("Arial", 9, "bold"), cursor="hand2")
                vid_lbl.pack(side="left", padx=2)
                if has_vid:
                    vid_lbl.bind("<Button-1>", lambda e, s=sid: self._open_vid(s))

                # Status
                if has_img and (has_vid or not vid_enabled):
                    status_txt, status_clr = "OK", '#00ff88'
                elif has_img:
                    status_txt, status_clr = "...", '#ffaa00'
                else:
                    status_txt, status_clr = "--", '#666'

                tk.Label(row, text=status_txt, width=4, bg=bg, fg=status_clr, font=("Arial", 9)).pack(side="left", padx=2)

            self.summary_var.set(f"Anh: {img_ok}/{len(scenes)} | Video: {vid_ok}/{vid_need}")

        except Exception as e:
            self.summary_var.set(f"Loi: {e}")

    def _load_excel_steps(self, wb):
        """Load và hiển thị trạng thái các bước Excel."""
        step_names = [
            ("step_1", "1. Story Analysis"),
            ("step_2", "2. Story Segments"),
            ("step_3", "3. Characters"),
            ("step_4", "4. Locations"),
            ("step_5", "5. Director Plan"),
            ("step_6", "6. Scene Planning"),
            ("step_7", "7. Scene Prompts")
        ]

        try:
            all_status = wb.get_all_step_status() if hasattr(wb, 'get_all_step_status') else []

            # Convert to dict
            status_dict = {}
            for s in all_status:
                if isinstance(s, dict):
                    status_dict[s.get('step_id', '')] = s

            for i, (step_id, step_name) in enumerate(step_names):
                bg = '#1a1a2e' if i % 2 == 0 else '#16213e'
                row = tk.Frame(self.steps_frame, bg=bg)
                row.pack(fill="x")

                step_status = status_dict.get(step_id, {})
                status = step_status.get('status', 'PENDING')
                last_updated = step_status.get('last_updated', '')
                notes_raw = step_status.get('notes', '') or ''

                # Extract duration from notes (format: "Xs - description")
                duration_txt = "--"
                notes_display = notes_raw[:30]
                if notes_raw and 's - ' in notes_raw:
                    parts = notes_raw.split('s - ', 1)
                    if parts[0].isdigit():
                        secs = int(parts[0])
                        if secs >= 60:
                            duration_txt = f"{secs//60}m{secs%60:02d}s"
                        else:
                            duration_txt = f"{secs}s"
                        notes_display = parts[1][:25] if len(parts) > 1 else ""

                # Status color and icon
                if status == 'COMPLETED':
                    status_txt = "OK"
                    status_clr = '#00ff88'
                elif status == 'IN_PROGRESS':
                    status_txt = "Dang chay"
                    status_clr = '#00d9ff'
                elif status == 'ERROR':
                    status_txt = "Loi"
                    status_clr = '#e94560'
                elif status == 'PARTIAL':
                    status_txt = "Mot phan"
                    status_clr = '#ffaa00'
                else:
                    status_txt = "Cho"
                    status_clr = '#666'

                tk.Label(row, text=step_name, width=20, bg=bg, fg='white',
                         font=("Consolas", 9), anchor="w").pack(side="left", padx=2)
                tk.Label(row, text=status_txt, width=10, bg=bg, fg=status_clr,
                         font=("Arial", 9, "bold")).pack(side="left", padx=2)
                tk.Label(row, text=duration_txt, width=8, bg=bg, fg='#ffcc00',
                         font=("Consolas", 9, "bold")).pack(side="left", padx=2)
                tk.Label(row, text=notes_display, width=28, bg=bg, fg='#888',
                         font=("Consolas", 8), anchor="w").pack(side="left", padx=2)

        except Exception as e:
            self._show_no_excel_steps()

    def _show_no_excel_steps(self):
        """Hiển thị steps khi chưa có Excel."""
        step_names = [
            "1. Story Analysis",
            "2. Story Segments",
            "3. Characters",
            "4. Locations",
            "5. Director Plan",
            "6. Scene Planning",
            "7. Scene Prompts"
        ]

        for i, name in enumerate(step_names):
            bg = '#1a1a2e' if i % 2 == 0 else '#16213e'
            row = tk.Frame(self.steps_frame, bg=bg)
            row.pack(fill="x")

            tk.Label(row, text=name, width=20, bg=bg, fg='#666',
                     font=("Consolas", 9), anchor="w").pack(side="left", padx=2)
            tk.Label(row, text="Cho", width=12, bg=bg, fg='#444',
                     font=("Arial", 9)).pack(side="left", padx=2)
            tk.Label(row, text="--", width=15, bg=bg, fg='#444',
                     font=("Consolas", 9)).pack(side="left", padx=2)
            tk.Label(row, text="", width=30, bg=bg, fg='#444',
                     font=("Consolas", 8)).pack(side="left", padx=2)

    def _find_dir(self):
        local = TOOL_DIR / "PROJECTS" / self.code
        if local.exists():
            return local
        for drive in ["Z:", "Y:", "X:"]:
            master = Path(f"{drive}/AUTO/ve3-tool-simple/PROJECTS/{self.code}")
            if master.exists():
                return master
        return None

    def _open_folder(self):
        d = self._find_dir()
        if d:
            os.startfile(str(d))

    def _open_img(self, sid):
        d = self._find_dir()
        if d:
            p = d / "img" / f"scene_{sid:03d}.png"
            if p.exists():
                os.startfile(str(p))

    def _open_vid(self, sid):
        d = self._find_dir()
        if d:
            p = d / "vid" / f"scene_{sid:03d}.mp4"
            if p.exists():
                os.startfile(str(p))


# ================================================================================
# PROJECT DETAIL - Xem chi tiết project + ảnh tham chiếu
# ================================================================================

class ProjectDetail(tk.Toplevel):
    """Popup hiển thị chi tiết project và ảnh tham chiếu."""

    def __init__(self, parent, code: str):
        super().__init__(parent)
        self.code = code
        self.title(f"Chi tiet: {code}")
        self.geometry("800x600")
        self.configure(bg='#1a1a2e')

        self._build()

    def _find_project_dir(self) -> Optional[Path]:
        """Tìm thư mục project."""
        local = TOOL_DIR / "PROJECTS" / self.code
        if local.exists():
            return local
        return None

    def _build(self):
        # Header
        header = tk.Frame(self, bg='#0f3460', height=40)
        header.pack(fill="x")
        tk.Label(header, text=f"Project: {self.code}", bg='#0f3460', fg='#00ff88',
                 font=("Arial", 14, "bold")).pack(side="left", padx=20, pady=8)

        # Main content - 2 panels
        main = tk.PanedWindow(self, orient=tk.HORIZONTAL, bg='#1a1a2e', sashwidth=4)
        main.pack(fill="both", expand=True, padx=5, pady=5)

        # Left - Info
        left = tk.Frame(main, bg='#16213e')
        main.add(left, width=300)

        info_frame = tk.LabelFrame(left, text=" THONG TIN ", bg='#16213e', fg='white',
                                   font=("Arial", 10, "bold"), padx=10, pady=10)
        info_frame.pack(fill="x", padx=5, pady=5)

        project_dir = self._find_project_dir()
        if project_dir:
            # Count images
            img_dir = project_dir / "img"
            img_count = len(list(img_dir.glob("*.png"))) + len(list(img_dir.glob("*.jpg"))) if img_dir.exists() else 0

            # Count videos
            vid_dir = project_dir / "vid"
            vid_count = len(list(vid_dir.glob("*.mp4"))) if vid_dir.exists() else 0

            # Count reference files
            ref_dir = project_dir / "NV"
            ref_count = len(list(ref_dir.glob("*.*"))) if ref_dir.exists() else 0

            info_text = f"""
Thu muc: {project_dir}

So anh: {img_count}
So video: {vid_count}
So anh tham chieu: {ref_count}
"""
            tk.Label(info_frame, text=info_text, bg='#16213e', fg='#c8d6e5',
                     font=("Consolas", 10), justify="left", anchor="w").pack(anchor="w")
        else:
            tk.Label(info_frame, text="Khong tim thay project", bg='#16213e', fg='#ff6b6b',
                     font=("Consolas", 10)).pack()

        # Right - Reference images
        right = tk.Frame(main, bg='#16213e')
        main.add(right, width=480)

        ref_frame = tk.LabelFrame(right, text=" ANH THAM CHIEU (NV/) ", bg='#16213e', fg='white',
                                  font=("Arial", 10, "bold"), padx=5, pady=5)
        ref_frame.pack(fill="both", expand=True, padx=5, pady=5)

        # Canvas with scrollbar for images
        canvas = tk.Canvas(ref_frame, bg='#1a1a2e', highlightthickness=0)
        scrollbar = ttk.Scrollbar(ref_frame, orient="vertical", command=canvas.yview)
        scrollable_frame = tk.Frame(canvas, bg='#1a1a2e')

        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        scrollbar.pack(side="right", fill="y")
        canvas.pack(side="left", fill="both", expand=True)

        # Load reference images
        if project_dir:
            nv_dir = project_dir / "NV"
            if nv_dir.exists():
                self._load_reference_images(scrollable_frame, nv_dir)
            else:
                tk.Label(scrollable_frame, text="Khong co thu muc NV/", bg='#1a1a2e', fg='#666',
                         font=("Consolas", 10)).pack(pady=20)

    def _load_reference_images(self, parent, nv_dir: Path):
        """Load và hiển thị ảnh tham chiếu từ thư mục NV."""
        if not PIL_AVAILABLE:
            tk.Label(parent, text="Can cai PIL de xem anh\npip install Pillow", bg='#1a1a2e', fg='#ffd93d',
                     font=("Consolas", 10)).pack(pady=20)
            return

        # Get all image files
        image_files = []
        for ext in ['*.png', '*.jpg', '*.jpeg', '*.webp']:
            image_files.extend(nv_dir.glob(ext))
        image_files.sort()

        if not image_files:
            tk.Label(parent, text="Khong co anh trong NV/", bg='#1a1a2e', fg='#666',
                     font=("Consolas", 10)).pack(pady=20)
            return

        # Display images in grid (3 columns)
        cols = 3
        row_frame = None
        self.photo_refs = []  # Keep references to prevent garbage collection

        for i, img_path in enumerate(image_files[:30]):  # Limit to 30 images
            if i % cols == 0:
                row_frame = tk.Frame(parent, bg='#1a1a2e')
                row_frame.pack(fill="x", pady=2)

            try:
                img = Image.open(img_path)
                img.thumbnail((140, 100))  # Thumbnail size
                photo = ImageTk.PhotoImage(img)
                self.photo_refs.append(photo)  # Keep reference

                # Frame for each image
                img_frame = tk.Frame(row_frame, bg='#0f3460', padx=2, pady=2)
                img_frame.pack(side="left", padx=3, pady=3)

                # Image label
                lbl = tk.Label(img_frame, image=photo, bg='#0f3460')
                lbl.pack()

                # Filename
                name = img_path.stem[:15] + "..." if len(img_path.stem) > 15 else img_path.stem
                tk.Label(img_frame, text=name, bg='#0f3460', fg='#aaa',
                         font=("Consolas", 7)).pack()

                # Click to open full size
                lbl.bind("<Button-1>", lambda e, p=img_path: os.startfile(str(p)))
                lbl.config(cursor="hand2")

            except Exception as e:
                pass

        # Info label
        tk.Label(parent, text=f"Tong: {len(image_files)} anh (click de mo)", bg='#1a1a2e', fg='#666',
                 font=("Consolas", 9)).pack(pady=10)


# ================================================================================
# MAIN GUI - Dashboard với LOG
# ================================================================================

class SimpleGUI(tk.Tk):
    """Dashboard don gian - click project de xem chi tiet."""

    def __init__(self):
        super().__init__()
        self.title("VE3 Dashboard")
        self.geometry("1600x900")
        self.configure(bg='#1a1a2e')
        self.minsize(1400, 800)

        # Hide current CMD window if running from CMD
        self._hide_current_cmd_window()

        # Khoi tao manager ngay de hien projects
        self.manager = VMManager(num_chrome_workers=2)
        self.running = False
        self.selected_project = None  # Project dang xem chi tiet
        self.scene_photo_refs = []  # Keep references for thumbnails
        self.windows_visible = False  # Track CMD/Chrome visibility

        self._build()
        self._load_projects_on_startup()  # Load projects ngay khi mo
        self._update_loop()

    def _hide_current_cmd_window(self):
        """Hide the CMD window that launched this GUI."""
        try:
            import win32gui
            import win32con
            import ctypes

            # Get console window handle
            kernel32 = ctypes.windll.kernel32
            hwnd = kernel32.GetConsoleWindow()

            if hwnd:
                # Hide it
                win32gui.ShowWindow(hwnd, win32con.SW_HIDE)
                print("[GUI] Hidden launch CMD window")
        except Exception as e:
            print(f"[GUI] Could not hide CMD window: {e}")

    def _build(self):
        # === TOP: Controls ===
        top = tk.Frame(self, bg='#0f3460', height=60)
        top.pack(fill="x")
        top.pack_propagate(False)

        # Start/Stop
        self.start_btn = tk.Button(
            top, text="BAT DAU", command=self._start,
            bg='#00ff88', fg='#1a1a2e', font=("Arial", 12, "bold"),
            relief="flat", padx=20, pady=5
        )
        self.start_btn.pack(side="left", padx=20, pady=12)

        self.stop_btn = tk.Button(
            top, text="DUNG", command=self._stop,
            bg='#e94560', fg='white', font=("Arial", 12, "bold"),
            relief="flat", padx=20, pady=5
        )
        self.stop_btn.pack(side="left", padx=5, pady=12)

        # Reset Workers button
        self.reset_btn = tk.Button(
            top, text="RESET", command=self._reset_workers,
            bg='#ff6348', fg='white', font=("Arial", 12, "bold"),
            relief="flat", padx=15, pady=5
        )
        self.reset_btn.pack(side="left", padx=5, pady=12)

        # Mode
        mode_frame = tk.Frame(top, bg='#0f3460')
        mode_frame.pack(side="left", padx=30)
        tk.Label(mode_frame, text="Mode:", bg='#0f3460', fg='white', font=("Arial", 10)).pack(side="left")
        self.mode_var = tk.StringVar(value="basic")
        tk.Radiobutton(mode_frame, text="Basic", variable=self.mode_var, value="basic",
                       bg='#0f3460', fg='white', selectcolor='#1a1a2e', font=("Arial", 10)).pack(side="left")
        tk.Radiobutton(mode_frame, text="Full", variable=self.mode_var, value="full",
                       bg='#0f3460', fg='white', selectcolor='#1a1a2e', font=("Arial", 10)).pack(side="left")

        # Show/Hide Chrome button (CMD already hidden, shown in log viewer)
        self.windows_visible = False  # Start hidden
        self.toggle_btn = tk.Button(top, text="HIEN CHROME", command=self._toggle_windows,
                  bg='#6c5ce7', fg='white', font=("Arial", 9, "bold"), relief="flat", padx=10)
        self.toggle_btn.pack(side="left", padx=10)

        # IPv6 toggle
        self.ipv6_var = tk.BooleanVar(value=self._get_ipv6_setting())
        self.ipv6_check = tk.Checkbutton(top, text="IPv6", variable=self.ipv6_var,
                  command=self._toggle_ipv6, bg='#0f3460', fg='#00ff88',
                  selectcolor='#1a1a2e', font=("Arial", 10, "bold"),
                  activebackground='#0f3460', activeforeground='#00ff88')
        self.ipv6_check.pack(side="left", padx=15)

        # Settings button - cau hinh
        self.settings_btn = tk.Button(top, text="SETTINGS", command=self._open_settings,
                  bg='#ff9f43', fg='white', font=("Arial", 9, "bold"), relief="flat", padx=10)
        self.settings_btn.pack(side="left", padx=5)

        # Update button - cap nhat tu GitHub
        self.update_btn = tk.Button(top, text="UPDATE", command=self._run_update,
                  bg='#0984e3', fg='white', font=("Arial", 9, "bold"), relief="flat", padx=10)
        self.update_btn.pack(side="left", padx=5)

        # Setup VM button - cai dat SMB share cho may ao
        self.setup_vm_btn = tk.Button(top, text="SETUP VM", command=self._setup_vm,
                  bg='#a29bfe', fg='white', font=("Arial", 9, "bold"), relief="flat", padx=10)
        self.setup_vm_btn.pack(side="left", padx=5)

        # Git version info
        git_info = self._get_git_version()
        tk.Label(top, text=git_info, bg='#0f3460', fg='#888',
                 font=("Consolas", 8)).pack(side="right", padx=10)

        # Status
        self.status_var = tk.StringVar(value="San sang")
        tk.Label(top, textvariable=self.status_var, bg='#0f3460', fg='#00d9ff',
                 font=("Consolas", 11, "bold")).pack(side="right", padx=20)

        # === MAIN CONTENT: Left (Workers + Projects) | Right (LOG) ===
        main = tk.PanedWindow(self, orient=tk.HORIZONTAL, bg='#1a1a2e', sashwidth=5)
        main.pack(fill="both", expand=True, padx=5, pady=5)

        # Left panel
        left = tk.Frame(main, bg='#1a1a2e')
        main.add(left, width=450)

        # === WORKERS - Compact inline ===
        workers = tk.Frame(left, bg='#16213e', padx=5, pady=5)
        workers.pack(fill="x", padx=5, pady=5)

        self.worker_vars = {}
        self.worker_labels = {}
        self.worker_rows = {}  # Store row frames

        # Tao 3 workers tren 1 dong
        for wid, name in [("excel", "EXCEL"), ("chrome_1", "CHR1"), ("chrome_2", "CHR2")]:
            frame = tk.Frame(workers, bg='#0f3460', padx=8, pady=4)
            frame.pack(side="left", padx=3, pady=2)

            self.worker_vars[f"{wid}_project"] = tk.StringVar(value="")
            self.worker_vars[f"{wid}_status"] = tk.StringVar(value="")

            # Name label
            name_lbl = tk.Label(frame, text=name, bg='#0f3460', fg='#888',
                     font=("Consolas", 9, "bold"))
            name_lbl.pack(side="left")

            # Project label - chi hien khi co project
            proj_lbl = tk.Label(frame, textvariable=self.worker_vars[f"{wid}_project"], bg='#0f3460', fg='#00d9ff',
                     font=("Consolas", 10, "bold"))
            proj_lbl.pack(side="left", padx=3)

            # Status label - chi hien khi dang chay
            status_lbl = tk.Label(frame, textvariable=self.worker_vars[f"{wid}_status"], bg='#0f3460', fg='#ffd93d',
                     font=("Consolas", 9))
            status_lbl.pack(side="left", padx=2)

            self.worker_labels[wid] = {'name': name_lbl, 'project': proj_lbl, 'status': status_lbl}
            self.worker_rows[wid] = frame

        # === WORKER LOGS VIEWER ===
        logs_frame = tk.LabelFrame(left, text=" WORKER LOGS ", bg='#16213e', fg='#00d9ff',
                                   font=("Arial", 10, "bold"), padx=5, pady=5)
        logs_frame.pack(fill="both", expand=True, padx=5, pady=(0, 5))

        # Tab selector
        tab_frame = tk.Frame(logs_frame, bg='#0f3460')
        tab_frame.pack(fill="x", pady=(0, 5))

        self.log_tab_var = tk.StringVar(value="excel")
        self.log_tab_buttons = {}

        for wid, label in [("excel", "EXCEL"), ("chrome_1", "CHROME 1"), ("chrome_2", "CHROME 2")]:
            btn = tk.Radiobutton(tab_frame, text=label, variable=self.log_tab_var, value=wid,
                                command=lambda w=wid: self._switch_log_tab(w),
                                bg='#0f3460', fg='white', selectcolor='#1a1a2e',
                                font=("Arial", 9, "bold"), indicatoron=False,
                                padx=10, pady=3)
            btn.pack(side="left", padx=2)
            self.log_tab_buttons[wid] = btn

        # Log text area
        log_text_frame = tk.Frame(logs_frame, bg='#1a1a2e')
        log_text_frame.pack(fill="both", expand=True)

        log_scrollbar = tk.Scrollbar(log_text_frame)
        log_scrollbar.pack(side="right", fill="y")

        self.log_text = tk.Text(log_text_frame, bg='#0a0a0a', fg='#00ff88',
                               font=("Consolas", 8), wrap="none",
                               yscrollcommand=log_scrollbar.set,
                               height=15)
        self.log_text.pack(fill="both", expand=True)
        log_scrollbar.config(command=self.log_text.yview)

        # Horizontal scrollbar
        log_scrollbar_x = tk.Scrollbar(logs_frame, orient="horizontal", command=self.log_text.xview)
        log_scrollbar_x.pack(fill="x")
        self.log_text.config(xscrollcommand=log_scrollbar_x.set)

        # === PROJECTS ===
        projects = tk.LabelFrame(left, text=" PROJECTS (click de xem) ", bg='#16213e', fg='white',
                                 font=("Arial", 10, "bold"), padx=5, pady=5)
        projects.pack(fill="both", expand=True, padx=5, pady=(0, 5))

        # Header
        header = tk.Frame(projects, bg='#0f3460')
        header.pack(fill="x")
        for txt, w in [("Code", 10), ("Excel", 6), ("NV", 5), ("Anh", 7), ("Video", 7), ("Status", 8)]:
            tk.Label(header, text=txt, width=w, bg='#0f3460', fg='white',
                     font=("Arial", 9, "bold")).pack(side="left", padx=2, pady=5)

        # List
        canvas = tk.Canvas(projects, bg='#16213e', highlightthickness=0)
        scrollbar = ttk.Scrollbar(projects, orient="vertical", command=canvas.yview)
        self.projects_frame = tk.Frame(canvas, bg='#16213e')

        canvas.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side="right", fill="y")
        canvas.pack(side="left", fill="both", expand=True)
        canvas.create_window((0, 0), window=self.projects_frame, anchor="nw")
        self.projects_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))

        self.project_rows: Dict[str, dict] = {}

        # Right panel - CHI TIET PROJECT (hien khi click project) - width lon hon de hien prompt
        right = tk.Frame(main, bg='#1a1a2e')
        main.add(right, width=750)

        # Header - ten project dang xem
        self.detail_header = tk.Frame(right, bg='#0f3460', height=40)
        self.detail_header.pack(fill="x")
        self.detail_header.pack_propagate(False)

        self.detail_title_var = tk.StringVar(value="Click vao project de xem chi tiet")
        tk.Label(self.detail_header, textvariable=self.detail_title_var, bg='#0f3460', fg='#00ff88',
                 font=("Arial", 12, "bold")).pack(side="left", padx=15, pady=8)

        # === EXCEL STEPS ===
        self.excel_frame = tk.LabelFrame(right, text=" EXCEL (7 STEPS) ", bg='#16213e', fg='#00ff88',
                                    font=("Arial", 10, "bold"), padx=10, pady=5)
        self.excel_frame.pack(fill="x", padx=5, pady=5)

        self.excel_step_labels = []
        step_names = ["1.Story", "2.Segments", "3.Characters", "4.Locations", "5.Director", "6.Plans", "7.Prompts"]
        for name in step_names:
            lbl = tk.Label(self.excel_frame, text=name, bg='#16213e', fg='#666',
                          font=("Consolas", 9), padx=8)
            lbl.pack(side="left")
            self.excel_step_labels.append(lbl)

        # === CHARACTERS - dang danh sach tu Excel ===
        ref_frame = tk.LabelFrame(right, text=" CHARACTERS (tu Excel) ", bg='#16213e', fg='#ff6b6b',
                                  font=("Arial", 9, "bold"), padx=5, pady=3)
        ref_frame.pack(fill="x", padx=5, pady=3, ipady=2)

        # Header
        ref_header = tk.Frame(ref_frame, bg='#0f3460')
        ref_header.pack(fill="x")
        tk.Label(ref_header, text="ID", width=6, bg='#0f3460', fg='white', font=("Consolas", 9, "bold")).pack(side="left", padx=2)
        tk.Label(ref_header, text="Thumb", width=6, bg='#0f3460', fg='white', font=("Consolas", 9, "bold")).pack(side="left", padx=2)
        tk.Label(ref_header, text="Character Description", width=50, bg='#0f3460', fg='white', font=("Consolas", 9, "bold")).pack(side="left", padx=2)

        # Scrollable list
        self.ref_canvas = tk.Canvas(ref_frame, bg='#1a1a2e', highlightthickness=0, height=150)
        ref_scrollbar = ttk.Scrollbar(ref_frame, orient="vertical", command=self.ref_canvas.yview)
        self.ref_images_frame = tk.Frame(self.ref_canvas, bg='#1a1a2e')

        self.ref_images_frame.bind("<Configure>",
            lambda e: self.ref_canvas.configure(scrollregion=self.ref_canvas.bbox("all")))
        self.ref_canvas.create_window((0, 0), window=self.ref_images_frame, anchor="nw")
        self.ref_canvas.configure(yscrollcommand=ref_scrollbar.set)

        ref_scrollbar.pack(side="right", fill="y")
        self.ref_canvas.pack(fill="both", expand=True)

        self.ref_photo_refs = []

        # === SCENES LIST ===
        scenes_frame = tk.LabelFrame(right, text=" SCENES ", bg='#16213e', fg='#ffd93d',
                                     font=("Arial", 10, "bold"), padx=5, pady=5)
        scenes_frame.pack(fill="both", expand=True, padx=5, pady=5)

        # Header row - font to hon
        header_row = tk.Frame(scenes_frame, bg='#0f3460')
        header_row.pack(fill="x")
        tk.Label(header_row, text="ID", width=4, bg='#0f3460', fg='white', font=("Consolas", 10, "bold")).pack(side="left", padx=2)
        tk.Label(header_row, text="Thumb", width=6, bg='#0f3460', fg='white', font=("Consolas", 10, "bold")).pack(side="left", padx=2)
        tk.Label(header_row, text="SRT Time", width=18, bg='#0f3460', fg='white', font=("Consolas", 10, "bold")).pack(side="left", padx=2)
        tk.Label(header_row, text="Prompt", width=45, bg='#0f3460', fg='white', font=("Consolas", 10, "bold")).pack(side="left", padx=2)
        tk.Label(header_row, text="Img", width=5, bg='#0f3460', fg='white', font=("Consolas", 10, "bold")).pack(side="left", padx=2)
        tk.Label(header_row, text="Vid", width=5, bg='#0f3460', fg='white', font=("Consolas", 10, "bold")).pack(side="left", padx=2)

        # Scrollable scene list - them horizontal scrollbar
        scene_scroll_frame = tk.Frame(scenes_frame, bg='#1a1a2e')
        scene_scroll_frame.pack(fill="both", expand=True)

        self.scene_canvas = tk.Canvas(scene_scroll_frame, bg='#1a1a2e', highlightthickness=0)
        scene_scrollbar_y = ttk.Scrollbar(scene_scroll_frame, orient="vertical", command=self.scene_canvas.yview)
        scene_scrollbar_x = ttk.Scrollbar(scenes_frame, orient="horizontal", command=self.scene_canvas.xview)
        self.scenes_list_frame = tk.Frame(self.scene_canvas, bg='#1a1a2e')

        self.scenes_list_frame.bind("<Configure>",
            lambda e: self.scene_canvas.configure(scrollregion=self.scene_canvas.bbox("all")))
        self.scene_canvas.create_window((0, 0), window=self.scenes_list_frame, anchor="nw")
        self.scene_canvas.configure(yscrollcommand=scene_scrollbar_y.set, xscrollcommand=scene_scrollbar_x.set)

        scene_scrollbar_y.pack(side="right", fill="y")
        self.scene_canvas.pack(fill="both", expand=True)
        scene_scrollbar_x.pack(fill="x")

        # Enable mouse wheel scrolling
        def _on_mousewheel(event):
            self.scene_canvas.yview_scroll(int(-1*(event.delta/120)), "units")
        self.scene_canvas.bind_all("<MouseWheel>", _on_mousewheel)

        # === STATUS BAR - Hien thi dang lam gi ===
        status_bar = tk.Frame(self, bg='#0f3460', height=35)
        status_bar.pack(fill="x", side="bottom")
        status_bar.pack_propagate(False)

        tk.Label(status_bar, text="DANG LAM:", bg='#0f3460', fg='#888',
                 font=("Consolas", 10)).pack(side="left", padx=10)

        self.current_action_var = tk.StringVar(value="Cho bat dau...")
        tk.Label(status_bar, textvariable=self.current_action_var, bg='#0f3460', fg='#ffd93d',
                 font=("Consolas", 11, "bold")).pack(side="left", padx=5)

    def _select_project(self, code: str):
        """Chon project de hien chi tiet."""
        self.selected_project = code
        self.detail_title_var.set(f"Project: {code}")
        self._load_project_detail(code)

    def _switch_log_tab(self, worker_id: str):
        """Switch log tab to show different worker."""
        self._update_worker_logs(worker_id)

    def _update_worker_logs(self, worker_id: str = None):
        """Update log display for current worker tab."""
        if worker_id is None:
            worker_id = self.log_tab_var.get()

        if not self.manager:
            self.log_text.delete('1.0', tk.END)
            self.log_text.insert('1.0', f"[{worker_id}] Manager not started\n")
            return

        # Read log file
        log_file = TOOL_DIR / ".agent" / "logs" / f"{worker_id}.log"
        if not log_file.exists():
            self.log_text.delete('1.0', tk.END)
            self.log_text.insert('1.0', f"[{worker_id}] No log file yet\n")
            return

        try:
            with open(log_file, 'r', encoding='utf-8', errors='ignore') as f:
                # Read last 500 lines
                lines = f.readlines()
                recent_lines = lines[-500:] if len(lines) > 500 else lines
                log_content = ''.join(recent_lines)

            self.log_text.delete('1.0', tk.END)
            self.log_text.insert('1.0', log_content)
            self.log_text.see(tk.END)  # Auto-scroll to bottom

        except Exception as e:
            self.log_text.delete('1.0', tk.END)
            self.log_text.insert('1.0', f"[{worker_id}] Error reading log: {e}\n")

    def _show_excel_detail(self, code: str):
        """Show popup with detailed Excel status."""
        if not self.manager:
            return

        status = self.manager.quality_checker.get_project_status(code)
        if not status:
            return

        popup = tk.Toplevel(self)
        popup.title(f"Excel Detail - {code}")
        popup.geometry("450x380")
        popup.configure(bg='#1a1a2e')
        popup.transient(self)
        popup.grab_set()

        # Header
        tk.Label(popup, text=f"EXCEL STATUS: {code}", bg='#0f3460', fg='#00ff88',
                 font=("Arial", 12, "bold"), pady=10).pack(fill="x")

        # Content frame
        content = tk.Frame(popup, bg='#1a1a2e', padx=20, pady=10)
        content.pack(fill="both", expand=True)

        # Get values from ProjectStatus
        excel_status = getattr(status, 'excel_status', 'none')
        total_scenes = getattr(status, 'total_scenes', 0)
        img_prompts = getattr(status, 'img_prompts_count', 0)
        video_prompts = getattr(status, 'video_prompts_count', 0)
        fallback_prompts = getattr(status, 'fallback_prompts', 0)
        missing_img = getattr(status, 'missing_img_prompts', [])
        characters_count = getattr(status, 'characters_count', 0)
        characters_with_ref = getattr(status, 'characters_with_ref', 0)

        # Status rows
        rows_data = [
            ("SRT file", "OK" if getattr(status, 'srt_exists', False) else "--"),
            ("Excel file", "OK" if getattr(status, 'excel_exists', False) else "--"),
            ("Total Scenes", str(total_scenes)),
            ("Characters", f"{characters_count}" if characters_count else "--"),
            ("Char with NV", f"{characters_with_ref}/{characters_count}" if characters_count else "--"),
            ("Img Prompts", f"{img_prompts}/{total_scenes}" if total_scenes else "--"),
            ("Video Prompts", f"{video_prompts}/{total_scenes}" if total_scenes else "--"),
            ("Fallback", str(fallback_prompts) if fallback_prompts > 0 else "0"),
        ]

        for name, value in rows_data:
            row = tk.Frame(content, bg='#1a1a2e')
            row.pack(fill="x", pady=3)

            # Determine color based on value
            if value == "OK" or (value.endswith(f"/{total_scenes}") and value.startswith(str(total_scenes))):
                color = '#00ff88'
            elif value == "--" or value == "0":
                color = '#666'
            elif "/" in value:  # partial like "5/10"
                color = '#00d9ff'
            else:
                color = 'white'

            tk.Label(row, text=name, width=15, bg='#1a1a2e', fg='white',
                     font=("Consolas", 10), anchor="w").pack(side="left")
            tk.Label(row, text=value, width=12, bg='#1a1a2e', fg=color,
                     font=("Consolas", 10, "bold")).pack(side="left")

        # Summary
        if excel_status == "complete":
            summary = "HOAN THANH" + (" (co Fallback)" if fallback_prompts > 0 else "")
            summary_color = '#00ff88'
        elif excel_status == "partial":
            pct = int(img_prompts * 100 / total_scenes) if total_scenes > 0 else 0
            summary = f"Dang tao: {pct}%"
            summary_color = '#00d9ff'
        elif excel_status == "fallback":
            summary = "Co Fallback - can API"
            summary_color = '#ffd93d'
        else:
            summary = "Chua co Excel"
            summary_color = '#ff6b6b'

        tk.Label(content, text=summary, bg='#1a1a2e', fg=summary_color,
                 font=("Arial", 11, "bold"), pady=10).pack()

        # Close button
        tk.Button(popup, text="Dong", command=popup.destroy, bg='#0f3460', fg='white',
                  font=("Arial", 10), padx=20).pack(pady=10)

    def _show_nv_detail(self, code: str):
        """Show popup with detailed NV (reference images) status."""
        if not self.manager:
            return

        status = self.manager.quality_checker.get_project_status(code)
        project_dir = TOOL_DIR / "PROJECTS" / code
        nv_dir = project_dir / "nv"

        popup = tk.Toplevel(self)
        popup.title(f"NV Detail - {code}")
        popup.geometry("500x400")
        popup.configure(bg='#1a1a2e')
        popup.transient(self)
        popup.grab_set()

        # Header
        tk.Label(popup, text=f"ANH THAM CHIEU (NV): {code}", bg='#0f3460', fg='#ff6b6b',
                 font=("Arial", 12, "bold"), pady=10).pack(fill="x")

        # Content frame with scrollbar
        canvas = tk.Canvas(popup, bg='#1a1a2e', highlightthickness=0)
        scrollbar = ttk.Scrollbar(popup, orient="vertical", command=canvas.yview)
        content = tk.Frame(canvas, bg='#1a1a2e')

        canvas.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side="right", fill="y")
        canvas.pack(fill="both", expand=True, padx=10, pady=5)
        canvas.create_window((0, 0), window=content, anchor="nw")
        content.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))

        # Get characters from Excel
        excel_path = project_dir / f"{code}_prompts.xlsx"
        characters = []
        if excel_path.exists():
            try:
                from openpyxl import load_workbook
                wb = load_workbook(excel_path, read_only=True)
                if "Characters" in wb.sheetnames:
                    ws = wb["Characters"]
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        if row and row[0]:
                            char_id = str(row[0]).strip()
                            if char_id:
                                characters.append(char_id)
                wb.close()
            except:
                pass

        if not characters:
            tk.Label(content, text="Khong co Characters trong Excel", bg='#1a1a2e', fg='#666',
                     font=("Consolas", 10), pady=20).pack()
        else:
            # Header row
            header = tk.Frame(content, bg='#0f3460')
            header.pack(fill="x", pady=(0, 5))
            tk.Label(header, text="ID", width=15, bg='#0f3460', fg='white',
                     font=("Consolas", 9, "bold")).pack(side="left", padx=5)
            tk.Label(header, text="Status", width=10, bg='#0f3460', fg='white',
                     font=("Consolas", 9, "bold")).pack(side="left", padx=5)

            # List each character
            done_count = 0
            for i, char_id in enumerate(characters):
                bg = '#1a1a2e' if i % 2 == 0 else '#16213e'
                row = tk.Frame(content, bg=bg)
                row.pack(fill="x", pady=1)

                # Check if image exists
                has_image = False
                for ext in ['.png', '.jpg', '.jpeg', '.webp']:
                    if (nv_dir / f"{char_id}{ext}").exists():
                        has_image = True
                        done_count += 1
                        break

                tk.Label(row, text=char_id, width=15, bg=bg, fg='white',
                         font=("Consolas", 10), anchor="w").pack(side="left", padx=5)

                status_text = "OK" if has_image else "Thieu"
                status_color = '#00ff88' if has_image else '#ff6b6b'
                tk.Label(row, text=status_text, width=10, bg=bg, fg=status_color,
                         font=("Consolas", 10, "bold")).pack(side="left", padx=5)

            # Summary
            total = len(characters)
            summary_frame = tk.Frame(content, bg='#1a1a2e')
            summary_frame.pack(fill="x", pady=10)

            if done_count >= total:
                summary = f"HOAN THANH: {done_count}/{total}"
                summary_color = '#00ff88'
            else:
                summary = f"CON THIEU: {total - done_count}/{total}"
                summary_color = '#ff6b6b'

            tk.Label(summary_frame, text=summary, bg='#1a1a2e', fg=summary_color,
                     font=("Arial", 11, "bold")).pack()

        # Close button
        tk.Button(popup, text="Dong", command=popup.destroy, bg='#0f3460', fg='white',
                  font=("Arial", 10), padx=20).pack(pady=10)

    def _show_prompt_popup(self, scene_id, prompt_text: str):
        """Show popup with full prompt text and copy button."""
        popup = tk.Toplevel(self)
        popup.title(f"Prompt - Scene {scene_id}")
        popup.geometry("600x400")
        popup.configure(bg='#1a1a2e')
        popup.transient(self)
        popup.grab_set()

        # Header
        tk.Label(popup, text=f"PROMPT - Scene {scene_id}", bg='#0f3460', fg='#ffd93d',
                 font=("Arial", 12, "bold"), pady=10).pack(fill="x")

        # Text widget with scrollbar
        text_frame = tk.Frame(popup, bg='#1a1a2e')
        text_frame.pack(fill="both", expand=True, padx=10, pady=10)

        scrollbar = ttk.Scrollbar(text_frame)
        scrollbar.pack(side="right", fill="y")

        text_widget = tk.Text(text_frame, wrap="word", bg='#16213e', fg='white',
                              font=("Consolas", 10), yscrollcommand=scrollbar.set,
                              padx=10, pady=10)
        text_widget.pack(fill="both", expand=True)
        text_widget.insert("1.0", prompt_text)
        text_widget.config(state="disabled")

        scrollbar.config(command=text_widget.yview)

        # Button frame
        btn_frame = tk.Frame(popup, bg='#1a1a2e')
        btn_frame.pack(pady=10)

        def copy_to_clipboard():
            self.clipboard_clear()
            self.clipboard_append(prompt_text)
            copy_btn.config(text="Da copy!")
            popup.after(1500, lambda: copy_btn.config(text="Copy"))

        copy_btn = tk.Button(btn_frame, text="Copy", command=copy_to_clipboard,
                             bg='#00d9ff', fg='black', font=("Arial", 10), padx=20)
        copy_btn.pack(side="left", padx=5)

        tk.Button(btn_frame, text="Dong", command=popup.destroy,
                  bg='#0f3460', fg='white', font=("Arial", 10), padx=20).pack(side="left", padx=5)

    def _load_project_detail(self, code: str):
        """Load chi tiet project vao panel phai."""
        project_dir = TOOL_DIR / "PROJECTS" / code
        if not project_dir.exists():
            return

        # Load reference images
        self._load_reference_images(code)

        # Load scenes from Excel
        self._load_scenes_list(code)

    def _load_reference_images(self, project_code: str):
        """Load characters tu Excel - hien thi ID, ten file va character lock."""
        for widget in self.ref_images_frame.winfo_children():
            widget.destroy()
        self.ref_photo_refs = []

        project_dir = TOOL_DIR / "PROJECTS" / project_code
        excel_path = project_dir / f"{project_code}_prompts.xlsx"
        nv_dir = project_dir / "nv"

        if not excel_path.exists():
            tk.Label(self.ref_images_frame, text="Chua co Excel", bg='#1a1a2e', fg='#666',
                     font=("Consolas", 10)).pack(pady=5)
            return

        try:
            from openpyxl import load_workbook
            wb = load_workbook(str(excel_path))

            # Tim sheet characters
            char_sheet = None
            for name in wb.sheetnames:
                if name.lower() == 'characters':
                    char_sheet = name
                    break

            if not char_sheet:
                tk.Label(self.ref_images_frame, text="Khong co sheet characters", bg='#1a1a2e', fg='#666',
                         font=("Consolas", 10)).pack(pady=5)
                wb.close()
                return

            ws = wb[char_sheet]
            headers = [cell.value for cell in ws[1]]

            # Tim index cua cac cot quan trong
            id_idx = headers.index('id') if 'id' in headers else 0
            name_idx = headers.index('name') if 'name' in headers else 1
            lock_idx = headers.index('character_lock') if 'character_lock' in headers else 2
            file_idx = headers.index('image_file') if 'image_file' in headers else -1

            characters = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[id_idx] is None:
                    continue
                characters.append({
                    'id': row[id_idx],
                    'name': row[name_idx] if name_idx < len(row) else '',
                    'lock': row[lock_idx] if lock_idx < len(row) else '',
                    'file': row[file_idx] if file_idx >= 0 and file_idx < len(row) else ''
                })

            wb.close()

            if not characters:
                tk.Label(self.ref_images_frame, text="Chua co characters", bg='#1a1a2e', fg='#666',
                         font=("Consolas", 10)).pack(pady=5)
                return

            # Hien thi header info
            tk.Label(self.ref_images_frame, text=f"Tim thay {len(characters)} characters",
                     bg='#1a1a2e', fg='#00ff88', font=("Consolas", 9, "bold")).pack(pady=2)

            ROW_WIDTH = 700

            for i, char in enumerate(characters[:20]):  # Max 20
                bg_color = '#1a1a2e' if i % 2 == 0 else '#16213e'
                row = tk.Frame(self.ref_images_frame, bg=bg_color, height=40, width=ROW_WIDTH)
                row.pack(fill="x", pady=1)
                row.pack_propagate(False)

                # ID
                tk.Label(row, text=str(char['id']), width=6, bg=bg_color, fg='#00d9ff',
                         font=("Consolas", 9, "bold")).pack(side="left", padx=3)

                # Thumbnail neu co file
                thumb_frame = tk.Frame(row, bg=bg_color, width=40, height=35)
                thumb_frame.pack(side="left", padx=2)
                thumb_frame.pack_propagate(False)

                img_file = char.get('file', '')
                if img_file and nv_dir.exists() and PIL_AVAILABLE:
                    img_path = nv_dir / img_file
                    if img_path.exists():
                        try:
                            img = Image.open(img_path)
                            img.thumbnail((35, 35))
                            photo = ImageTk.PhotoImage(img)
                            self.ref_photo_refs.append(photo)
                            thumb_lbl = tk.Label(thumb_frame, image=photo, bg=bg_color, cursor="hand2")
                            thumb_lbl.pack(expand=True)
                            thumb_lbl.bind("<Button-1>", lambda e, p=img_path: os.startfile(str(p)))
                        except:
                            tk.Label(thumb_frame, text="--", bg=bg_color, fg='#444',
                                     font=("Consolas", 8)).pack(expand=True)
                    else:
                        tk.Label(thumb_frame, text="--", bg=bg_color, fg='#444',
                                 font=("Consolas", 8)).pack(expand=True)
                else:
                    tk.Label(thumb_frame, text="--", bg=bg_color, fg='#444',
                             font=("Consolas", 8)).pack(expand=True)

                # Character lock (truncated) - noi dung chinh
                lock_text = char.get('lock', '') or ''
                lock_display = lock_text[:50] + "..." if len(lock_text) > 50 else lock_text or "--"
                tk.Label(row, text=lock_display, width=50, bg=bg_color, fg='#c8d6e5',
                         font=("Consolas", 8), anchor="w").pack(side="left", padx=3)

        except Exception as e:
            tk.Label(self.ref_images_frame, text=f"Loi: {str(e)[:30]}", bg='#1a1a2e', fg='#ff6b6b',
                     font=("Consolas", 9)).pack(pady=5)
            print(f"[ERROR] Load characters: {e}")

    def _load_scenes_list(self, project_code: str):
        """Load danh sach scenes tu Excel - voi thumbnail va SRT time."""
        for widget in self.scenes_list_frame.winfo_children():
            widget.destroy()
        self.scene_photo_refs = []

        project_dir = TOOL_DIR / "PROJECTS" / project_code
        excel_path = project_dir / f"{project_code}_prompts.xlsx"

        if not excel_path.exists():
            tk.Label(self.scenes_list_frame, text="Chua co Excel", bg='#1a1a2e', fg='#666',
                     font=("Consolas", 12)).pack(pady=20)
            return

        try:
            from modules.excel_manager import PromptWorkbook
            wb = PromptWorkbook(str(excel_path))
            wb.load_or_create()
            scenes = wb.get_scenes()

            # Hien thi so luong scenes
            if not scenes:
                tk.Label(self.scenes_list_frame, text="Khong tim thay scenes trong Excel",
                         bg='#1a1a2e', fg='#ffd93d', font=("Consolas", 10)).pack(pady=10)
                return

            info_label = tk.Label(self.scenes_list_frame, text=f"Tim thay {len(scenes)} scenes",
                                  bg='#1a1a2e', fg='#00ff88', font=("Consolas", 10, "bold"))
            info_label.pack(pady=3)

            img_dir = project_dir / "img"
            vid_dir = project_dir / "vid"

            # Tinh toan width cho row: ID(4) + Thumb(6) + SRT(18) + Prompt(45) + Img(5) + Vid(5) = 83 chars
            # Font Consolas 9 ~ 7px/char => 83*7 + padding = ~650px
            ROW_WIDTH = 750

            for i, scene in enumerate(scenes[:150]):  # Max 150 scenes
                bg = '#1a1a2e' if i % 2 == 0 else '#16213e'
                row = tk.Frame(self.scenes_list_frame, bg=bg, height=50, width=ROW_WIDTH)
                row.pack(fill="x", pady=1)
                row.pack_propagate(False)  # Fixed height and width

                # Scene ID - font to hon
                tk.Label(row, text=str(scene.scene_id), width=4, bg=bg, fg='#00d9ff',
                         font=("Consolas", 11, "bold")).pack(side="left", padx=3)

                # Thumbnail (45x45)
                img_path = find_scene_image(img_dir, scene.scene_id)
                thumb_frame = tk.Frame(row, bg=bg, width=50, height=45)
                thumb_frame.pack(side="left", padx=3)
                thumb_frame.pack_propagate(False)

                if img_path and PIL_AVAILABLE:
                    try:
                        img = Image.open(str(img_path))
                        img.thumbnail((45, 45))
                        photo = ImageTk.PhotoImage(img)
                        self.scene_photo_refs.append(photo)

                        thumb_lbl = tk.Label(thumb_frame, image=photo, bg=bg, cursor="hand2")
                        thumb_lbl.pack(expand=True)
                        thumb_lbl.bind("<Button-1>", lambda e, p=img_path: os.startfile(str(p)))
                    except:
                        tk.Label(thumb_frame, text="IMG", bg=bg, fg='#1dd1a1',
                                 font=("Consolas", 9)).pack(expand=True)
                else:
                    tk.Label(thumb_frame, text="--", bg=bg, fg='#444',
                             font=("Consolas", 9)).pack(expand=True)

                # SRT time - lay tu srt_start va srt_end
                srt_start = getattr(scene, 'srt_start', '') or ''
                srt_end = getattr(scene, 'srt_end', '') or ''

                # Debug 3 scenes dau
                if i < 3:
                    print(f"[DEBUG SRT] Scene {scene.scene_id}: start={repr(srt_start)}, end={repr(srt_end)}")

                # Convert to string neu la datetime.time hoac timedelta
                if hasattr(srt_start, 'strftime'):
                    srt_start = srt_start.strftime('%H:%M:%S')
                elif hasattr(srt_start, 'total_seconds'):  # timedelta
                    total = int(srt_start.total_seconds())
                    h, m, s = total // 3600, (total % 3600) // 60, total % 60
                    srt_start = f"{h:02d}:{m:02d}:{s:02d}"

                if hasattr(srt_end, 'strftime'):
                    srt_end = srt_end.strftime('%H:%M:%S')
                elif hasattr(srt_end, 'total_seconds'):  # timedelta
                    total = int(srt_end.total_seconds())
                    h, m, s = total // 3600, (total % 3600) // 60, total % 60
                    srt_end = f"{h:02d}:{m:02d}:{s:02d}"

                srt_start = str(srt_start) if srt_start else ''
                srt_end = str(srt_end) if srt_end else ''

                if srt_start and srt_end:
                    # Format: "00:01:23" thay vi "00:01:23,456"
                    start_short = srt_start.split(',')[0] if ',' in str(srt_start) else str(srt_start)
                    end_short = srt_end.split(',')[0] if ',' in str(srt_end) else str(srt_end)
                    srt_text = f"{start_short} - {end_short}"
                else:
                    srt_text = "--"

                tk.Label(row, text=srt_text, width=18, bg=bg, fg='#ffd93d',
                         font=("Consolas", 9)).pack(side="left", padx=3)

                # Prompt (truncated) - click de xem day du
                prompt_text = scene.img_prompt or ""
                prompt_display = prompt_text[:50] + "..." if len(prompt_text) > 50 else prompt_text or "--"
                prompt_label = tk.Label(row, text=prompt_display, width=45, bg=bg, fg='#c8d6e5',
                         font=("Consolas", 9), anchor="w", cursor="hand2" if prompt_text else "")
                prompt_label.pack(side="left", padx=3)
                if prompt_text:
                    prompt_label.bind("<Button-1>", lambda e, p=prompt_text, sid=scene.scene_id: self._show_prompt_popup(sid, p))

                # Image status
                if img_path:
                    img_status = tk.Label(row, text="OK", width=5, bg=bg, fg='#1dd1a1',
                             font=("Consolas", 10, "bold"), cursor="hand2")
                    img_status.pack(side="left", padx=2)
                    img_status.bind("<Button-1>", lambda e, p=img_path: os.startfile(str(p)))
                else:
                    tk.Label(row, text="--", width=5, bg=bg, fg='#666',
                             font=("Consolas", 10)).pack(side="left", padx=2)

                # Video status
                vid_path = find_scene_video(vid_dir, scene.scene_id)
                if vid_path:
                    vid_status = tk.Label(row, text="OK", width=5, bg=bg, fg='#1dd1a1',
                             font=("Consolas", 10, "bold"), cursor="hand2")
                    vid_status.pack(side="left", padx=2)
                    # Copy vid_path to avoid closure issue
                    vid_status.bind("<Button-1>", lambda e, p=str(vid_path): os.startfile(p))
                else:
                    tk.Label(row, text="--", width=5, bg=bg, fg='#666',
                             font=("Consolas", 10)).pack(side="left", padx=2)

        except Exception as e:
            tk.Label(self.scenes_list_frame, text=f"Loi: {e}", bg='#1a1a2e', fg='#ff6b6b',
                     font=("Consolas", 10)).pack(pady=20)

    def _update_detail_panel(self):
        """Update detail panel neu co project duoc chon."""
        if self.selected_project:
            # Update Excel steps status
            self._update_excel_steps(self.selected_project)

    def _update_excel_steps(self, project_code: str):
        """Update mau cua Excel step labels dua tren trang thai."""
        project_dir = TOOL_DIR / "PROJECTS" / project_code
        excel_path = project_dir / f"{project_code}_prompts.xlsx"

        if not excel_path.exists():
            # Reset all to gray
            for lbl in self.excel_step_labels:
                lbl.config(fg='#666')
            return

        try:
            from modules.excel_manager import PromptWorkbook
            wb = PromptWorkbook(str(excel_path))
            wb.load_or_create()

            # Get step status
            step_ids = ["step_1", "step_2", "step_3", "step_4", "step_5", "step_6", "step_7"]

            if hasattr(wb, 'get_all_step_status'):
                all_status = wb.get_all_step_status()
                status_dict = {}
                for s in all_status:
                    if isinstance(s, dict):
                        status_dict[s.get('step_id', '')] = s.get('status', 'PENDING')

                for i, step_id in enumerate(step_ids):
                    status = status_dict.get(step_id, 'PENDING')
                    if status == 'COMPLETED':
                        self.excel_step_labels[i].config(fg='#00ff88')  # Green
                    elif status == 'IN_PROGRESS':
                        self.excel_step_labels[i].config(fg='#ffd93d')  # Yellow
                    elif status == 'ERROR':
                        self.excel_step_labels[i].config(fg='#ff6b6b')  # Red
                    else:
                        self.excel_step_labels[i].config(fg='#666')  # Gray
            else:
                # Fallback - check if sheets exist
                for i, lbl in enumerate(self.excel_step_labels):
                    lbl.config(fg='#666')

        except Exception:
            for lbl in self.excel_step_labels:
                lbl.config(fg='#666')

    def _load_projects_on_startup(self):
        """Load projects ngay khi mo GUI (truoc khi bat dau chay)."""
        if not self.manager:
            return

        try:
            projects = self.manager.scan_projects()[:12]

            for i, code in enumerate(projects):
                if code not in self.project_rows:
                    self._create_row(code, i)

                # Load basic info
                status = self.manager.quality_checker.get_project_status(code)
                if status:
                    labels = self.project_rows[code]['labels']

                    # Excel status - show OK or %
                    excel_status = getattr(status, 'excel_status', '')
                    fallback_prompts = getattr(status, 'fallback_prompts', 0)

                    if excel_status == "complete":
                        text = "OK*" if fallback_prompts > 0 else "OK"
                        labels['excel'].config(text=text, fg='#00ff88')
                    elif excel_status == "partial":
                        img_prompts = getattr(status, 'img_prompts_count', 0)
                        total = getattr(status, 'total_scenes', 0)
                        if total > 0:
                            pct = int(img_prompts * 100 / total)
                            labels['excel'].config(text=f"{pct}%", fg='#00d9ff')
                        else:
                            labels['excel'].config(text="--", fg='#666')
                    elif excel_status == "fallback":
                        labels['excel'].config(text="FB", fg='#ffd93d')
                    else:
                        labels['excel'].config(text="--", fg='#666')

                    # NV status - reference images (characters with ref)
                    nv_done = getattr(status, 'characters_with_ref', 0)
                    nv_total = getattr(status, 'characters_count', 0)
                    if nv_total > 0:
                        if nv_done >= nv_total:
                            labels['nv'].config(text="OK", fg='#00ff88')
                        else:
                            labels['nv'].config(text=f"{nv_done}/{nv_total}", fg='#ff6b6b')
                    else:
                        labels['nv'].config(text="--", fg='#666')

                    # Images
                    img_done = getattr(status, 'images_done', 0)
                    img_total = getattr(status, 'total_scenes', 0)
                    if img_total > 0:
                        pct = int(img_done * 100 / img_total)
                        labels['images'].config(text=f"{img_done}/{img_total}", fg='#00d9ff' if pct > 0 else '#666')
                    else:
                        labels['images'].config(text="--", fg='#666')

                    # Videos
                    vid_done = getattr(status, 'videos_done', 0)
                    vid_total = getattr(status, 'total_scenes', 0)
                    if vid_total > 0:
                        labels['videos'].config(text=f"{vid_done}/{vid_total}", fg='#00d9ff' if vid_done > 0 else '#666')
                    else:
                        labels['videos'].config(text="--", fg='#666')

                    labels['status'].config(text="Ready", fg='#aaa')

            # Auto-select first project
            if projects and not self.selected_project:
                self._select_project(projects[0])

        except Exception as e:
            print(f"Error loading projects: {e}")

    def _start(self):
        if not self.manager:
            self.manager = VMManager(num_chrome_workers=2)

        self.manager.settings.excel_mode = self.mode_var.get()
        self.manager.settings.video_mode = self.mode_var.get()

        self.running = True
        self.status_var.set(f"Dang chay ({self.mode_var.get().upper()})")
        self.start_btn.config(bg='#666', state="disabled")

        # Log start
        if LOGGER_AVAILABLE:
            from modules.central_logger import log
            log("main", f"=== STARTED === Mode: {self.mode_var.get()}", "INFO")

        def run():
            self.manager.start_all(gui_mode=True)  # True = ẩn CMD windows
            threading.Thread(target=self.manager.orchestrate, daemon=True).start()

            # Auto-hide CMD windows after workers start
            time.sleep(5)  # Wait for workers to fully start
            self.after(0, lambda: self._auto_hide_windows())

            # Retry after another 3 seconds to make sure
            time.sleep(3)
            self.after(0, lambda: self._auto_hide_windows())

        threading.Thread(target=run, daemon=True).start()

    def _stop(self):
        if self.manager:
            self.running = False
            self.status_var.set("Dang dung...")

            # Log stop
            if LOGGER_AVAILABLE:
                from modules.central_logger import log
                log("main", "=== STOPPING ===", "INFO")

            # Kill all CMD and Chrome processes
            def stop_and_kill():
                self.manager.stop_all()
                self.manager.kill_all_chrome()
            threading.Thread(target=stop_and_kill, daemon=True).start()
            self.start_btn.config(bg='#00ff88', state="normal")

    def _reset_workers(self):
        """Reset workers: Kill all Chrome + CMD, then restart workers."""
        if not self.manager:
            return

        from tkinter import messagebox
        if not messagebox.askyesno("Reset Workers",
                                   "Reset tat ca workers?\n\n"
                                   "- Tat tat ca CMD (Excel + Chrome)\n"
                                   "- Kill Chrome processes\n"
                                   "- Khoi dong lai tat ca workers\n\n"
                                   "Tiep tuc?"):
            return

        self.status_var.set("Dang reset workers...")
        self.reset_btn.config(bg='#666', state="disabled")

        def do_reset():
            try:
                # Log
                if LOGGER_AVAILABLE:
                    from modules.central_logger import log
                    log("main", "=== RESET WORKERS ===", "INFO")

                # 1. Stop all workers
                self.manager.stop_all()
                time.sleep(2)

                # 2. Kill all Chrome + CMD
                self.manager.kill_all_chrome()
                time.sleep(2)

                # 3. Restart ALL workers (Excel + Chrome)
                for wid in self.manager.workers:
                    self.manager.start_worker(wid)
                    time.sleep(2)

                # Update status
                self.status_var.set("Reset xong!")
                messagebox.showinfo("Reset Complete", "Tat ca workers da duoc reset thanh cong!")

            except Exception as e:
                self.status_var.set(f"Loi reset: {str(e)[:40]}")
                messagebox.showerror("Reset Error", f"Loi: {e}")

            finally:
                self.reset_btn.config(bg='#ff6348', state="normal")

        threading.Thread(target=do_reset, daemon=True).start()

    def _toggle_windows(self):
        """Toggle Chrome windows only (CMD stays hidden, logs in GUI)."""
        if not self.manager:
            return

        if self.windows_visible:
            # Hide Chrome only
            self.manager.hide_chrome_windows()
            self.toggle_btn.config(text="HIEN CHROME", bg='#6c5ce7')
            self.windows_visible = False
        else:
            # Show Chrome and arrange nicely
            self._arrange_chrome_windows()
            self.toggle_btn.config(text="AN CHROME", bg='#00b894')
            self.windows_visible = True

    def _arrange_chrome_windows(self):
        """Arrange Chrome windows in good positions on screen."""
        if not self.manager:
            return

        import win32gui
        import win32con

        screen_width = win32gui.GetSystemMetrics(0)
        screen_height = win32gui.GetSystemMetrics(1)

        # Chrome window size - TO HƠN ĐỂ DỄ QUAN SÁT
        chrome_width = int(screen_width * 0.55)  # 55% màn hình
        chrome_height = int(screen_height * 0.45)  # 45% màn hình

        # Tối thiểu 1200x800
        chrome_width = max(chrome_width, 1200)
        chrome_height = max(chrome_height, 800)

        # Position Chrome 1: Top-right
        # Position Chrome 2: Bottom-right

        def enum_handler(hwnd, count):
            if win32gui.IsWindowVisible(hwnd):
                title = win32gui.GetWindowText(hwnd).lower()

                # Only Chrome windows (skip CMD)
                if "chrome" in title and "chrome.exe" not in title:
                    class_name = win32gui.GetClassName(hwnd)

                    # Check if it's actually Chrome browser window
                    if class_name.startswith("Chrome_WidgetWin"):
                        # Determine which Chrome (1 or 2)
                        if count[0] == 0:
                            # Chrome 1 - Top-right (phía trên bên phải)
                            x = screen_width - chrome_width - 10
                            y = 10
                            count[0] += 1
                        else:
                            # Chrome 2 - Bottom-right (phía dưới bên phải)
                            x = screen_width - chrome_width - 10
                            y = screen_height - chrome_height - 50
                            count[0] += 1

                        # Show window và set position/size
                        win32gui.ShowWindow(hwnd, win32con.SW_RESTORE)  # Restore nếu minimized
                        win32gui.SetWindowPos(hwnd, win32con.HWND_TOPMOST,
                                            x, y, chrome_width, chrome_height,
                                            win32con.SWP_SHOWWINDOW)
                        # Remove TOPMOST sau khi hiện
                        win32gui.SetWindowPos(hwnd, win32con.HWND_NOTOPMOST,
                                            x, y, chrome_width, chrome_height,
                                            win32con.SWP_SHOWWINDOW | win32con.SWP_NOMOVE | win32con.SWP_NOSIZE)

        try:
            count = [0]  # Chrome counter
            win32gui.EnumWindows(enum_handler, count)
        except Exception as e:
            print(f"Error arranging Chrome windows: {e}")

    def _auto_hide_windows(self):
        """Auto-hide Chrome and CMD windows when GUI starts."""
        if self.manager:
            try:
                print("[GUI] Auto-hiding CMD and Chrome windows...")
                self.manager.hide_cmd_windows()
                self.manager.hide_chrome_windows()
                self.toggle_btn.config(text="HIEN CHROME", bg='#6c5ce7')
                self.windows_visible = False
                print("[GUI] Windows hidden successfully")
            except Exception as e:
                print(f"[GUI] Error hiding windows: {e}")

    def _get_git_version(self) -> str:
        """Lay thong tin git commit cuoi cung."""
        try:
            import subprocess
            # Get commit hash (short)
            result = subprocess.run(['git', 'rev-parse', '--short', 'HEAD'],
                                  capture_output=True, text=True, cwd=str(TOOL_DIR), timeout=2)
            if result.returncode == 0:
                commit_hash = result.stdout.strip()

                # Get commit date
                result2 = subprocess.run(['git', 'log', '-1', '--format=%cd', '--date=format:%Y-%m-%d_%H:%M'],
                                       capture_output=True, text=True, cwd=str(TOOL_DIR), timeout=2)
                if result2.returncode == 0:
                    commit_date = result2.stdout.strip()
                    return f"v{commit_hash} | {commit_date}"
                return f"v{commit_hash}"
        except Exception as e:
            print(f"[GUI] Git version error: {e}")
        return "unknown"

    def _get_ipv6_setting(self) -> bool:
        """Doc IPv6 enabled tu settings.yaml. Mac dinh la True."""
        try:
            import yaml
            config_path = TOOL_DIR / "config" / "settings.yaml"
            if config_path.exists():
                with open(config_path, "r", encoding="utf-8") as f:
                    config = yaml.safe_load(f) or {}
                ipv6_cfg = config.get('ipv6_rotation', {})
                # Mac dinh la True neu khong co setting
                return ipv6_cfg.get('enabled', True)
        except:
            pass
        return True  # Mac dinh enabled

    def _run_update(self):
        """Cap nhat code tu GitHub - ho tro ca khi khong co Git."""
        import subprocess
        import urllib.request
        import zipfile
        import shutil

        GITHUB_ZIP_URL = "https://github.com/nguyenvantuong161978-dotcom/ve3-tool-simple/archive/refs/heads/main.zip"
        GITHUB_GIT_URL = "https://github.com/nguyenvantuong161978-dotcom/ve3-tool-simple.git"

        def do_update():
            self.update_btn.config(state="disabled", text="DANG CAP NHAT...", bg='#666')
            self.status_var.set("Dang kiem tra...")

            try:
                # Kiem tra git co san khong
                git_available = False
                try:
                    result = subprocess.run(
                        ["git", "--version"],
                        capture_output=True,
                        text=True,
                        timeout=10
                    )
                    git_available = (result.returncode == 0)
                except:
                    git_available = False

                if git_available:
                    # === DUNG GIT ===
                    self.status_var.set("Dang cap nhat qua Git...")

                    # Kiem tra remote origin
                    result = subprocess.run(
                        ["git", "remote", "get-url", "origin"],
                        cwd=str(TOOL_DIR),
                        capture_output=True,
                        text=True,
                        timeout=10
                    )

                    if result.returncode != 0:
                        subprocess.run(
                            ["git", "remote", "add", "origin", GITHUB_GIT_URL],
                            cwd=str(TOOL_DIR),
                            capture_output=True,
                            timeout=10
                        )

                    # Fetch va reset
                    cmds = [
                        ["git", "fetch", "origin", "main"],
                        ["git", "checkout", "main"],
                        ["git", "reset", "--hard", "origin/main"]
                    ]

                    for cmd in cmds:
                        result = subprocess.run(
                            cmd,
                            cwd=str(TOOL_DIR),
                            capture_output=True,
                            text=True,
                            timeout=120
                        )
                else:
                    # === KHONG CO GIT - TAI ZIP ===
                    self.status_var.set("Dang tai ZIP tu GitHub...")

                    # Tai file zip
                    zip_path = TOOL_DIR / "update_temp.zip"
                    extract_dir = TOOL_DIR / "update_temp"

                    # Download - bo qua SSL certificate
                    import ssl
                    import time
                    ssl_context = ssl.create_default_context()
                    ssl_context.check_hostname = False
                    ssl_context.verify_mode = ssl.CERT_NONE

                    # Cache-busting: them timestamp vao URL
                    cache_buster = f"?t={int(time.time())}"
                    download_url = GITHUB_ZIP_URL + cache_buster

                    with urllib.request.urlopen(download_url, context=ssl_context) as response:
                        with open(str(zip_path), 'wb') as out_file:
                            out_file.write(response.read())

                    self.status_var.set("Dang giai nen...")

                    # Giai nen
                    with zipfile.ZipFile(str(zip_path), 'r') as zip_ref:
                        zip_ref.extractall(str(extract_dir))

                    # Tim thu muc da giai nen (ve3-tool-simple-main)
                    extracted_folder = extract_dir / "ve3-tool-simple-main"

                    self.status_var.set("Dang cap nhat files...")

                    # Copy files moi (chi copy .py va modules/)
                    files_to_update = [
                        "vm_manager.py",
                        "vm_manager_gui.py",
                        "run_excel_api.py",
                        "run_worker.py",
                        "START.py",
                        "START.bat",
                        "requirements.txt",
                        "_run_chrome1.py",
                        "_run_chrome2.py",
                        "google_login.py",
                    ]

                    for f in files_to_update:
                        src = extracted_folder / f
                        dst = TOOL_DIR / f
                        if src.exists():
                            shutil.copy2(str(src), str(dst))

                    # Copy modules folder
                    src_modules = extracted_folder / "modules"
                    dst_modules = TOOL_DIR / "modules"
                    if src_modules.exists():
                        for py_file in src_modules.glob("*.py"):
                            shutil.copy2(str(py_file), str(dst_modules / py_file.name))

                    # Xoa temp files
                    if zip_path.exists():
                        zip_path.unlink()
                    if extract_dir.exists():
                        shutil.rmtree(str(extract_dir))

                # Lay version moi sau khi update
                new_version = self._get_git_version()

                self.status_var.set("Cap nhat xong! Khoi dong lai tool.")
                self.update_btn.config(text="XONG", bg='#00ff88')

                # Hoi co muon khoi dong lai khong
                from tkinter import messagebox
                update_msg = f"Da cap nhat xong!\n\nPhien ban moi: {new_version}\n\nBan co muon khoi dong lai tool?"
                if messagebox.askyesno("Cap nhat xong", update_msg):
                    import os
                    os.execv(sys.executable, [sys.executable] + sys.argv)

            except Exception as e:
                self.status_var.set(f"Loi: {str(e)[:40]}")
                self.update_btn.config(text="LOI", bg='#e94560')
                print(f"Update error: {e}")

                from tkinter import messagebox
                messagebox.showerror("Loi cap nhat", f"Loi: {e}\n\nThu tai thu cong:\n{GITHUB_ZIP_URL}")
            finally:
                self.after(3000, lambda: self.update_btn.config(state="normal", text="UPDATE", bg='#0984e3'))

        threading.Thread(target=do_update, daemon=True).start()

    def _open_settings(self):
        """Mo cua so Settings."""
        SettingsWindow(self)

    def _setup_vm(self):
        """Setup SMB share cho may ao - ket noi Z: den may chu."""
        import tkinter.messagebox as msgbox

        # Default settings - co the thay doi trong popup
        default_ip = "192.168.88.254"
        default_share = "D"
        default_user = "smbuser"
        default_pass = "159753"

        # Tao popup de nhap thong tin
        popup = tk.Toplevel(self)
        popup.title("Setup VM - Ket noi o mang")
        popup.geometry("400x300")
        popup.configure(bg='#1a1a2e')
        popup.transient(self)
        popup.grab_set()

        # Center popup
        popup.update_idletasks()
        x = self.winfo_x() + (self.winfo_width() - 400) // 2
        y = self.winfo_y() + (self.winfo_height() - 300) // 2
        popup.geometry(f"+{x}+{y}")

        # Title
        tk.Label(popup, text="KET NOI O MANG (SMB SHARE)",
                bg='#1a1a2e', fg='#00ff88', font=("Arial", 12, "bold")).pack(pady=10)

        tk.Label(popup, text="Map o Z: den may chu de copy du lieu on dinh hon",
                bg='#1a1a2e', fg='#aaa', font=("Arial", 9)).pack()

        # Form frame
        form = tk.Frame(popup, bg='#1a1a2e')
        form.pack(pady=15, padx=20, fill="x")

        # IP
        tk.Label(form, text="IP May chu:", bg='#1a1a2e', fg='white', font=("Arial", 10)).grid(row=0, column=0, sticky="e", pady=5)
        ip_var = tk.StringVar(value=default_ip)
        tk.Entry(form, textvariable=ip_var, width=25, font=("Arial", 10)).grid(row=0, column=1, pady=5, padx=5)

        # Share name
        tk.Label(form, text="Ten Share:", bg='#1a1a2e', fg='white', font=("Arial", 10)).grid(row=1, column=0, sticky="e", pady=5)
        share_var = tk.StringVar(value=default_share)
        tk.Entry(form, textvariable=share_var, width=25, font=("Arial", 10)).grid(row=1, column=1, pady=5, padx=5)

        # Username
        tk.Label(form, text="Username:", bg='#1a1a2e', fg='white', font=("Arial", 10)).grid(row=2, column=0, sticky="e", pady=5)
        user_var = tk.StringVar(value=default_user)
        tk.Entry(form, textvariable=user_var, width=25, font=("Arial", 10)).grid(row=2, column=1, pady=5, padx=5)

        # Password
        tk.Label(form, text="Password:", bg='#1a1a2e', fg='white', font=("Arial", 10)).grid(row=3, column=0, sticky="e", pady=5)
        pass_var = tk.StringVar(value=default_pass)
        tk.Entry(form, textvariable=pass_var, width=25, font=("Arial", 10), show="*").grid(row=3, column=1, pady=5, padx=5)

        # Status label
        status_var = tk.StringVar(value="")
        status_lbl = tk.Label(popup, textvariable=status_var, bg='#1a1a2e', fg='#ffd93d', font=("Arial", 9))
        status_lbl.pack(pady=5)

        def do_setup():
            """Thuc hien ket noi SMB."""
            ip = ip_var.get().strip()
            share = share_var.get().strip()
            user = user_var.get().strip()
            passwd = pass_var.get().strip()

            if not all([ip, share, user, passwd]):
                status_var.set("Vui long nhap day du thong tin!")
                status_lbl.config(fg='#e94560')
                return

            status_var.set("Dang ket noi...")
            status_lbl.config(fg='#ffd93d')
            popup.update()

            try:
                import subprocess

                # Xoa mapping cu neu co
                subprocess.run(['net', 'use', 'Z:', '/delete', '/y'],
                             capture_output=True, text=True)

                # Tao mapping moi
                cmd = ['net', 'use', 'Z:', f'\\\\{ip}\\{share}',
                       f'/user:{user}', passwd, '/persistent:yes']
                result = subprocess.run(cmd, capture_output=True, text=True)

                if result.returncode == 0:
                    # Kiem tra Z:\AUTO
                    from pathlib import Path
                    auto_path = Path("Z:\\AUTO")
                    if auto_path.exists():
                        status_var.set("THANH CONG! Z:\\AUTO da san sang")
                        status_lbl.config(fg='#00ff88')
                        msgbox.showinfo("Thanh cong",
                            f"Da ket noi Z: den \\\\{ip}\\{share}\n\n"
                            f"Z:\\AUTO da san sang su dung!")
                    else:
                        status_var.set(f"Da ket noi Z: nhung khong tim thay AUTO")
                        status_lbl.config(fg='#ffd93d')
                        msgbox.showwarning("Canh bao",
                            f"Da ket noi Z: den \\\\{ip}\\{share}\n\n"
                            f"Nhung khong tim thay thu muc AUTO.\n"
                            f"Kiem tra lai may chu.")
                else:
                    error_msg = result.stderr or result.stdout or "Loi khong xac dinh"
                    status_var.set(f"LOI: {error_msg[:50]}")
                    status_lbl.config(fg='#e94560')
                    msgbox.showerror("Loi ket noi",
                        f"Khong the ket noi den \\\\{ip}\\{share}\n\n"
                        f"Loi: {error_msg}\n\n"
                        f"Kiem tra:\n"
                        f"1. May chu {ip} co dang chay?\n"
                        f"2. Thu muc {share} da share chua?\n"
                        f"3. User {user} co quyen truy cap?")

            except Exception as e:
                status_var.set(f"LOI: {str(e)[:50]}")
                status_lbl.config(fg='#e94560')
                msgbox.showerror("Loi", f"Loi khi setup: {e}")

        # Buttons
        btn_frame = tk.Frame(popup, bg='#1a1a2e')
        btn_frame.pack(pady=15)

        tk.Button(btn_frame, text="KET NOI", command=do_setup,
                 bg='#00ff88', fg='#1a1a2e', font=("Arial", 10, "bold"),
                 relief="flat", padx=20, pady=5).pack(side="left", padx=10)

        tk.Button(btn_frame, text="DONG", command=popup.destroy,
                 bg='#e94560', fg='white', font=("Arial", 10, "bold"),
                 relief="flat", padx=20, pady=5).pack(side="left", padx=10)

    def _toggle_ipv6(self):
        """Bat/tat IPv6 va luu vao settings.yaml."""
        try:
            import yaml
            config_path = TOOL_DIR / "config" / "settings.yaml"

            # Doc config hien tai
            config = {}
            if config_path.exists():
                with open(config_path, "r", encoding="utf-8") as f:
                    config = yaml.safe_load(f) or {}

            # Update IPv6 setting
            if 'ipv6_rotation' not in config:
                config['ipv6_rotation'] = {}
            config['ipv6_rotation']['enabled'] = self.ipv6_var.get()

            # Luu lai
            with open(config_path, "w", encoding="utf-8") as f:
                yaml.safe_dump(config, f, default_flow_style=False, allow_unicode=True)

            # Update manager settings neu dang chay
            if self.manager and hasattr(self.manager, 'settings'):
                if hasattr(self.manager.settings, 'ipv6_rotation'):
                    self.manager.settings.ipv6_rotation['enabled'] = self.ipv6_var.get()

            status = "BAT" if self.ipv6_var.get() else "TAT"
            print(f"[GUI] IPv6: {status}")

        except Exception as e:
            print(f"[GUI] Error toggling IPv6: {e}")

    def _update_loop(self):
        self._update_workers()
        self._update_projects()
        self._update_detail_panel()
        self._update_worker_logs()  # Auto-update logs
        self.after(1000, self._update_loop)

    def _update_workers(self):
        if not self.manager:
            return

        actions = []  # Thu thap cac hanh dong dang lam

        for wid in ["excel", "chrome_1", "chrome_2"]:
            try:
                status = self.manager.get_worker_status(wid)
                if status:
                    proj = status.get('current_project', '') or ''
                    state = status.get('state', 'idle')
                    step = status.get('current_step', 0)
                    step_name = status.get('step_name', '')
                    task = status.get('current_task', '')

                    # Chi hien project khi dang lam viec
                    if proj and state.lower() in ['working', 'busy']:
                        self.worker_vars[f"{wid}_project"].set(proj)
                        self.worker_labels[wid]['project'].config(fg='#00d9ff')
                        self.worker_labels[wid]['name'].config(fg='#00ff88')  # Green khi active

                        # Status text
                        if wid == "excel" and step > 0:
                            self.worker_vars[f"{wid}_status"].set(f"S{step}/7")
                            actions.append(f"EXCEL {proj}: Step {step}/7 - {step_name}")
                        else:
                            self.worker_vars[f"{wid}_status"].set("")
                            # Chrome worker
                            if 'image' in task.lower():
                                actions.append(f"{wid.upper()}: {proj} tao anh")
                            elif 'video' in task.lower():
                                actions.append(f"{wid.upper()}: {proj} tao video")
                            else:
                                actions.append(f"{wid.upper()}: {proj}")
                    else:
                        # Idle - chi hien ten, bo project va status
                        self.worker_vars[f"{wid}_project"].set("")
                        self.worker_vars[f"{wid}_status"].set("")
                        self.worker_labels[wid]['name'].config(fg='#666')  # Gray khi idle
            except:
                pass

        # Update status bar
        if actions:
            self.current_action_var.set(" | ".join(actions[:3]))  # Max 3 actions
        else:
            if self.running:
                self.current_action_var.set("Dang tim project...")
            else:
                self.current_action_var.set("Cho bat dau...")

    def _update_projects(self):
        if not self.manager:
            return

        try:
            projects = self.manager.scan_projects()[:12]

            for i, code in enumerate(projects):
                if code not in self.project_rows:
                    self._create_row(code, i)

                status = self.manager.quality_checker.get_project_status(code)
                if status:
                    labels = self.project_rows[code]['labels']

                    # Excel - show OK or %
                    excel_status = getattr(status, 'excel_status', '')
                    fallback_prompts = getattr(status, 'fallback_prompts', 0)

                    if excel_status == "complete":
                        text = "OK*" if fallback_prompts > 0 else "OK"
                        labels['excel'].config(text=text, fg='#00ff88')
                    elif excel_status == "partial":
                        img_prompts = getattr(status, 'img_prompts_count', 0)
                        total = getattr(status, 'total_scenes', 0)
                        if total > 0:
                            pct = int(img_prompts * 100 / total)
                            labels['excel'].config(text=f"{pct}%", fg='#00d9ff')
                        else:
                            labels['excel'].config(text="--", fg='#666')
                    elif excel_status == "fallback":
                        labels['excel'].config(text="FB", fg='#ffd93d')
                    else:
                        labels['excel'].config(text="--", fg='#666')

                    # NV status - reference images (characters with ref)
                    nv_done = getattr(status, 'characters_with_ref', 0)
                    nv_total = getattr(status, 'characters_count', 0)
                    if nv_total > 0:
                        if nv_done >= nv_total:
                            labels['nv'].config(text="OK", fg='#00ff88')
                        else:
                            labels['nv'].config(text=f"{nv_done}/{nv_total}", fg='#ff6b6b')
                    else:
                        labels['nv'].config(text="--", fg='#666')

                    # Images
                    img_done = getattr(status, 'images_done', 0)
                    img_total = getattr(status, 'total_scenes', 0)
                    if img_total > 0:
                        if img_done >= img_total:
                            labels['images'].config(text=f"{img_done}/{img_total}", fg='#00ff88')
                        else:
                            labels['images'].config(text=f"{img_done}/{img_total}", fg='#00d9ff')
                    else:
                        labels['images'].config(text="--", fg='#666')

                    # Videos
                    vid_done = getattr(status, 'videos_done', 0)
                    videos_needed = getattr(status, 'videos_needed', [])
                    vid_total = len(videos_needed) if videos_needed else 0
                    if vid_total > 0:
                        if vid_done >= vid_total:
                            labels['videos'].config(text=f"{vid_done}/{vid_total}", fg='#00ff88')
                        else:
                            labels['videos'].config(text=f"{vid_done}/{vid_total}", fg='#00d9ff')
                    else:
                        labels['videos'].config(text="--", fg='#666')

                    # Status
                    next_action = getattr(status, 'next_action', '')
                    if next_action == 'create_excel':
                        labels['status'].config(text="Excel")
                    elif next_action == 'create_images':
                        labels['status'].config(text="Anh")
                    elif next_action == 'create_videos':
                        labels['status'].config(text="Video")
                    elif next_action == 'copy_to_visual':
                        labels['status'].config(text="XONG")
                    else:
                        labels['status'].config(text="--")
        except Exception as e:
            pass

    def _create_row(self, code: str, i: int):
        bg = '#1a1a2e' if i % 2 == 0 else '#16213e'

        row = tk.Frame(self.projects_frame, bg=bg, cursor="hand2")
        row.pack(fill="x", pady=1)
        row.bind("<Button-1>", lambda e, c=code: self._select_project(c))

        labels = {}

        labels['code'] = tk.Label(row, text=code, width=10, bg=bg, fg='white', font=("Consolas", 10, "bold"), anchor="w")
        labels['code'].pack(side="left", padx=2)
        labels['code'].bind("<Button-1>", lambda e, c=code: self._select_project(c))

        labels['excel'] = tk.Label(row, text="--", width=6, bg=bg, fg='#666', font=("Consolas", 10), cursor="hand2")
        labels['excel'].pack(side="left", padx=2)
        labels['excel'].bind("<Button-1>", lambda e, c=code: self._show_excel_detail(c))

        labels['nv'] = tk.Label(row, text="--", width=5, bg=bg, fg='#666', font=("Consolas", 10), cursor="hand2")
        labels['nv'].pack(side="left", padx=2)
        labels['nv'].bind("<Button-1>", lambda e, c=code: self._show_nv_detail(c))

        labels['images'] = tk.Label(row, text="--", width=7, bg=bg, fg='#666', font=("Consolas", 10))
        labels['images'].pack(side="left", padx=2)

        labels['videos'] = tk.Label(row, text="--", width=7, bg=bg, fg='#666', font=("Consolas", 10))
        labels['videos'].pack(side="left", padx=2)

        labels['status'] = tk.Label(row, text="--", width=8, bg=bg, fg='#aaa', font=("Consolas", 10))
        labels['status'].pack(side="left", padx=2)

        self.project_rows[code] = {'row': row, 'labels': labels, 'bg': bg}

    def _show_detail_popup(self, code: str):
        """Double click - mo popup chi tiet."""
        ProjectDetail(self, code)

    def destroy(self):
        # Kill all processes khi dong GUI
        if self.manager:
            try:
                self.manager.stop_all()
                self.manager.kill_all_chrome()
            except:
                pass

        # Remove callback when closing
        if LOGGER_AVAILABLE:
            try:
                remove_callback(self._on_new_log)
            except:
                pass
        super().destroy()


# ================================================================================
# MAIN
# ================================================================================

if __name__ == "__main__":
    if not VM_AVAILABLE:
        print("Loi: Khong tim thay vm_manager.py!")
        input("Nhan Enter de thoat...")
        sys.exit(1)

    app = SimpleGUI()
    app.mainloop()
