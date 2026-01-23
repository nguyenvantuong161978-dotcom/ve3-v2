"""
Check step status trong Excel
"""
import sys
import io
from pathlib import Path
from openpyxl import load_workbook

# Fix encoding
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

excel_file = Path(__file__).parent / "PROJECTS" / "AR8-0003" / "AR8-0003_prompts.xlsx"

print("=" * 80)
print("CHECK STEP STATUS - AR8-0003")
print("=" * 80)
print()

if not excel_file.exists():
    print("[ERROR] Excel file not found!")
    sys.exit(1)

# Load workbook
wb = load_workbook(str(excel_file), read_only=True)

print(f"Sheets: {wb.sheetnames}")
print()

# Check step_status sheet
if "step_status" in wb.sheetnames:
    ws = wb["step_status"]
    rows = list(ws.values)

    print("Step Status:")
    print("-" * 80)

    for row in rows:
        print(f"  {row}")

    print()

# Check scenes count
if "scenes" in wb.sheetnames:
    ws = wb["scenes"]
    rows = list(ws.values)
    print(f"Scenes sheet: {len(rows)-1} rows (excluding header)")
    print()

# Check director_plan count
if "director_plan" in wb.sheetnames:
    ws = wb["director_plan"]
    rows = list(ws.values)
    print(f"Director plan sheet: {len(rows)-1} rows")
    print()

wb.close()
print("=" * 80)
