"""
VE3 Tool - Excel Manager Module
===============================
Quản lý file Excel chứa prompts và thông tin nhân vật.
"""

import json
from pathlib import Path
from typing import List, Dict, Any, Optional, Union
from datetime import datetime

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from modules.utils import get_logger


# ============================================================================
# CONSTANTS
# ============================================================================

# Cột cho sheet Characters
CHARACTERS_COLUMNS = [
    "id",               # ID nhân vật (nvc, nvp1, nvp2, ...)
    "role",             # Vai trò (main/supporting)
    "name",             # Tên nhân vật trong truyện
    "english_prompt",   # Prompt tiếng Anh mô tả ngoại hình
    "vietnamese_prompt", # Prompt tiếng Việt (nếu cần)
    "character_lock",   # Mô tả cố định nhân vật (dùng cho scene prompts)
    "image_file",       # Tên file ảnh tham chiếu (nvc.png, nvp1.png, ...)
    "status",           # Trạng thái (pending/done/skip/error) - skip = trẻ em, không tạo ảnh
    "is_child",         # True nếu là trẻ vị thành niên (bỏ qua tạo ảnh)
    "media_id",         # Media ID từ Google Flow API (dùng cho reference)
]

# Cột cho sheet Director Plan (lưu kế hoạch từ SRT trước khi tạo prompts)
# Dùng để detect gaps và resume, có backup prompts với character/location refs
DIRECTOR_PLAN_COLUMNS = [
    "plan_id",          # ID theo thứ tự (1, 2, 3, ...)
    "srt_start",        # Thời gian bắt đầu (HH:MM:SS,mmm)
    "srt_end",          # Thời gian kết thúc (HH:MM:SS,mmm)
    "duration",         # Độ dài (giây)
    "srt_text",         # Nội dung text
    "characters_used",  # JSON list nhân vật trong scene (backup)
    "location_used",    # Location ID (backup)
    "reference_files",  # JSON list reference files (backup)
    "img_prompt",       # Backup prompt (dùng nếu director fail)
    "status",           # backup/pending/done
]

# Cột cho sheet Scenes
# QUAN TRONG: srt_start/srt_end la timestamp chinh (khong dung start_time/end_time nua)
SCENES_COLUMNS = [
    "scene_id",         # ID scene (1, 2, 3, ...)
    "srt_start",        # Thời gian bắt đầu (HH:MM:SS,mmm) - từ SRT hoặc do đạo diễn set
    "srt_end",          # Thời gian kết thúc (HH:MM:SS,mmm) - từ SRT hoặc do đạo diễn set
    "duration",         # Độ dài tính từ srt_start/srt_end (giây) - auto-calculate
    "planned_duration", # Thời lượng đạo diễn lên kế hoạch cho ảnh này (giây) - dùng khi edit video
    "srt_text",         # Nội dung text của scene
    "img_prompt",       # Prompt tạo ảnh (text)
    "prompt_json",      # JSON prompt đầy đủ (có seed, imageInputs) - dùng khi tạo ảnh
    "video_prompt",     # Prompt tạo video
    "img_path",         # Path đến ảnh đã tạo
    "video_path",       # Path đến video đã tạo
    "status_img",       # Trạng thái ảnh (pending/done/error)
    "status_vid",       # Trạng thái video (pending/done/error)
    # Thông tin reference
    "characters_used",  # JSON list nhân vật trong scene
    "location_used",    # Location ID
    "reference_files",  # JSON list reference files cho image generation
    "media_id",         # Media ID từ Google Flow API (dùng cho I2V - Image to Video)
]

# Cột cho sheet Backup Characters (nhân vật narrator cố định cho fallback)
BACKUP_CHARACTERS_COLUMNS = [
    "id",               # ID nhân vật (nvc)
    "name",             # Tên nhân vật
    "character_lock",   # Mô tả cố định nhân vật (tuổi, tóc, mặt...)
    "costume_lock",     # Mô tả cố định trang phục
    "image_file",       # File ảnh tham chiếu (nvc.png)
]

# Cột cho sheet Backup Locations (location kể chuyện cố định cho fallback)
BACKUP_LOCATIONS_COLUMNS = [
    "id",               # ID location (loc)
    "name",             # Tên location
    "location_lock",    # Mô tả cố định location (phòng, ánh sáng, đồ vật...)
    "image_file",       # File ảnh tham chiếu (loc.png)
]


# ============================================================================
# CHARACTER DATA CLASS
# ============================================================================

class Character:
    """Đại diện cho một nhân vật trong truyện."""

    def __init__(
        self,
        id: str,
        role: str = "supporting",
        name: str = "",
        english_prompt: str = "",
        vietnamese_prompt: str = "",
        character_lock: str = "",
        image_file: str = "",
        status: str = "pending",
        is_child: bool = False,
        media_id: str = ""
    ):
        self.id = id
        self.role = role
        self.name = name
        self.english_prompt = english_prompt
        self.vietnamese_prompt = vietnamese_prompt
        self.character_lock = character_lock
        self.image_file = image_file
        self.status = status
        self.is_child = is_child
        self.media_id = media_id

    def to_dict(self) -> Dict[str, Any]:
        """Chuyển đổi thành dictionary."""
        return {
            "id": self.id,
            "role": self.role,
            "name": self.name,
            "english_prompt": self.english_prompt,
            "vietnamese_prompt": self.vietnamese_prompt,
            "character_lock": self.character_lock,
            "image_file": self.image_file,
            "status": self.status,
            "is_child": self.is_child,
            "media_id": self.media_id,
        }

    @classmethod
    def from_dict(cls, data: Dict[str, Any]) -> "Character":
        """Tạo Character từ dictionary."""
        return cls(
            id=str(data.get("id", "")),
            role=str(data.get("role", "supporting")),
            name=str(data.get("name", "")),
            english_prompt=str(data.get("english_prompt", "")),
            vietnamese_prompt=str(data.get("vietnamese_prompt", "")),
            character_lock=str(data.get("character_lock", "")),
            image_file=str(data.get("image_file", "")),
            status=str(data.get("status", "pending")),
            is_child=bool(data.get("is_child", False)),
            media_id=str(data.get("media_id", "")),
        )


# ============================================================================
# LOCATION DATA CLASS
# ============================================================================

class Location:
    """Dai dien cho mot dia diem trong truyen."""

    def __init__(
        self,
        id: str,
        name: str = "",
        english_prompt: str = "",
        location_lock: str = "",
        lighting_default: str = "",
        image_file: str = "",
        status: str = "pending",
        media_id: str = ""
    ):
        self.id = id
        self.name = name
        self.english_prompt = english_prompt
        self.location_lock = location_lock
        self.lighting_default = lighting_default
        self.image_file = image_file
        self.status = status
        self.media_id = media_id

    def to_dict(self) -> Dict[str, Any]:
        """Chuyen doi thanh dictionary."""
        return {
            "id": self.id,
            "name": self.name,
            "english_prompt": self.english_prompt,
            "location_lock": self.location_lock,
            "lighting_default": self.lighting_default,
            "image_file": self.image_file,
            "status": self.status,
            "media_id": self.media_id,
        }

    @classmethod
    def from_dict(cls, data: Dict[str, Any]) -> "Location":
        """Tao Location tu dictionary."""
        return cls(
            id=str(data.get("id", "")),
            name=str(data.get("name", "")),
            english_prompt=str(data.get("english_prompt", "")),
            location_lock=str(data.get("location_lock", "")),
            lighting_default=str(data.get("lighting_default", "")),
            image_file=str(data.get("image_file", "")),
            status=str(data.get("status", "pending")),
            media_id=str(data.get("media_id", "")),
        )


# ============================================================================
# SCENE DATA CLASS
# ============================================================================

class Scene:
    """Đại diện cho một scene trong video."""

    def __init__(
        self,
        scene_id: int,
        srt_start: str = "",            # Timestamp bắt đầu (HH:MM:SS,mmm)
        srt_end: str = "",              # Timestamp kết thúc (HH:MM:SS,mmm)
        duration: float = 0.0,          # Độ dài tính từ timestamps (giây) - auto
        planned_duration: float = 0.0,  # Thời lượng đạo diễn lên kế hoạch (giây) - dùng khi edit
        srt_text: str = "",
        img_prompt: str = "",
        prompt_json: str = "",          # JSON prompt đầy đủ (seed, imageInputs)
        video_prompt: str = "",
        img_path: str = "",
        video_path: str = "",
        status_img: str = "pending",
        status_vid: str = "pending",
        # Reference info
        characters_used: str = "",      # JSON list of character IDs used
        location_used: str = "",        # Location ID used
        reference_files: str = "",      # JSON list of reference files
        media_id: str = "",             # Media ID từ Google Flow API (dùng cho I2V)
        # DEPRECATED - giữ để backward compatible, sẽ map sang srt_start/srt_end
        start_time: str = "",
        end_time: str = ""
    ):
        self.scene_id = scene_id
        # Ưu tiên srt_start/srt_end, fallback sang start_time/end_time (backward compatible)
        self.srt_start = str(srt_start) if srt_start else (str(start_time) if start_time else "")
        self.srt_end = str(srt_end) if srt_end else (str(end_time) if end_time else "")
        self.duration = duration
        self.planned_duration = planned_duration  # Thời lượng đạo diễn lên kế hoạch
        self.srt_text = srt_text
        self.img_prompt = img_prompt
        self.prompt_json = prompt_json
        self.video_prompt = video_prompt
        self.img_path = img_path
        self.video_path = video_path
        self.status_img = status_img
        self.status_vid = status_vid
        # References
        self.characters_used = characters_used
        self.location_used = location_used
        self.reference_files = reference_files
        self.media_id = media_id  # Media ID cho I2V

        # DEPRECATED aliases (để code cũ không bị lỗi)
        self.start_time = self.srt_start
        self.end_time = self.srt_end
    
    def to_dict(self) -> Dict[str, Any]:
        """Chuyển đổi thành dictionary."""
        return {
            "scene_id": self.scene_id,
            "srt_start": self.srt_start,
            "srt_end": self.srt_end,
            "duration": self.duration,
            "planned_duration": self.planned_duration,  # Thời lượng đạo diễn
            "srt_text": self.srt_text,
            "img_prompt": self.img_prompt,
            "prompt_json": self.prompt_json,
            "video_prompt": self.video_prompt,
            "img_path": self.img_path,
            "video_path": self.video_path,
            "status_img": self.status_img,
            "status_vid": self.status_vid,
            "characters_used": self.characters_used,
            "location_used": self.location_used,
            "reference_files": self.reference_files,
            "media_id": self.media_id,  # Media ID cho I2V
        }
    
    @classmethod
    def from_dict(cls, data: Dict[str, Any]) -> "Scene":
        """Tao Scene tu dictionary."""
        def safe_int(val, default=0):
            """Convert to int safely, handling time strings like '00:00'."""
            if val is None or val == "":
                return default
            if isinstance(val, int):
                return val
            if isinstance(val, float):
                return int(val)
            if isinstance(val, str):
                # Handle time format "HH:MM" or "MM:SS"
                if ":" in val:
                    return default  # Skip time strings
                try:
                    return int(val)
                except ValueError:
                    return default
            return default

        def safe_float(val, default=0.0):
            """Convert to float safely."""
            if val is None or val == "":
                return default
            try:
                return float(val)
            except (ValueError, TypeError):
                return default

        return cls(
            scene_id=safe_int(data.get("scene_id", 0)),
            srt_start=str(data.get("srt_start", "") or ""),  # Timestamp string
            srt_end=str(data.get("srt_end", "") or ""),      # Timestamp string
            duration=safe_float(data.get("duration", 0.0)),
            planned_duration=safe_float(data.get("planned_duration", 0.0)),  # Thời lượng đạo diễn
            srt_text=str(data.get("srt_text", "") or ""),
            img_prompt=str(data.get("img_prompt", "") or ""),
            prompt_json=str(data.get("prompt_json", "") or ""),
            video_prompt=str(data.get("video_prompt", "") or ""),
            img_path=str(data.get("img_path", "") or ""),
            video_path=str(data.get("video_path", "") or ""),
            status_img=str(data.get("status_img", "pending") or "pending"),
            status_vid=str(data.get("status_vid", "pending") or "pending"),
            # DEPRECATED: backward compatible - sẽ được map vào srt_start/srt_end
            start_time=str(data.get("start_time", "") or ""),
            end_time=str(data.get("end_time", "") or ""),
            characters_used=str(data.get("characters_used", "") or ""),
            location_used=str(data.get("location_used", "") or ""),
            reference_files=str(data.get("reference_files", "") or ""),
            media_id=str(data.get("media_id", "") or ""),  # Media ID cho I2V
        )


# ============================================================================
# PROMPT WORKBOOK CLASS
# ============================================================================

class PromptWorkbook:
    """
    Class quản lý file Excel chứa prompts.
    
    Attributes:
        path: Path đến file Excel
        workbook: Workbook object
        characters_sheet: Sheet chứa thông tin nhân vật
        scenes_sheet: Sheet chứa thông tin các scene
    """
    
    CHARACTERS_SHEET = "characters"
    SCENES_SHEET = "scenes"
    DIRECTOR_PLAN_SHEET = "director_plan"
    STORY_ANALYSIS_SHEET = "story_analysis"
    STORY_SEGMENTS_SHEET = "story_segments"  # Nội dung con của câu chuyện
    SCENE_PLANNING_SHEET = "scene_planning"  # Kế hoạch chi tiết từng scene
    LOCATIONS_SHEET = "locations"
    BACKUP_CHARACTERS_SHEET = "backup_characters"
    BACKUP_LOCATIONS_SHEET = "backup_locations"
    SRT_COVERAGE_SHEET = "srt_coverage"  # Đối chiếu SRT entries với segments/scenes
    PROCESSING_STATUS_SHEET = "processing_status"  # Trạng thái xử lý từng step

    # Step definitions for tracking
    STEPS = [
        ("step_1", "Story Analysis", "Phân tích tổng quan câu chuyện"),
        ("step_1.5", "Story Segments", "Chia câu chuyện thành segments"),
        ("step_2", "Characters", "Tạo danh sách nhân vật"),
        ("step_3", "Locations", "Tạo danh sách bối cảnh"),
        ("step_4", "Director Plan", "Tạo kế hoạch đạo diễn"),
        ("step_4.5", "Scene Planning", "Lên ý đồ nghệ thuật"),
        ("step_5", "Scene Prompts", "Tạo prompts cho từng scene"),
    ]

    def __init__(self, path: Union[str, Path]):
        """
        Khởi tạo PromptWorkbook.

        Args:
            path: Path đến file Excel (có thể là str hoặc Path)
        """
        # Chuyển str thành Path để đảm bảo tương thích
        self.path = Path(path) if isinstance(path, str) else path
        self.workbook: Optional[Workbook] = None
        self.logger = get_logger("excel_manager")
    
    def load_or_create(self) -> "PromptWorkbook":
        """
        Load file Excel nếu tồn tại, hoặc tạo mới nếu chưa có.
        
        Returns:
            self để hỗ trợ method chaining
        """
        if self.path.exists():
            self.logger.info(f"Loading existing Excel file: {self.path}")
            self.workbook = load_workbook(self.path)
        else:
            self.logger.info(f"Creating new Excel file: {self.path}")
            self._create_new_workbook()
        
        return self
    
    def _create_new_workbook(self) -> None:
        """Tạo workbook mới với cấu trúc chuẩn."""
        self.workbook = Workbook()
        
        # Xóa sheet mặc định
        default_sheet = self.workbook.active
        
        # Tạo sheet Characters
        self._create_characters_sheet()

        # Tạo sheet Scenes
        self._create_scenes_sheet()

        # Tạo sheet Director Plan
        self._create_director_plan_sheet()

        # Xóa sheet mặc định
        if default_sheet and default_sheet.title == "Sheet":
            self.workbook.remove(default_sheet)
        
        # Lưu file
        self.save()
    
    def _create_characters_sheet(self) -> None:
        """Tạo sheet Characters với header."""
        ws = self.workbook.create_sheet(self.CHARACTERS_SHEET)
        
        # Header style
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Thêm header
        for col, column_name in enumerate(CHARACTERS_COLUMNS, start=1):
            cell = ws.cell(row=1, column=col, value=column_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Điều chỉnh độ rộng cột
        column_widths = {
            "id": 10,
            "role": 12,
            "name": 20,
            "english_prompt": 60,
            "vietnamese_prompt": 40,
            "image_file": 15,
            "status": 10,
        }
        
        for col, column_name in enumerate(CHARACTERS_COLUMNS, start=1):
            ws.column_dimensions[get_column_letter(col)].width = column_widths.get(column_name, 15)
    
    def _create_scenes_sheet(self) -> None:
        """Tạo sheet Scenes với header."""
        ws = self.workbook.create_sheet(self.SCENES_SHEET)
        
        # Header style
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Thêm header
        for col, column_name in enumerate(SCENES_COLUMNS, start=1):
            cell = ws.cell(row=1, column=col, value=column_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Điều chỉnh độ rộng cột
        column_widths = {
            "scene_id": 10,
            "srt_start": 10,
            "srt_end": 10,
            "srt_text": 50,
            "img_prompt": 70,
            "video_prompt": 70,
            "img_path": 30,
            "video_path": 30,
            "status_img": 12,
            "status_vid": 12,
        }
        
        for col, column_name in enumerate(SCENES_COLUMNS, start=1):
            ws.column_dimensions[get_column_letter(col)].width = column_widths.get(column_name, 15)

    def _create_director_plan_sheet(self) -> None:
        """Tạo sheet Director Plan với header."""
        ws = self.workbook.create_sheet(self.DIRECTOR_PLAN_SHEET)

        # Header style - màu cam để phân biệt
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        # Thêm header
        for col, column_name in enumerate(DIRECTOR_PLAN_COLUMNS, start=1):
            cell = ws.cell(row=1, column=col, value=column_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Điều chỉnh độ rộng cột
        column_widths = {
            "plan_id": 10,
            "srt_start": 15,
            "srt_end": 15,
            "duration": 10,
            "srt_text": 50,
            "characters_used": 25,
            "location_used": 15,
            "reference_files": 30,
            "img_prompt": 60,
            "status": 10,
        }

        for col, column_name in enumerate(DIRECTOR_PLAN_COLUMNS, start=1):
            ws.column_dimensions[get_column_letter(col)].width = column_widths.get(column_name, 15)

    def save(self) -> None:
        """Lưu workbook ra file."""
        if self.workbook is None:
            raise RuntimeError("Workbook chưa được load hoặc tạo")
        
        # Đảm bảo thư mục tồn tại
        self.path.parent.mkdir(parents=True, exist_ok=True)
        
        self.workbook.save(self.path)
        self.logger.debug(f"Saved Excel file: {self.path}")
    
    # ========================================================================
    # CHARACTERS METHODS
    # ========================================================================
    
    def get_characters(self) -> List[Character]:
        """
        Lấy danh sách tất cả nhân vật.
        
        Returns:
            List các Character objects
        """
        if self.workbook is None:
            self.load_or_create()
        
        ws = self.workbook[self.CHARACTERS_SHEET]
        characters = []
        
        # Đọc từ dòng 2 (skip header)
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is None:  # Skip empty rows
                continue
            
            data = dict(zip(CHARACTERS_COLUMNS, row))
            characters.append(Character.from_dict(data))
        
        return characters
    
    def add_character(self, character: Character) -> None:
        """
        Thêm nhân vật mới vào sheet.
        
        Args:
            character: Character object
        """
        if self.workbook is None:
            self.load_or_create()
        
        ws = self.workbook[self.CHARACTERS_SHEET]
        
        # Tìm dòng trống tiếp theo
        next_row = ws.max_row + 1
        
        # Thêm dữ liệu
        data = character.to_dict()
        for col, column_name in enumerate(CHARACTERS_COLUMNS, start=1):
            ws.cell(row=next_row, column=col, value=data.get(column_name, ""))
        
        self.logger.debug(f"Added character: {character.id}")
    
    def update_character(self, character_id: str, **kwargs) -> bool:
        """
        Cập nhật thông tin nhân vật.
        
        Args:
            character_id: ID của nhân vật cần cập nhật
            **kwargs: Các field cần cập nhật
            
        Returns:
            True nếu cập nhật thành công, False nếu không tìm thấy
        """
        if self.workbook is None:
            self.load_or_create()
        
        ws = self.workbook[self.CHARACTERS_SHEET]
        
        # Tìm dòng có character_id
        for row_idx in range(2, ws.max_row + 1):
            if ws.cell(row=row_idx, column=1).value == character_id:
                # Cập nhật các field
                for key, value in kwargs.items():
                    if key in CHARACTERS_COLUMNS:
                        col_idx = CHARACTERS_COLUMNS.index(key) + 1
                        ws.cell(row=row_idx, column=col_idx, value=value)
                
                self.logger.debug(f"Updated character: {character_id}")
                return True
        
        self.logger.warning(f"Character not found: {character_id}")
        return False
    
    def clear_characters(self) -> None:
        """Xóa tất cả nhân vật (giữ lại header)."""
        if self.workbook is None:
            self.load_or_create()

        ws = self.workbook[self.CHARACTERS_SHEET]

        # Xóa tất cả dòng trừ header
        ws.delete_rows(2, ws.max_row)
        self.logger.debug("Cleared all characters")

    def get_media_ids(self) -> Dict[str, str]:
        """
        Lấy tất cả media_id từ characters sheet.
        (Cả nhân vật nv* và bối cảnh loc* đều nằm trong characters sheet)

        Returns:
            Dict mapping id -> media_id
            VD: {"nvc": "CAMSJDZiYzQ2...", "loc_01": "CAMSJDZiYzQ1..."}
        """
        if self.workbook is None:
            self.load_or_create()

        result = {}
        ws = self.workbook[self.CHARACTERS_SHEET]

        # Tìm cột media_id
        media_id_col = None
        for col_idx, col_name in enumerate(CHARACTERS_COLUMNS, start=1):
            if col_name == "media_id":
                media_id_col = col_idx
                break

        if media_id_col is None:
            return result

        # Đọc từ dòng 2 (skip header)
        for row_idx in range(2, ws.max_row + 1):
            char_id = ws.cell(row=row_idx, column=1).value
            media_id = ws.cell(row=row_idx, column=media_id_col).value

            if char_id and media_id:
                result[str(char_id)] = str(media_id)

        self.logger.debug(f"Loaded {len(result)} media_ids from Excel")
        return result

    def get_scene_media_ids(self) -> Dict[str, str]:
        """
        Lấy tất cả media_id từ scenes sheet (cho I2V).

        Returns:
            Dict mapping scene_id (string) -> media_id
            VD: {"1": "CAMSJDZiYzQ2...", "2": "CAMSJDZiYzQ1..."}
        """
        if self.workbook is None:
            self.load_or_create()

        result = {}
        ws = self.workbook[self.SCENES_SHEET]

        # Tìm cột media_id
        media_id_col = None
        for col_idx, col_name in enumerate(SCENES_COLUMNS, start=1):
            if col_name == "media_id":
                media_id_col = col_idx
                break

        if media_id_col is None:
            return result

        # Đọc từ dòng 2 (skip header)
        for row_idx in range(2, ws.max_row + 1):
            scene_id = ws.cell(row=row_idx, column=1).value
            media_id = ws.cell(row=row_idx, column=media_id_col).value

            if scene_id and media_id:
                result[str(scene_id)] = str(media_id)

        self.logger.debug(f"Loaded {len(result)} scene media_ids from Excel")
        return result

    # ========================================================================
    # SCENES METHODS
    # ========================================================================
    
    def get_scenes(self) -> List[Scene]:
        """
        Lấy danh sách tất cả scenes.
        
        Returns:
            List các Scene objects
        """
        if self.workbook is None:
            self.load_or_create()
        
        ws = self.workbook[self.SCENES_SHEET]
        scenes = []
        
        # Đọc từ dòng 2 (skip header)
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is None:  # Skip empty rows
                continue
            
            data = dict(zip(SCENES_COLUMNS, row))
            scenes.append(Scene.from_dict(data))
        
        return scenes
    
    def add_scene(self, scene: Scene) -> None:
        """
        Thêm scene mới vào sheet.
        
        Args:
            scene: Scene object
        """
        if self.workbook is None:
            self.load_or_create()
        
        ws = self.workbook[self.SCENES_SHEET]
        
        # Tìm dòng trống tiếp theo
        next_row = ws.max_row + 1
        
        # Thêm dữ liệu
        data = scene.to_dict()
        for col, column_name in enumerate(SCENES_COLUMNS, start=1):
            ws.cell(row=next_row, column=col, value=data.get(column_name, ""))
        
        self.logger.debug(f"Added scene: {scene.scene_id}")
    
    def update_scene(self, scene_id: int, **kwargs) -> bool:
        """
        Cập nhật thông tin scene.
        
        Args:
            scene_id: ID của scene cần cập nhật
            **kwargs: Các field cần cập nhật
            
        Returns:
            True nếu cập nhật thành công, False nếu không tìm thấy
        """
        if self.workbook is None:
            self.load_or_create()
        
        ws = self.workbook[self.SCENES_SHEET]
        
        # Tìm dòng có scene_id
        for row_idx in range(2, ws.max_row + 1):
            cell_value = ws.cell(row=row_idx, column=1).value
            if cell_value is not None and int(cell_value) == scene_id:
                # Cập nhật các field
                for key, value in kwargs.items():
                    if key in SCENES_COLUMNS:
                        col_idx = SCENES_COLUMNS.index(key) + 1
                        ws.cell(row=row_idx, column=col_idx, value=value)
                
                self.logger.debug(f"Updated scene: {scene_id}")
                return True
        
        self.logger.warning(f"Scene not found: {scene_id}")
        return False
    
    def clear_scenes(self) -> None:
        """Xóa tất cả scenes (giữ lại header)."""
        if self.workbook is None:
            self.load_or_create()
        
        ws = self.workbook[self.SCENES_SHEET]
        
        # Xóa tất cả dòng trừ header
        ws.delete_rows(2, ws.max_row)
        self.logger.debug("Cleared all scenes")
    
    def get_pending_image_scenes(self) -> List[Scene]:
        """Lấy danh sách scenes chưa tạo ảnh."""
        scenes = self.get_scenes()
        return [s for s in scenes if s.status_img != "done" and s.img_prompt]
    
    def get_pending_video_scenes(self) -> List[Scene]:
        """Lấy danh sách scenes chưa tạo video (nhưng đã có ảnh)."""
        scenes = self.get_scenes()
        return [s for s in scenes if s.status_vid != "done" and s.img_path and s.video_prompt]

    # ========================================================================
    # DIRECTOR PLAN METHODS
    # ========================================================================

    def _ensure_director_plan_sheet(self) -> None:
        """Đảm bảo sheet director_plan tồn tại (cho Excel cũ)."""
        if self.workbook is None:
            self.load_or_create()

        if self.DIRECTOR_PLAN_SHEET not in self.workbook.sheetnames:
            self._create_director_plan_sheet()
            self.save()

    def save_director_plan(self, scenes_data: List[Dict]) -> None:
        """
        Lưu kế hoạch scenes từ SRT vào sheet director_plan.
        Gọi hàm này TRƯỚC KHI tạo prompts để có thể detect gaps.

        Args:
            scenes_data: List các scene dict với keys:
                - scene_id, srt_start, srt_end, duration, text (required)
                - characters_used, location_used, reference_files, img_prompt (optional - backup)
        """
        self._ensure_director_plan_sheet()

        ws = self.workbook[self.DIRECTOR_PLAN_SHEET]

        # Xóa dữ liệu cũ (giữ header)
        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row)

        # Thêm scenes
        for scene in scenes_data:
            next_row = ws.max_row + 1
            # Đảm bảo scene_id là integer, không phải float (1.0 -> 1)
            scene_id = scene.get("scene_id", 0)
            ws.cell(row=next_row, column=1, value=int(scene_id) if scene_id else 0)
            ws.cell(row=next_row, column=2, value=scene.get("srt_start", ""))
            ws.cell(row=next_row, column=3, value=scene.get("srt_end", ""))
            # Duration: handle cả "duration" và "duration_seconds" (3-8s từ SRT timing)
            duration = scene.get("duration") or scene.get("duration_seconds") or 0
            ws.cell(row=next_row, column=4, value=round(duration, 2) if duration else 0)
            ws.cell(row=next_row, column=5, value=scene.get("text", "")[:500])
            # New columns for backup
            # Handle characters_used - convert list to JSON string if needed
            chars_used = scene.get("characters_used", "[]")
            if isinstance(chars_used, list):
                chars_used = json.dumps(chars_used)
            ws.cell(row=next_row, column=6, value=chars_used)

            ws.cell(row=next_row, column=7, value=scene.get("location_used", ""))

            # Handle reference_files - convert list to JSON string if needed
            ref_files = scene.get("reference_files", "[]")
            if isinstance(ref_files, list):
                ref_files = json.dumps(ref_files)
            ws.cell(row=next_row, column=8, value=ref_files)
            ws.cell(row=next_row, column=9, value=scene.get("img_prompt", "")[:1000])
            ws.cell(row=next_row, column=10, value=scene.get("status", "backup"))

        self.save()
        self.logger.info(f"Saved {len(scenes_data)} scenes to director_plan")

    def get_director_plan(self) -> List[Dict]:
        """
        Lấy kế hoạch scenes từ sheet director_plan.

        Returns:
            List các scene dict với backup info
        """
        self._ensure_director_plan_sheet()

        ws = self.workbook[self.DIRECTOR_PLAN_SHEET]
        plans = []

        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue

            # Handle both old format (6 cols) and new format (10 cols)
            plans.append({
                "plan_id": row[0],
                "scene_id": row[0],  # Alias cho plan_id (step 5 dùng scene_id)
                "srt_start": row[1] or "",
                "srt_end": row[2] or "",
                "duration": row[3] or 0,
                "srt_text": row[4] or "",
                "characters_used": row[5] if len(row) > 5 else "[]",
                "location_used": row[6] if len(row) > 6 else "",
                "reference_files": row[7] if len(row) > 7 else "[]",
                "img_prompt": row[8] if len(row) > 8 else "",
                "status": row[9] if len(row) > 9 else "pending",
            })

        return plans

    def update_director_plan_status(self, plan_id: int, status: str) -> bool:
        """Cập nhật status của một plan entry."""
        self._ensure_director_plan_sheet()

        ws = self.workbook[self.DIRECTOR_PLAN_SHEET]

        for row_idx in range(2, ws.max_row + 1):
            if ws.cell(row=row_idx, column=1).value == plan_id:
                ws.cell(row=row_idx, column=6, value=status)
                return True

        return False

    # ========== STORY ANALYSIS SHEET ==========

    def _ensure_story_analysis_sheet(self) -> None:
        """Đảm bảo sheet story_analysis tồn tại."""
        if self.workbook is None:
            self.load_or_create()

        if self.STORY_ANALYSIS_SHEET not in self.workbook.sheetnames:
            self._create_story_analysis_sheet()
            self.save()

    def _create_story_analysis_sheet(self) -> None:
        """Tạo sheet story_analysis với header."""
        ws = self.workbook.create_sheet(self.STORY_ANALYSIS_SHEET)

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="8B4513", end_color="8B4513", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        columns = ["key", "value"]
        for col, column_name in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col, value=column_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 100

    def save_story_analysis(self, data: dict) -> None:
        """
        Lưu phân tích story vào sheet story_analysis.

        Args:
            data: Dict với các keys như setting, themes, visual_style, context_lock
        """
        self._ensure_story_analysis_sheet()
        ws = self.workbook[self.STORY_ANALYSIS_SHEET]

        # Xóa dữ liệu cũ (giữ header)
        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row)

        # Flatten nested dict và lưu
        def flatten_dict(d, parent_key=''):
            items = []
            for k, v in d.items():
                new_key = f"{parent_key}.{k}" if parent_key else k
                if isinstance(v, dict):
                    items.extend(flatten_dict(v, new_key))
                elif isinstance(v, list):
                    items.append((new_key, json.dumps(v)))
                else:
                    items.append((new_key, str(v) if v else ""))
            return items

        for key, value in flatten_dict(data):
            next_row = ws.max_row + 1
            ws.cell(row=next_row, column=1, value=key)
            ws.cell(row=next_row, column=2, value=value[:1000] if value else "")

        self.save()
        self.logger.info(f"Saved story_analysis to Excel")

    def get_story_analysis(self) -> dict:
        """
        Đọc phân tích story từ sheet story_analysis.

        Returns:
            Dict với các keys từ sheet
        """
        self._ensure_story_analysis_sheet()
        ws = self.workbook[self.STORY_ANALYSIS_SHEET]

        data = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue
            key = row[0]
            value = row[1] or ""

            # Try to parse JSON for lists
            if value.startswith("["):
                try:
                    value = json.loads(value)
                except:
                    pass

            # Handle nested keys (e.g., "setting.era")
            if "." in key:
                parts = key.split(".")
                current = data
                for part in parts[:-1]:
                    if part not in current:
                        current[part] = {}
                    current = current[part]
                current[parts[-1]] = value
            else:
                data[key] = value

        return data

    # ========== STORY SEGMENTS SHEET ==========

    def _ensure_story_segments_sheet(self) -> None:
        """Đảm bảo sheet story_segments tồn tại."""
        if self.workbook is None:
            self.load_or_create()

        if self.STORY_SEGMENTS_SHEET not in self.workbook.sheetnames:
            self._create_story_segments_sheet()
            self.save()

    def _create_story_segments_sheet(self) -> None:
        """Tạo sheet story_segments với header."""
        ws = self.workbook.create_sheet(self.STORY_SEGMENTS_SHEET)

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="9932CC", end_color="9932CC", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        columns = [
            "segment_id", "segment_name", "message", "key_elements",
            "image_count", "estimated_duration", "srt_range_start",
            "srt_range_end", "importance", "status"
        ]
        for col, column_name in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col, value=column_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

    def save_story_segments(self, segments: list, total_images: int = 0, summary: str = "") -> None:
        """
        Lưu story segments vào Excel.

        Args:
            segments: List các segment dict
            total_images: Tổng số ảnh cần tạo
            summary: Tóm tắt cấu trúc câu chuyện
        """
        self._ensure_story_segments_sheet()
        ws = self.workbook[self.STORY_SEGMENTS_SHEET]

        # Xóa dữ liệu cũ (giữ header)
        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row)

        # Thêm segments
        for seg in segments:
            next_row = ws.max_row + 1
            ws.cell(row=next_row, column=1, value=int(seg.get("segment_id", 0)))
            ws.cell(row=next_row, column=2, value=seg.get("segment_name", ""))
            ws.cell(row=next_row, column=3, value=seg.get("message", "")[:500])

            # key_elements - convert list to JSON string
            key_elements = seg.get("key_elements", [])
            if isinstance(key_elements, list):
                key_elements = json.dumps(key_elements)
            ws.cell(row=next_row, column=4, value=key_elements)

            ws.cell(row=next_row, column=5, value=int(seg.get("image_count", 1)))
            ws.cell(row=next_row, column=6, value=round(seg.get("estimated_duration", 0), 2))
            ws.cell(row=next_row, column=7, value=int(seg.get("srt_range_start", 0)))
            ws.cell(row=next_row, column=8, value=int(seg.get("srt_range_end", 0)))
            ws.cell(row=next_row, column=9, value=seg.get("importance", "medium"))
            ws.cell(row=next_row, column=10, value="pending")

        self.logger.info(f"Saved {len(segments)} story segments (total {total_images} images)")

    def get_story_segments(self) -> list:
        """
        Đọc story segments từ Excel.

        Returns:
            List các segment dict
        """
        self._ensure_story_segments_sheet()
        ws = self.workbook[self.STORY_SEGMENTS_SHEET]

        segments = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue

            seg = {
                "segment_id": int(row[0]) if row[0] else 0,
                "segment_name": row[1] or "",
                "message": row[2] or "",
                "key_elements": row[3] or "[]",
                "image_count": int(row[4]) if row[4] else 1,
                "estimated_duration": float(row[5]) if row[5] else 0,
                "srt_range_start": int(row[6]) if row[6] else 0,
                "srt_range_end": int(row[7]) if row[7] else 0,
                "importance": row[8] or "medium",
                "status": row[9] or "pending"
            }

            # Parse key_elements JSON
            if isinstance(seg["key_elements"], str) and seg["key_elements"].startswith("["):
                try:
                    seg["key_elements"] = json.loads(seg["key_elements"])
                except:
                    pass

            segments.append(seg)

        return segments

    # ========== SCENE PLANNING SHEET ==========

    def _ensure_scene_planning_sheet(self) -> None:
        """Đảm bảo sheet scene_planning tồn tại."""
        if self.workbook is None:
            self.load_or_create()

        if self.SCENE_PLANNING_SHEET not in self.workbook.sheetnames:
            self._create_scene_planning_sheet()
            self.save()

    def _create_scene_planning_sheet(self) -> None:
        """Tạo sheet scene_planning với header."""
        ws = self.workbook.create_sheet(self.SCENE_PLANNING_SHEET)

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="FF6347", end_color="FF6347", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        columns = [
            "scene_id", "artistic_intent", "shot_type", "character_action",
            "mood", "lighting", "color_palette", "key_focus"
        ]
        for col, column_name in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col, value=column_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

    def save_scene_planning(self, plans: list) -> None:
        """
        Lưu scene planning vào Excel.

        Args:
            plans: List các plan dict từ API
        """
        self._ensure_scene_planning_sheet()
        ws = self.workbook[self.SCENE_PLANNING_SHEET]

        # Xóa dữ liệu cũ (giữ header)
        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row)

        # Thêm plans
        for plan in plans:
            next_row = ws.max_row + 1
            ws.cell(row=next_row, column=1, value=int(plan.get("scene_id", 0)))
            ws.cell(row=next_row, column=2, value=plan.get("artistic_intent", "")[:500])
            ws.cell(row=next_row, column=3, value=plan.get("shot_type", ""))
            ws.cell(row=next_row, column=4, value=plan.get("character_action", "")[:500])
            ws.cell(row=next_row, column=5, value=plan.get("mood", ""))
            ws.cell(row=next_row, column=6, value=plan.get("lighting", ""))
            ws.cell(row=next_row, column=7, value=plan.get("color_palette", ""))
            ws.cell(row=next_row, column=8, value=plan.get("key_focus", "")[:300])

        self.logger.info(f"Saved {len(plans)} scene plans")

    def get_scene_planning(self) -> list:
        """
        Đọc scene planning từ Excel.

        Returns:
            List các plan dict
        """
        self._ensure_scene_planning_sheet()
        ws = self.workbook[self.SCENE_PLANNING_SHEET]

        plans = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue

            plan = {
                "scene_id": int(row[0]) if row[0] else 0,
                "artistic_intent": row[1] or "",
                "shot_type": row[2] or "",
                "character_action": row[3] or "",
                "mood": row[4] or "",
                "lighting": row[5] or "",
                "color_palette": row[6] or "",
                "key_focus": row[7] or "" if len(row) > 7 else ""
            }
            plans.append(plan)

        return plans

    # ========== LOCATIONS SHEET ==========

    def _ensure_locations_sheet(self) -> None:
        """Đảm bảo sheet locations tồn tại."""
        if self.workbook is None:
            self.load_or_create()

        if self.LOCATIONS_SHEET not in self.workbook.sheetnames:
            self._create_locations_sheet()
            self.save()

    def _create_locations_sheet(self) -> None:
        """Tạo sheet locations với header."""
        ws = self.workbook.create_sheet(self.LOCATIONS_SHEET)

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2E8B57", end_color="2E8B57", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        columns = ["id", "name", "english_prompt", "location_lock", "lighting_default", "image_file", "status"]
        for col, column_name in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col, value=column_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        column_widths = {"id": 15, "name": 25, "english_prompt": 80, "location_lock": 50, "lighting_default": 30, "image_file": 20, "status": 12}
        for col, column_name in enumerate(columns, start=1):
            ws.column_dimensions[get_column_letter(col)].width = column_widths.get(column_name, 15)

    def add_location(self, location: "Location") -> None:
        """
        Thêm một location mới vào sheet locations.

        Args:
            location: Location object
        """
        self._ensure_locations_sheet()
        ws = self.workbook[self.LOCATIONS_SHEET]

        next_row = ws.max_row + 1
        ws.cell(row=next_row, column=1, value=location.id)
        ws.cell(row=next_row, column=2, value=location.name)
        ws.cell(row=next_row, column=3, value=location.english_prompt[:500] if location.english_prompt else "")
        ws.cell(row=next_row, column=4, value=getattr(location, 'location_lock', '')[:200])
        ws.cell(row=next_row, column=5, value=getattr(location, 'lighting_default', ''))
        ws.cell(row=next_row, column=6, value=getattr(location, 'image_file', ''))
        ws.cell(row=next_row, column=7, value="pending")

    def get_locations(self) -> List["Location"]:
        """
        Đọc locations từ sheet characters (role="location" hoặc id bắt đầu bằng "loc_").

        Returns:
            List[Location]
        """
        # Đọc từ characters sheet thay vì locations sheet riêng
        characters = self.get_characters()

        locations = []
        for char in characters:
            # Check nếu là location (role="location" hoặc id bắt đầu bằng "loc_")
            if char.role == "location" or (char.id and char.id.startswith("loc_")):
                loc = Location(
                    id=char.id,
                    name=char.name,
                    english_prompt=char.english_prompt,
                    location_lock=char.character_lock,  # character_lock chứa location_lock
                    lighting_default=char.vietnamese_prompt,  # vietnamese_prompt chứa lighting_default
                    image_file=char.image_file,
                )
                locations.append(loc)

        return locations

    # ========== BACKUP CHARACTERS SHEET ==========

    def _ensure_backup_characters_sheet(self) -> None:
        """Đảm bảo sheet backup_characters tồn tại."""
        if self.workbook is None:
            self.load_or_create()

        if self.BACKUP_CHARACTERS_SHEET not in self.workbook.sheetnames:
            self._create_backup_characters_sheet()
            self.save()

    def _create_backup_characters_sheet(self) -> None:
        """Tạo sheet backup_characters với header."""
        ws = self.workbook.create_sheet(self.BACKUP_CHARACTERS_SHEET)

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        for col, column_name in enumerate(BACKUP_CHARACTERS_COLUMNS, start=1):
            cell = ws.cell(row=1, column=col, value=column_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Điều chỉnh độ rộng cột
        column_widths = {"id": 10, "name": 20, "character_lock": 80, "costume_lock": 60, "image_file": 15}
        for col, column_name in enumerate(BACKUP_CHARACTERS_COLUMNS, start=1):
            ws.column_dimensions[get_column_letter(col)].width = column_widths.get(column_name, 15)

    def save_backup_characters(self, characters: List[Dict]) -> None:
        """
        Lưu nhân vật dự phòng (narrator) vào backup_characters sheet.

        Args:
            characters: List dict với keys: id, name, character_lock, costume_lock, image_file
        """
        self._ensure_backup_characters_sheet()
        ws = self.workbook[self.BACKUP_CHARACTERS_SHEET]

        # Xóa dữ liệu cũ (giữ header)
        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row)

        for char in characters:
            next_row = ws.max_row + 1
            ws.cell(row=next_row, column=1, value=char.get("id", "nvc"))
            ws.cell(row=next_row, column=2, value=char.get("name", "Narrator"))
            ws.cell(row=next_row, column=3, value=char.get("character_lock", ""))
            ws.cell(row=next_row, column=4, value=char.get("costume_lock", ""))
            ws.cell(row=next_row, column=5, value=char.get("image_file", "nvc.png"))

        self.save()
        self.logger.info(f"Saved {len(characters)} backup characters")

    def get_backup_characters(self) -> List[Dict]:
        """Lấy danh sách nhân vật dự phòng từ backup_characters sheet."""
        self._ensure_backup_characters_sheet()
        ws = self.workbook[self.BACKUP_CHARACTERS_SHEET]
        characters = []

        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue
            characters.append({
                "id": row[0],
                "name": row[1] or "",
                "character_lock": row[2] or "",
                "costume_lock": row[3] or "",
                "image_file": row[4] or "nvc.png",
            })

        return characters

    # ========== BACKUP LOCATIONS SHEET ==========

    def _ensure_backup_locations_sheet(self) -> None:
        """Đảm bảo sheet backup_locations tồn tại."""
        if self.workbook is None:
            self.load_or_create()

        if self.BACKUP_LOCATIONS_SHEET not in self.workbook.sheetnames:
            self._create_backup_locations_sheet()
            self.save()

    def _create_backup_locations_sheet(self) -> None:
        """Tạo sheet backup_locations với header."""
        ws = self.workbook.create_sheet(self.BACKUP_LOCATIONS_SHEET)

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        for col, column_name in enumerate(BACKUP_LOCATIONS_COLUMNS, start=1):
            cell = ws.cell(row=1, column=col, value=column_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Điều chỉnh độ rộng cột
        column_widths = {"id": 10, "name": 25, "location_lock": 80, "image_file": 15}
        for col, column_name in enumerate(BACKUP_LOCATIONS_COLUMNS, start=1):
            ws.column_dimensions[get_column_letter(col)].width = column_widths.get(column_name, 15)

    def save_backup_locations(self, locations: List[Dict]) -> None:
        """
        Lưu location dự phòng (nơi kể chuyện) vào backup_locations sheet.

        Args:
            locations: List dict với keys: id, name, location_lock, image_file
        """
        self._ensure_backup_locations_sheet()
        ws = self.workbook[self.BACKUP_LOCATIONS_SHEET]

        # Xóa dữ liệu cũ (giữ header)
        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row)

        for loc in locations:
            next_row = ws.max_row + 1
            ws.cell(row=next_row, column=1, value=loc.get("id", "loc"))
            ws.cell(row=next_row, column=2, value=loc.get("name", "Storytelling Location"))
            ws.cell(row=next_row, column=3, value=loc.get("location_lock", ""))
            ws.cell(row=next_row, column=4, value=loc.get("image_file", "loc.png"))

        self.save()
        self.logger.info(f"Saved {len(locations)} backup locations")

    def get_backup_locations(self) -> List[Dict]:
        """Lấy danh sách location dự phòng từ backup_locations sheet."""
        self._ensure_backup_locations_sheet()
        ws = self.workbook[self.BACKUP_LOCATIONS_SHEET]
        locations = []

        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue
            locations.append({
                "id": row[0],
                "name": row[1] or "",
                "location_lock": row[2] or "",
                "image_file": row[3] or "loc.png",
            })

        return locations

    def detect_scene_gaps(self) -> List[Dict]:
        """
        So sánh director_plan với scenes để detect gaps (scenes thiếu).

        Returns:
            List các gap dict: {plan_id, srt_start, srt_end, reason}
        """
        plans = self.get_director_plan()
        scenes = self.get_scenes()

        # Tạo set scene_ids đã có prompts
        scene_ids_with_prompts = {
            s.scene_id for s in scenes
            if s.img_prompt and s.img_prompt.strip()
        }

        gaps = []
        for plan in plans:
            plan_id = plan["plan_id"]
            if plan_id not in scene_ids_with_prompts:
                gaps.append({
                    "plan_id": plan_id,
                    "srt_start": plan["srt_start"],
                    "srt_end": plan["srt_end"],
                    "srt_text": plan["srt_text"][:100],
                    "reason": "missing_prompt"
                })

        return gaps

    def detect_timeline_gaps(self, video_duration_seconds: float = None) -> List[Dict]:
        """
        Phát hiện gaps trong timeline của scenes (khoảng thời gian không có scene nào cover).

        Args:
            video_duration_seconds: Tổng thời lượng video (giây). Nếu None sẽ dùng scene cuối.

        Returns:
            List các gap dict: {start_seconds, end_seconds, start_time, end_time, duration}
        """
        scenes = self.get_scenes()
        if not scenes:
            return []

        def parse_timestamp(ts: str) -> float:
            """Convert HH:MM:SS,mmm to seconds"""
            if not ts:
                return 0
            try:
                # Handle both HH:MM:SS,mmm and HH:MM:SS formats
                ts = ts.replace(',', '.')
                parts = ts.split(':')
                if len(parts) == 3:
                    h, m, s = parts
                    return int(h) * 3600 + int(m) * 60 + float(s)
                elif len(parts) == 2:
                    m, s = parts
                    return int(m) * 60 + float(s)
                return float(ts)
            except:
                return 0

        def seconds_to_timestamp(secs: float) -> str:
            """Convert seconds to HH:MM:SS,mmm"""
            h = int(secs // 3600)
            m = int((secs % 3600) // 60)
            s = secs % 60
            return f"{h:02d}:{m:02d}:{s:06.3f}".replace('.', ',')

        # Sort scenes by start time
        scene_times = []
        for s in scenes:
            if s.img_prompt and s.img_prompt.strip():  # Chỉ đếm scenes có prompt
                start = parse_timestamp(s.srt_start)
                end = parse_timestamp(s.srt_end)
                if end > start:
                    scene_times.append((start, end))

        if not scene_times:
            return []

        scene_times.sort(key=lambda x: x[0])

        # Determine video end time
        if video_duration_seconds:
            video_end = video_duration_seconds
        else:
            video_end = max(end for _, end in scene_times)

        # Find gaps
        gaps = []
        MIN_GAP_SECONDS = 3  # Ignore gaps smaller than 3 seconds

        # Gap from start?
        first_start = scene_times[0][0]
        if first_start > MIN_GAP_SECONDS:
            gaps.append({
                "start_seconds": 0,
                "end_seconds": first_start,
                "start_time": "00:00:00,000",
                "end_time": seconds_to_timestamp(first_start),
                "duration": first_start
            })

        # Gaps between scenes
        current_end = scene_times[0][1]
        for start, end in scene_times[1:]:
            if start > current_end + MIN_GAP_SECONDS:
                gaps.append({
                    "start_seconds": current_end,
                    "end_seconds": start,
                    "start_time": seconds_to_timestamp(current_end),
                    "end_time": seconds_to_timestamp(start),
                    "duration": start - current_end
                })
            current_end = max(current_end, end)

        # Gap at end?
        if video_end > current_end + MIN_GAP_SECONDS:
            gaps.append({
                "start_seconds": current_end,
                "end_seconds": video_end,
                "start_time": seconds_to_timestamp(current_end),
                "end_time": seconds_to_timestamp(video_end),
                "duration": video_end - current_end
            })

        return gaps

    # ========================================================================
    # SRT COVERAGE TRACKING - Đối chiếu SRT với Segments/Scenes
    # ========================================================================

    def _ensure_srt_coverage_sheet(self) -> None:
        """Tạo sheet srt_coverage nếu chưa có."""
        if self.SRT_COVERAGE_SHEET not in self.workbook.sheetnames:
            ws = self.workbook.create_sheet(self.SRT_COVERAGE_SHEET)
            headers = [
                "srt_index", "start_time", "end_time", "text_preview",
                "segment_id", "segment_name", "scene_id", "status"
            ]
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")

            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill

            # Set column widths
            ws.column_dimensions['A'].width = 10
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 50
            ws.column_dimensions['E'].width = 12
            ws.column_dimensions['F'].width = 30
            ws.column_dimensions['G'].width = 12
            ws.column_dimensions['H'].width = 12

    def init_srt_coverage(self, srt_entries: list) -> None:
        """
        Khởi tạo SRT coverage tracking với tất cả SRT entries.
        Gọi ở đầu quy trình để có baseline.

        Args:
            srt_entries: List các SRT entry objects
        """
        self._ensure_srt_coverage_sheet()
        ws = self.workbook[self.SRT_COVERAGE_SHEET]

        # Clear existing data (keep header)
        for row in range(ws.max_row, 1, -1):
            ws.delete_rows(row)

        # Add all SRT entries
        uncovered_fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")

        for i, entry in enumerate(srt_entries, 1):
            row = i + 1  # +1 for header
            ws.cell(row=row, column=1, value=i)  # srt_index
            ws.cell(row=row, column=2, value=entry.start_time)
            ws.cell(row=row, column=3, value=entry.end_time)
            ws.cell(row=row, column=4, value=entry.text[:50] + "..." if len(entry.text) > 50 else entry.text)
            ws.cell(row=row, column=5, value="")  # segment_id - empty
            ws.cell(row=row, column=6, value="")  # segment_name - empty
            ws.cell(row=row, column=7, value="")  # scene_id - empty
            status_cell = ws.cell(row=row, column=8, value="UNCOVERED")
            status_cell.fill = uncovered_fill

        self.save()
        self.logger.info(f"Initialized SRT coverage tracking for {len(srt_entries)} entries")

    def update_srt_coverage_segments(self, segments: list) -> dict:
        """
        Cập nhật coverage sau Step 1.5 (segments).

        Args:
            segments: List segments từ Step 1.5

        Returns:
            Dict với coverage statistics
        """
        self._ensure_srt_coverage_sheet()
        ws = self.workbook[self.SRT_COVERAGE_SHEET]

        segment_fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")  # Yellow
        covered = 0
        uncovered = 0

        for row in range(2, ws.max_row + 1):
            srt_index = ws.cell(row=row, column=1).value
            if srt_index is None:
                continue

            # Find which segment covers this SRT
            segment_found = None
            for seg in segments:
                start = seg.get("srt_range_start", 0)
                end = seg.get("srt_range_end", 0)
                if start <= srt_index <= end:
                    segment_found = seg
                    break

            if segment_found:
                ws.cell(row=row, column=5, value=segment_found.get("segment_id", ""))
                ws.cell(row=row, column=6, value=segment_found.get("segment_name", ""))
                ws.cell(row=row, column=8, value="SEGMENT_OK")
                ws.cell(row=row, column=8).fill = segment_fill
                covered += 1
            else:
                uncovered += 1

        self.save()

        total = covered + uncovered
        coverage_pct = (covered / total * 100) if total > 0 else 0

        return {
            "total_srt": total,
            "covered_by_segment": covered,
            "uncovered": uncovered,
            "coverage_percent": round(coverage_pct, 1)
        }

    def update_srt_coverage_scenes(self, director_plan: list) -> dict:
        """
        Cập nhật coverage sau Step 4 (director_plan).

        Args:
            director_plan: List scenes từ Step 4

        Returns:
            Dict với coverage statistics
        """
        self._ensure_srt_coverage_sheet()
        ws = self.workbook[self.SRT_COVERAGE_SHEET]

        scene_fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")  # Green
        covered = 0
        uncovered = 0

        for row in range(2, ws.max_row + 1):
            srt_index = ws.cell(row=row, column=1).value
            if srt_index is None:
                continue

            # Find which scene covers this SRT
            scene_ids = []
            for scene in director_plan:
                indices = scene.get("srt_indices", [])
                if srt_index in indices:
                    scene_ids.append(scene.get("scene_id", ""))

            if scene_ids:
                ws.cell(row=row, column=7, value=", ".join(map(str, scene_ids)))
                ws.cell(row=row, column=8, value="COVERED")
                ws.cell(row=row, column=8).fill = scene_fill
                covered += 1
            else:
                current_status = ws.cell(row=row, column=8).value
                if current_status != "COVERED":
                    uncovered += 1

        self.save()

        total = covered + uncovered
        coverage_pct = (covered / total * 100) if total > 0 else 0

        return {
            "total_srt": total,
            "covered_by_scene": covered,
            "uncovered": uncovered,
            "coverage_percent": round(coverage_pct, 1)
        }

    def get_srt_coverage_summary(self) -> dict:
        """
        Lấy tổng hợp coverage hiện tại.

        Returns:
            Dict với statistics
        """
        self._ensure_srt_coverage_sheet()
        ws = self.workbook[self.SRT_COVERAGE_SHEET]

        total = 0
        covered = 0
        segment_only = 0
        uncovered = 0

        for row in range(2, ws.max_row + 1):
            srt_index = ws.cell(row=row, column=1).value
            if srt_index is None:
                continue

            total += 1
            status = ws.cell(row=row, column=8).value

            if status == "COVERED":
                covered += 1
            elif status == "SEGMENT_OK":
                segment_only += 1
            else:
                uncovered += 1

        return {
            "total_srt": total,
            "fully_covered": covered,
            "segment_only": segment_only,
            "uncovered": uncovered,
            "coverage_percent": round((covered / total * 100) if total > 0 else 0, 1)
        }

    def get_uncovered_srt_entries(self) -> list:
        """
        Lấy danh sách SRT entries chưa được cover.

        Returns:
            List of dicts với SRT info
        """
        self._ensure_srt_coverage_sheet()
        ws = self.workbook[self.SRT_COVERAGE_SHEET]

        uncovered = []
        for row in range(2, ws.max_row + 1):
            status = ws.cell(row=row, column=8).value
            if status == "UNCOVERED":
                uncovered.append({
                    "srt_index": ws.cell(row=row, column=1).value,
                    "start_time": ws.cell(row=row, column=2).value,
                    "end_time": ws.cell(row=row, column=3).value,
                    "text_preview": ws.cell(row=row, column=4).value
                })

        return uncovered

    # ========================================================================
    # PROCESSING STATUS - Theo dõi trạng thái từng step
    # ========================================================================

    def _ensure_processing_status_sheet(self) -> None:
        """Tạo sheet processing_status nếu chưa có."""
        if self.PROCESSING_STATUS_SHEET not in self.workbook.sheetnames:
            ws = self.workbook.create_sheet(self.PROCESSING_STATUS_SHEET)
            headers = [
                "step_id", "step_name", "description", "status",
                "items_total", "items_done", "coverage_pct", "notes", "last_updated"
            ]
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")

            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill

            # Initialize all steps
            for i, (step_id, step_name, desc) in enumerate(self.STEPS, 2):
                ws.cell(row=i, column=1, value=step_id)
                ws.cell(row=i, column=2, value=step_name)
                ws.cell(row=i, column=3, value=desc)
                ws.cell(row=i, column=4, value="PENDING")
                ws.cell(row=i, column=5, value=0)
                ws.cell(row=i, column=6, value=0)
                ws.cell(row=i, column=7, value=0)
                ws.cell(row=i, column=8, value="")
                ws.cell(row=i, column=9, value="")

            # Set column widths
            ws.column_dimensions['A'].width = 10
            ws.column_dimensions['B'].width = 18
            ws.column_dimensions['C'].width = 35
            ws.column_dimensions['D'].width = 12
            ws.column_dimensions['E'].width = 12
            ws.column_dimensions['F'].width = 12
            ws.column_dimensions['G'].width = 12
            ws.column_dimensions['H'].width = 50
            ws.column_dimensions['I'].width = 20

    def update_step_status(self, step_id: str, status: str, items_total: int = 0,
                           items_done: int = 0, notes: str = "") -> None:
        """
        Cập nhật trạng thái của một step.

        Args:
            step_id: ID của step (step_1, step_1.5, etc.)
            status: PENDING, IN_PROGRESS, COMPLETED, PARTIAL, ERROR
            items_total: Tổng số items cần xử lý
            items_done: Số items đã xong
            notes: Ghi chú
        """
        from datetime import datetime

        self._ensure_processing_status_sheet()
        ws = self.workbook[self.PROCESSING_STATUS_SHEET]

        # Color mapping
        status_colors = {
            "PENDING": "E0E0E0",      # Gray
            "IN_PROGRESS": "FFF9C4",  # Yellow
            "COMPLETED": "C8E6C9",    # Green
            "PARTIAL": "FFE0B2",      # Orange
            "ERROR": "FFCDD2"         # Red
        }

        # Find the row for this step
        for row in range(2, ws.max_row + 1):
            if ws.cell(row=row, column=1).value == step_id:
                ws.cell(row=row, column=4, value=status)
                ws.cell(row=row, column=4).fill = PatternFill(
                    start_color=status_colors.get(status, "FFFFFF"),
                    end_color=status_colors.get(status, "FFFFFF"),
                    fill_type="solid"
                )
                ws.cell(row=row, column=5, value=items_total)
                ws.cell(row=row, column=6, value=items_done)

                # Calculate coverage percentage
                coverage = (items_done / items_total * 100) if items_total > 0 else 0
                ws.cell(row=row, column=7, value=round(coverage, 1))

                # Add notes
                if notes:
                    existing_notes = ws.cell(row=row, column=8).value or ""
                    if existing_notes and notes not in existing_notes:
                        notes = f"{existing_notes}; {notes}"
                    ws.cell(row=row, column=8, value=notes[:500])  # Limit notes length

                ws.cell(row=row, column=9, value=datetime.now().strftime("%Y-%m-%d %H:%M"))
                break

        self.save()

    def get_step_status(self, step_id: str) -> dict:
        """Lấy trạng thái của một step."""
        self._ensure_processing_status_sheet()
        ws = self.workbook[self.PROCESSING_STATUS_SHEET]

        for row in range(2, ws.max_row + 1):
            if ws.cell(row=row, column=1).value == step_id:
                return {
                    "step_id": step_id,
                    "step_name": ws.cell(row=row, column=2).value,
                    "status": ws.cell(row=row, column=4).value,
                    "items_total": ws.cell(row=row, column=5).value or 0,
                    "items_done": ws.cell(row=row, column=6).value or 0,
                    "coverage_pct": ws.cell(row=row, column=7).value or 0,
                    "notes": ws.cell(row=row, column=8).value or "",
                    "last_updated": ws.cell(row=row, column=9).value or ""
                }
        return {}

    def get_all_step_status(self) -> list:
        """Lấy trạng thái của tất cả steps."""
        self._ensure_processing_status_sheet()
        ws = self.workbook[self.PROCESSING_STATUS_SHEET]

        statuses = []
        for row in range(2, ws.max_row + 1):
            step_id = ws.cell(row=row, column=1).value
            if step_id:
                statuses.append({
                    "step_id": step_id,
                    "step_name": ws.cell(row=row, column=2).value,
                    "status": ws.cell(row=row, column=4).value,
                    "items_total": ws.cell(row=row, column=5).value or 0,
                    "items_done": ws.cell(row=row, column=6).value or 0,
                    "coverage_pct": ws.cell(row=row, column=7).value or 0,
                    "notes": ws.cell(row=row, column=8).value or ""
                })
        return statuses

    def get_incomplete_steps(self) -> list:
        """Lấy danh sách các steps chưa hoàn thành (PARTIAL hoặc ERROR)."""
        statuses = self.get_all_step_status()
        return [s for s in statuses if s["status"] in ("PENDING", "PARTIAL", "ERROR", "IN_PROGRESS")]

    def get_processing_summary(self) -> dict:
        """
        Lấy tổng hợp trạng thái xử lý.

        Returns:
            Dict với thông tin tổng quan
        """
        statuses = self.get_all_step_status()

        completed = sum(1 for s in statuses if s["status"] == "COMPLETED")
        partial = sum(1 for s in statuses if s["status"] == "PARTIAL")
        pending = sum(1 for s in statuses if s["status"] == "PENDING")
        error = sum(1 for s in statuses if s["status"] == "ERROR")

        # Get SRT coverage if available
        srt_summary = {}
        try:
            srt_summary = self.get_srt_coverage_summary()
        except:
            pass

        return {
            "total_steps": len(statuses),
            "completed": completed,
            "partial": partial,
            "pending": pending,
            "error": error,
            "completion_pct": round(completed / len(statuses) * 100, 1) if statuses else 0,
            "srt_coverage": srt_summary,
            "needs_attention": [s for s in statuses if s["status"] in ("PARTIAL", "ERROR")]
        }

    # ========================================================================
    # UTILITY METHODS
    # ========================================================================

    def has_prompts(self) -> bool:
        """Kiểm tra xem đã có prompt nào chưa."""
        scenes = self.get_scenes()
        return any(s.img_prompt for s in scenes)
    
    def get_stats(self) -> Dict[str, Any]:
        """
        Lấy thống kê tổng quan.

        Returns:
            Dictionary chứa các thống kê
        """
        characters = self.get_characters()
        scenes = self.get_scenes()

        # Director plan stats
        try:
            plans = self.get_director_plan()
            total_planned = len(plans)
            plans_done = sum(1 for p in plans if p.get("status") == "done")
        except:
            total_planned = 0
            plans_done = 0

        # Count invalid prompts
        invalid_prompts = self.detect_invalid_prompts()

        return {
            "total_characters": len(characters),
            "total_scenes": len(scenes),
            "scenes_with_prompts": sum(1 for s in scenes if s.img_prompt),
            "images_done": sum(1 for s in scenes if s.status_img == "done"),
            "images_error": sum(1 for s in scenes if s.status_img == "error"),
            "videos_done": sum(1 for s in scenes if s.status_vid == "done"),
            "videos_error": sum(1 for s in scenes if s.status_vid == "error"),
            # Director plan
            "total_planned": total_planned,
            "plans_done": plans_done,
            "scenes_missing": total_planned - len(scenes) if total_planned > 0 else 0,
            # Invalid prompts
            "invalid_prompts": len(invalid_prompts),
            "invalid_scene_ids": [s.scene_id for s in invalid_prompts],
        }

    def detect_invalid_prompts(self) -> List:
        """
        Detect scenes có prompts bị lỗi (chỉ có template, không có nội dung).

        Returns:
            List các Scene objects có prompts invalid
        """
        # Template patterns - prompts chỉ có style mà không có nội dung cụ thể
        TEMPLATE_PATTERNS = [
            "cinematic, 4k photorealistic",
            "shot on arri alexa",
            "medium shot, dramatic scene",
            "subtle film grain",
            "natural lighting with dramatic shadows",
        ]

        # Minimum meaningful content length (sau khi loại bỏ template text)
        MIN_CONTENT_LENGTH = 50

        invalid_scenes = []
        scenes = self.get_scenes()

        for scene in scenes:
            prompt = (scene.img_prompt or "").lower().strip()

            if not prompt:
                invalid_scenes.append(scene)
                continue

            # Loại bỏ template text để kiểm tra nội dung thực
            content = prompt
            for pattern in TEMPLATE_PATTERNS:
                content = content.replace(pattern.lower(), "")

            # Loại bỏ các ký tự thừa
            content = content.replace(",", " ").replace(".", " ").strip()
            content = " ".join(content.split())  # Normalize whitespace

            # Kiểm tra nếu prompt chỉ có template (không đủ nội dung)
            if len(content) < MIN_CONTENT_LENGTH:
                invalid_scenes.append(scene)

        return invalid_scenes

    def fix_invalid_prompts_from_backup(self) -> Dict[str, int]:
        """
        Tự động fix prompts bị lỗi bằng cách lấy từ director_plan backup.
        Match theo timestamp (srt_start) hoặc scene_id.

        Returns:
            Dict với số lượng fixed, skipped, no_backup
        """
        result = {"fixed": 0, "skipped": 0, "no_backup": 0}

        # Get invalid scenes và backup plans
        invalid_scenes = self.detect_invalid_prompts()
        if not invalid_scenes:
            return result

        try:
            backup_plans = self.get_director_plan()
        except:
            backup_plans = []

        if not backup_plans:
            result["no_backup"] = len(invalid_scenes)
            return result

        # Build lookup dicts cho backup
        # 1. By scene_id/plan_id
        backup_by_id = {p["plan_id"]: p for p in backup_plans}
        # 2. By timestamp
        backup_by_time = {}
        for p in backup_plans:
            ts = p.get("srt_start", "")
            if ts:
                backup_by_time[ts] = p

        for scene in invalid_scenes:
            backup = None

            # Strategy 1: Match by scene_id
            if scene.scene_id in backup_by_id:
                backup = backup_by_id[scene.scene_id]

            # Strategy 2: Match by timestamp
            if not backup and scene.srt_start:
                backup = backup_by_time.get(scene.srt_start)

            # Strategy 3: Fuzzy timestamp match (within 2 seconds)
            if not backup and scene.srt_start:
                scene_seconds = self._timestamp_to_seconds(scene.srt_start)
                best_match = None
                best_diff = float('inf')

                for p in backup_plans:
                    plan_seconds = self._timestamp_to_seconds(p.get("srt_start", ""))
                    diff = abs(scene_seconds - plan_seconds)
                    if diff < 2 and diff < best_diff:  # Within 2 seconds
                        best_diff = diff
                        best_match = p

                backup = best_match

            if backup and backup.get("img_prompt"):
                backup_prompt = backup["img_prompt"]

                # Validate backup prompt is not also template-only
                if len(backup_prompt) > 80:  # Reasonable prompt length
                    # Update scene with backup data
                    self.update_scene(
                        scene.scene_id,
                        img_prompt=backup_prompt,
                        characters_used=backup.get("characters_used", ""),
                        location_used=backup.get("location_used", ""),
                        reference_files=backup.get("reference_files", "")
                    )
                    result["fixed"] += 1
                    self.logger.info(f"Scene {scene.scene_id}: Fixed from backup")
                else:
                    result["skipped"] += 1
            else:
                result["no_backup"] += 1

        self.save()
        return result

    def _timestamp_to_seconds(self, ts: str) -> float:
        """Convert SRT timestamp (HH:MM:SS,mmm) to seconds."""
        if not ts:
            return 0
        try:
            ts = ts.replace(",", ".")
            parts = ts.split(":")
            if len(parts) == 3:
                h, m, s = parts
                return int(h) * 3600 + int(m) * 60 + float(s)
            elif len(parts) == 2:
                m, s = parts
                return int(m) * 60 + float(s)
            else:
                return float(ts)
        except:
            return 0
