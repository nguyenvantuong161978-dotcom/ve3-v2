#!/usr/bin/env python3
"""
Parallel Reference Generator
============================
Tạo ảnh tham chiếu (nv/loc) song song với N Chrome workers.

Key Features:
- Dynamic work stealing: Worker nào xong grab task tiếp theo
- Thread-safe work queue
- Barrier synchronization: Đợi TẤT CẢ workers xong
- Resume capability: Skip items đã có media_id
- Health monitoring: Auto-restart Chrome khi cần

Architecture:
- ParallelReferenceGenerator: Quản lý N workers
- WorkItem/WorkResult: Type-safe work items
- queue.Queue(): Thread-safe task distribution
- threading.Barrier(): Synchronization point
"""

import threading
import queue
import time
from pathlib import Path
from dataclasses import dataclass
from typing import List, Optional, Dict, Any, Callable
from openpyxl import load_workbook

from modules.drission_flow_api import DrissionFlowAPI
from modules.chrome_manager import ChromeManager


@dataclass
class WorkItem:
    """Một task tạo ảnh tham chiếu"""
    scene_name: str  # e.g., "nv1", "loc3"
    prompt: str
    aspect_ratio: str  # "16:9", "9:16", etc.
    row_idx: int  # Row index trong Excel để update


@dataclass
class WorkResult:
    """Kết quả sau khi tạo ảnh"""
    scene_name: str
    success: bool
    media_ids: List[str]  # Có thể có nhiều ảnh
    error_msg: str
    row_idx: int


class ParallelReferenceGenerator:
    """
    Parallel generator cho reference images (nv/loc).

    Sử dụng N Chrome workers để tạo ảnh song song:
    - Work queue chứa tất cả tasks
    - Worker threads grab tasks từ queue
    - Barrier đảm bảo TẤT CẢ workers hoàn thành
    """

    def __init__(
        self,
        num_workers: int,
        chrome_managers: List[ChromeManager],
        excel_path: Path,
        bearer_token: Optional[str] = None,
        log_callback: Optional[Callable] = None
    ):
        """
        Args:
            num_workers: Số lượng Chrome workers (2, 4, 8, etc.)
            chrome_managers: List of ChromeManager instances (pre-initialized)
            excel_path: Path to Excel file
            bearer_token: Optional bearer token
            log_callback: Optional logging callback
        """
        self.num_workers = num_workers
        self.chrome_managers = chrome_managers
        self.excel_path = excel_path
        self.bearer_token = bearer_token
        self.log_callback = log_callback or print

        # Work queue và results
        self.work_queue: queue.Queue[WorkItem] = queue.Queue()
        self.results: List[WorkResult] = []
        self.results_lock = threading.Lock()

        # Barrier để đồng bộ workers
        self.barrier = threading.Barrier(num_workers)

        # Stop flag
        self.stop_flag = threading.Event()

    def log(self, msg: str, level: str = "INFO"):
        """Log with prefix"""
        prefix = f"[PARALLEL-REF] [{level}]"
        self.log_callback(f"{prefix} {msg}")

    def _load_work_items(self) -> List[WorkItem]:
        """
        Load work items from Excel.
        Chỉ load items chưa có media_id (cần tạo).

        Returns:
            List of WorkItem
        """
        work_items = []

        try:
            wb = load_workbook(self.excel_path)
            ws = wb['scenes']

            for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
                scene_name = row[0].value  # Column A
                if not scene_name:
                    continue

                # Chỉ xử lý reference scenes (nv*, loc*)
                if not (scene_name.startswith('nv') or scene_name.startswith('loc')):
                    continue

                # Check xem đã có media_id chưa
                media_id = row[7].value  # Column H
                if media_id:
                    continue  # Skip - đã có rồi

                # Extract info
                prompt = row[2].value  # Column C
                aspect_ratio = row[6].value or "16:9"  # Column G

                if prompt:
                    work_items.append(WorkItem(
                        scene_name=scene_name,
                        prompt=prompt,
                        aspect_ratio=aspect_ratio,
                        row_idx=row_idx
                    ))

            wb.close()

        except Exception as e:
            self.log(f"Error loading work items: {e}", "ERROR")
            return []

        return work_items

    def _worker_thread(self, worker_id: int):
        """
        Worker thread function.
        Continuously grabs tasks from queue until empty.

        Args:
            worker_id: Worker ID (0-based)
        """
        chrome_mgr = self.chrome_managers[worker_id]
        chrome_api = chrome_mgr.api

        self.log(f"Worker #{worker_id} started")

        while not self.stop_flag.is_set():
            try:
                # Grab next task (non-blocking with timeout)
                try:
                    work_item = self.work_queue.get(timeout=1)
                except queue.Empty:
                    # No more work - done
                    break

                self.log(f"Worker #{worker_id} processing: {work_item.scene_name}")

                # Generate image
                try:
                    media_ids, error_msg = chrome_api.generate_image_forward(
                        prompt=work_item.prompt,
                        aspect_ratio=work_item.aspect_ratio,
                        num_variations=1,  # Reference images chỉ cần 1
                        image_references=[],  # No references for reference images
                        force_model=None,
                        timeout=90
                    )

                    success = bool(media_ids and not error_msg)

                    result = WorkResult(
                        scene_name=work_item.scene_name,
                        success=success,
                        media_ids=media_ids,
                        error_msg=error_msg,
                        row_idx=work_item.row_idx
                    )

                    # Save result (thread-safe)
                    with self.results_lock:
                        self.results.append(result)

                    if success:
                        self.log(f"Worker #{worker_id} ✅ {work_item.scene_name}: {len(media_ids)} image(s)")
                    else:
                        self.log(f"Worker #{worker_id} ❌ {work_item.scene_name}: {error_msg}", "WARN")

                except Exception as e:
                    self.log(f"Worker #{worker_id} exception: {e}", "ERROR")
                    result = WorkResult(
                        scene_name=work_item.scene_name,
                        success=False,
                        media_ids=[],
                        error_msg=str(e),
                        row_idx=work_item.row_idx
                    )
                    with self.results_lock:
                        self.results.append(result)

                finally:
                    self.work_queue.task_done()

            except Exception as e:
                self.log(f"Worker #{worker_id} outer exception: {e}", "ERROR")
                break

        self.log(f"Worker #{worker_id} finished")

        # Wait at barrier for all workers to finish
        try:
            self.barrier.wait(timeout=300)  # 5 min timeout
        except Exception as e:
            self.log(f"Worker #{worker_id} barrier error: {e}", "ERROR")

    def _update_excel_results(self):
        """
        Update Excel with all results.
        Write media_ids back to Excel.
        """
        try:
            wb = load_workbook(self.excel_path)
            ws = wb['scenes']

            for result in self.results:
                if result.success and result.media_ids:
                    # Update media_id in Excel
                    row_idx = result.row_idx
                    media_id_str = ",".join(result.media_ids)
                    ws.cell(row=row_idx, column=8).value = media_id_str  # Column H

                    self.log(f"Excel updated: {result.scene_name} → {media_id_str}")

            wb.save(self.excel_path)
            wb.close()

            self.log("Excel updated successfully")

        except Exception as e:
            self.log(f"Error updating Excel: {e}", "ERROR")

    def generate_all(self) -> bool:
        """
        Generate all reference images in parallel.

        Returns:
            True if all successful, False if any errors
        """
        # 1. Load work items
        self.log("Loading work items from Excel...")
        work_items = self._load_work_items()

        if not work_items:
            self.log("No work items to process (all references already created)")
            return True

        self.log(f"Found {len(work_items)} reference image(s) to create")

        # 2. Add to queue
        for item in work_items:
            self.work_queue.put(item)

        # 3. Start worker threads
        threads = []
        for worker_id in range(self.num_workers):
            thread = threading.Thread(
                target=self._worker_thread,
                args=(worker_id,),
                name=f"RefWorker-{worker_id}"
            )
            thread.start()
            threads.append(thread)

        self.log(f"Started {self.num_workers} worker threads")

        # 4. Wait for all threads to complete
        for thread in threads:
            thread.join()

        self.log("All workers finished")

        # 5. Update Excel with results
        self._update_excel_results()

        # 6. Check results
        total = len(self.results)
        successful = sum(1 for r in self.results if r.success)
        failed = total - successful

        self.log(f"Results: {successful}/{total} successful, {failed} failed")

        if failed > 0:
            self.log(f"Failed scenes:", "WARN")
            for r in self.results:
                if not r.success:
                    self.log(f"  - {r.scene_name}: {r.error_msg}", "WARN")

        return failed == 0
