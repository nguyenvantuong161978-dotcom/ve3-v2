"""
VE3 Tool - Flow Image Generator Module
======================================
Tích hợp Google Flow API vào pipeline để tự động tạo ảnh từ Excel prompts.

Workflow:
1. Đọc Excel prompts (characters + scenes sheets)
2. Tạo ảnh NV (nhân vật) trước - lưu vào thư mục nv/
3. Tạo ảnh scenes sau - lưu vào thư mục img/
4. Cập nhật Excel với đường dẫn ảnh và status
"""

import sys
import os

# Fix Windows encoding issues
if sys.platform == "win32":
    if sys.stdout and hasattr(sys.stdout, 'reconfigure'):
        try:
            sys.stdout.reconfigure(encoding='utf-8', errors='replace')
        except:
            pass
    if sys.stderr and hasattr(sys.stderr, 'reconfigure'):
        try:
            sys.stderr.reconfigure(encoding='utf-8', errors='replace')
        except:
            pass


import os
import time
import yaml
from pathlib import Path
from typing import Optional, List, Dict, Tuple, Any
from datetime import datetime
import openpyxl
from openpyxl import load_workbook

from .google_flow_api import GoogleFlowAPI, AspectRatio


class FlowImageGenerator:
    """
    Generator ảnh sử dụng Google Flow API.
    Đọc prompts từ Excel và tạo ảnh tự động.
    """
    
    def __init__(
        self,
        project_path: Path,
        bearer_token: str,
        project_id: Optional[str] = None,
        aspect_ratio: str = "landscape",
        delay_between_requests: float = 3.0,
        verbose: bool = True
    ):
        """
        Khởi tạo Flow Image Generator.
        
        Args:
            project_path: Đường dẫn đến thư mục project (PROJECTS/{CODE}/)
            bearer_token: Google Flow Bearer token
            project_id: Flow Project ID (optional)
            aspect_ratio: Tỷ lệ khung hình (landscape/portrait/square)
            delay_between_requests: Thời gian chờ giữa các request (giây)
            verbose: In log chi tiết
        """
        self.project_path = Path(project_path)
        self.bearer_token = bearer_token
        self.project_id = project_id
        self.delay = delay_between_requests
        self.verbose = verbose
        
        # Map aspect ratio
        ar_map = {
            "landscape": AspectRatio.LANDSCAPE,
            "portrait": AspectRatio.PORTRAIT,
            "square": AspectRatio.SQUARE,
            "16:9": AspectRatio.LANDSCAPE,
            "9:16": AspectRatio.PORTRAIT,
            "1:1": AspectRatio.SQUARE,
        }
        self.aspect_ratio = ar_map.get(aspect_ratio.lower(), AspectRatio.LANDSCAPE)
        
        # Tạo Flow API client
        self.flow_client = GoogleFlowAPI(
            bearer_token=bearer_token,
            project_id=project_id,
            verbose=verbose
        )
        
        # Paths
        self.nv_path = self.project_path / "nv"
        self.img_path = self.project_path / "img"
        self.prompts_path = self.project_path / "prompts"
        
        # Tạo thư mục nếu chưa có
        self.nv_path.mkdir(parents=True, exist_ok=True)
        self.img_path.mkdir(parents=True, exist_ok=True)
        
        # Stats
        self.stats = {
            "characters_total": 0,
            "characters_success": 0,
            "characters_failed": 0,
            "scenes_total": 0,
            "scenes_success": 0,
            "scenes_failed": 0,
        }
    
    def _log(self, message: str) -> None:
        """Print log message."""
        if self.verbose:
            timestamp = datetime.now().strftime("%H:%M:%S")
            print(f"[{timestamp}] {message}")
    
    def _find_excel_file(self) -> Optional[Path]:
        """Tìm file Excel prompts trong thư mục project."""
        # Tìm trong thư mục prompts/
        for pattern in ["*_prompts.xlsx", "*.xlsx"]:
            files = list(self.prompts_path.glob(pattern))
            if files:
                return files[0]
        
        # Tìm trực tiếp trong project
        for pattern in ["*_prompts.xlsx", "*.xlsx"]:
            files = list(self.project_path.glob(pattern))
            if files:
                return files[0]
        
        return None
    
    def generate_character_images(
        self,
        excel_path: Optional[Path] = None,
        overwrite: bool = False
    ) -> Tuple[int, int, List[str]]:
        """
        Tạo ảnh cho tất cả nhân vật trong sheet "characters".
        
        Args:
            excel_path: Đường dẫn file Excel (tự tìm nếu không chỉ định)
            overwrite: Ghi đè ảnh đã có
            
        Returns:
            Tuple[success_count, failed_count, error_messages]
        """
        self._log("=" * 60)
        self._log("GENERATING CHARACTER IMAGES")
        self._log("=" * 60)
        
        # Tìm file Excel
        if excel_path is None:
            excel_path = self._find_excel_file()
        
        if excel_path is None or not excel_path.exists():
            return 0, 0, ["Excel file not found"]
        
        self._log(f"Excel file: {excel_path}")
        
        errors = []
        success_count = 0
        failed_count = 0
        
        try:
            # Load workbook
            wb = load_workbook(excel_path)
            
            if "characters" not in wb.sheetnames:
                return 0, 0, ["Sheet 'characters' not found in Excel"]
            
            ws = wb["characters"]
            
            # Get header row
            headers = [cell.value for cell in ws[1]]
            
            # Find column indices
            col_idx = {
                "id": headers.index("id") if "id" in headers else -1,
                "english_prompt": headers.index("english_prompt") if "english_prompt" in headers else -1,
                "image_file": headers.index("image_file") if "image_file" in headers else -1,
                "status": headers.index("status") if "status" in headers else -1,
            }
            
            if col_idx["english_prompt"] == -1:
                return 0, 0, ["Column 'english_prompt' not found"]
            
            # Process each character
            for row_num, row in enumerate(ws.iter_rows(min_row=2), start=2):
                char_id = row[col_idx["id"]].value if col_idx["id"] >= 0 else f"char_{row_num}"
                prompt = row[col_idx["english_prompt"]].value
                image_file = row[col_idx["image_file"]].value if col_idx["image_file"] >= 0 else f"{char_id}.png"
                status = row[col_idx["status"]].value if col_idx["status"] >= 0 else "pending"
                
                if not prompt:
                    continue

                # Skip children (status="skip" or english_prompt="DO_NOT_GENERATE")
                if status == "skip" or prompt == "DO_NOT_GENERATE":
                    self._log(f"  [SKIP]  {char_id}: Child character, skipping (will use inline description)")
                    continue

                self.stats["characters_total"] += 1

                # Check if already done
                output_file = self.nv_path / image_file
                if output_file.exists() and not overwrite:
                    if status == "done":
                        self._log(f"  [SKIP]  {char_id}: Already done, skipping")
                        success_count += 1
                        self.stats["characters_success"] += 1
                        continue

                self._log(f"\n[GEN] Generating image for character: {char_id}")
                self._log(f"   Prompt: {prompt[:80]}...")
                
                # Generate image
                success, images, error = self.flow_client.generate_images(
                    prompt=prompt,
                    count=1,
                    aspect_ratio=self.aspect_ratio
                )
                
                if success and images:
                    # Download image
                    filename = image_file.replace(".png", "")
                    downloaded = self.flow_client.download_image(
                        images[0],
                        self.nv_path,
                        filename
                    )
                    
                    if downloaded:
                        self._log(f"   [OK] Saved to: {downloaded}")
                        success_count += 1
                        self.stats["characters_success"] += 1
                        
                        # Update status in Excel
                        if col_idx["status"] >= 0:
                            row[col_idx["status"]].value = "done"
                    else:
                        self._log(f"   [FAIL] Download failed")
                        failed_count += 1
                        self.stats["characters_failed"] += 1
                        errors.append(f"{char_id}: Download failed")
                else:
                    self._log(f"   [FAIL] Generation failed: {error}")
                    failed_count += 1
                    self.stats["characters_failed"] += 1
                    errors.append(f"{char_id}: {error}")
                
                # Delay between requests
                if self.delay > 0:
                    time.sleep(self.delay)
            
            # Save workbook
            wb.save(excel_path)
            self._log(f"\n[SAVE] Excel updated: {excel_path}")
            
        except Exception as e:
            errors.append(f"Excel error: {str(e)}")
            self._log(f"[FAIL] Error: {e}")
        
        self._log(f"\n[STATS] Characters: {success_count} success, {failed_count} failed")
        return success_count, failed_count, errors
    
    def generate_scene_images(
        self,
        excel_path: Optional[Path] = None,
        start_scene: int = 1,
        end_scene: Optional[int] = None,
        overwrite: bool = False
    ) -> Tuple[int, int, List[str]]:
        """
        Tạo ảnh cho các scenes trong sheet "scenes".
        
        Args:
            excel_path: Đường dẫn file Excel
            start_scene: Scene bắt đầu (1-indexed)
            end_scene: Scene kết thúc (None = tất cả)
            overwrite: Ghi đè ảnh đã có
            
        Returns:
            Tuple[success_count, failed_count, error_messages]
        """
        self._log("=" * 60)
        self._log("GENERATING SCENE IMAGES")
        self._log("=" * 60)
        
        # Tìm file Excel
        if excel_path is None:
            excel_path = self._find_excel_file()
        
        if excel_path is None or not excel_path.exists():
            return 0, 0, ["Excel file not found"]
        
        self._log(f"Excel file: {excel_path}")
        
        errors = []
        success_count = 0
        failed_count = 0
        
        try:
            # Load workbook
            wb = load_workbook(excel_path)
            
            if "scenes" not in wb.sheetnames:
                return 0, 0, ["Sheet 'scenes' not found in Excel"]
            
            ws = wb["scenes"]
            
            # Get header row
            headers = [cell.value for cell in ws[1]]
            
            # Find column indices
            col_idx = {
                "scene_id": headers.index("scene_id") if "scene_id" in headers else -1,
                "img_prompt": headers.index("img_prompt") if "img_prompt" in headers else -1,
                "img_path": headers.index("img_path") if "img_path" in headers else -1,
                "status_img": headers.index("status_img") if "status_img" in headers else -1,
            }
            
            if col_idx["img_prompt"] == -1:
                return 0, 0, ["Column 'img_prompt' not found"]
            
            # Process each scene
            for row_num, row in enumerate(ws.iter_rows(min_row=2), start=2):
                scene_id = row[col_idx["scene_id"]].value if col_idx["scene_id"] >= 0 else row_num - 1
                
                # Filter by scene range
                if isinstance(scene_id, int):
                    if scene_id < start_scene:
                        continue
                    if end_scene is not None and scene_id > end_scene:
                        break
                
                prompt = row[col_idx["img_prompt"]].value
                img_path_val = row[col_idx["img_path"]].value if col_idx["img_path"] >= 0 else None
                status = row[col_idx["status_img"]].value if col_idx["status_img"] >= 0 else "pending"
                
                if not prompt:
                    continue
                
                self.stats["scenes_total"] += 1
                
                # Generate filename
                filename = f"scene_{scene_id:03d}"
                output_file = self.img_path / f"{filename}.png"
                
                # Check if already done
                if output_file.exists() and not overwrite:
                    if status == "done":
                        self._log(f"  [SKIP]  Scene {scene_id}: Already done, skipping")
                        success_count += 1
                        self.stats["scenes_success"] += 1
                        continue
                
                self._log(f"\n[VIDEO] Generating image for Scene {scene_id}")
                self._log(f"   Prompt: {prompt[:100]}...")
                
                # Generate image
                success, images, error = self.flow_client.generate_images(
                    prompt=prompt,
                    count=1,
                    aspect_ratio=self.aspect_ratio
                )
                
                if success and images:
                    # Download image
                    downloaded = self.flow_client.download_image(
                        images[0],
                        self.img_path,
                        filename
                    )
                    
                    if downloaded:
                        self._log(f"   [OK] Saved to: {downloaded}")
                        success_count += 1
                        self.stats["scenes_success"] += 1
                        
                        # Update Excel
                        if col_idx["img_path"] >= 0:
                            row[col_idx["img_path"]].value = str(downloaded.relative_to(self.project_path))
                        if col_idx["status_img"] >= 0:
                            row[col_idx["status_img"]].value = "done"
                    else:
                        self._log(f"   [FAIL] Download failed")
                        failed_count += 1
                        self.stats["scenes_failed"] += 1
                        errors.append(f"Scene {scene_id}: Download failed")
                else:
                    self._log(f"   [FAIL] Generation failed: {error}")
                    failed_count += 1
                    self.stats["scenes_failed"] += 1
                    errors.append(f"Scene {scene_id}: {error}")
                
                # Delay between requests
                if self.delay > 0:
                    time.sleep(self.delay)
            
            # Save workbook
            wb.save(excel_path)
            self._log(f"\n[SAVE] Excel updated: {excel_path}")
            
        except Exception as e:
            errors.append(f"Excel error: {str(e)}")
            self._log(f"[FAIL] Error: {e}")
        
        self._log(f"\n[STATS] Scenes: {success_count} success, {failed_count} failed")
        return success_count, failed_count, errors
    
    def generate_all(
        self,
        excel_path: Optional[Path] = None,
        characters: bool = True,
        scenes: bool = True,
        start_scene: int = 1,
        end_scene: Optional[int] = None,
        overwrite: bool = False
    ) -> Dict[str, Any]:
        """
        Tạo tất cả ảnh (characters + scenes).
        
        Args:
            excel_path: Đường dẫn file Excel
            characters: Tạo ảnh characters
            scenes: Tạo ảnh scenes
            start_scene: Scene bắt đầu
            end_scene: Scene kết thúc
            overwrite: Ghi đè
            
        Returns:
            Dict với kết quả
        """
        results = {
            "characters": {"success": 0, "failed": 0, "errors": []},
            "scenes": {"success": 0, "failed": 0, "errors": []},
        }
        
        if characters:
            s, f, e = self.generate_character_images(excel_path, overwrite)
            results["characters"] = {"success": s, "failed": f, "errors": e}
        
        if scenes:
            s, f, e = self.generate_scene_images(excel_path, start_scene, end_scene, overwrite)
            results["scenes"] = {"success": s, "failed": f, "errors": e}
        
        # Print summary
        self._log("\n" + "=" * 60)
        self._log("SUMMARY")
        self._log("=" * 60)
        self._log(f"Characters: {results['characters']['success']} success, {results['characters']['failed']} failed")
        self._log(f"Scenes: {results['scenes']['success']} success, {results['scenes']['failed']} failed")
        
        return results
    
    def get_stats(self) -> Dict[str, int]:
        """Lấy thống kê."""
        return self.stats.copy()


# =============================================================================
# HELPER FUNCTIONS
# =============================================================================

def load_config(config_path: str = "config/settings.yaml") -> Dict[str, Any]:
    """Load config từ file YAML."""
    with open(config_path, "r", encoding="utf-8") as f:
        return yaml.safe_load(f)


def create_generator_from_config(
    project_path: str,
    config_path: str = "config/settings.yaml",
    verbose: bool = True
) -> FlowImageGenerator:
    """
    Tạo FlowImageGenerator từ config file.
    
    Args:
        project_path: Đường dẫn thư mục project
        config_path: Đường dẫn file config
        verbose: In log
        
    Returns:
        FlowImageGenerator instance
    """
    config = load_config(config_path)
    
    return FlowImageGenerator(
        project_path=Path(project_path),
        bearer_token=config.get("flow_bearer_token", ""),
        project_id=config.get("flow_project_id"),
        aspect_ratio=config.get("flow_aspect_ratio", "landscape"),
        delay_between_requests=config.get("flow_delay", 3.0),
        verbose=verbose
    )


# =============================================================================
# CLI
# =============================================================================

if __name__ == "__main__":
    import sys
    
    print("""
╔══════════════════════════════════════════════════════════════╗
║           FLOW IMAGE GENERATOR - VE3 TOOL                   ║
╠══════════════════════════════════════════════════════════════╣
║  Usage:                                                      ║
║    python flow_image_generator.py <project_path> [options]   ║
║                                                              ║
║  Options:                                                    ║
║    --characters    Generate character images only            ║
║    --scenes        Generate scene images only                ║
║    --all           Generate all (default)                    ║
║    --start N       Start from scene N                        ║
║    --end N         End at scene N                            ║
║    --overwrite     Overwrite existing images                 ║
╚══════════════════════════════════════════════════════════════╝
""")
    
    if len(sys.argv) < 2:
        print("Error: Please provide project path")
        print("Example: python flow_image_generator.py ./PROJECTS/KA1-0001")
        sys.exit(1)
    
    project_path = sys.argv[1]
    
    # Parse options
    do_characters = "--all" in sys.argv or "--characters" in sys.argv or (
        "--scenes" not in sys.argv and "--characters" not in sys.argv
    )
    do_scenes = "--all" in sys.argv or "--scenes" in sys.argv or (
        "--scenes" not in sys.argv and "--characters" not in sys.argv
    )
    overwrite = "--overwrite" in sys.argv
    
    start_scene = 1
    end_scene = None
    
    for i, arg in enumerate(sys.argv):
        if arg == "--start" and i + 1 < len(sys.argv):
            start_scene = int(sys.argv[i + 1])
        if arg == "--end" and i + 1 < len(sys.argv):
            end_scene = int(sys.argv[i + 1])
    
    # Create generator
    try:
        generator = create_generator_from_config(project_path)
        
        # Run
        results = generator.generate_all(
            characters=do_characters,
            scenes=do_scenes,
            start_scene=start_scene,
            end_scene=end_scene,
            overwrite=overwrite
        )
        
        # Exit code
        total_failed = results["characters"]["failed"] + results["scenes"]["failed"]
        sys.exit(0 if total_failed == 0 else 1)
        
    except FileNotFoundError as e:
        print(f"[FAIL] Config file not found: {e}")
        sys.exit(1)
    except Exception as e:
        print(f"[FAIL] Error: {e}")
        sys.exit(1)
