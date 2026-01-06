"""
VE3 Tool - Browser Flow Generator Module
========================================
Tich hop browser automation voi Excel workflow.

Workflow:
1. Doc Excel prompts (PromptWorkbook)
2. Lay cac scenes chua tao anh (status_img != 'done')
3. Mo trinh duyet, inject JS
4. Goi VE3.run() voi [{sceneId, prompt}]
5. Di chuyen file tu Downloads -> project/img/
6. Cap nhat Excel (img_path, status_img = 'done')

Usage:
    from modules.browser_flow_generator import BrowserFlowGenerator

    gen = BrowserFlowGenerator("PROJECTS/KA1-0001")
    gen.generate_all_images()
"""

import os
import sys
import time
import json
import shutil
import glob
import base64
from pathlib import Path
from typing import Optional, List, Dict, Any, Tuple
from datetime import datetime

# Import PromptWorkbook
from modules.excel_manager import PromptWorkbook, Scene
from modules.utils import get_logger, load_settings

# Browser driver imports - PREFER SELENIUM (more stable)
DRIVER_TYPE = None

try:
    from selenium import webdriver
    from selenium.webdriver.common.by import By
    from selenium.webdriver.chrome.options import Options
    from selenium.webdriver.chrome.service import Service
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    from selenium.common.exceptions import (
        TimeoutException,
        WebDriverException,
        JavascriptException
    )
    DRIVER_TYPE = "selenium"
except ImportError:
    DRIVER_TYPE = None

SELENIUM_AVAILABLE = DRIVER_TYPE is not None


class BrowserFlowGenerator:
    """
    Tao anh tu Excel bang browser automation.

    Su dung JavaScript injection de dieu khien Google Flow.
    Tu dong di chuyen file tu Downloads va cap nhat Excel.
    """

    FLOW_URL = "https://labs.google/fx/vi/tools/flow"

    def __init__(
        self,
        project_path: str,
        profile_name: str = "main",
        headless: bool = False,
        verbose: bool = True,
        config_path: str = "config/settings.yaml",
        worker_id: int = 0
    ):
        """
        Khoi tao BrowserFlowGenerator.

        Args:
            project_path: Duong dan den thu muc project (PROJECTS/KA1-0001)
            profile_name: Ten Chrome profile (luu trong chrome_profiles/)
            headless: Chay an (khong hien UI) - nen False lan dau de dang nhap
            verbose: In log chi tiet
            config_path: Duong dan file config
            worker_id: Worker ID for parallel processing (affects proxy, Chrome port)
        """
        if not SELENIUM_AVAILABLE:
            raise ImportError(
                "Selenium chua duoc cai dat. "
                "Chay: pip install selenium undetected-chromedriver"
            )

        self.project_path = Path(project_path)
        self.profile_name = profile_name
        self.headless = headless
        self.verbose = verbose
        self.worker_id = worker_id  # For parallel processing

        # Load config
        self.config = {}
        config_file = Path(config_path)
        if config_file.exists():
            self.config = load_settings(config_file)  # Pass Path object, not string

        # Paths
        self.img_path = self.project_path / "img"
        self.prompts_path = self.project_path / "prompts"
        self.nv_path = self.project_path / "nv"

        # Tao thu muc neu chua co
        self.img_path.mkdir(parents=True, exist_ok=True)
        self.nv_path.mkdir(parents=True, exist_ok=True)

        # Chrome profile
        base_dir = Path(__file__).parent.parent
        profiles_dir = self.config.get("browser_profiles_dir", "./chrome_profiles")
        if not os.path.isabs(profiles_dir):
            profiles_dir = base_dir / profiles_dir
        self.profile_dir = Path(profiles_dir) / profile_name
        self.profile_dir.mkdir(parents=True, exist_ok=True)

        # Downloads folder - noi browser tai anh ve
        self.downloads_dir = Path.home() / "Downloads"

        # Project code (dung cho ten file)
        self.project_code = self.project_path.name  # VD: KA1-0001

        # Driver
        self.driver = None
        self._js_injected = False
        self._project_url = ""  # Luu project URL de giu nguyen phien lam viec

        # Logger
        self.logger = get_logger("browser_flow")

        # Stats
        self.stats = {
            "total": 0,
            "success": 0,
            "failed": 0,
            "skipped": 0,
        }

    def _log(self, message: str, level: str = "info") -> None:
        """Print log message."""
        if self.verbose:
            timestamp = datetime.now().strftime("%H:%M:%S")
            icons = {
                "info": "[INFO]",
                "success": "[OK]",
                "error": "[ERROR]",
                "warn": "[WARN]",
            }
            print(f"[{timestamp}] {icons.get(level, '')} {message}")

    def _get_profile_path(self) -> Optional[str]:
        """
        Lấy Chrome profile path cho worker này.
        Mỗi worker cần profile riêng để chạy song song.
        """
        base_dir = Path(__file__).parent.parent
        profiles_dir = base_dir / "chrome_profiles"

        # === PARALLEL MODE: Mỗi worker dùng profile khác nhau ===
        if profiles_dir.exists():
            # Lấy danh sách profiles có sẵn (đã đăng nhập từ GUI)
            available_profiles = sorted([
                p for p in profiles_dir.iterdir()
                if p.is_dir() and not p.name.startswith('.')
            ])

            if available_profiles:
                # Worker 0 → profile[0], Worker 1 → profile[1], ...
                worker_id = getattr(self, 'worker_id', 0) or 0
                profile_idx = worker_id % len(available_profiles)
                selected_profile = available_profiles[profile_idx]
                self._log(f"[Worker {worker_id}] Dùng profile: {selected_profile.name}")
                return str(selected_profile)

        # Fallback: từ settings.yaml
        chrome_profile = self.config.get('chrome_profile', '')
        if chrome_profile:
            profile_path = Path(chrome_profile)
            if profile_path.exists():
                return str(profile_path)
            abs_path = Path.cwd() / chrome_profile
            if abs_path.exists():
                return str(abs_path)

        # Fallback: dùng tool profile
        if hasattr(self, 'profile_dir') and self.profile_dir:
            return str(self.profile_dir)

        return None

    def _find_excel_file(self) -> Optional[Path]:
        """Tim file Excel prompts trong project."""
        # Tim trong prompts/
        for pattern in ["*_prompts.xlsx", "*.xlsx"]:
            files = list(self.prompts_path.glob(pattern))
            if files:
                return files[0]

        # Tim trong project root
        for pattern in ["*_prompts.xlsx", "*.xlsx"]:
            files = list(self.project_path.glob(pattern))
            if files:
                return files[0]

        return None

    def _get_js_script(self) -> str:
        """Doc file JavaScript automation."""
        script_path = Path(__file__).parent.parent / "scripts" / "ve3_browser_automation.js"

        if script_path.exists():
            with open(script_path, "r", encoding="utf-8") as f:
                return f.read()

        raise FileNotFoundError(f"JS script khong tim thay: {script_path}")

    def _create_driver(self):
        """
        Tao Chrome WebDriver - PARALLEL SAFE.
        - Moi instance dung port rieng (khong xung dot)
        - Dung working profile rieng (giu nguyen settings nhu download permission)
        - Mac dinh headless (chay an)
        """
        import random

        # Download prefs
        prefs = {
            "download.default_directory": str(self.downloads_dir),
            "download.prompt_for_download": False,
            "download.directory_upgrade": True,
            "safebrowsing.enabled": True,
        }

        self._log(f"Headless: {self.headless}")
        self._log("Su dung Selenium WebDriver (Parallel Safe)")

        try:
            options = Options()

            # Tim Chrome binary
            chrome_paths = [
                r"C:\Program Files\Google\Chrome\Application\chrome.exe",
                r"C:\Program Files (x86)\Google\Chrome\Application\chrome.exe",
                "/usr/bin/google-chrome",
                "/usr/bin/chromium-browser",
            ]
            for chrome_path in chrome_paths:
                if os.path.exists(chrome_path):
                    options.binary_location = chrome_path
                    self._log(f"Chrome: {chrome_path}")
                    break

            # Tao working profile rieng (khong phai temp, giu nguyen settings)
            # Vi tri: ~/.ve3_chrome_profiles/{profile_name}
            self.profile_dir.mkdir(parents=True, exist_ok=True)
            working_profile_base = Path.home() / ".ve3_chrome_profiles"
            working_profile_base.mkdir(parents=True, exist_ok=True)
            working_profile = working_profile_base / self.profile_name

            self._log(f"Profile goc: {self.profile_dir}")
            self._log(f"Working profile: {working_profile}")

            # LUÔN sync cookies/login data từ profile gốc (đảm bảo dùng đúng account đã đăng nhập)
            import shutil
            if not working_profile.exists():
                working_profile.mkdir(parents=True, exist_ok=True)

            # Sync các file quan trọng (cookies, login data) mỗi lần chạy
            critical_files = [
                "Cookies", "Login Data", "Web Data",
                "Network/Cookies", "Network/TransportSecurity"
            ]
            critical_dirs = ["Default", "Network"]

            if any(self.profile_dir.iterdir()):
                # Sync critical dirs first
                for dir_name in critical_dirs:
                    src_dir = self.profile_dir / dir_name
                    if src_dir.exists() and src_dir.is_dir():
                        dest_dir = working_profile / dir_name
                        dest_dir.mkdir(parents=True, exist_ok=True)
                        for item in src_dir.iterdir():
                            try:
                                dest = dest_dir / item.name
                                if item.is_file():
                                    shutil.copy2(item, dest)
                            except Exception:
                                pass  # Skip locked files

                # Sync root level files
                for item in self.profile_dir.iterdir():
                    try:
                        dest = working_profile / item.name
                        if item.is_file():
                            shutil.copy2(item, dest)
                        elif item.is_dir() and item.name not in critical_dirs:
                            if not dest.exists():
                                shutil.copytree(item, dest, dirs_exist_ok=True)
                    except Exception:
                        pass  # Skip locked files

                self._log(f"Da sync profile data tu profile goc")

            self._working_profile = str(working_profile)  # Luu de reference
            options.add_argument(f"--user-data-dir={working_profile}")

            # PARALLEL SAFE: Moi instance dung port rieng
            debug_port = random.randint(9222, 9999)
            options.add_argument(f"--remote-debugging-port={debug_port}")

            # Headless mac dinh (chay an, toi uu cho auto)
            if self.headless:
                options.add_argument("--headless=new")
                options.add_argument("--disable-gpu")

            # Cac options cho automation
            options.add_argument("--no-sandbox")
            options.add_argument("--disable-dev-shm-usage")
            options.add_argument("--disable-blink-features=AutomationControlled")
            options.add_argument("--disable-extensions")
            options.add_argument("--disable-infobars")
            options.add_argument("--window-size=1920,1080")
            options.add_experimental_option("excludeSwitches", ["enable-automation", "enable-logging"])
            options.add_experimental_option("useAutomationExtension", False)
            options.add_experimental_option("prefs", prefs)

            self._log(f"Dang khoi dong Chrome (port {debug_port})...")
            driver = webdriver.Chrome(options=options)

            # An webdriver flag
            driver.execute_script(
                "Object.defineProperty(navigator, 'webdriver', {get: () => undefined})"
            )

            self._log("Chrome da san sang!", "success")

            return driver

        except Exception as e:
            self._log(f"Loi khoi dong Chrome: {e}", "error")
            import traceback
            traceback.print_exc()
            raise

    def start_browser(self, use_cached_project: bool = True) -> bool:
        """
        Khoi dong trinh duyet va navigate den Google Flow.

        Args:
            use_cached_project: Neu True, se vao project cu (tu cache) de giu media_name valid

        QUAN TRONG: Khi tao img, phai vao DUNG project da tao nv
        de media_name cua nv con valid cho reference.

        Returns:
            True neu thanh cong
        """
        self._log("Khoi dong trinh duyet...")

        try:
            self.driver = self._create_driver()
            self._log("Da khoi dong Chrome", "success")

            # Tang timeout cho async script (mac dinh 30s, can nhieu hon cho upload/generate)
            self.driver.set_script_timeout(300)  # 5 phut
            self._log("Set script timeout: 300s")

            # QUAN TRONG: Kiem tra cached project URL
            # Neu co, vao project cu de giu media_name valid
            target_url = self.FLOW_URL
            cached_url = getattr(self, '_cached_project_url', None)

            if use_cached_project and cached_url:
                self._log(f"[REUSE PROJECT] Vao project cu de giu media_name valid")
                self._log(f"  -> URL: {cached_url[:60]}...")
                target_url = cached_url
            else:
                self._log(f"Navigate den: {self.FLOW_URL}")

            self.driver.get(target_url)

            # Cho page load
            time.sleep(5)

            return True

        except Exception as e:
            self._log(f"Loi khoi dong: {e}", "error")
            return False

    def stop_browser(self) -> None:
        """Dong trinh duyet (giu nguyen working profile de luu settings)."""
        if self.driver:
            self._log("Dong trinh duyet...")
            try:
                self.driver.quit()
            except:
                pass
            self.driver = None
            self._js_injected = False
            # Khong xoa working profile - giu nguyen de luu settings (download permission, etc.)

    def wait_for_login(self, timeout: int = 300) -> bool:
        """
        Cho nguoi dung dang nhap.

        Args:
            timeout: Thoi gian cho (giay)

        Returns:
            True neu da dang nhap
        """
        self._log(f"Cho dang nhap (timeout: {timeout}s)...")
        self._log("Vui long dang nhap Google account tren trinh duyet")

        start_time = time.time()

        while time.time() - start_time < timeout:
            try:
                textarea = self.driver.find_element(By.CSS_SELECTOR, "textarea")
                if textarea:
                    self._log("Da phat hien dang nhap thanh cong!", "success")
                    return True
            except:
                pass

            time.sleep(2)

        self._log("Timeout - chua dang nhap", "error")
        return False

    def _inject_js(self) -> bool:
        """Inject JavaScript automation script."""
        if self._js_injected:
            return True

        try:
            self._log("Buoc 1: Inject JavaScript script...")
            js_code = self._get_js_script()
            self.driver.execute_script(js_code)

            # Init VE3 voi project name
            self._log("Buoc 2: Init VE3...")
            self.driver.execute_script(f'VE3.init("{self.project_code}")')

            # Setup UI: Click "Du an moi" + Chon "Tao hinh anh"
            self._log("Buoc 3: Setup UI (Du an moi + Tao hinh anh)...")
            setup_result = self.driver.execute_async_script("""
                const callback = arguments[arguments.length - 1];
                (async () => {
                    try {
                        await VE3.setup();
                        callback({success: true});
                    } catch(e) {
                        callback({success: false, error: e.message});
                    }
                })();
            """)

            if setup_result and setup_result.get('success'):
                self._log("Setup UI thanh cong!", "success")
                # Luu project URL de giu nguyen phien
                self._project_url = self._get_project_url_from_js()
                if self._project_url:
                    self._log(f"Project URL: {self._project_url}", "info")
                    # === QUAN TRONG: Extract project_id va luu vao config ===
                    # De API mode co the dung chung project
                    if '/project/' in self._project_url:
                        try:
                            project_id = self._project_url.split('/project/')[1].split('/')[0].split('?')[0]
                            if project_id:
                                self.config['flow_project_id'] = project_id
                                self._log(f"  -> Saved project_id: {project_id[:8]}...", "info")
                        except Exception as e:
                            self._log(f"  -> Could not extract project_id: {e}", "warn")
            else:
                error = setup_result.get('error', 'Unknown') if setup_result else 'No response'
                self._log(f"Setup UI that bai: {error}", "warn")

            self._js_injected = True
            self._log("Da san sang tao anh!", "success")
            return True

        except Exception as e:
            self._log(f"Loi inject JS: {e}", "error")
            import traceback
            traceback.print_exc()
            return False

    # =========================================================================
    # MEDIA NAMES CACHE - Luu media_name de reference
    # =========================================================================

    def _get_media_cache_path(self) -> Path:
        """Duong dan file cache media_names."""
        return self.project_path / "prompts" / ".media_cache.json"

    def _load_media_cache(self) -> Dict[str, Any]:
        """
        Load media_names tu cache file.

        Format moi: {
            "_project_url": "https://...",
            "_project_id": "xxx",
            "_bearer_token": "ya29.xxx",
            "_token_time": 1234567890.0,
            "nvc": {mediaName: str, seed: int|null},
            ...
        }
        Backward compatible voi format cu: {id: str}
        """
        cache_path = self._get_media_cache_path()
        if cache_path.exists():
            try:
                with open(cache_path, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    # Extract project info
                    project_url = data.pop('_project_url', None)
                    project_id = data.pop('_project_id', None)
                    bearer_token = data.pop('_bearer_token', None)
                    token_time = data.pop('_token_time', None)

                    if project_url:
                        self._cached_project_url = project_url
                        self._log(f"[CACHE] Project URL: {project_url[:50]}...")
                    if project_id:
                        self._cached_project_id = project_id
                        self.config['flow_project_id'] = project_id  # Set vào config
                        self._log(f"[CACHE] Project ID: {project_id[:20]}...")

                    # LUON dung cached token truoc - neu API tra ve 401 thi moi refresh
                    # Khong dua vao thoi gian vi khong dang tin cay
                    if bearer_token:
                        import time
                        self._cached_bearer_token = bearer_token
                        self._cached_token_time = token_time or time.time()
                        self.config['flow_bearer_token'] = bearer_token
                        if token_time:
                            age_minutes = (time.time() - token_time) / 60
                            self._log(f"[CACHE] Token loaded ({age_minutes:.1f} phút) - TRY FIRST, refresh if API fails")
                        else:
                            self._log(f"[CACHE] Token loaded - TRY FIRST, refresh if API fails")

                    self._log(f"[CACHE] Loaded {len(data)} media_names")
                    return data
            except Exception as e:
                self._log(f"Error loading cache: {e}", "warn")
        return {}

    def _save_media_cache(self, media_names: Dict[str, Any]) -> None:
        """
        Luu media_names vao cache file.

        Format: {
            "_project_url": "https://...",
            "_project_id": "xxx",
            "_bearer_token": "ya29.xxx",
            "_token_time": 1234567890.0,
            "nvc": {mediaName: str, seed: int|null},
            ...
        }

        QUAN TRONG: Luu project_url VA token de khi tao img co the:
        1. Vao dung project (giu media_name valid)
        2. Reuse token (khong can mo Chrome lai)
        """
        import time
        cache_path = self._get_media_cache_path()
        try:
            cache_path.parent.mkdir(parents=True, exist_ok=True)

            # Add project info to cache
            cache_data = dict(media_names)  # Copy to avoid modifying original

            # Get project URL from current session or config
            project_url = getattr(self, '_project_url', None)
            project_id = self.config.get('flow_project_id', '')
            bearer_token = self.config.get('flow_bearer_token', '')

            if project_url:
                cache_data['_project_url'] = project_url
                self._log(f"[CACHE] Saving project_url: {project_url[:50]}...")
            if project_id:
                cache_data['_project_id'] = project_id
                self._log(f"[CACHE] Saving project_id: {project_id[:20]}...")

            # QUAN TRONG: Luu token de reuse (khong can mo Chrome lai)
            if bearer_token:
                cache_data['_bearer_token'] = bearer_token
                cache_data['_token_time'] = time.time()  # Luu thoi diem lay token
                self._log(f"[CACHE] Saving bearer_token: {bearer_token[:30]}... (de reuse)")

            # Lưu thêm recaptcha_token và x_browser_validation cho I2V
            recaptcha_token = self.config.get('flow_recaptcha_token', '')
            x_browser_val = self.config.get('flow_x_browser_validation', '')
            if recaptcha_token:
                cache_data['_recaptcha_token'] = recaptcha_token
                self._log(f"[CACHE] Saving recaptcha_token (cho I2V)")
            if x_browser_val:
                cache_data['_x_browser_validation'] = x_browser_val

            # Lưu Chrome profile path để VIDEO worker có thể mở đúng Chrome
            chrome_profile = self.config.get('chrome_profile_path', '') or str(getattr(self, 'chrome_profile', ''))
            if chrome_profile:
                cache_data['_chrome_profile_path'] = chrome_profile
                self._log(f"[CACHE] Saving chrome_profile_path: {chrome_profile}")

            with open(cache_path, 'w', encoding='utf-8') as f:
                json.dump(cache_data, f, indent=2)
            self._log(f"[CACHE] Saved {len(media_names)} media_names + token + project")
        except Exception as e:
            self._log(f"Loi save cache: {e}", "warn")

    def _get_media_names_from_js(self) -> Dict[str, Any]:
        """
        Lay tat ca media_names tu JS.

        Returns:
            Dict voi format: {id: {mediaName: str, seed: int|null}}
        """
        if not self.driver:
            return {}
        try:
            return self.driver.execute_script("return VE3.getMediaNames();") or {}
        except:
            return {}

    def _get_project_url_from_js(self) -> str:
        """Lay project URL tu JS."""
        if not self.driver:
            return ""
        try:
            return self.driver.execute_script("return VE3.getProjectUrl();") or ""
        except:
            return ""

    def _load_media_names_to_js(self, media_names: Dict[str, Any]) -> None:
        """
        Load media_names vao JS tu cache.

        Supports ca format cu (string) va moi ({mediaName, seed}).
        JS se xu ly chinh xac dua vao format.
        """
        if not self.driver or not media_names:
            return
        try:
            self.driver.execute_script(f"VE3.setMediaNames({json.dumps(media_names)});")
        except Exception as e:
            self._log(f"Loi load media_names to JS: {e}", "warn")

    def _is_child_character(self, char_id: str) -> bool:
        """
        Check if a character ID represents a child (under 15 years old).
        Children cannot use reference images due to API policy.

        Child patterns:
        - nvc1 (exactly) = narrator as child
        - nv1c (exactly) = character 1 as child
        - *_child, *-child = any character with child suffix
        - *child* (but not just containing 'child' in middle of word)
        """
        if not char_id:
            return False

        char_id_clean = char_id.replace('.png', '').replace('.jpg', '').replace('.jpeg', '').replace('.webp', '').lower().strip()

        # Exact child character IDs (narrator/character as child)
        exact_child_ids = ['nvc1', 'nv1c', 'nvc_child', 'nv1_child', 'child']

        if char_id_clean in exact_child_ids:
            self._log(f"[CHILD] {char_id} matched exact child ID: {char_id_clean}", "info")
            return True

        # Suffix patterns for child characters
        if char_id_clean.endswith('_child') or char_id_clean.endswith('-child'):
            self._log(f"[CHILD] {char_id} matched child suffix pattern", "info")
            return True

        # Pattern: nvc followed by single digit 1 (nvc1 but not nvc10, nvc11, etc.)
        # This is for narrator-child-version-1
        import re
        if re.match(r'^nvc1$', char_id_clean):  # Exactly nvc1
            self._log(f"[CHILD] {char_id} matched nvc1 pattern", "info")
            return True

        return False

    def _filter_children_from_refs(self, ref_files: List[str]) -> List[str]:
        """
        Filter out child characters from reference_files list.
        Children under 15 should not be uploaded as reference.
        """
        if not ref_files:
            return []

        filtered = []
        for ref in ref_files:
            if self._is_child_character(ref):
                self._log(f"[FILTER] Bo qua tre em: {ref}", "info")
                continue
            filtered.append(ref)

        return filtered

    def _upload_reference_images(self, reference_files: List[str]) -> bool:
        """
        Upload cac anh reference truoc khi tao scene.

        Args:
            reference_files: List ten file (vd: ["nvc.png", "nv1.png"])

        Returns:
            True neu upload thanh cong
        """
        if not reference_files or not self.driver:
            return True

        self._log(f"[UPLOAD] Input reference_files: {reference_files}", "info")
        self._log(f"[UPLOAD] nv_path: {self.nv_path}", "info")

        # Filter out children under 15
        filtered_refs = self._filter_children_from_refs(reference_files)
        if not filtered_refs:
            self._log("[UPLOAD] Khong con anh nao sau khi filter tre em", "info")
            return True

        self._log(f"[UPLOAD] After filter children: {filtered_refs}", "info")

        images_to_upload = []

        for ref_file in filtered_refs:
            # Xac dinh duong dan file - ref_file co the la "nvc.png" hoac "nvc"
            ref_id = ref_file.replace('.png', '').replace('.jpg', '').replace('.jpeg', '').replace('.webp', '')

            # Tim file voi nhieu extension
            file_path = None
            filename = None

            # Thu cac extension khac nhau
            extensions = ['.png', '.jpg', '.jpeg', '.webp', '']
            search_dirs = [self.nv_path, self.project_path / "img"]

            for search_dir in search_dirs:
                if file_path:
                    break
                for ext in extensions:
                    test_path = search_dir / f"{ref_id}{ext}"
                    if test_path.exists():
                        file_path = test_path
                        filename = f"{ref_id}{ext}" if ext else ref_id
                        self._log(f"[UPLOAD] Found: {test_path}", "info")
                        break

            if not file_path:
                self._log(f"[UPLOAD] Khong tim thay file: {ref_id} (searched in nv/ and img/)", "warn")
                # List available files in nv/ for debugging
                if self.nv_path.exists():
                    available = list(self.nv_path.glob("*.*"))
                    self._log(f"[UPLOAD] Files in nv/: {[f.name for f in available[:10]]}", "info")
                continue

            # Doc file va convert sang base64
            try:
                with open(file_path, 'rb') as f:
                    image_data = f.read()
                    base64_data = base64.b64encode(image_data).decode('utf-8')
                    images_to_upload.append({
                        'base64': base64_data,
                        'filename': filename
                    })
                    self._log(f"[UPLOAD] Doc file: {filename} ({len(image_data)/1024:.1f} KB)")
            except Exception as e:
                self._log(f"[UPLOAD] Loi doc file {filename}: {e}", "error")
                continue

        if not images_to_upload:
            self._log("[UPLOAD] Khong co anh nao de upload", "warn")
            return True

        # Upload qua JS
        try:
            self._log(f"[UPLOAD] Goi JS VE3.uploadReferences() voi {len(images_to_upload)} files...", "info")
            images_json = json.dumps(images_to_upload)

            # Timeout dai hon cho upload nhieu file
            timeout_ms = 60000 + (len(images_to_upload) * 30000)  # 60s + 30s per file
            self._log(f"[UPLOAD] Timeout: {timeout_ms/1000:.0f}s", "info")

            result = self.driver.execute_async_script(f"""
                const callback = arguments[arguments.length - 1];
                const timeout = setTimeout(() => {{
                    callback({{ success: false, error: 'JS timeout' }});
                }}, {timeout_ms});

                try {{
                    VE3.uploadReferences({images_json}).then(success => {{
                        clearTimeout(timeout);
                        callback({{ success: success, message: 'Upload completed' }});
                    }}).catch(e => {{
                        clearTimeout(timeout);
                        callback({{ success: false, error: e.message || 'Upload exception' }});
                    }});
                }} catch (e) {{
                    clearTimeout(timeout);
                    callback({{ success: false, error: 'JS error: ' + e.message }});
                }}
            """)

            self._log(f"[UPLOAD] JS result: {result}", "info")

            if result and result.get('success'):
                success_count = result.get('successCount', 0)
                total_count = result.get('totalCount', len(images_to_upload))
                self._log(f"[UPLOAD] Da upload {success_count}/{total_count} anh reference", "success")
                return True
            else:
                # Log chi tiet loi tu JS
                errors = result.get('errors', []) if result else []
                if errors:
                    for err in errors:
                        self._log(f"[UPLOAD] - {err.get('file', '?')}: {err.get('error', 'Unknown')}", "error")
                else:
                    error = result.get('error', 'Unknown') if result else 'No response from JS'
                    self._log(f"[UPLOAD] Loi upload: {error}", "error")
                # Tiep tuc du co loi - khong block generation
                return False

        except Exception as e:
            self._log(f"[UPLOAD] Python Exception: {e}", "error")
            import traceback
            self._log(f"[UPLOAD] Traceback: {traceback.format_exc()}", "error")
            return False

    def _find_downloaded_files(self, pattern: str, wait_timeout: int = 30) -> List[Path]:
        """
        Tim file vua download trong Downloads folder.

        Args:
            pattern: Pattern de match (vd: KA1-0001_scene_*)
            wait_timeout: Thoi gian cho file xuat hien

        Returns:
            List cac file tim duoc
        """
        start_time = time.time()
        search_pattern = str(self.downloads_dir / pattern)

        while time.time() - start_time < wait_timeout:
            files = glob.glob(search_pattern)
            # Loai bo file .crdownload (dang tai)
            files = [f for f in files if not f.endswith('.crdownload')]

            if files:
                return [Path(f) for f in files]

            time.sleep(0.5)

        return []

    def _select_best_image(self, files: List[Path], is_character: bool = False) -> Tuple[Path, float]:
        """
        Chon anh tot nhat tu nhieu files.
        Su dung ImageEvaluator de danh gia chat luong anh (sharpness, brightness, contrast, faces).
        Fallback: dung file size neu khong co opencv.

        Args:
            files: List cac file anh
            is_character: Co phai anh nhan vat (nvc/nv*/loc*) - uu tien face detection

        Returns:
            Tuple[Path den file tot nhat, score]
        """
        if len(files) == 1:
            # Van danh gia de biet score
            try:
                from modules.image_evaluator import ImageEvaluator
                evaluator = ImageEvaluator(verbose=False)
                _, score = evaluator.evaluate(files[0], is_character)
                return files[0], score.total_score
            except ImportError:
                return files[0], 100.0  # Assume good if can't evaluate

        # Thu dung ImageEvaluator (tot hon)
        try:
            from modules.image_evaluator import ImageEvaluator
            evaluator = ImageEvaluator(verbose=False)
            best_path, best_score = evaluator.select_best(files, is_character)

            self._log(f"Chon anh tot nhat: {best_path.name} (score={best_score.total_score}, grade={best_score.grade})")

            # Log comparison
            if len(files) > 1:
                scores_str = []
                for f in files:
                    _, score = evaluator.evaluate(f, is_character)
                    scores_str.append(f"{f.name}={score.total_score}")
                self._log(f"  So sanh: {', '.join(scores_str)}")

            return best_path, best_score.total_score

        except ImportError:
            self._log("ImageEvaluator khong co, dung file size", "warn")

        # Fallback: Lay file size cua moi file
        file_sizes = []
        for f in files:
            try:
                size = f.stat().st_size
                file_sizes.append((f, size))
            except:
                file_sizes.append((f, 0))

        # Sort theo size giam dan (lon nhat = tot nhat)
        file_sizes.sort(key=lambda x: x[1], reverse=True)

        best_file = file_sizes[0][0]
        best_size = file_sizes[0][1]

        self._log(f"Chon anh tot nhat: {best_file.name} ({best_size/1024:.1f}KB)")

        # Log comparison
        if len(file_sizes) > 1:
            sizes_str = ", ".join([f"{f.name}={s/1024:.1f}KB" for f, s in file_sizes])
            self._log(f"  So sanh: {sizes_str}")

        return best_file, 70.0  # Assume decent score for fallback

    def _move_downloaded_images(
        self,
        scene_id: str,
        min_score: float = 50.0
    ) -> Tuple[Optional[Path], float, bool]:
        """
        Di chuyen anh vua download tu Downloads vao project/img/ hoac nv/.
        Neu co 2 anh, chon anh tot nhat bang ImageEvaluator.
        Tra ve score de biet co can tao lai khong.

        Args:
            scene_id: ID cua scene (1, 2, ... hoac nvc, nv1, loc1...)
            min_score: Diem toi thieu de pass (0-100)

        Returns:
            Tuple[Path da di chuyen, score, needs_regeneration]
        """
        # Pattern: {project_code}_{scene_id}*.png
        pattern = f"{self.project_code}_{scene_id}*.png"

        files = self._find_downloaded_files(pattern, wait_timeout=120)

        if not files:
            self._log(f"Khong tim thay file: {pattern}", "warn")
            return None, 0.0, True

        # Xac dinh co phai nhan vat/dia diem khong (uu tien face detection)
        scene_id_str = str(scene_id)
        is_character = scene_id_str.startswith('nv') or scene_id_str.startswith('loc')

        # QUAN TRONG: Chon anh tot nhat va danh gia chat luong
        best_file, score = self._select_best_image(files, is_character)

        # Check neu can tao lai
        needs_regeneration = score < min_score
        if needs_regeneration:
            self._log(f"Anh {scene_id} chua dat chuan: {score:.1f} < {min_score}", "warn")

        # Xac dinh thu muc dich: nv/ cho nvc/nv*/loc*, img/ cho scenes
        if is_character:
            dst_dir = self.nv_path
        else:
            dst_dir = self.img_path

        dst_dir.mkdir(parents=True, exist_ok=True)
        dst_file = dst_dir / f"{scene_id}.png"

        try:
            shutil.move(str(best_file), str(dst_file))
            self._log(f"Da di chuyen: {best_file.name} -> {dst_file} (score={score:.1f})", "success")

            # Xoa cac file con lai (khong can nua)
            for f in files:
                if f != best_file and f.exists():
                    try:
                        os.remove(f)
                        self._log(f"  Xoa file du: {f.name}")
                    except:
                        pass

            return dst_file, score, needs_regeneration

        except Exception as e:
            self._log(f"Loi di chuyen file: {e}", "error")
            return None, 0.0, True

    def generate_scene_images(
        self,
        excel_path: Optional[Path] = None,
        start_scene: int = 1,
        end_scene: Optional[int] = None,
        overwrite: bool = False
    ) -> Dict[str, Any]:
        """
        Tao anh cho cac scenes trong Excel.

        Args:
            excel_path: Duong dan file Excel (tu tim neu khong chi dinh)
            start_scene: Scene bat dau (1-indexed)
            end_scene: Scene ket thuc (None = tat ca)
            overwrite: Ghi de anh da co

        Returns:
            Dict voi ket qua
        """
        self._log("=" * 60)
        self._log("BROWSER FLOW GENERATOR - TAO ANH TU EXCEL")
        self._log("=" * 60)

        # Tim file Excel
        if excel_path is None:
            excel_path = self._find_excel_file()

        if excel_path is None or not excel_path.exists():
            return {"success": False, "error": "Khong tim thay file Excel"}

        self._log(f"Excel: {excel_path}")
        self._log(f"Project: {self.project_code}")

        # Load Excel
        workbook = PromptWorkbook(excel_path)
        workbook.load_or_create()

        # Lay cac scene can tao anh
        all_scenes = workbook.get_scenes()
        scenes_to_process = []

        for scene in all_scenes:
            # Filter theo range
            if scene.scene_id < start_scene:
                continue
            if end_scene is not None and scene.scene_id > end_scene:
                break

            # Skip neu khong co prompt
            if not scene.img_prompt:
                continue

            # Skip neu da done va khong overwrite
            if scene.status_img == "done" and not overwrite:
                self.stats["skipped"] += 1
                continue

            scenes_to_process.append(scene)

        if not scenes_to_process:
            self._log("Khong co scene nao can tao anh", "warn")
            return {"success": True, "message": "No scenes to process"}

        self._log(f"Se tao {len(scenes_to_process)} anh")
        self.stats["total"] = len(scenes_to_process)

        # QUAN TRONG: Load cache TRUOC khi start browser
        # De lay _cached_project_url va vao dung project (giu media_name valid)
        cached_media_names = self._load_media_cache()
        has_cached_project = hasattr(self, '_cached_project_url') and self._cached_project_url

        if cached_media_names:
            self._log(f"[CACHE] Loaded {len(cached_media_names)} media references (nv/loc)")
            if has_cached_project:
                self._log(f"[CACHE] Co project URL -> se reuse de giu media_name valid")
        else:
            self._log("Khong co media cache - scenes se khong co reference", "warn")

        # Khoi dong browser - vao dung project neu co cache
        if not self.driver:
            if not self.start_browser(use_cached_project=has_cached_project):
                return {"success": False, "error": "Khong khoi dong duoc browser"}

            # Cho dang nhap
            if not self.wait_for_login(timeout=120):
                self.stop_browser()
                return {"success": False, "error": "Chua dang nhap"}

        # Inject JS
        if not self._inject_js():
            return {"success": False, "error": "Khong inject duoc JS"}

        # Load media_names vao JS sau khi inject
        if cached_media_names:
            self._load_media_names_to_js(cached_media_names)
            self._log(f"Loaded media references vao JS")

        # Chuan bi prompts cho VE3.run()
        # QUAN TRONG: Dung numeric ID (1, 2, 3) de khop voi SmartEngine video composer
        prompts_data = []
        for scene in scenes_to_process:
            scene_id = str(scene.scene_id)  # Dung numeric ID, khong phai scene_001
            prompts_data.append({
                "sceneId": scene_id,
                "prompt": scene.img_prompt
            })

        self._log(f"\nBat dau tao {len(prompts_data)} anh...")

        # DEBUG: List files trong thu muc nv/
        self._log("\n" + "=" * 60)
        self._log("[DEBUG] FILES TRONG THU MUC NV/:")
        self._log(f"  nv_path: {self.nv_path}")
        if self.nv_path.exists():
            nv_files = list(self.nv_path.glob("*.*"))
            self._log(f"  Found {len(nv_files)} files:")
            for f in nv_files:
                self._log(f"    - {f.name} ({f.stat().st_size / 1024:.1f} KB)")
        else:
            self._log(f"  [ERROR] Thu muc nv/ KHONG TON TAI!")
        self._log("=" * 60)

        # DEBUG: Hien thi cac scene de kiem tra reference_files
        self._log("\n" + "=" * 60)
        self._log("[DEBUG] KIEM TRA REFERENCE_FILES TU EXCEL:")
        self._log("=" * 60)
        for idx, s in enumerate(scenes_to_process[:5]):  # Hien thi 5 scene dau
            ref_raw = getattr(s, 'reference_files', None)
            self._log(f"  Scene {s.scene_id}: reference_files = '{ref_raw}'")
        self._log("=" * 60 + "\n")

        # Goi VE3.run() - xu ly tung prompt mot de cap nhat Excel theo thoi gian thuc
        for i, item in enumerate(prompts_data):
            scene = scenes_to_process[i]
            scene_id = item["sceneId"]
            prompt = item["prompt"]

            # Lay reference_files tu scene (JSON string hoac list)
            reference_files = []
            ref_str = getattr(scene, 'reference_files', '') or ''

            self._log(f"\n[DEBUG] Scene {scene_id} raw reference_files: '{ref_str}' (type={type(ref_str).__name__})")

            if ref_str:
                try:
                    # Thu parse JSON truoc
                    if ref_str.startswith('['):
                        parsed = json.loads(ref_str)
                        reference_files = parsed if isinstance(parsed, list) else [parsed]
                        self._log(f"[DEBUG] Parsed JSON: {reference_files}")
                    else:
                        # Khong phai JSON, split by comma
                        reference_files = [f.strip() for f in str(ref_str).split(',') if f.strip()]
                        self._log(f"[DEBUG] Split by comma: {reference_files}")
                except Exception as e:
                    self._log(f"[DEBUG] Parse error: {e}, trying split...")
                    reference_files = [f.strip() for f in str(ref_str).split(',') if f.strip()]

            self._log(f"\n[{i+1}/{len(prompts_data)}] Scene {scene_id}")
            self._log(f"Prompt ({len(prompt)} chars): {prompt[:100]}...")
            self._log(f"[REF] Final reference_files: {reference_files}")

            # VERIFY: Check if prompt has filename annotations
            has_annotations = False
            for ref in reference_files:
                ref_name = ref.replace('.png', '').replace('.jpg', '')
                if f"({ref})" in prompt or f"({ref_name}.png)" in prompt:
                    has_annotations = True
                    break
            if "(reference:" in prompt:
                has_annotations = True

            if reference_files:
                if has_annotations:
                    self._log(f"[ANNOTATION] ✓ Prompt DA CO annotations", "success")
                else:
                    self._log(f"[ANNOTATION] ⚠️ Prompt CHUA CO annotations - them vao cuoi...", "warn")
                    # Them annotation neu chua co
                    refs_str = ", ".join(reference_files)
                    prompt = prompt.rstrip('. ') + f" (reference: {refs_str})."
                    self._log(f"[ANNOTATION] Prompt sau khi them: ...{prompt[-80:]}", "info")

            try:
                # QUAN TRONG: Upload reference images TRUOC KHI tao anh
                if reference_files:
                    self._log(f"[UPLOAD] Dang upload {len(reference_files)} anh reference...")
                    upload_success = self._upload_reference_images(reference_files)
                    if upload_success:
                        self._log(f"[UPLOAD] Upload thanh cong!", "success")
                    else:
                        self._log("[UPLOAD] Upload that bai, tiep tuc khong co reference", "warn")

                # Goi VE3.run() cho 1 prompt (voi reference_files)
                ref_files_json = json.dumps(reference_files)
                result = self.driver.execute_async_script(f"""
                    const callback = arguments[arguments.length - 1];
                    const timeout = setTimeout(() => {{
                        callback({{ success: false, error: 'Timeout 120s' }});
                    }}, 120000);

                    VE3.run([{{
                        sceneId: "{scene_id}",
                        prompt: `{self._escape_js_string(prompt)}`,
                        referenceFiles: {ref_files_json}
                    }}]).then(r => {{
                        clearTimeout(timeout);
                        callback({{ success: true, result: r }});
                    }}).catch(e => {{
                        clearTimeout(timeout);
                        callback({{ success: false, error: e.message }});
                    }});
                """)

                if result and result.get("success"):
                    # Di chuyen file - scene_id la numeric ("1", "2", ...)
                    img_file, score, needs_regen = self._move_downloaded_images(scene_id)

                    if img_file:
                        # Cap nhat Excel - dung numeric ID
                        relative_path = f"img/{scene_id}.png"
                        workbook.update_scene(
                            scene.scene_id,  # scene.scene_id la int
                            img_path=relative_path,
                            status_img="done" if not needs_regen else "low_quality"
                        )
                        workbook.save()

                        if needs_regen:
                            self._log(f"Anh {scene_id} chua dat chuan (score={score:.1f}), can tao lai", "warn")
                            self.stats["low_quality"] = self.stats.get("low_quality", 0) + 1
                        else:
                            self._log(f"Da cap nhat Excel: {scene_id} = done (score={score:.1f})", "success")
                        self.stats["success"] += 1
                    else:
                        workbook.update_scene(scene.scene_id, status_img="error")
                        workbook.save()
                        self.stats["failed"] += 1
                else:
                    error = result.get("error", "Unknown") if result else "No response"
                    self._log(f"Loi: {error}", "error")
                    workbook.update_scene(scene.scene_id, status_img="error")
                    workbook.save()
                    self.stats["failed"] += 1

                # Delay giua cac prompt
                if i < len(prompts_data) - 1:
                    time.sleep(2)

            except Exception as e:
                self._log(f"Exception: {e}", "error")
                workbook.update_scene(scene.scene_id, status_img="error")
                workbook.save()
                self.stats["failed"] += 1

        # Summary
        self._log("\n" + "=" * 60)
        self._log("HOAN THANH")
        self._log("=" * 60)
        self._log(f"Tong: {self.stats['total']}")
        self._log(f"Thanh cong: {self.stats['success']}")
        self._log(f"That bai: {self.stats['failed']}")
        self._log(f"Bo qua: {self.stats['skipped']}")

        return {
            "success": True,
            "stats": self.stats.copy()
        }

    def _process_single_prompt(self, prompt_data: Dict, index: int, total: int) -> Tuple[bool, Optional[Path], float, str]:
        """
        Xu ly mot prompt don le.

        Returns:
            Tuple[success, image_path, score, prompt_json]
        """
        pid = str(prompt_data.get('id', index + 1))
        prompt = prompt_data.get('prompt', '')

        # Lay reference_files tu prompt_data (JSON string hoac list)
        reference_files = []
        ref_str = prompt_data.get('reference_files', '')
        if ref_str:
            try:
                parsed = json.loads(ref_str) if isinstance(ref_str, str) else ref_str
                reference_files = parsed if isinstance(parsed, list) else [parsed]
            except:
                reference_files = [f.strip() for f in str(ref_str).split(',') if f.strip()]

        self._log(f"\n[{index+1}/{total}] ID: {pid}")
        self._log(f"Prompt ({len(prompt)} chars): {prompt[:100]}...")
        if reference_files:
            self._log(f"[REF] reference_files: {reference_files}")
        else:
            self._log(f"[REF] ⚠️ NO REFERENCES - Excel cot 'reference_files' trong!")
            self._log(f"[REF] raw value from Excel: '{ref_str}'")

        if not prompt:
            self._log("Skip - prompt rong", "warn")
            return False, None, 0.0

        try:
            # UPLOAD REFERENCE IMAGES TRUOC KHI TAO ANH
            if reference_files:
                self._log(f"[UPLOAD] Dang upload {len(reference_files)} anh reference...")
                upload_success = self._upload_reference_images(reference_files)
                if not upload_success:
                    self._log("[UPLOAD] Upload that bai, tiep tuc khong co reference", "warn")

            # Goi VE3.run() cho 1 prompt (voi reference_files)
            ref_files_json = json.dumps(reference_files)
            result = self.driver.execute_async_script(f"""
                const callback = arguments[arguments.length - 1];
                const timeout = setTimeout(() => {{
                    callback({{ success: false, error: 'Timeout 120s' }});
                }}, 120000);

                VE3.run([{{
                    sceneId: "{pid}",
                    prompt: `{self._escape_js_string(prompt)}`,
                    referenceFiles: {ref_files_json}
                }}]).then(r => {{
                    clearTimeout(timeout);
                    callback({{ success: true, result: r }});
                }}).catch(e => {{
                    clearTimeout(timeout);
                    callback({{ success: false, error: e.message }});
                }});
            """)

            if result and result.get("success"):
                # Lay prompt_json va images tu result
                js_result = result.get("result", {})
                prompt_json = js_result.get("prompt_json", "") if isinstance(js_result, dict) else ""
                js_images = js_result.get("images", []) if isinstance(js_result, dict) else []

                self._log(f"[DEBUG] prompt_json: {prompt_json[:100] if prompt_json else '(empty)'}...")
                self._log(f"[DEBUG] JS returned {len(js_images)} images with mediaNames")

                # Di chuyen file tu Downloads (timeout 2 phut)
                img_file, score, needs_regen = self._move_downloaded_images(pid)

                if img_file:
                    # TIM MEDIA_NAME VA SEED DUNG CHO ANH DA CHON
                    selected_media_name = ""
                    selected_seed = None
                    if js_images:
                        # Neu chi co 1 anh, lay mediaName va seed cua no
                        if len(js_images) == 1:
                            selected_media_name = js_images[0].get("mediaName", "")
                            selected_seed = js_images[0].get("seed")
                            self._log(f"[MEDIA] 1 image -> mediaName: {selected_media_name[:50] if selected_media_name else 'NONE'}, seed={selected_seed}")
                        else:
                            # Neu co nhieu anh, lay anh dau tien (thuong la best)
                            # TODO: Match filename de lay dung anh duoc chon
                            selected_media_name = js_images[0].get("mediaName", "")
                            selected_seed = js_images[0].get("seed")
                            self._log(f"[MEDIA] {len(js_images)} images, selected first -> mediaName: {selected_media_name[:50] if selected_media_name else 'NONE'}, seed={selected_seed}")

                    # Set mediaName + seed vao JS STATE de reference sau
                    if selected_media_name and self.driver:
                        try:
                            # Pass ca mediaName va seed
                            seed_arg = f", {selected_seed}" if selected_seed else ", null"
                            self.driver.execute_script(
                                f"VE3.setMediaName('{pid}', '{selected_media_name}'{seed_arg});"
                            )
                            self._log(f"[MEDIA] Saved mediaInfo for {pid}: mediaName + seed={selected_seed}")
                        except Exception as e:
                            self._log(f"[MEDIA] Warning: Could not set mediaName: {e}", "warn")

                    if needs_regen:
                        self._log(f"OK - Da tao anh nhung chua dat chuan (score={score:.1f})", "warn")
                    else:
                        self._log(f"OK - Da tao va luu anh (score={score:.1f})", "success")
                    return True, img_file, score, prompt_json
                else:
                    self._log(f"Khong tim thay file download sau 2 phut", "warn")
                    return False, None, 0.0, prompt_json
            else:
                error = result.get("error", "Unknown") if result else "No response"
                self._log(f"Loi: {error}", "error")
                return False, None, 0.0, ""

        except Exception as e:
            self._log(f"Exception: {e}", "error")
            return False, None, 0.0, ""

    def _restart_browser_and_setup(self) -> bool:
        """
        Khoi dong lai browser va setup (dung khi setup that bai).
        Neu da co project URL, navigate ve do thay vi tao project moi.

        Returns:
            True neu thanh cong
        """
        self._log("Khoi dong lai browser...", "warn")

        # Luu project URL truoc khi dong browser
        saved_project_url = self._project_url

        # Dong browser cu
        self.stop_browser()
        self._js_injected = False

        # Doi 3 giay
        time.sleep(3)

        # Khoi dong lai
        if not self.start_browser():
            return False

        if not self.wait_for_login(timeout=120):
            self.stop_browser()
            return False

        # Neu co project URL cu, navigate ve do thay vi setup moi
        if saved_project_url and '/project/' in saved_project_url:
            self._log(f"Navigate ve project cu: {saved_project_url}", "info")
            self.driver.get(saved_project_url)
            time.sleep(5)  # Cho page load

            # Inject JS nhung KHONG goi VE3.setup()
            if not self._inject_js_without_setup():
                return False

            # Restore project URL
            self._project_url = saved_project_url
        else:
            # Inject JS va setup moi
            if not self._inject_js():
                return False

        return True

    def _inject_js_without_setup(self) -> bool:
        """Inject JS ma khong goi VE3.setup() - dung khi navigate ve project cu."""
        if self._js_injected:
            return True

        try:
            self._log("Inject JavaScript (khong setup)...")
            js_code = self._get_js_script()
            self.driver.execute_script(js_code)

            # Init VE3 voi project name
            self.driver.execute_script(f'VE3.init("{self.project_code}")')

            # Danh dau da setup (vi da co project)
            self.driver.execute_script("VE3.markSetupDone();")

            # Load media_names tu cache
            cached_media_names = self._load_media_cache()
            if cached_media_names:
                self._load_media_names_to_js(cached_media_names)

            self._js_injected = True
            self._log("Da san sang tao anh (project cu)!", "success")
            return True

        except Exception as e:
            self._log(f"Loi inject JS: {e}", "error")
            return False

    def generate_from_prompts(
        self,
        prompts: List[Dict],
        excel_path: Optional[Path] = None,
        max_setup_retries: int = 3
    ) -> Dict[str, Any]:
        """
        Tao anh tu danh sach prompts da load san (tu smart_engine._load_prompts).
        Method nay nhan prompts truc tiep thay vi doc lai tu Excel.

        Features:
        - Neu prompt dau tien that bai (co the do setup loi), tu dong restart browser va thu lai (toi da 3 lan)
        - Theo doi cac prompt that bai va retry o cuoi
        - Timeout 2 phut cho moi anh

        Args:
            prompts: List cac dict co dang {'id': '1', 'prompt': '...', 'output_path': '...'}
            excel_path: Duong dan Excel (de cap nhat status)
            max_setup_retries: So lan retry toi da neu setup that bai (default: 3)

        Returns:
            Dict voi ket qua
        """
        self._log("=" * 60)
        self._log("BROWSER FLOW GENERATOR - TAO ANH TU PROMPTS")
        self._log("=" * 60)

        if not prompts:
            return {"success": False, "error": "Khong co prompts"}

        self._log(f"Tong: {len(prompts)} prompts")
        self._log(f"Project: {self.project_code}")

        # Reset stats
        self.stats = {"total": len(prompts), "success": 0, "failed": 0, "skipped": 0, "low_quality": 0}

        # Load Excel workbook de cap nhat prompt_json
        workbook = None
        if excel_path and Path(excel_path).exists():
            try:
                workbook = PromptWorkbook(excel_path)
                workbook.load_or_create()
                self._log(f"[Excel] Loaded: {excel_path}")
            except Exception as e:
                self._log(f"[Excel] Warning: Khong load duoc Excel: {e}", "warn")

        # QUAN TRONG: Load cache TRUOC khi start browser
        # De lay _cached_project_url va vao dung project (giu media_name valid)
        cached_media_names = self._load_media_cache()
        has_cached_project = hasattr(self, '_cached_project_url') and self._cached_project_url

        if cached_media_names:
            self._log(f"[CACHE] Loaded {len(cached_media_names)} media_names")
            if has_cached_project:
                self._log(f"[CACHE] Co project URL -> se reuse de giu media_name valid")
        else:
            self._log("[CACHE] ⚠️ EMPTY - Characters (nv/loc) chua duoc tao!", "warn")

        # Khoi dong browser - vao dung project neu co cache
        if not self.driver:
            if not self.start_browser(use_cached_project=has_cached_project):
                return {"success": False, "error": "Khong khoi dong duoc browser"}

            if not self.wait_for_login(timeout=120):
                self.stop_browser()
                return {"success": False, "error": "Chua dang nhap"}

        # Inject JS
        if not self._inject_js():
            return {"success": False, "error": "Khong inject duoc JS"}

        # Load media_names vao JS sau khi inject
        if cached_media_names:
            for key, val in cached_media_names.items():
                if isinstance(val, dict):
                    mn = val.get('mediaName', '')
                    seed = val.get('seed')
                    self._log(f"  {key} -> mediaName:{mn[:40] if mn else 'None'}..., seed:{seed}")
                else:
                    self._log(f"  {key} -> {val[:50] if val else 'None'}...")
            self._load_media_names_to_js(cached_media_names)

        self._log(f"\nBat dau tao {len(prompts)} anh...")

        # DEBUG: Hien thi prompt dau tien
        if prompts:
            p = prompts[0]
            self._log(f"[DEBUG] Prompt dau tien: id={p.get('id')}")
            self._log(f"[DEBUG] prompt = '{str(p.get('prompt', ''))[:100]}'")

        # Track failed prompts de retry sau
        failed_prompts = []  # List of (prompt_data, original_index)

        # === XU LY PROMPT DAU TIEN VOI SETUP RETRY ===
        first_prompt_success = False
        setup_attempts = 0

        while not first_prompt_success and setup_attempts < max_setup_retries:
            setup_attempts += 1

            if setup_attempts > 1:
                self._log(f"\n=== SETUP RETRY {setup_attempts}/{max_setup_retries} ===", "warn")
                if not self._restart_browser_and_setup():
                    self._log(f"Khong the khoi dong lai browser", "error")
                    continue

            # Thu prompt dau tien
            success, img_file, score, prompt_json = self._process_single_prompt(prompts[0], 0, len(prompts))

            if success:
                first_prompt_success = True
                self.stats["success"] += 1
                if score < 50.0:
                    self.stats["low_quality"] += 1
                # Luu prompt_json vao prompt_data
                if prompt_json:
                    prompts[0]['prompt_json'] = prompt_json
                # Cap nhat Excel (prompt_json, img_path)
                if workbook:
                    try:
                        pid = prompts[0].get('id', '1')
                        # Chi cap nhat cho scenes (so), khong phai nv/loc
                        if pid.isdigit():
                            scene_id = int(pid)
                            relative_path = f"img/{pid}.png" if img_file else ""
                            workbook.update_scene(
                                scene_id,
                                img_path=relative_path,
                                status_img="done" if score >= 50.0 else "low_quality",
                                prompt_json=prompt_json
                            )
                            workbook.save()
                            self._log(f"[Excel] Updated scene {scene_id}: prompt_json saved")
                    except Exception as e:
                        self._log(f"[Excel] Warning: {e}", "warn")
            else:
                self._log(f"Prompt dau tien that bai (lan {setup_attempts})", "error")

        if not first_prompt_success:
            # Da thu het so lan retry, ghi nhan that bai
            self._log(f"Prompt dau tien that bai sau {max_setup_retries} lan thu", "error")
            failed_prompts.append((prompts[0], 0))
            self.stats["failed"] += 1

        # === XU LY CAC PROMPT CON LAI ===
        for i, prompt_data in enumerate(prompts[1:], start=1):
            pid = str(prompt_data.get('id', i + 1))
            prompt = prompt_data.get('prompt', '')

            if not prompt:
                self._log(f"\n[{i+1}/{len(prompts)}] ID: {pid} - Skip (prompt rong)", "warn")
                self.stats["skipped"] += 1
                continue

            success, img_file, score, prompt_json = self._process_single_prompt(prompt_data, i, len(prompts))

            if success:
                self.stats["success"] += 1
                if score < 50.0:
                    self.stats["low_quality"] += 1
                # Luu prompt_json vao prompt_data
                if prompt_json:
                    prompt_data['prompt_json'] = prompt_json
                # Cap nhat Excel (prompt_json, img_path)
                if workbook:
                    try:
                        # Chi cap nhat cho scenes (so), khong phai nv/loc
                        if pid.isdigit():
                            scene_id = int(pid)
                            relative_path = f"img/{pid}.png" if img_file else ""
                            workbook.update_scene(
                                scene_id,
                                img_path=relative_path,
                                status_img="done" if score >= 50.0 else "low_quality",
                                prompt_json=prompt_json
                            )
                            workbook.save()
                            self._log(f"[Excel] Updated scene {scene_id}: prompt_json saved")
                    except Exception as e:
                        self._log(f"[Excel] Warning: {e}", "warn")
            else:
                failed_prompts.append((prompt_data, i))
                self.stats["failed"] += 1

            # Delay giua cac prompt
            if i < len(prompts) - 1:
                time.sleep(2)

        # === RETRY FAILED PROMPTS ===
        if failed_prompts:
            self._log("\n" + "=" * 60)
            self._log(f"RETRY {len(failed_prompts)} ANH THAT BAI")
            self._log("=" * 60)

            retry_success = 0
            for prompt_data, original_index in failed_prompts:
                pid = str(prompt_data.get('id', original_index + 1))
                self._log(f"\nRetry ID: {pid}")

                success, img_file, score, prompt_json = self._process_single_prompt(
                    prompt_data, original_index, len(prompts)
                )

                if success:
                    retry_success += 1
                    self.stats["success"] += 1
                    self.stats["failed"] -= 1  # Giam failed vi da thanh cong
                    if score < 50.0:
                        self.stats["low_quality"] += 1
                    # Luu prompt_json vao prompt_data
                    if prompt_json:
                        prompt_data['prompt_json'] = prompt_json
                    # Cap nhat Excel (prompt_json, img_path)
                    if workbook:
                        try:
                            # Chi cap nhat cho scenes (so), khong phai nv/loc
                            if pid.isdigit():
                                scene_id = int(pid)
                                relative_path = f"img/{pid}.png" if img_file else ""
                                workbook.update_scene(
                                    scene_id,
                                    img_path=relative_path,
                                    status_img="done" if score >= 50.0 else "low_quality",
                                    prompt_json=prompt_json
                                )
                                workbook.save()
                                self._log(f"[Excel] Updated scene {scene_id}: prompt_json saved (retry)")
                        except Exception as e:
                            self._log(f"[Excel] Warning: {e}", "warn")

                # Delay
                time.sleep(2)

            self._log(f"\nRetry: {retry_success}/{len(failed_prompts)} thanh cong")

        # Luu media_names tu JS vao cache (cho cac lan chay sau)
        js_media_names = self._get_media_names_from_js()
        if js_media_names:
            # Merge voi cached (uu tien moi)
            all_media_names = {**cached_media_names, **js_media_names}
            self._save_media_cache(all_media_names)

        # Summary
        self._log("\n" + "=" * 60)
        self._log("HOAN THANH")
        self._log("=" * 60)
        self._log(f"Tong: {self.stats['total']}")
        self._log(f"Thanh cong: {self.stats['success']}")
        self._log(f"That bai: {self.stats['failed']}")
        self._log(f"Bo qua: {self.stats['skipped']}")
        if self.stats.get('low_quality', 0) > 0:
            self._log(f"Chat luong thap: {self.stats['low_quality']}")
        if js_media_names:
            self._log(f"Media names saved: {len(js_media_names)}")

        return {
            "success": True,
            "stats": self.stats.copy()
        }

    def generate_character_images(
        self,
        excel_path: Optional[Path] = None,
        overwrite: bool = False
    ) -> Dict[str, Any]:
        """
        Tao anh cho cac nhan vat trong Excel.

        Args:
            excel_path: Duong dan file Excel
            overwrite: Ghi de anh da co

        Returns:
            Dict voi ket qua
        """
        self._log("=" * 60)
        self._log("TAO ANH NHAN VAT")
        self._log("=" * 60)

        # Tim file Excel
        if excel_path is None:
            excel_path = self._find_excel_file()

        if excel_path is None or not excel_path.exists():
            return {"success": False, "error": "Khong tim thay file Excel"}

        # Load Excel
        workbook = PromptWorkbook(excel_path)
        workbook.load_or_create()

        # Lay cac nhan vat can tao anh
        characters = workbook.get_characters()
        chars_to_process = []

        for char in characters:
            if not char.english_prompt:
                continue

            # Skip children (status="skip" or english_prompt="DO_NOT_GENERATE" or is_child=True)
            if char.status == "skip" or char.english_prompt == "DO_NOT_GENERATE" or getattr(char, 'is_child', False):
                self._log(f"  ⏭️  {char.id}: Child character, skipping (will use inline description)")
                continue

            if char.status == "done" and not overwrite:
                continue

            chars_to_process.append(char)

        if not chars_to_process:
            self._log("Khong co nhan vat nao can tao anh", "warn")
            return {"success": True, "message": "No characters to process"}

        self._log(f"Se tao {len(chars_to_process)} anh nhan vat")

        # Khoi dong browser neu chua
        if not self.driver:
            if not self.start_browser():
                return {"success": False, "error": "Khong khoi dong duoc browser"}

            if not self.wait_for_login(timeout=120):
                self.stop_browser()
                return {"success": False, "error": "Chua dang nhap"}

        if not self._inject_js():
            return {"success": False, "error": "Khong inject duoc JS"}

        # Load media cache de co the reference characters da tao truoc do
        cached_media_names = self._load_media_cache()
        if cached_media_names:
            self._load_media_names_to_js(cached_media_names)

        success_count = 0
        failed_count = 0

        for i, char in enumerate(chars_to_process):
            char_id = char.id or f"char_{i+1}"
            prompt = char.english_prompt

            self._log(f"\n[{i+1}/{len(chars_to_process)}] Nhan vat: {char_id}")

            try:
                result = self.driver.execute_async_script(f"""
                    const callback = arguments[arguments.length - 1];
                    VE3.run([{{
                        sceneId: "{char_id}",
                        prompt: `{self._escape_js_string(prompt)}`
                    }}]).then(r => {{
                        callback({{ success: true, result: r }});
                    }}).catch(e => {{
                        callback({{ success: false, error: e.message }});
                    }});
                """)

                if result and result.get("success"):
                    # Di chuyen file vao nv/
                    pattern = f"{self.project_code}_{char_id}*.png"
                    files = self._find_downloaded_files(pattern, wait_timeout=120)

                    if files:
                        dst_file = self.nv_path / f"{char_id}.png"
                        shutil.move(str(files[0]), str(dst_file))

                        workbook.update_character(char_id, status="done", image_file=f"{char_id}.png")
                        workbook.save()

                        self._log(f"Da luu: {dst_file}", "success")
                        success_count += 1
                    else:
                        workbook.update_character(char_id, status="error")
                        workbook.save()
                        failed_count += 1
                else:
                    failed_count += 1

                if i < len(chars_to_process) - 1:
                    time.sleep(2)

            except Exception as e:
                self._log(f"Loi: {e}", "error")
                failed_count += 1

        self._log(f"\nNhan vat: {success_count} thanh cong, {failed_count} that bai")

        # IMPORTANT: Luu mediaNames vao cache sau khi tao xong characters
        # De khi tao scenes co the reference den characters
        try:
            media_names = self._get_media_names_from_js()
            if media_names:
                self._save_media_cache(media_names)
                self._log(f"Da luu {len(media_names)} media_names cho reference", "success")
        except Exception as e:
            self._log(f"Loi luu media cache: {e}", "warn")

        return {
            "success": True,
            "characters_success": success_count,
            "characters_failed": failed_count
        }

    def _get_generation_mode(self) -> str:
        """
        Lay generation mode tu config: 'chrome' hoac 'api'.
        Mac dinh: 'api' (user preference).
        """
        return self.config.get('generation_mode', 'api')

    def _auto_extract_token(self, force_refresh: bool = False) -> Optional[str]:
        """
        Tu dong lay bearer token tu Chrome bang profile hien tai.
        Su dung ChromeAutoToken de mo Chrome, navigate den Flow va capture token.

        Args:
            force_refresh: True = bo qua cache, luon lay token moi (dung khi API tra 401)

        QUAN TRONG:
        - KHONG check thoi gian token
        - Chi refresh khi force_refresh=True (API tra 401)
        - Neu co cached project_url -> mo truc tiep project do -> skip 'Tao du an moi'

        Returns:
            Bearer token (ya29.xxx) hoac None neu that bai
        """
        import time

        # =====================================================================
        # LOAD CACHE TRUOC - de co token va project_url
        # =====================================================================
        if not hasattr(self, '_cached_project_url') or not self._cached_project_url:
            self._log("Loading media cache de lay token va project URL...")
            self._load_media_cache()

        # =====================================================================
        # REUSE CACHED TOKEN (tru khi force_refresh)
        # Chi refresh khi API tra loi 401 -> caller goi force_refresh=True
        # =====================================================================
        if not force_refresh:
            cached_token = self.config.get('flow_bearer_token', '')

            if cached_token:
                # Luon dung token da co - khong can biet tuoi
                # Neu het han, API se tra 401 -> caller goi force_refresh=True
                cached_token_time = getattr(self, '_cached_token_time', 0)
                if cached_token_time:
                    age_minutes = (time.time() - cached_token_time) / 60
                    self._log(f"=== REUSE CACHED TOKEN (tuoi: {age_minutes:.1f} phut) ===")
                else:
                    self._log(f"=== REUSE CACHED TOKEN ===")
                self._log(f"Token: {cached_token[:20]}...{cached_token[-10:]}")
                return cached_token
        else:
            self._log("=== FORCE REFRESH TOKEN (API tra 401) ===")
            # Clear cached token
            self.config['flow_bearer_token'] = ''

        self._log("=== TU DONG LAY BEARER TOKEN (mo Chrome) ===")

        # QUAN TRONG: Lay token LUON phai chay Chrome HIEN THI (khong an)
        # Vi Google Flow detect headless mode va block!
        # Headless chi dung cho viec TAO ANH, khong dung cho LAY TOKEN.
        use_headless = False  # LUON False khi lay token
        self._log("⚠️ Lay token: Chrome se HIEN THI (Google block headless)")

        # Chon extractor phu hop:
        TokenExtractor = None
        extractor_name = None
        if use_headless:
            try:
                from modules.chrome_token_extractor import ChromeTokenExtractor
                TokenExtractor = ChromeTokenExtractor
                extractor_name = "ChromeTokenExtractor"
                self._log("Su dung ChromeTokenExtractor (Selenium - headless)")
            except ImportError as e:
                self._log(f"ChromeTokenExtractor khong kha dung: {e}", "warn")
                self._log("Fallback sang PyAutoGUI...", "warn")
            except Exception as e:
                self._log(f"Loi import ChromeTokenExtractor: {e}", "error")

        if TokenExtractor is None:
            try:
                from modules.auto_token import ChromeAutoToken
                TokenExtractor = ChromeAutoToken
                extractor_name = "ChromeAutoToken"
                self._log("Su dung ChromeAutoToken (PyAutoGUI - can cua so)")
            except ImportError as e:
                self._log(f"Khong import duoc token extractor: {e}", "error")
                return None
            except Exception as e:
                self._log(f"Loi import ChromeAutoToken: {e}", "error")
                return None

        # Lay chrome_path tu config
        chrome_path = self.config.get('chrome_path', '')
        if not chrome_path:
            # Default paths
            import platform
            if platform.system() == 'Windows':
                chrome_path = r"C:\Program Files\Google\Chrome\Application\chrome.exe"
            else:
                chrome_path = "/usr/bin/google-chrome"

        # Lay profile path
        # Uu tien: 1. chrome_profiles/ directory (GUI tao)
        #          2. chrome_profiles tu accounts.json
        #          3. chrome_profile tu settings.yaml (fallback)
        #          4. browser_profiles_dir/profile_name (fallback cuoi)
        chrome_profile = ''
        root_dir = Path(__file__).parent.parent

        # 1. UU TIEN NHAT: chrome_profiles/ directory (tao tu GUI)
        profiles_dir = root_dir / "chrome_profiles"
        if profiles_dir.exists():
            for profile_path in sorted(profiles_dir.iterdir()):
                if profile_path.is_dir() and not profile_path.name.startswith('.'):
                    chrome_profile = str(profile_path)
                    self._log(f"Got chrome_profile from chrome_profiles/ dir: {chrome_profile}")
                    break

        # 2. Fallback: accounts.json
        if not chrome_profile:
            try:
                accounts_file = root_dir / "config" / "accounts.json"
                if accounts_file.exists():
                    import json
                    with open(accounts_file, 'r', encoding='utf-8') as f:
                        accounts = json.load(f)
                    profiles = accounts.get('chrome_profiles', [])
                    for p in profiles:
                        path = p if isinstance(p, str) else p.get('path', '')
                        if path and not path.startswith('THAY_BANG') and Path(path).exists():
                            chrome_profile = path
                            self._log(f"Got chrome_profile from accounts.json: {chrome_profile}")
                            break
            except Exception as e:
                self._log(f"[DEBUG] Cannot read accounts.json: {e}")

        # 3. Fallback: settings.yaml
        if not chrome_profile:
            chrome_profile = self.config.get('chrome_profile', '')

        self._log(f"[DEBUG] chrome_profile: '{chrome_profile}'")

        if chrome_profile:
            chrome_profile_path = Path(chrome_profile)
            self._log(f"[DEBUG] chrome_profile exists: {chrome_profile_path.exists()}")

            if chrome_profile_path.exists():
                # Dung profile tu settings.yaml
                profile_path = str(chrome_profile_path)
                self._log(f"Su dung Chrome profile tu settings: {profile_path}")
            else:
                # Path khong ton tai - thu resolve relative path
                base_dir = Path(__file__).parent.parent
                abs_chrome_profile = base_dir / chrome_profile
                self._log(f"[DEBUG] Trying absolute: {abs_chrome_profile}")

                if abs_chrome_profile.exists():
                    profile_path = str(abs_chrome_profile)
                    self._log(f"Su dung Chrome profile (resolved): {profile_path}")
                else:
                    # Van khong ton tai - dung profile_dir da setup
                    profile_path = str(self.profile_dir)
                    self._log(f"[WARN] chrome_profile path khong ton tai, dung: {profile_path}")
        else:
            # Khong co chrome_profile - dung profile_dir da setup
            profile_path = str(self.profile_dir)
            self._log(f"Su dung tool profile: {profile_path}")

        self._log(f"Chrome: {chrome_path}")
        self._log(f"Profile: {profile_path}")
        self._log(f"Headless: {'ON' if use_headless else 'OFF'}")

        try:
            # Callback de log
            def log_callback(msg, level="info"):
                self._log(f"[TokenExtract] {msg}", level)

            # QUAN TRONG: Reuse project URL/ID da co de:
            # 1. Share media_ids giua nv va img
            # 2. Skip buoc "Tao du an moi" -> lay token nhanh hon
            existing_project_url = getattr(self, '_cached_project_url', None)
            existing_project_id = self.config.get('flow_project_id', '')

            if existing_project_url:
                self._log(f"  -> Reuse project URL: {existing_project_url[:50]}...")
            elif existing_project_id:
                self._log(f"  -> Reuse project_id: {existing_project_id[:20]}...")
            else:
                self._log(f"  -> Chua co project -> se tao moi")

            # Tao extractor va goi extract_token
            # ChromeTokenExtractor (Selenium) va ChromeAutoToken (PyAutoGUI) co interface khac nhau
            self._log(f"Creating extractor: {extractor_name}...")

            if extractor_name == "ChromeTokenExtractor":
                # Selenium-based: headless OK, nhung khong co project_id/url param
                self._log("Khoi tao Selenium extractor...")
                extractor = TokenExtractor(
                    chrome_path=chrome_path,
                    profile_path=profile_path,
                    headless=use_headless,
                    timeout=90
                )
                self._log("Goi extract_token (Selenium)...")
                token, proj_id, error = extractor.extract_token(callback=log_callback)
            else:
                # PyAutoGUI-based: can cua so, nhung co project reuse
                self._log("Khoi tao PyAutoGUI extractor...")
                extractor = TokenExtractor(
                    chrome_path=chrome_path,
                    profile_path=profile_path,
                    headless=use_headless
                )
                self._log("Goi extract_token (PyAutoGUI)...")
                token, proj_id, error = extractor.extract_token(
                    project_id=existing_project_id,
                    project_url=existing_project_url,
                    callback=log_callback
                )

            if token:
                self._log(f"OK - Da lay duoc token: {token[:20]}...{token[-10:]}", "success")
                # Luu project_id va token vao config
                if proj_id:
                    self.config['flow_project_id'] = proj_id
                    self._log(f"  -> Project ID: {proj_id[:20]}...")

                # Luu token vao config va cache
                import time
                self.config['flow_bearer_token'] = token
                self._cached_bearer_token = token
                self._cached_token_time = time.time()

                # Save vao cache file de reuse lan sau
                try:
                    cached_media_names = self._load_media_cache() or {}
                    self._save_media_cache(cached_media_names)
                    self._log("  -> Token da luu vao cache")
                except Exception as e:
                    self._log(f"  -> Luu cache that bai: {e}")

                return token
            else:
                self._log(f"FAIL - Khong lay duoc token: {error}", "error")
                return None

        except Exception as e:
            self._log(f"Exception khi lay token: {e}", "error")
            import traceback
            traceback.print_exc()
            return None

    def _capture_headers_via_chrome(self) -> Optional[Dict[str, str]]:
        """
        Capture headers (x-browser-validation, authorization, etc) tu Chrome bang Selenium + CDP.
        TAO 1 ANH TEST de trigger API request va capture headers.

        Returns:
            Dict chua headers hoac None neu that bai
        """
        self._log("=== CAPTURE HEADERS TU CHROME (SELENIUM) ===")

        try:
            from modules.chrome_headers_extractor import ChromeHeadersExtractor
        except ImportError as e:
            self._log(f"Khong import duoc ChromeHeadersExtractor: {e}", "error")
            return None

        # Lay profile path - UU TIEN accounts.json (GUI) truoc settings.yaml
        chrome_profile = ''

        # 1. UU TIEN accounts.json (user them profile qua GUI)
        try:
            accounts_file = Path(__file__).parent.parent / "config" / "accounts.json"
            if accounts_file.exists():
                import json
                with open(accounts_file, 'r', encoding='utf-8') as f:
                    accounts = json.load(f)
                profiles = accounts.get('chrome_profiles', [])
                for p in profiles:
                    path = p if isinstance(p, str) else p.get('path', '')
                    if path and not path.startswith('THAY_BANG') and Path(path).exists():
                        chrome_profile = path
                        self._log(f"Got chrome_profile from accounts.json (GUI): {chrome_profile}")
                        break
        except:
            pass

        # 2. Fallback: settings.yaml
        if not chrome_profile:
            chrome_profile = self.config.get('chrome_profile', '')

        if chrome_profile:
            chrome_profile_path = Path(chrome_profile)
            if chrome_profile_path.exists():
                profile_path = str(chrome_profile_path)
                self._log(f"Su dung Chrome profile: {profile_path}")
            else:
                base_dir = Path(__file__).parent.parent
                abs_chrome_profile = base_dir / chrome_profile
                if abs_chrome_profile.exists():
                    profile_path = str(abs_chrome_profile)
                    self._log(f"Su dung Chrome profile (resolved): {profile_path}")
                else:
                    profile_path = str(self.profile_dir)
                    self._log(f"[WARN] chrome_profile khong ton tai, dung: {profile_path}")
        else:
            profile_path = str(self.profile_dir)
            self._log(f"Su dung tool profile: {profile_path}")

        self._log(f"Profile: {profile_path}")

        try:
            extractor = ChromeHeadersExtractor(
                chrome_profile_path=profile_path,
                headless=False,
                verbose=True
            )

            # Start browser
            if not extractor.start_browser():
                self._log("Khong khoi dong duoc Chrome", "error")
                return None

            # Navigate to Flow
            if not extractor.navigate_to_flow():
                self._log("Khong navigate duoc den Flow", "error")
                extractor.stop_browser()
                return None

            # Wait for page load (12s like auto_token.py)
            self._log("Doi trang load (12s)...")
            import time
            time.sleep(12)

            # Trigger API request and capture headers
            self._log("Trigger API request de capture headers...")
            captured = extractor.trigger_api_and_capture()

            if captured.is_valid():
                self._log(f"✅ x-browser-validation: {captured.x_browser_validation[:40]}...", "success")
                self._log(f"✅ Authorization: {captured.authorization[:50]}...", "success")

                # Extract token and project_id
                if captured.authorization.startswith("Bearer "):
                    token = captured.authorization[7:]
                    self.config['flow_bearer_token'] = token

                    # Extract project_id from URL if available
                    # (will be captured in next request)

                # Convert to dict for API use
                headers_dict = captured.to_dict()

                # Store for later use
                self.config['captured_headers'] = headers_dict

                extractor.stop_browser()
                return headers_dict
            else:
                self._log("Khong capture duoc headers day du", "error")
                self._log(f"  authorization: {bool(captured.authorization)}")
                self._log(f"  x-browser-validation: {bool(captured.x_browser_validation)}")

                # Fallback: try capture from network logs
                self._log("Thu capture tu network logs...")
                captured = extractor.capture_headers_from_network(timeout=30)

                if captured.is_valid():
                    headers_dict = captured.to_dict()
                    self.config['captured_headers'] = headers_dict
                    extractor.stop_browser()
                    return headers_dict

                extractor.stop_browser()
                return None

        except Exception as e:
            self._log(f"Exception khi capture headers: {e}", "error")
            import traceback
            traceback.print_exc()
            return None

    def generate_images_auto(
        self,
        excel_path: Optional[Path] = None,
        start_scene: int = 1,
        end_scene: Optional[int] = None,
        overwrite: bool = False,
        bearer_token: Optional[str] = None
    ) -> Dict[str, Any]:
        """
        Tao anh tu dong - chon mode dua tren config.

        Mode duoc cau hinh trong settings.yaml (generation_mode: 'chrome' hoac 'api')

        Args:
            excel_path: Duong dan file Excel
            start_scene: Scene bat dau
            end_scene: Scene ket thuc
            overwrite: Ghi de anh da co
            bearer_token: Bearer token (chi can cho API mode)

        Returns:
            Dict voi ket qua
        """
        mode = self._get_generation_mode()
        self._log(f"[AUTO] Generation mode: {mode.upper()}")

        # Check proxy support for API mode
        proxy_api_token = self.config.get('proxy_api_token', '')

        if mode == 'api':
            if proxy_api_token:
                self._log("[AUTO] API mode voi proxy support - su dung proxy de bypass captcha")
                return self.generate_scene_images_api(
                    excel_path=excel_path,
                    start_scene=start_scene,
                    end_scene=end_scene,
                    overwrite=overwrite
                )
            else:
                self._log("[AUTO] API mode khong co proxy token, chuyen sang Chrome mode...")

        # Fall back to Chrome mode
        return self.generate_scene_images(
            excel_path=excel_path,
            start_scene=start_scene,
            end_scene=end_scene,
            overwrite=overwrite
        )

    def generate_from_prompts_auto(
        self,
        prompts: List[Dict],
        excel_path: Optional[Path] = None,
        bearer_token: Optional[str] = None,
        max_setup_retries: int = 3
    ) -> Dict[str, Any]:
        """
        Tao anh tu prompts - tu dong chon mode dua tren config.

        Args:
            prompts: List prompts [{'id': '1', 'prompt': '...'}]
            excel_path: Duong dan Excel
            bearer_token: Bearer token (chi can cho API mode)
            max_setup_retries: So lan retry setup (chi cho Chrome mode)

        Returns:
            Dict voi ket qua
        """
        mode = self._get_generation_mode()
        self._log(f"[AUTO] Generation mode: {mode.upper()}")

        # Check proxy support for API mode
        proxy_api_token = self.config.get('proxy_api_token', '')

        if mode == 'api':
            if proxy_api_token:
                self._log("[AUTO] API mode voi proxy support - su dung proxy de bypass captcha")
                # Use API mode with proxy for prompts
                return self.generate_from_prompts_api(
                    prompts=prompts,
                    excel_path=excel_path,
                    bearer_token=bearer_token
                )
            else:
                self._log("[AUTO] API mode khong co proxy token, chuyen sang Chrome mode...")

        # Fall back to Chrome mode
        return self.generate_from_prompts(
            prompts=prompts,
            excel_path=excel_path,
            max_setup_retries=max_setup_retries
        )

    def generate_all(
        self,
        characters: bool = True,
        scenes: bool = True,
        start_scene: int = 1,
        end_scene: Optional[int] = None,
        overwrite: bool = False,
        bearer_token: Optional[str] = None
    ) -> Dict[str, Any]:
        """
        Tao tat ca anh (nhan vat + scenes).
        Tu dong chon mode dua tren config (generation_mode).

        Args:
            characters: Tao anh nhan vat
            scenes: Tao anh scenes
            start_scene: Scene bat dau
            end_scene: Scene ket thuc
            overwrite: Ghi de
            bearer_token: Bearer token (chi can cho API mode)

        Returns:
            Dict voi ket qua
        """
        results = {
            "characters": {},
            "scenes": {},
        }

        mode = self._get_generation_mode()
        self._log(f"[GENERATE_ALL] Mode: {mode.upper()}")

        try:
            # AUTO-UPDATE: Them filename annotations vao prompts (neu chua co)
            # Giup Flow match uploaded reference images voi prompt
            excel_path = self._find_excel_file()
            if scenes and excel_path and excel_path.exists():
                self._log("[AUTO] Kiem tra va cap nhat filename annotations trong prompts...", "info")
                try:
                    from modules.prompts_generator import PromptsGenerator
                    pg = PromptsGenerator()
                    updated = pg.update_excel_prompts_with_annotations(str(excel_path))
                    if updated:
                        self._log("[AUTO] Da cap nhat prompts voi filename annotations", "success")
                    else:
                        self._log("[AUTO] Khong can cap nhat annotations (da co san hoac khong co ref_files)", "info")
                except ImportError:
                    self._log("[AUTO] Khong the import PromptsGenerator, bo qua auto-update", "warn")
                except Exception as e:
                    self._log(f"[AUTO] Loi khi cap nhat annotations: {e}", "warn")

            if characters:
                # Characters luon dung Chrome mode (can UI de handle consent)
                results["characters"] = self.generate_character_images(overwrite=overwrite)

            if scenes:
                # Scenes su dung mode tu config
                results["scenes"] = self.generate_images_auto(
                    start_scene=start_scene,
                    end_scene=end_scene,
                    overwrite=overwrite,
                    bearer_token=bearer_token
                )
        finally:
            # Dong browser khi xong (chi can cho Chrome mode)
            if mode == 'chrome':
                self.stop_browser()

        return results

    def _escape_js_string(self, s: str) -> str:
        """Escape string cho JavaScript."""
        return (s
            .replace("\\", "\\\\")
            .replace("`", "\\`")
            .replace("$", "\\$")
            .replace("\n", "\\n")
            .replace("\r", "\\r")
            .replace("\t", "\\t"))

    # =========================================================================
    # API MODE - Direct API call without browser
    # =========================================================================

    def generate_scene_images_api(
        self,
        excel_path: Optional[Path] = None,
        start_scene: int = 1,
        end_scene: Optional[int] = None,
        overwrite: bool = False,
        bearer_token: Optional[str] = None
    ) -> Dict[str, Any]:
        """
        Tao anh bang API mode - goi truc tiep batchGenerateImages API.

        Uu diem:
        - Nhanh hon Chrome mode (khong can khoi dong browser)
        - On dinh hon (khong bi loi UI)

        Nhuoc diem:
        - Can bearer token (het han sau ~1h)
        - Khong tu dong xu ly consent dialogs

        Args:
            excel_path: Duong dan file Excel
            start_scene: Scene bat dau (1-indexed)
            end_scene: Scene ket thuc (None = tat ca)
            overwrite: Ghi de anh da co
            bearer_token: Bearer token (bat buoc)

        Returns:
            Dict voi ket qua
        """
        self._log("=" * 60)
        self._log("API MODE - TAO ANH BANG DIRECT API CALL")
        self._log("=" * 60)

        # Import GoogleFlowAPI
        try:
            from modules.google_flow_api import GoogleFlowAPI, AspectRatio
        except ImportError as e:
            return {"success": False, "error": f"Khong import duoc GoogleFlowAPI: {e}"}

        # Check proxy support
        proxy_api_token = self.config.get('proxy_api_token', '')
        use_proxy = bool(proxy_api_token)

        if use_proxy:
            self._log("Proxy API token co san - se su dung proxy de bypass captcha")

        # Check bearer token
        if not bearer_token:
            # Thu lay tu config
            bearer_token = self.config.get('flow_bearer_token', '')

        # Neu van chua co token, capture tu Chrome (PyAutoGUI)
        if not bearer_token:
            self._log("Khong co bearer token, thu tu dong lay...")
            bearer_token = self._auto_extract_token()

        if not bearer_token:
            return {
                "success": False,
                "error": "Can bearer token cho API mode. Chua co token va khong the tu dong lay."
            }

        # Tim file Excel
        if excel_path is None:
            excel_path = self._find_excel_file()

        if excel_path is None or not excel_path.exists():
            return {"success": False, "error": "Khong tim thay file Excel"}

        self._log(f"Excel: {excel_path}")

        # Use extracted flow_project_id if available, otherwise fallback to project_code
        flow_project_id = self.config.get('flow_project_id', self.project_code)
        self._log(f"Project ID: {flow_project_id}")
        self._log(f"Token: {bearer_token[:20]}...{bearer_token[-10:]}")
        if use_proxy:
            self._log(f"Proxy mode: ENABLED (nanoai.pics)")
        else:
            self._log(f"Proxy mode: DISABLED (direct API call)")

        # Create API client with proxy support
        api = GoogleFlowAPI(
            bearer_token=bearer_token,
            project_id=flow_project_id,
            timeout=self.config.get('flow_timeout', 120),
            verbose=self.verbose,
            proxy_api_token=proxy_api_token,
            use_proxy=use_proxy
        )

        # Map aspect ratio
        ar_setting = self.config.get('flow_aspect_ratio', 'landscape')
        ar_map = {
            'landscape': AspectRatio.LANDSCAPE,
            'portrait': AspectRatio.PORTRAIT,
            'square': AspectRatio.SQUARE,
        }
        aspect_ratio = ar_map.get(ar_setting, AspectRatio.LANDSCAPE)

        # Load Excel
        workbook = PromptWorkbook(excel_path)
        workbook.load_or_create()

        # Lay cac scene can tao anh
        all_scenes = workbook.get_scenes()
        scenes_to_process = []

        for scene in all_scenes:
            if scene.scene_id < start_scene:
                continue
            if end_scene is not None and scene.scene_id > end_scene:
                break
            if not scene.img_prompt:
                continue
            if scene.status_img == "done" and not overwrite:
                self.stats["skipped"] += 1
                continue
            scenes_to_process.append(scene)

        if not scenes_to_process:
            self._log("Khong co scene nao can tao anh", "warn")
            return {"success": True, "message": "No scenes to process"}

        self._log(f"Se tao {len(scenes_to_process)} anh bang API")
        self.stats["total"] = len(scenes_to_process)

        # Load media cache cho reference
        cached_media_names = self._load_media_cache()
        if cached_media_names:
            self._log(f"Loaded {len(cached_media_names)} media references")

        # Process tung scene
        for i, scene in enumerate(scenes_to_process):
            scene_id = str(scene.scene_id)
            prompt = scene.img_prompt

            self._log(f"\n[{i+1}/{len(scenes_to_process)}] Scene {scene_id}")
            self._log(f"Prompt ({len(prompt)} chars): {prompt[:100]}...")

            try:
                # Build image inputs from references
                image_inputs = []
                ref_str = getattr(scene, 'reference_files', '') or ''
                if ref_str:
                    try:
                        ref_files = json.loads(ref_str) if ref_str.startswith('[') else [f.strip() for f in ref_str.split(',') if f.strip()]
                    except:
                        ref_files = [f.strip() for f in str(ref_str).split(',') if f.strip()]

                    for ref_file in ref_files:
                        ref_id = ref_file.replace('.png', '').replace('.jpg', '')
                        # Check cache for media_name
                        if ref_id in cached_media_names:
                            media_info = cached_media_names[ref_id]
                            media_name = media_info.get('mediaName') if isinstance(media_info, dict) else media_info
                            if media_name:
                                from modules.google_flow_api import ImageInput, ImageInputType
                                image_inputs.append(ImageInput(
                                    name=media_name,
                                    input_type=ImageInputType.REFERENCE
                                ))
                                self._log(f"[REF] Using cached media: {ref_id}")

                # Generate image - co retry khi token het han
                success, images, error = api.generate_images(
                    prompt=prompt,
                    count=self.config.get('flow_image_count', 1),  # Default 1 ảnh
                    aspect_ratio=aspect_ratio,
                    image_inputs=[inp.to_dict() for inp in image_inputs] if image_inputs else None
                )

                # Check token expired (401) - auto refresh and retry
                if not success and error:
                    error_lower = str(error).lower()
                    is_token_expired = '401' in error_lower or 'expired' in error_lower or 'authentication' in error_lower or 'unauthenticated' in error_lower
                    if is_token_expired:
                        self._log(f"Token het han (401)! Dang mo Chrome lay token moi...", "warn")
                        new_token = self._auto_extract_token(force_refresh=True)  # FORCE lay token moi
                        if new_token:
                            # Update API instance with new token
                            api.bearer_token = new_token
                            bearer_token = new_token
                            self._log(f"Token moi OK - dang retry scene {scene_id}...")

                            # Retry generation
                            success, images, error = api.generate_images(
                                prompt=prompt,
                                count=self.config.get('flow_image_count', 1),  # Default 1 ảnh
                                aspect_ratio=aspect_ratio,
                                image_inputs=[inp.to_dict() for inp in image_inputs] if image_inputs else None
                            )
                        else:
                            self._log(f"Khong the refresh token!", "error")

                if success and images:
                    # Xác định thư mục lưu: nv*/loc* -> nv/, còn lại -> img/
                    scene_id_str = str(scene_id)
                    if scene_id_str.lower().startswith('nv') or scene_id_str.lower().startswith('loc'):
                        save_dir = self.nv_path
                        relative_folder = "nv"
                    else:
                        save_dir = self.img_path
                        relative_folder = "img"

                    output_file = save_dir / f"{scene_id}.png"

                    downloaded = api.download_image(
                        images[0],  # Take first image
                        save_dir,
                        scene_id
                    )

                    if downloaded:
                        # Update Excel
                        relative_path = f"{relative_folder}/{scene_id}.png"
                        workbook.update_scene(
                            scene.scene_id,
                            img_path=relative_path,
                            status_img="done"
                        )
                        workbook.save()

                        # Save media_name to cache
                        if images[0].media_name:
                            cached_media_names[scene_id] = {
                                'mediaName': images[0].media_name,
                                'seed': images[0].seed
                            }

                        self._log(f"OK - Da tao va luu anh: {downloaded}", "success")
                        self.stats["success"] += 1
                    else:
                        self._log("Loi download anh", "error")
                        workbook.update_scene(scene.scene_id, status_img="error")
                        workbook.save()
                        self.stats["failed"] += 1
                else:
                    self._log(f"Loi: {error}", "error")
                    workbook.update_scene(scene.scene_id, status_img="error")
                    workbook.save()
                    self.stats["failed"] += 1

                # Delay giua cac prompt
                delay = self.config.get('flow_delay', 3.0)
                if i < len(scenes_to_process) - 1:
                    time.sleep(delay)

            except Exception as e:
                self._log(f"Exception: {e}", "error")
                import traceback
                traceback.print_exc()
                workbook.update_scene(scene.scene_id, status_img="error")
                workbook.save()
                self.stats["failed"] += 1

        # Save updated media cache
        if cached_media_names:
            self._save_media_cache(cached_media_names)

        # Summary
        self._log("\n" + "=" * 60)
        self._log("HOAN THANH (API MODE)")
        self._log("=" * 60)
        self._log(f"Tong: {self.stats['total']}")
        self._log(f"Thanh cong: {self.stats['success']}")
        self._log(f"That bai: {self.stats['failed']}")
        self._log(f"Bo qua: {self.stats['skipped']}")

        return {
            "success": True,
            "stats": self.stats.copy()
        }

    def generate_scene_videos_api(
        self,
        excel_path: Optional[Path] = None,
        start_scene: int = 1,
        end_scene: Optional[int] = None,
        max_videos: int = 10,
        bearer_token: Optional[str] = None
    ) -> Dict[str, Any]:
        """
        Tao video cho scenes bang API mode (proxy).

        Args:
            excel_path: Duong dan file Excel
            start_scene: Scene bat dau (1-indexed)
            end_scene: Scene ket thuc (None = tat ca)
            max_videos: So video toi da can tao (default 10)
            bearer_token: Bearer token (bat buoc)

        Returns:
            Dict voi ket qua
        """
        self._log("=" * 60)
        self._log("API MODE - TAO VIDEO CHO SCENES")
        self._log("=" * 60)

        # Import GoogleFlowAPI
        try:
            from modules.google_flow_api import GoogleFlowAPI, VideoAspectRatio, VideoModel
        except ImportError as e:
            return {"success": False, "error": f"Khong import duoc GoogleFlowAPI: {e}"}

        # Check proxy support
        proxy_api_token = self.config.get('proxy_api_token', '')
        use_proxy = bool(proxy_api_token)

        if not use_proxy:
            return {"success": False, "error": "Can proxy_api_token de tao video"}

        self._log("Proxy API token co san - su dung proxy de tao video")

        # Check bearer token
        if not bearer_token:
            bearer_token = self.config.get('flow_bearer_token', '')

        if not bearer_token:
            self._log("Khong co bearer token, thu tu dong lay...")
            bearer_token = self._auto_extract_token()

        if not bearer_token:
            return {
                "success": False,
                "error": "Can bearer token cho API mode"
            }

        # Tim file Excel
        if excel_path is None:
            excel_path = self._find_excel_file()

        if excel_path is None or not excel_path.exists():
            return {"success": False, "error": "Khong tim thay file Excel"}

        self._log(f"Excel: {excel_path}")

        # Use extracted flow_project_id if available
        flow_project_id = self.config.get('flow_project_id', self.project_code)
        self._log(f"Project ID: {flow_project_id}")
        self._log(f"Token: {bearer_token[:20]}...{bearer_token[-10:]}")
        self._log(f"Max videos: {max_videos}")

        # Create API client with proxy support
        api = GoogleFlowAPI(
            bearer_token=bearer_token,
            project_id=flow_project_id,
            timeout=self.config.get('flow_timeout', 300),  # Longer timeout for video
            verbose=self.verbose,
            proxy_api_token=proxy_api_token,
            use_proxy=True
        )

        # Map aspect ratio
        ar_setting = self.config.get('flow_aspect_ratio', 'landscape')
        ar_map = {
            'landscape': VideoAspectRatio.LANDSCAPE,
            'portrait': VideoAspectRatio.PORTRAIT,
            'square': VideoAspectRatio.SQUARE,
        }
        aspect_ratio = ar_map.get(ar_setting, VideoAspectRatio.LANDSCAPE)

        # Load Excel
        workbook = PromptWorkbook(excel_path)
        workbook.load_or_create()

        # Lay cac scene can tao video
        all_scenes = workbook.get_scenes()
        scenes_to_process = []

        for scene in all_scenes:
            if scene.scene_id < start_scene:
                continue
            if end_scene is not None and scene.scene_id > end_scene:
                break
            # Chi tao video cho scenes da co anh
            if not scene.img_path:
                continue
            # Bo qua scenes da tao video
            if scene.status_vid == "done":
                self.stats["skipped"] += 1
                continue
            # Lay video_prompt hoac img_prompt
            video_prompt = scene.video_prompt or scene.img_prompt
            if not video_prompt:
                continue
            scenes_to_process.append((scene, video_prompt))
            # Gioi han so video
            if len(scenes_to_process) >= max_videos:
                break

        if not scenes_to_process:
            self._log("Khong co scene nao can tao video", "warn")
            return {"success": True, "message": "No scenes to process"}

        self._log(f"Se tao {len(scenes_to_process)} video bang API")
        self.stats["total"] = len(scenes_to_process)

        # Output folder
        output_folder = excel_path.parent / "img"
        output_folder.mkdir(parents=True, exist_ok=True)

        # Generate videos
        for i, (scene, video_prompt) in enumerate(scenes_to_process):
            scene_id = str(scene.scene_id)
            self._log(f"\n[{i+1}/{len(scenes_to_process)}] Scene: {scene_id}")
            self._log(f"Prompt ({len(video_prompt)} chars): {video_prompt[:80]}...")

            try:
                # Generate video
                success, result, error = api.generate_video(
                    prompt=video_prompt,
                    aspect_ratio=aspect_ratio,
                    model=VideoModel.VEO3_FAST,
                    scene_id=scene_id
                )

                if success and result.video_url:
                    # Download video
                    video_filename = f"scene_{scene_id}"
                    downloaded = api.download_video(result, output_folder, video_filename)

                    if downloaded:
                        # Update Excel - chi luu filename, khong luu full path
                        relative_path = downloaded.name
                        workbook.update_scene(
                            scene.scene_id,
                            video_path=relative_path,
                            status_vid="done"
                        )
                        workbook.save()

                        self._log(f"OK - Da tao va luu video: {downloaded}", "success")
                        self.stats["success"] += 1
                    else:
                        self._log("Loi download video", "error")
                        workbook.update_scene(scene.scene_id, status_vid="error")
                        workbook.save()
                        self.stats["failed"] += 1
                else:
                    self._log(f"Loi: {error}", "error")
                    workbook.update_scene(scene.scene_id, status_vid="error")
                    workbook.save()
                    self.stats["failed"] += 1

                # Delay giua cac video (video mat nhieu thoi gian hon)
                delay = self.config.get('video_delay', 5.0)
                if i < len(scenes_to_process) - 1:
                    time.sleep(delay)

            except Exception as e:
                self._log(f"Exception: {e}", "error")
                import traceback
                traceback.print_exc()
                workbook.update_scene(scene.scene_id, status_vid="error")
                workbook.save()
                self.stats["failed"] += 1

        # Summary
        self._log("\n" + "=" * 60)
        self._log("HOAN THANH VIDEO (API MODE)")
        self._log("=" * 60)
        self._log(f"Tong: {self.stats['total']}")
        self._log(f"Thanh cong: {self.stats['success']}")
        self._log(f"That bai: {self.stats['failed']}")
        self._log(f"Bo qua: {self.stats['skipped']}")

        return {
            "success": True,
            "stats": self.stats.copy()
        }

    def generate_from_prompts_api(
        self,
        prompts: List[Dict],
        excel_path: Optional[Path] = None,
        bearer_token: Optional[str] = None
    ) -> Dict[str, Any]:
        """
        Tao anh tu danh sach prompts bang API mode.

        Args:
            prompts: List prompts [{'id': '1', 'prompt': '...'}]
            excel_path: Duong dan Excel (de cap nhat status)
            bearer_token: Bearer token (bat buoc)

        Returns:
            Dict voi ket qua
        """
        self._log("=" * 60)
        self._log("API MODE - TAO ANH TU PROMPTS")
        self._log("=" * 60)

        # === DRISSION MODE ONLY ===
        # Sử dụng DrissionPage + Interceptor để tạo ảnh
        if not prompts:
            return {"success": False, "error": "Khong co prompts"}

        try:
            from modules.drission_flow_api import DrissionFlowAPI
        except ImportError as e:
            return {"success": False, "error": f"Khong import duoc DrissionFlowAPI: {e}. Cài đặt: pip install DrissionPage"}

        # Webshare proxy config (nested dict from GUI settings)
        # IPv6 proxy đã bị bỏ - chỉ dùng Webshare
        webshare_cfg = self.config.get('webshare_proxy', {})
        webshare_api_key = webshare_cfg.get('api_key', '')
        webshare_proxy_file = webshare_cfg.get('proxy_file', 'config/proxies.txt')  # Default file
        use_webshare = webshare_cfg.get('enabled', True)  # Default ON - Webshare enabled by default

        # === PROXY MODE CONFIG ===
        # "direct" = Direct Proxy List (100 IP cố định)
        # "rotating" = Rotating Residential (IP tự động đổi)
        proxy_mode = webshare_cfg.get('proxy_mode', 'direct')
        rotating_host = webshare_cfg.get('rotating_host', 'p.webshare.io')
        rotating_port = webshare_cfg.get('rotating_port', 80)
        # Ưu tiên base_username, fallback sang username cũ
        rotating_username = webshare_cfg.get('rotating_base_username') or webshare_cfg.get('rotating_username', 'jhvbehdf-residential')
        rotating_password = webshare_cfg.get('rotating_password', 'cf1bi3yvq0t1')
        machine_id = webshare_cfg.get('machine_id', 1)  # Máy số mấy (1-99)

        # Khởi tạo Webshare Proxy Manager nếu enabled
        if use_webshare:

            try:
                from webshare_proxy import init_proxy_manager, get_proxy_manager

                # === ROTATING RESIDENTIAL MODE ===
                if proxy_mode == "rotating":
                    self._log(f"🌍 ROTATING RESIDENTIAL mode")
                    self._log(f"   → {rotating_host}:{rotating_port}")
                    self._log(f"   → User: {rotating_username}")
                    self._log(f"   → Mỗi request sẽ tự động đổi IP!")

                    manager = init_proxy_manager(
                        username=rotating_username,
                        password=rotating_password,
                        rotating_endpoint=True,
                        rotating_host=rotating_host,
                        rotating_port=rotating_port
                    )
                else:
                    # === DIRECT PROXY LIST MODE ===
                    # Load proxy list từ file hoặc API
                    if webshare_proxy_file:
                        self._log(f"Loading proxies from: {webshare_proxy_file}")
                    else:
                        # Kiểm tra file mặc định
                        default_proxy_file = Path("config/proxies.txt")
                        if default_proxy_file.exists():
                            webshare_proxy_file = str(default_proxy_file)
                            self._log(f"Found default proxy file: {webshare_proxy_file}")

                    manager = init_proxy_manager(
                        api_key=webshare_api_key,
                        proxy_file=webshare_proxy_file if webshare_proxy_file else None
                    )

                # Verify initialization
                if manager.is_rotating_mode():
                    self._log(f"✓ Rotating Endpoint ready")
                elif manager.proxies:
                    self._log(f"✓ Loaded {len(manager.proxies)} proxies")
                    self._log(f"  Current: {manager.current_proxy.endpoint}")
                else:
                    self._log("⚠️ No proxies loaded - chạy không có proxy", "WARN")
                    use_webshare = False

            except Exception as e:
                self._log(f"⚠️ Webshare init error: {e} - chạy không có proxy", "WARN")
                use_webshare = False

        # === ĐỌC CONFIG TỪ EXCEL/CACHE TRƯỚC (để biết profile nào đã dùng) ===
        saved_project_url = None
        saved_chrome_profile = None  # Chrome profile đã dùng cho dự án này
        self._log(f"[DEBUG] excel_path = {excel_path}")

        # 1. Thử đọc từ Excel sheet 'config'
        if excel_path and Path(excel_path).exists():
            self._log(f"[DEBUG] Excel tồn tại, kiểm tra sheet 'config'...")
            try:
                import openpyxl
                wb = openpyxl.load_workbook(excel_path)
                self._log(f"[DEBUG] Sheets có: {wb.sheetnames}")
                if 'config' in wb.sheetnames:
                    ws = wb['config']
                    config_keys_found = []
                    for row in ws.iter_rows(min_row=2, max_row=20, values_only=True):
                        if row and len(row) >= 2:
                            key = str(row[0] or '').strip().lower()
                            val = str(row[1] or '').strip() if row[1] else ''
                            config_keys_found.append(key)
                            if key == 'flow_project_url' and val and '/project/' in val:
                                saved_project_url = val
                                self._log(f"📂 Tìm thấy project URL từ Excel: {saved_project_url[:50]}...")
                            elif key == 'flow_project_id' and val and not saved_project_url:
                                # Nếu chỉ có project_id, tạo URL
                                saved_project_url = f"https://labs.google/fx/vi/tools/flow/project/{val}"
                                self._log(f"📂 Tìm thấy project_id từ Excel: {val[:20]}...")
                            elif key == 'chrome_profile_path' and val:
                                # Đọc Chrome profile đã dùng cho dự án này
                                if Path(val).exists():
                                    saved_chrome_profile = val
                                    self._log(f"📂 Tìm thấy Chrome profile từ Excel: {val}")
                    if not saved_project_url:
                        self._log(f"[DEBUG] Config keys: {config_keys_found}")
                else:
                    self._log(f"[DEBUG] Không có sheet 'config' trong Excel")
                wb.close()
            except Exception as e:
                self._log(f"⚠️ Không đọc được config từ Excel: {e}", "warn")

        # Chọn profile: ưu tiên saved profile từ Excel, fallback về default
        if saved_chrome_profile:
            profile_to_use = saved_chrome_profile
            self._log(f"🔄 Dùng Chrome profile đã lưu: {profile_to_use}")
        else:
            profile_to_use = self._get_profile_path() or "./chrome_profile"
            self._log(f"📁 Dùng Chrome profile mặc định: {profile_to_use}")

        # Đọc setting headless từ config (default: True = chạy ẩn)
        # Dùng chung setting 'browser_headless' với Selenium mode
        drission_headless = self.config.get('browser_headless', True)

        drission_api = DrissionFlowAPI(
            profile_dir=profile_to_use,
            verbose=self.verbose,
            log_callback=self._log,
            webshare_enabled=use_webshare,
            worker_id=self.worker_id,  # Parallel mode - mỗi worker có proxy riêng
            headless=drission_headless,  # Chạy Chrome ẩn (default: True)
            machine_id=machine_id  # Máy số mấy - tránh trùng session giữa các máy
        )

        self._log("🚀 DrissionPage + Interceptor")
        if use_webshare:
            manager = get_proxy_manager()
            if manager.is_rotating_mode():
                self._log(f"   Proxy: 🔄 ROTATING ENDPOINT (auto IP change)")
            else:
                self._log(f"   Proxy: Webshare Pool ({len(manager.proxies)} proxies)")
        else:
            self._log("   Proxy: OFF (không có proxy)")

        # 2. Fallback: Thử đọc từ cache file
        if not saved_project_url and excel_path:
            cache_path = Path(excel_path).parent / ".media_cache.json"
            self._log(f"[DEBUG] Cache path: {cache_path}")
            if cache_path.exists():
                try:
                    import json
                    with open(cache_path, 'r', encoding='utf-8') as f:
                        cache_data = json.load(f)
                    self._log(f"[DEBUG] Cache keys: {list(cache_data.keys())[:10]}")
                    cached_url = cache_data.get('_project_url', '')
                    cached_id = cache_data.get('_project_id', '')
                    if cached_url and '/project/' in cached_url:
                        saved_project_url = cached_url
                        self._log(f"📂 Tìm thấy project URL từ cache: {saved_project_url[:50]}...")
                    elif cached_id:
                        saved_project_url = f"https://labs.google/fx/vi/tools/flow/project/{cached_id}"
                        self._log(f"📂 Tìm thấy project_id từ cache: {cached_id[:20]}...")
                except Exception as e:
                    self._log(f"⚠️ Không đọc được cache: {e}", "warn")
            else:
                self._log(f"[DEBUG] Cache file không tồn tại")

        if saved_project_url:
            self._log(f"🔄 Sẽ vào lại project cũ để giữ media_id...")
        else:
            self._log(f"📝 Sẽ tạo project mới...")

        # Setup Chrome và đợi user chọn project (với retry + IP rotation)
        MAX_SETUP_RETRIES = 3
        setup_success = False

        for setup_attempt in range(MAX_SETUP_RETRIES):
            if drission_api.setup(project_url=saved_project_url):
                setup_success = True

                # === LƯU PROJECT URL VÀO EXCEL NGAY SAU KHI SETUP ===
                # Đảm bảo 1 voice = 1 project link (không bị mất nếu fail giữa chừng)
                new_project_url = getattr(drission_api, '_current_project_url', '')
                if new_project_url and '/project/' in new_project_url and excel_path:
                    try:
                        import openpyxl
                        wb = openpyxl.load_workbook(excel_path)
                        if 'config' not in wb.sheetnames:
                            wb.create_sheet('config')
                        ws = wb['config']

                        # Tìm hoặc thêm row cho flow_project_url
                        found = False
                        for row_num in range(2, ws.max_row + 1):
                            if ws.cell(row=row_num, column=1).value == 'flow_project_url':
                                ws.cell(row=row_num, column=2, value=new_project_url)
                                found = True
                                break
                        if not found:
                            next_row = ws.max_row + 1
                            ws.cell(row=next_row, column=1, value='flow_project_url')
                            ws.cell(row=next_row, column=2, value=new_project_url)

                        # Cũng lưu chrome_profile_path
                        profile_path = str(drission_api.profile_dir) if hasattr(drission_api, 'profile_dir') else ''
                        if profile_path:
                            found = False
                            for row_num in range(2, ws.max_row + 1):
                                if ws.cell(row=row_num, column=1).value == 'chrome_profile_path':
                                    ws.cell(row=row_num, column=2, value=profile_path)
                                    found = True
                                    break
                            if not found:
                                next_row = ws.max_row + 1
                                ws.cell(row=next_row, column=1, value='chrome_profile_path')
                                ws.cell(row=next_row, column=2, value=profile_path)

                        wb.save(excel_path)
                        wb.close()
                        self._log(f"✓ Lưu project URL vào Excel: {new_project_url[:50]}...")
                    except Exception as e:
                        self._log(f"⚠️ Không lưu được project URL: {e}", "warn")

                break
            else:
                self._log(f"❌ Setup failed (attempt {setup_attempt + 1}/{MAX_SETUP_RETRIES})", "error")

                if setup_attempt < MAX_SETUP_RETRIES - 1:
                    # Rotate IP và restart Chrome
                    self._log("🔄 Đang rotate IP và restart Chrome...", "warn")
                    if use_webshare:
                        try:
                            manager = get_proxy_manager()
                            success, msg = manager.rotate_worker_proxy(self.worker_id, "setup_timeout")
                            self._log(f"   → {msg}")
                            if success and drission_api.restart_chrome():
                                self._log("✓ Chrome restarted với IP mới")
                                import time
                                time.sleep(3)
                                continue
                        except Exception as e:
                            self._log(f"   → Rotate error: {e}", "warn")

        if not setup_success:
            self._log("❌ DrissionFlowAPI setup failed sau tất cả retries!", "error")
            return {"success": False, "error": "DrissionFlowAPI setup failed"}

        self._log(f"Tong: {len(prompts)} prompts")

        # Store reference
        self._drission_api = drission_api

        # Generate images - use self.img_path as output directory
        return self._generate_images_drission_mode(prompts, self.img_path, excel_path)

    def _generate_images_drission_mode(
        self,
        prompts: List[Dict[str, Any]],
        output_dir: Path,
        excel_path: Optional[Path] = None
    ) -> Dict[str, Any]:
        """
        Generate images sử dụng DrissionFlowAPI (DrissionPage + Interceptor).

        Flow:
        1. DrissionFlowAPI đã setup sẵn Chrome + interceptor
        2. Mỗi prompt: gửi qua Chrome để capture tokens → gọi API → save ảnh
        3. Lưu vào output_dir và cập nhật Excel nếu có

        Args:
            prompts: Danh sách prompts
            output_dir: Thư mục lưu ảnh
            excel_path: Path đến Excel file (optional)

        Returns:
            Dict với stats
        """
        from modules.excel_manager import PromptWorkbook

        self._log("=" * 60)
        self._log("DRISSION MODE - Generate Images")
        self._log("=" * 60)

        drission_api = getattr(self, '_drission_api', None)
        if not drission_api:
            return {"success": False, "error": "DrissionFlowAPI chưa được khởi tạo"}

        # Reset stats
        self.stats = {"total": len(prompts), "success": 0, "failed": 0, "skipped": 0}

        # Track failed prompts để retry sau
        failed_prompts = []  # List[Tuple[prompt_data, index, error]]

        # Load Excel workbook
        workbook = None
        if excel_path and Path(excel_path).exists():
            try:
                workbook = PromptWorkbook(excel_path)
                workbook.load_or_create()
            except Exception as e:
                self._log(f"Warning: Khong load duoc Excel: {e}", "warn")

        # === LOAD MEDIA_IDs từ Excel (thay vì cache file) ===
        excel_media_ids = {}
        if workbook:
            try:
                excel_media_ids = workbook.get_media_ids()
                if excel_media_ids:
                    self._log(f"[EXCEL] Loaded {len(excel_media_ids)} media_ids: {list(excel_media_ids.keys())}")
                else:
                    self._log("[EXCEL] ⚠️ KHÔNG CÓ MEDIA_ID TRONG EXCEL - ảnh nv/loc sẽ được tạo lại", "warn")
            except Exception as e:
                self._log(f"Warning: Cannot load media_ids from Excel: {e}", "warn")

        # Fallback to cache file nếu Excel không có data
        cached_media_names = {}
        if not excel_media_ids:
            cached_media_names = self._load_media_cache()
            if cached_media_names:
                self._log(f"[CACHE] Fallback: Loaded {len(cached_media_names)} media references from cache")

        # Ensure output dir exists
        output_dir = Path(output_dir)
        output_dir.mkdir(parents=True, exist_ok=True)

        # Log danh sách prompts sẽ xử lý
        prompt_ids = [str(p.get('id', idx+1)) for idx, p in enumerate(prompts)]
        ref_ids = [pid for pid in prompt_ids if pid.lower().startswith('nv') or pid.lower().startswith('loc')]
        if ref_ids:
            self._log(f"[INFO] Reference images (nv/loc): {ref_ids}")

        for i, prompt_data in enumerate(prompts):
            pid = str(prompt_data.get('id', i + 1))
            prompt = prompt_data.get('prompt', '')

            if not prompt:
                self._log(f"[{i+1}/{len(prompts)}] ID: {pid} - Skip (prompt rong)", "warn")
                self.stats["skipped"] += 1
                continue

            # Skip DO_NOT_GENERATE markers
            if prompt.strip().upper() == "DO_NOT_GENERATE":
                self._log(f"[{i+1}/{len(prompts)}] ID: {pid} - Skip (DO_NOT_GENERATE)")
                self.stats["skipped"] += 1
                continue

            # Xác định là ảnh tham chiếu (nv*/loc*) hay ảnh scene
            is_reference_image = pid.lower().startswith('nv') or pid.lower().startswith('loc')

            # Xác định thư mục lưu: nv*/loc* -> nv_path (tham chiếu), còn lại -> img_path
            if is_reference_image:
                save_dir = self.nv_path
            else:
                save_dir = output_dir

            # Check if image already exists
            output_file = save_dir / f"{pid}.png"
            if output_file.exists():
                # === CHECK MEDIA_ID FOR REFERENCE IMAGES ===
                # Nếu là ảnh nv*/loc* nhưng KHÔNG có media_id → xóa và tạo lại
                # Normalize key để so sánh (case-insensitive)
                pid_lower = pid.lower()
                has_media_id = any(k.lower() == pid_lower for k in excel_media_ids.keys())

                if is_reference_image and not has_media_id:
                    self._log(f"[{i+1}/{len(prompts)}] ID: {pid} - ⚠️ ANH TON TAI NHUNG KHONG CO MEDIA_ID")
                    self._log(f"   → Dang xoa {output_file.name} de tao lai...")
                    try:
                        output_file.unlink()  # Xóa file
                        self._log(f"   → Da xoa! Se tao lai de co media_id", "success")
                    except Exception as e:
                        self._log(f"   → Khong the xoa file: {e}", "warn")
                        self.stats["skipped"] += 1
                        continue
                    # Tiếp tục generate (không skip)
                else:
                    self._log(f"[{i+1}/{len(prompts)}] ID: {pid} - Skip (da co anh)")
                    self.stats["skipped"] += 1
                    continue

            self._log(f"[{i+1}/{len(prompts)}] ID: {pid}")
            self._log(f"   Prompt: {prompt[:60]}...")

            # === BUILD IMAGE_INPUTS từ reference_files và media_ids ===
            image_inputs = []
            # is_reference_image đã được định nghĩa ở trên

            # Merge Excel và cache media_ids (Excel ưu tiên)
            all_media_ids = {**cached_media_names, **excel_media_ids}

            if not is_reference_image and all_media_ids:
                # Parse reference_files từ prompt_data
                ref_str = prompt_data.get('reference_files', '')
                ref_files = []
                if ref_str:
                    try:
                        parsed = json.loads(ref_str) if ref_str.startswith('[') else None
                        if isinstance(parsed, list):
                            ref_files = parsed
                        else:
                            ref_files = [f.strip() for f in str(ref_str).split(',') if f.strip()]
                    except:
                        ref_files = [f.strip() for f in str(ref_str).split(',') if f.strip()]

                # Fallback: nếu không có reference, dùng nvc.png mặc định
                if not ref_files:
                    ref_files = ["nvc.png"]
                    self._log(f"   [REF] No reference, using default nvc.png")

                # Build image_inputs từ media_ids (Excel hoặc cache)
                # QUAN TRỌNG: Dùng "imageInputType" (không phải "inputType") với giá trị đầy đủ
                for ref_file in ref_files:
                    ref_id = ref_file.replace('.png', '').replace('.jpg', '')
                    # Thử Excel media_id trước
                    if ref_id in excel_media_ids:
                        media_id = excel_media_ids[ref_id]
                        image_inputs.append({
                            "name": media_id,
                            "imageInputType": "IMAGE_INPUT_TYPE_REFERENCE"
                        })
                        self._log(f"   [REF] Using (Excel): {ref_id} → {media_id[:30]}...")
                    elif ref_id in cached_media_names:
                        # Fallback to cache
                        media_info = cached_media_names[ref_id]
                        media_name = media_info.get('mediaName') if isinstance(media_info, dict) else media_info
                        if media_name:
                            image_inputs.append({
                                "name": media_name,
                                "imageInputType": "IMAGE_INPUT_TYPE_REFERENCE"
                            })
                            self._log(f"   [REF] Using (cache): {ref_id} → {media_name[:30]}...")

                if image_inputs:
                    self._log(f"   [REF] Total: {len(image_inputs)} reference images")
                else:
                    self._log(f"   [REF] No media_id found for references", "warn")

            try:
                # Generate image using DrissionFlowAPI with reference images
                success, images, error = drission_api.generate_image(
                    prompt=prompt,
                    save_dir=save_dir,
                    filename=pid,
                    image_inputs=image_inputs if image_inputs else None
                )

                if success and images:
                    self._log(f"   ✓ Thành công! Saved {len(images)} image(s)")
                    self.stats["success"] += 1
                    consecutive_403 = 0  # Reset counter on success

                    # Update Excel if available - SAVE MEDIA_ID for SCENE images
                    if workbook and not is_reference_image:
                        try:
                            # Dùng update_scene với tất cả fields cần update
                            workbook.update_scene(
                                int(pid),
                                img_path=str(images[0].local_path) if images[0].local_path else None,
                                status_img="done",
                                media_id=images[0].media_name if images[0].media_name else None
                            )
                            workbook.save()
                            if images[0].media_name:
                                self._log(f"   [EXCEL] Saved scene {pid}: media_id={images[0].media_name[:40]}...")
                            elif pid.isdigit():
                                # Scene image but no media_name - this will cause I2V to skip
                                self._log(f"   ⚠️ Scene {pid}: API không trả về media_name (I2V sẽ không hoạt động)", "warn")
                        except Exception as e:
                            self._log(f"   [EXCEL] Cannot update scene {pid}: {e}", "warn")

                    # === SAVE MEDIA_ID to Excel for nv/loc images ===
                    # Cả nv* và loc* đều nằm trong sheet "characters"
                    if images[0].media_name and is_reference_image:
                        media_id_saved = False
                        if workbook:
                            try:
                                # update_character works for both nv* and loc* (same sheet)
                                if workbook.update_character(pid, media_id=images[0].media_name):
                                    workbook.save()
                                    self._log(f"   [EXCEL] Saved media_id for {pid}: {images[0].media_name[:40]}...")
                                    excel_media_ids[pid] = images[0].media_name
                                    media_id_saved = True
                                else:
                                    self._log(f"   [EXCEL] {pid} not found in characters sheet", "warn")
                            except Exception as e:
                                self._log(f"   [EXCEL] Cannot save media_id: {e}", "warn")

                        # Fallback: save to cache file if Excel update fails
                        if not media_id_saved:
                            try:
                                cached_media_names[pid] = {'mediaName': images[0].media_name}
                                self._save_media_cache(cached_media_names)
                                self._log(f"   [CACHE] Fallback - saved media_id for {pid}")
                                excel_media_ids[pid] = images[0].media_name
                            except Exception as e:
                                self._log(f"   [CACHE] Cannot save media_id: {e}", "warn")
                    elif images[0].media_name:
                        self._log(f"   Media name: {images[0].media_name[:40]}...")
                else:
                    self._log(f"   ✗ Thất bại: {error}", "error")
                    self.stats["failed"] += 1

                    # Check for token expiry - thử refresh và retry
                    if error and "401" in str(error):
                        self._log("⚠️ Bearer token hết hạn - thử restart Chrome...", "warn")
                        try:
                            if drission_api.restart_chrome():
                                self._log(f"→ Retry prompt: {pid}...", "info")
                                success2, images2, error2 = drission_api.generate_image(
                                    prompt=prompt,
                                    save_dir=save_dir,
                                    filename=pid,
                                    image_inputs=image_inputs if image_inputs else None
                                )
                                if success2 and images2:
                                    self._log(f"   ✓ Retry thành công!")
                                    self.stats["success"] += 1
                                    self.stats["failed"] -= 1
                                    consecutive_errors = 0
                                    continue
                                else:
                                    self._log(f"   ✗ Retry vẫn thất bại - token không hợp lệ", "error")
                                    break  # Token thực sự hết hạn
                            else:
                                self._log("✗ Không restart được Chrome - dừng", "error")
                                break
                        except Exception as e:
                            self._log(f"✗ Refresh token error: {e}", "error")
                            break

                    # Check for 429 - Quota exceeded, cần đổi proxy/tài khoản
                    if error and "429" in str(error):
                        self._log(f"⚠️ Lỗi 429 (Quota) - Restart Chrome + đổi proxy...", "warn")
                        try:
                            if drission_api.restart_chrome():
                                self._log(f"→ Retry prompt: {pid}...", "info")
                                success2, images2, error2 = drission_api.generate_image(
                                    prompt=prompt,
                                    save_dir=save_dir,
                                    filename=pid,
                                    image_inputs=image_inputs if image_inputs else None
                                )
                                if success2 and images2:
                                    self._log(f"   ✓ Retry thành công!")
                                    self.stats["success"] += 1
                                    self.stats["failed"] -= 1
                                    consecutive_errors = 0
                                    continue
                                else:
                                    self._log(f"   ✗ Retry vẫn thất bại: {error2}", "warn")
                            else:
                                self._log("✗ Không restart được Chrome", "error")
                        except Exception as e:
                            self._log(f"✗ Restart error: {e}", "error")

                    # Check for 400 - Invalid argument (reference image expired or invalid prompt)
                    if error and "400" in str(error):
                        self._log(f"⚠️ Lỗi 400 - Restart Chrome + retry không có reference...", "warn")
                        try:
                            # Restart Chrome trước khi retry
                            if drission_api.restart_chrome():
                                # Retry without reference images
                                success2, images2, error2 = drission_api.generate_image(
                                    prompt=prompt,
                                    save_dir=save_dir,
                                    filename=pid,
                                    image_inputs=None  # No reference images
                                )
                                if success2 and images2:
                                    self._log(f"   ✓ Retry (no ref) thành công!")
                                    self.stats["success"] += 1
                                    self.stats["failed"] -= 1  # Undo fail count
                                    consecutive_errors = 0
                                    continue  # Move to next prompt
                                else:
                                    self._log(f"   ✗ Retry (no ref) thất bại: {error2}", "warn")
                        except Exception as e:
                            self._log(f"   ✗ Retry exception: {e}", "error")

                    # Check for 403 - restart Chrome với proxy mới
                    if error and "403" in str(error):
                        self._log(f"⚠️ Lỗi 403 - Restart Chrome với proxy mới...", "warn")
                        try:
                            # Restart Chrome (clear blocked IPs + restart với proxy)
                            if drission_api.restart_chrome():
                                # Retry current prompt with new Chrome session
                                self._log(f"→ Retry prompt: {pid}...", "info")
                                success2, images2, error2 = drission_api.generate_image(
                                    prompt=prompt,
                                    save_dir=save_dir,
                                    filename=pid,
                                    image_inputs=image_inputs if image_inputs else None
                                )
                                if success2 and images2:
                                    self._log(f"   ✓ Retry thành công! Saved {len(images2)} image(s)")
                                    self.stats["success"] += 1
                                    self.stats["failed"] -= 1  # Undo the fail count
                                    # Save media_id for scene images
                                    if images2[0].media_name and not is_reference_image:
                                        try:
                                            if workbook:
                                                workbook.update_scene(int(pid), media_id=images2[0].media_name)
                                                workbook.save()
                                                self._log(f"   [EXCEL] Saved media_id for scene {pid}")
                                        except:
                                            pass
                                    # Save media_id for nv/loc images
                                    if images2[0].media_name and is_reference_image:
                                        media_id_saved = False
                                        if workbook:
                                            try:
                                                if workbook.update_character(pid, media_id=images2[0].media_name):
                                                    workbook.save()
                                                    self._log(f"   [EXCEL] Saved media_id for {pid}")
                                                    excel_media_ids[pid] = images2[0].media_name
                                                    media_id_saved = True
                                            except:
                                                pass
                                        # Fallback to cache
                                        if not media_id_saved:
                                            try:
                                                cached_media_names[pid] = {'mediaName': images2[0].media_name}
                                                self._save_media_cache(cached_media_names)
                                                excel_media_ids[pid] = images2[0].media_name
                                            except:
                                                pass
                                    elif images2[0].media_name:
                                        self._log(f"   Media name: {images2[0].media_name[:40]}...")
                                else:
                                    self._log(f"   ✗ Retry vẫn thất bại: {error2}", "error")
                            else:
                                self._log("✗ Không restart được Chrome", "error")
                                break
                        except Exception as e:
                            self._log(f"✗ Restart error: {e}", "error")
                            break

            except Exception as e:
                self._log(f"   ✗ Exception: {e}", "error")
                self.stats["failed"] += 1

            # Rate limit
            time.sleep(1)

        # Save workbook (trước retry phase)
        if workbook:
            try:
                workbook.save()
            except:
                pass

        # === RETRY PHASE: Tìm và retry những ảnh còn thiếu ===
        if self.stats["failed"] > 0:
            self._log("\n" + "=" * 60)
            self._log("RETRY PHASE - Tìm ảnh còn thiếu")
            self._log("=" * 60)

            # Tìm những prompts chưa có ảnh
            missing_prompts = []
            for prompt_data in prompts:
                pid = str(prompt_data.get('id', ''))
                if not pid:
                    continue

                is_reference = pid.lower().startswith('nv') or pid.lower().startswith('loc')
                save_dir = self.nv_path if is_reference else output_dir
                output_file = save_dir / f"{pid}.png"

                if not output_file.exists():
                    missing_prompts.append(prompt_data)

            if missing_prompts:
                self._log(f"Tìm thấy {len(missing_prompts)} ảnh thiếu, đang retry...")

                # Retry up to 3 rounds
                MAX_RETRY_ROUNDS = 3
                for retry_round in range(MAX_RETRY_ROUNDS):
                    if not missing_prompts:
                        break

                    self._log(f"\n--- Retry Round {retry_round + 1}/{MAX_RETRY_ROUNDS} ---")
                    still_missing = []

                    for prompt_data in missing_prompts:
                        pid = str(prompt_data.get('id', ''))
                        prompt = prompt_data.get('prompt', '')
                        is_reference = pid.lower().startswith('nv') or pid.lower().startswith('loc')
                        save_dir = self.nv_path if is_reference else output_dir

                        self._log(f"[RETRY] {pid}...")

                        try:
                            # Build reference images if needed
                            image_inputs = None
                            refs = prompt_data.get('references', [])
                            if refs and not is_reference:
                                image_inputs = []
                                for ref_id in refs:
                                    ref_key = ref_id.lower()
                                    media_id = None
                                    for k, v in excel_media_ids.items():
                                        if k.lower() == ref_key:
                                            media_id = v
                                            break
                                    if not media_id:
                                        for k, v in cached_media_names.items():
                                            if k.lower() == ref_key:
                                                media_id = v.get('mediaName', v) if isinstance(v, dict) else v
                                                break
                                    if media_id:
                                        image_inputs.append({
                                            "inputType": "IMAGE_INPUT_TYPE_REFERENCE",
                                            "referenceId": media_id,
                                            "referenceType": "REFERENCE_TYPE_STYLE"
                                        })

                            success, images, error = drission_api.generate_image(
                                prompt=prompt,
                                save_dir=save_dir,
                                filename=pid,
                                image_inputs=image_inputs
                            )

                            if success and images:
                                self._log(f"   ✓ Retry OK: {pid}")
                                self.stats["success"] += 1
                                self.stats["failed"] -= 1
                            else:
                                self._log(f"   ✗ Retry fail: {error}", "warn")
                                still_missing.append(prompt_data)

                        except Exception as e:
                            self._log(f"   ✗ Retry error: {e}", "error")
                            still_missing.append(prompt_data)

                        time.sleep(2)

                    missing_prompts = still_missing

                    # Wait before next round
                    if missing_prompts and retry_round < MAX_RETRY_ROUNDS - 1:
                        wait_time = 5 * (retry_round + 1)
                        self._log(f"Còn {len(missing_prompts)} ảnh thiếu, đợi {wait_time}s...")
                        time.sleep(wait_time)

                if missing_prompts:
                    self._log(f"⚠️ Vẫn còn {len(missing_prompts)} ảnh không tạo được sau {MAX_RETRY_ROUNDS} rounds", "warn")
            else:
                self._log("Tất cả ảnh đã có, không cần retry")

        # === LƯU TOKEN VÀO EXCEL + CACHE CHO VIDEO WORKER ===
        # Quan trọng: Video worker cần token để tạo I2V
        try:
            if drission_api.bearer_token and drission_api.project_id:
                bearer = drission_api.bearer_token
                if bearer.startswith("Bearer "):
                    bearer = bearer[7:]  # Remove "Bearer " prefix
                project_id = drission_api.project_id
                # Lấy recaptcha_token nếu có (quan trọng cho I2V!)
                recaptcha = getattr(drission_api, 'recaptcha_token', '') or ''
                x_browser_val = getattr(drission_api, 'x_browser_validation', '') or ''

                # 1. Lưu vào config để _save_media_cache có thể đọc
                self.config['flow_bearer_token'] = bearer
                self.config['flow_project_id'] = project_id
                self.config['flow_recaptcha_token'] = recaptcha
                self.config['flow_x_browser_validation'] = x_browser_val

                # 2. Lưu vào Excel (sheet config) để tái sử dụng
                if workbook:
                    try:
                        import openpyxl
                        wb = openpyxl.load_workbook(excel_path)

                        # Tạo hoặc lấy sheet 'config'
                        if 'config' not in wb.sheetnames:
                            ws = wb.create_sheet('config')
                            ws['A1'] = 'key'
                            ws['B1'] = 'value'
                            next_row = 2
                        else:
                            ws = wb['config']
                            next_row = ws.max_row + 1

                        # Lấy project URL để lưu (cho lần chạy tiếp theo vào đúng project)
                        project_url = getattr(drission_api, '_current_project_url', '')
                        if not project_url and project_id:
                            project_url = f"https://labs.google/fx/vi/tools/flow/project/{project_id}"

                        # Lưu Chrome profile path để resume đúng profile
                        chrome_profile_path = str(drission_api.profile_dir) if hasattr(drission_api, 'profile_dir') else ''

                        # Lưu các config - đầy đủ để tái sử dụng cho I2V
                        config_items = {
                            'flow_project_id': project_id,
                            'flow_project_url': project_url,  # URL để vào lại project cũ
                            'flow_bearer_token': bearer,  # Full token để video worker dùng
                            'flow_recaptcha_token': recaptcha,  # Quan trọng cho I2V!
                            'flow_x_browser_validation': x_browser_val,  # Auth header
                            'token_time': str(int(time.time())),
                            'chrome_profile_path': chrome_profile_path  # Profile để resume đúng Chrome
                        }

                        for key, value in config_items.items():
                            # Tìm row có key này để update
                            found = False
                            for row_num in range(2, ws.max_row + 1):
                                if ws.cell(row=row_num, column=1).value == key:
                                    ws.cell(row=row_num, column=2, value=value)
                                    found = True
                                    break
                            if not found:
                                ws.cell(row=next_row, column=1, value=key)
                                ws.cell(row=next_row, column=2, value=value)
                                next_row += 1

                        wb.save(excel_path)
                        wb.close()
                        self._log(f"[EXCEL] Saved project_id + token to Excel")
                    except Exception as e:
                        self._log(f"[EXCEL] Warning: Cannot save to Excel: {e}", "warn")

                # 3. Lưu full token vào media cache (để video worker dùng)
                self._save_media_cache(cached_media_names)
                self._log(f"[CACHE] Saved full token for video worker")
        except Exception as e:
            self._log(f"[CACHE] Warning: Cannot save token: {e}", "warn")

        # === I2V: TẠO VIDEO TỪ ẢNH (CÙNG SESSION CHROME) ===
        video_count_setting = self.config.get('video_count', 0)
        try:
            if video_count_setting == 'full':
                video_count = 999999  # Tất cả
            else:
                video_count = int(video_count_setting)
        except:
            video_count = 0

        # === KIỂM TRA: Tất cả scene images đã xong chưa? ===
        all_images_done = True
        pending_scenes = []
        if video_count > 0 and workbook:
            try:
                for scene in workbook.get_scenes():
                    scene_id = str(scene.scene_id) if hasattr(scene, 'scene_id') else ''
                    if not scene_id or not scene_id.isdigit():
                        continue  # Bỏ qua nv/loc

                    # Kiểm tra status_img
                    status_img = getattr(scene, 'status_img', '') or ''
                    prompt = getattr(scene, 'prompt', '') or ''

                    # Nếu có prompt và chưa done → chưa xong
                    if prompt and prompt.strip().upper() != 'DO_NOT_GENERATE' and status_img != 'done':
                        all_images_done = False
                        pending_scenes.append(scene_id)
            except Exception as e:
                self._log(f"[I2V] Warning: Cannot check pending scenes: {e}", "warn")

        if not all_images_done:
            self._log(f"[I2V] ⏳ SKIP - Còn {len(pending_scenes)} scene chưa có ảnh: {pending_scenes[:10]}...")
            self._log(f"[I2V] Video sẽ được tạo sau khi tất cả ảnh scene hoàn thành")
        elif video_count > 0 and drission_api._ready:
            self._log("")
            self._log("=" * 60)
            self._log(f"[I2V] TẠO VIDEO TỪ ẢNH (cùng session)")
            self._log("=" * 60)

            # Lấy danh sách scenes cần tạo video (có media_id, chưa có video)
            scenes_for_video = []
            scenes_without_media_id = []
            if workbook:
                try:
                    all_scenes = workbook.get_scenes()
                    self._log(f"[I2V] Loaded {len(all_scenes)} scenes from Excel")

                    for scene in all_scenes:
                        # Chỉ lấy scene (không phải character/location)
                        scene_id = str(scene.scene_id) if hasattr(scene, 'scene_id') else ''
                        if not scene_id or not scene_id.isdigit():
                            continue

                        # Kiểm tra có media_id và chưa có video
                        media_id = getattr(scene, 'media_id', '') or ''
                        video_path = getattr(scene, 'video_path', '') or ''
                        status_vid = getattr(scene, 'status_vid', '') or ''

                        if not media_id:
                            scenes_without_media_id.append(scene_id)
                        elif not video_path and status_vid != 'done':
                            video_prompt = getattr(scene, 'video_prompt', '') or 'Subtle cinematic motion'
                            scenes_for_video.append({
                                'scene_id': scene_id,
                                'media_id': media_id,
                                'video_prompt': video_prompt
                            })

                    if scenes_without_media_id:
                        self._log(f"[I2V] ⚠️ {len(scenes_without_media_id)} scenes KHÔNG CÓ media_id: {scenes_without_media_id[:5]}{'...' if len(scenes_without_media_id) > 5 else ''}", "warn")
                except Exception as e:
                    self._log(f"[I2V] Error loading scenes: {e}", "warn")

            # Fallback: Lấy từ cached_media_names
            if not scenes_for_video and cached_media_names:
                for pid, media_info in cached_media_names.items():
                    if pid.isdigit():  # Chỉ scenes (số)
                        media_id = media_info.get('mediaName', media_info) if isinstance(media_info, dict) else media_info
                        if media_id:
                            scenes_for_video.append({
                                'scene_id': pid,
                                'media_id': media_id,
                                'video_prompt': 'Subtle cinematic motion, slow camera movement'
                            })

            # Giới hạn số lượng
            scenes_for_video = scenes_for_video[:video_count]

            if scenes_for_video:
                self._log(f"[I2V] Tạo video cho {len(scenes_for_video)} ảnh...")
                video_success = 0
                video_failed = 0

                for i, scene_info in enumerate(scenes_for_video):
                    scene_id = scene_info['scene_id']
                    media_id = scene_info['media_id']
                    video_prompt = scene_info['video_prompt']

                    self._log(f"[I2V] [{i+1}/{len(scenes_for_video)}] Scene {scene_id}...")

                    try:
                        # generate_video sẽ tự refresh recaptcha token (one-time token)
                        success, video_url, error = drission_api.generate_video(
                            media_id=media_id,
                            prompt=video_prompt,
                            video_model="veo_3_0_r2v_fast_ultra"
                        )

                        if success and video_url:
                            # Download video - lưu vào img/ folder (giống smart_engine)
                            video_dir = output_dir  # img/ folder
                            video_file = video_dir / f"{scene_id}.mp4"

                            try:
                                import requests as req
                                resp = req.get(video_url, timeout=60)
                                if resp.status_code == 200:
                                    video_file.write_bytes(resp.content)
                                    self._log(f"   ✓ OK: {video_file.name}")
                                    video_success += 1

                                    # Update Excel
                                    if workbook:
                                        workbook.update_scene(int(scene_id), video_path=video_file.name, status_vid='done')
                                        workbook.save()
                                else:
                                    self._log(f"   ✗ Download failed: {resp.status_code}", "warn")
                                    video_failed += 1
                            except Exception as dl_err:
                                self._log(f"   ✗ Download error: {dl_err}", "warn")
                                video_failed += 1
                        else:
                            self._log(f"   ✗ Failed: {error}", "warn")
                            video_failed += 1

                    except Exception as e:
                        self._log(f"   ✗ Error: {e}", "error")
                        video_failed += 1

                    # Delay giữa các video
                    time.sleep(3)

                self._log(f"[I2V] Hoàn tất: {video_success} OK, {video_failed} failed")
            else:
                self._log(f"[I2V] Không có ảnh nào cần tạo video")
        elif video_count > 0:
            self._log(f"[I2V] Bỏ qua - DrissionAPI chưa sẵn sàng")

        # Cleanup (sau I2V)
        try:
            drission_api.close()
        except:
            pass

        # Save workbook final
        if workbook:
            try:
                workbook.save()
            except:
                pass

        # Summary
        self._log("=" * 60)
        self._log("HOAN THANH (DRISSION MODE)")
        self._log("=" * 60)
        self._log(f"Tong: {self.stats['total']}")
        self._log(f"Thanh cong: {self.stats['success']}")
        self._log(f"That bai: {self.stats['failed']}")
        self._log(f"Bo qua: {self.stats['skipped']}")

        return {
            "success": True,
            "stats": self.stats.copy()
        }

    def __enter__(self):
        """Context manager entry."""
        self.start_browser()
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        """Context manager exit."""
        self.stop_browser()


# =============================================================================
# HELPER FUNCTIONS
# =============================================================================

def create_browser_flow_generator(
    project_path: str,
    profile_name: str = "main",
    headless: bool = False,
    verbose: bool = True
) -> BrowserFlowGenerator:
    """
    Factory function de tao BrowserFlowGenerator.

    Args:
        project_path: Duong dan project
        profile_name: Ten Chrome profile
        headless: Chay an
        verbose: In log

    Returns:
        BrowserFlowGenerator instance
    """
    return BrowserFlowGenerator(
        project_path=project_path,
        profile_name=profile_name,
        headless=headless,
        verbose=verbose
    )


def generate_images_from_excel(
    project_path: str,
    profile_name: str = "main",
    headless: bool = False,
    characters: bool = True,
    scenes: bool = True,
    start_scene: int = 1,
    end_scene: Optional[int] = None,
    overwrite: bool = False
) -> Dict[str, Any]:
    """
    Ham tien ich de tao anh tu Excel.

    Args:
        project_path: Duong dan project (PROJECTS/KA1-0001)
        profile_name: Ten Chrome profile
        headless: Chay an
        characters: Tao anh nhan vat
        scenes: Tao anh scenes
        start_scene: Scene bat dau
        end_scene: Scene ket thuc
        overwrite: Ghi de anh cu

    Returns:
        Dict voi ket qua
    """
    generator = BrowserFlowGenerator(
        project_path=project_path,
        profile_name=profile_name,
        headless=headless,
        verbose=True
    )

    return generator.generate_all(
        characters=characters,
        scenes=scenes,
        start_scene=start_scene,
        end_scene=end_scene,
        overwrite=overwrite
    )


# =============================================================================
# CLI
# =============================================================================

if __name__ == "__main__":
    print("""
+============================================================+
|      BROWSER FLOW GENERATOR - VE3 TOOL                     |
+============================================================+
|  Tao anh tu Excel bang browser automation                  |
|                                                            |
|  Usage:                                                    |
|    python browser_flow_generator.py <project_path>         |
|                                                            |
|  Options:                                                  |
|    --profile <name>    Chrome profile name (default: main) |
|    --headless          Chay an (khong hien UI)             |
|    --characters        Chi tao anh nhan vat                |
|    --scenes            Chi tao anh scenes                  |
|    --start N           Bat dau tu scene N                  |
|    --end N             Ket thuc o scene N                  |
|    --overwrite         Ghi de anh da co                    |
+============================================================+
""")

    if not SELENIUM_AVAILABLE:
        print("Error: Selenium chua duoc cai dat")
        print("Chay: pip install selenium undetected-chromedriver")
        sys.exit(1)

    if len(sys.argv) < 2:
        print("Vui long cung cap duong dan project")
        print("Vi du: python browser_flow_generator.py ./PROJECTS/KA1-0001")
        sys.exit(1)

    project_path = sys.argv[1]

    # Parse options
    profile_name = "main"
    headless = "--headless" in sys.argv
    do_characters = "--characters" in sys.argv or "--scenes" not in sys.argv
    do_scenes = "--scenes" in sys.argv or "--characters" not in sys.argv
    overwrite = "--overwrite" in sys.argv
    start_scene = 1
    end_scene = None

    for i, arg in enumerate(sys.argv):
        if arg == "--profile" and i + 1 < len(sys.argv):
            profile_name = sys.argv[i + 1]
        if arg == "--start" and i + 1 < len(sys.argv):
            start_scene = int(sys.argv[i + 1])
        if arg == "--end" and i + 1 < len(sys.argv):
            end_scene = int(sys.argv[i + 1])

    # Run
    results = generate_images_from_excel(
        project_path=project_path,
        profile_name=profile_name,
        headless=headless,
        characters=do_characters,
        scenes=do_scenes,
        start_scene=start_scene,
        end_scene=end_scene,
        overwrite=overwrite
    )

    # Exit code
    scenes_failed = results.get("scenes", {}).get("stats", {}).get("failed", 0)
    chars_failed = results.get("characters", {}).get("characters_failed", 0)

    sys.exit(0 if (scenes_failed + chars_failed) == 0 else 1)
