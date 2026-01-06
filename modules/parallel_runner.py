"""
VE3 Tool - Parallel Runner
==========================
1 nut chay tat ca:
- Lay token cho TAT CA profiles (tuan tu vi can Chrome)
- Chay tao anh SONG SONG voi tat ca accounts
"""

import json
import time
import threading
from pathlib import Path
from typing import List, Dict, Optional, Callable
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass, field


@dataclass
class Account:
    """1 tai khoan = 1 Chrome profile + token."""
    profile_path: str
    token: str = ""
    project_id: str = ""
    name: str = ""
    
    def __post_init__(self):
        if not self.name:
            self.name = Path(self.profile_path).name


class ParallelRunner:
    """
    Chay song song voi nhieu tai khoan.
    """
    
    def __init__(self, config_path: str = None):
        self.config_path = Path(config_path or "config/accounts.json")
        self.chrome_path = r"C:\Program Files\Google\Chrome\Application\chrome.exe"
        self.accounts: List[Account] = []
        self.groq_keys: List[str] = []
        self.gemini_keys: List[str] = []
        self.parallel = 2
        self.delay = 2
        
        self.stop_flag = False
        self.callback = None
        
        self.load_config()
    
    def log(self, msg: str):
        print(f"[Runner] {msg}")
        if self.callback:
            self.callback(msg)
    
    def load_config(self):
        """Load accounts.json."""
        if not self.config_path.exists():
            self.log(f"File khong ton tai: {self.config_path}")
            return
        
        try:
            with open(self.config_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
            
            self.chrome_path = data.get('chrome_path', self.chrome_path)
            
            # Load profiles - support both string and dict format
            profiles = data.get('chrome_profiles', [])
            self.accounts = []
            for i, p in enumerate(profiles):
                path = None
                if isinstance(p, str):
                    path = p
                elif isinstance(p, dict):
                    path = p.get('path', '')
                
                if path and not path.startswith('C:\\Users\\admin\\AppData\\Local\\Google'):
                    # Real path (not placeholder)
                    self.accounts.append(Account(
                        profile_path=path,
                        name=f"Account{i+1}"
                    ))
                elif path and Path(path).exists():
                    self.accounts.append(Account(
                        profile_path=path,
                        name=f"Account{i+1}"
                    ))
            
            # API Keys
            api = data.get('api_keys', {})
            self.groq_keys = [k for k in api.get('groq', []) if k and not k.startswith('gsk_YOUR')]
            self.gemini_keys = [k for k in api.get('gemini', []) if k and not k.startswith('AIzaSy_YOUR')]
            
            # Settings
            settings = data.get('settings', {})
            self.parallel = settings.get('parallel', 2)
            self.delay = settings.get('delay_between_images', 2)
            
            self.log(f"Loaded: {len(self.accounts)} profiles, {len(self.groq_keys)} Groq keys, parallel={self.parallel}")
            
        except Exception as e:
            self.log(f"Load error: {e}")
    
    def save_config(self):
        """Save tokens back to file."""
        try:
            with open(self.config_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
            
            # Update profiles with tokens
            data['_tokens'] = {
                acc.profile_path: {
                    'token': acc.token,
                    'project_id': acc.project_id
                }
                for acc in self.accounts if acc.token
            }
            
            with open(self.config_path, 'w', encoding='utf-8') as f:
                json.dump(data, f, indent=4, ensure_ascii=False)
                
        except Exception as e:
            self.log(f"Save error: {e}")
    
    def get_all_tokens(self, callback: Callable = None) -> int:
        """
        Lay token cho TAT CA profiles.
        Chay TUAN TU vi can Chrome GUI.
        Return: so luong token thanh cong.
        """
        from modules.auto_token import ChromeAutoToken
        
        self.callback = callback
        success = 0
        
        self.log(f"=== LAY TOKEN CHO {len(self.accounts)} ACCOUNTS ===")
        
        for i, acc in enumerate(self.accounts):
            if self.stop_flag:
                break
            
            self.log(f"[{i+1}/{len(self.accounts)}] {acc.name}: {acc.profile_path}")
            
            # Skip if already has token
            if acc.token:
                self.log(f"  -> Da co token, skip")
                success += 1
                continue
            
            # Check path exists
            if not Path(acc.profile_path).exists():
                self.log(f"  -> Path khong ton tai, skip")
                continue
            
            try:
                extractor = ChromeAutoToken(
                    chrome_path=self.chrome_path,
                    profile_path=acc.profile_path
                )

                # Log project_id status
                if acc.project_id:
                    self.log(f"  -> Reuse project_id: {acc.project_id[:8]}...")

                # QUAN TRONG: Reuse project_id da co de share media_ids
                token, proj_id, error = extractor.extract_token(
                    project_id=acc.project_id,  # Reuse existing project
                    callback=callback
                )

                if token:
                    acc.token = token
                    # Chi update project_id neu co gia tri moi
                    if proj_id:
                        acc.project_id = proj_id
                    success += 1
                    self.log(f"  -> OK! Token: {token[:20]}...")
                    self.log(f"  -> Project ID: {acc.project_id[:8]}..." if acc.project_id else "")
                else:
                    self.log(f"  -> FAIL: {error}")

            except Exception as e:
                self.log(f"  -> ERROR: {e}")
            
            # Doi giua cac profile
            if i < len(self.accounts) - 1:
                self.log("Doi 3s truoc profile tiep theo...")
                time.sleep(3)
        
        self.save_config()
        self.log(f"=== XONG: {success}/{len(self.accounts)} tokens ===")
        return success
    
    def get_active_accounts(self) -> List[Account]:
        """Lay danh sach accounts co token."""
        return [acc for acc in self.accounts if acc.token]
    
    def generate_images_parallel(
        self,
        prompts: List[Dict],  # [{id, prompt, output_path}, ...]
        callback: Callable = None
    ) -> Dict:
        """
        Tao anh SONG SONG voi tat ca accounts co token.
        Chia deu prompts cho cac accounts.
        """
        from modules.google_flow_api import GoogleFlowAPI, AspectRatio
        
        self.callback = callback
        active = self.get_active_accounts()
        
        if not active:
            self.log("KHONG CO ACCOUNT NAO CO TOKEN!")
            return {"success": 0, "failed": len(prompts)}
        
        n_workers = min(len(active), self.parallel)
        self.log(f"=== CHAY {n_workers} ACCOUNTS SONG SONG ===")
        self.log(f"Tong: {len(prompts)} prompts")
        
        # Chia prompts cho cac workers
        chunks = [[] for _ in range(n_workers)]
        for i, p in enumerate(prompts):
            chunks[i % n_workers].append(p)
        
        # Stats
        results = {"success": 0, "failed": 0, "details": []}
        lock = threading.Lock()
        
        def worker(worker_id: int, account: Account, prompt_list: List[Dict]):
            """Worker thread."""
            self.log(f"[W{worker_id}] Start: {account.name} - {len(prompt_list)} prompts")
            
            api = GoogleFlowAPI(
                bearer_token=account.token,
                project_id=account.project_id
            )
            
            for pd in prompt_list:
                if self.stop_flag:
                    break
                
                pid = pd.get('id', '')
                prompt = pd.get('prompt', '')
                output = pd.get('output_path', '')
                
                self.log(f"[W{worker_id}] {pid}: {prompt[:40]}...")
                
                try:
                    # Generate
                    img_data = api.generate_image(
                        prompt=prompt,
                        aspect_ratio=AspectRatio.LANDSCAPE
                    )
                    
                    if img_data:
                        # Save
                        if output:
                            Path(output).parent.mkdir(parents=True, exist_ok=True)
                            with open(output, 'wb') as f:
                                f.write(img_data)
                        
                        with lock:
                            results["success"] += 1
                            results["details"].append({"id": pid, "status": "OK"})
                        self.log(f"[W{worker_id}] {pid}: OK!")
                    else:
                        with lock:
                            results["failed"] += 1
                            results["details"].append({"id": pid, "status": "NO_DATA"})
                        self.log(f"[W{worker_id}] {pid}: No data")
                        
                except Exception as e:
                    with lock:
                        results["failed"] += 1
                        results["details"].append({"id": pid, "status": str(e)})
                    self.log(f"[W{worker_id}] {pid}: ERROR - {e}")
                
                # Delay
                time.sleep(self.delay)
            
            self.log(f"[W{worker_id}] Done: {account.name}")
        
        # Start workers
        threads = []
        for i in range(n_workers):
            if chunks[i]:
                t = threading.Thread(
                    target=worker,
                    args=(i+1, active[i], chunks[i]),
                    daemon=True
                )
                t.start()
                threads.append(t)
        
        # Wait
        for t in threads:
            t.join()
        
        self.log(f"=== XONG: {results['success']} OK, {results['failed']} FAIL ===")
        return results
    
    def run_full_pipeline(
        self,
        prompts: List[Dict],
        callback: Callable = None
    ) -> Dict:
        """
        CHAY TAT CA:
        1. Lay token cho tat ca profiles
        2. Chay song song tao anh
        """
        self.callback = callback
        self.stop_flag = False
        
        # Step 1: Get all tokens
        self.log("========== BUOC 1: LAY TOKEN ==========")
        n_tokens = self.get_all_tokens(callback)
        
        if n_tokens == 0:
            self.log("KHONG LAY DUOC TOKEN NAO!")
            return {"success": 0, "failed": len(prompts), "error": "no_tokens"}
        
        if self.stop_flag:
            return {"success": 0, "failed": len(prompts), "error": "stopped"}
        
        # Step 2: Generate images parallel
        self.log("========== BUOC 2: TAO ANH SONG SONG ==========")
        results = self.generate_images_parallel(prompts, callback)
        
        return results
    
    def stop(self):
        """Dung."""
        self.stop_flag = True
        self.log("STOPPING...")


# ============================================================================
# HELPER
# ============================================================================

def load_prompts_from_excel(excel_path: str) -> List[Dict]:
    """Load prompts tu Excel."""
    import openpyxl
    
    prompts = []
    wb = openpyxl.load_workbook(excel_path)
    
    # Tim sheet "scenes" hoac "characters"
    for sheet_name in ['scenes', 'characters', 'Sheet1']:
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            headers = [c.value for c in ws[1]]
            
            id_col = None
            prompt_col = None
            
            for i, h in enumerate(headers):
                if h and 'id' in h.lower():
                    id_col = i
                elif h and 'english' in h.lower() and 'prompt' in h.lower():
                    prompt_col = i
                elif h and h.lower() == 'prompt':
                    prompt_col = i
            
            if id_col is not None and prompt_col is not None:
                for row in ws.iter_rows(min_row=2, values_only=True):
                    pid = row[id_col]
                    prompt = row[prompt_col]
                    if pid and prompt:
                        prompts.append({
                            'id': str(pid),
                            'prompt': str(prompt),
                            'output_path': ''
                        })
            break
    
    return prompts


if __name__ == "__main__":
    runner = ParallelRunner()
    print(f"Accounts: {len(runner.accounts)}")
    print(f"Groq keys: {len(runner.groq_keys)}")
    print(f"Parallel: {runner.parallel}")
