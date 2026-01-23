"""
Quick check AR8-0003 Excel sheets
"""
import sys
import io
from pathlib import Path
from openpyxl import load_workbook

# Fix encoding
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

excel_file = Path(__file__).parent / "PROJECTS" / "AR8-0003" / "AR8-0003_prompts.xlsx"

print("=" * 80)
print("QUICK CHECK AR8-0003 EXCEL")
print("=" * 80)
print()

if not excel_file.exists():
    print("[ERROR] Excel not found!")
    sys.exit(1)

wb = load_workbook(str(excel_file), read_only=True)

print(f"File size: {excel_file.stat().st_size / 1024:.1f} KB")
print(f"Sheets: {len(wb.sheetnames)}")
print()

# Count rows in each sheet
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    rows = list(ws.values)
    row_count = len(rows) - 1  # Exclude header

    print(f"  {sheet_name:20s}: {row_count:4d} rows")

    # Show first row for important sheets
    if sheet_name in ["director_plan", "scenes"] and len(rows) > 1:
        print(f"    First data row: {rows[1][:3]}...")

wb.close()

print()
print("=" * 80)

# Try to load with PromptWorkbook
sys.path.insert(0, str(Path(__file__).parent))
from modules.excel_manager import PromptWorkbook

print()
print("Loading with PromptWorkbook...")
workbook = PromptWorkbook(str(excel_file))

scenes = workbook.get_scenes()
print(f"  Scenes loaded: {len(scenes)}")

if len(scenes) > 0:
    # Check first scene
    scene = scenes[0]
    print(f"  First scene:")
    print(f"    scene_id: {scene.scene_id}")
    print(f"    video_note: '{getattr(scene, 'video_note', 'N/A')}'")
    print(f"    img_prompt: {getattr(scene, 'img_prompt', 'N/A')[:50]}...")

print()
print("=" * 80)
