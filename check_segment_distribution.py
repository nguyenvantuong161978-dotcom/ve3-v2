"""
Check segment distribution - verify segment_id assignment
"""
import sys
import io
from pathlib import Path

# Fix encoding
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
sys.path.insert(0, str(Path(__file__).parent))

from modules.excel_manager import PromptWorkbook

excel_file = Path(__file__).parent / "PROJECTS" / "AR8-0003" / "AR8-0003_prompts.xlsx"

print("=" * 80)
print("CHECK SEGMENT DISTRIBUTION")
print("=" * 80)
print()

workbook = PromptWorkbook(str(excel_file))

# Check segments sheet first
print("SEGMENTS SHEET:")
print("-" * 80)
segments = workbook.get_story_segments()
print(f"Total segments: {len(segments)}")
for seg in segments:
    seg_id = seg.get('segment_id', '?')
    srt_start = seg.get('srt_start', '?')
    srt_end = seg.get('srt_end', '?')
    scene_count = seg.get('scene_count', '?')
    print(f"  Segment {seg_id}: SRT {srt_start}-{srt_end}, {scene_count} scenes planned")
print()

# Check director_plan sheet
print("DIRECTOR_PLAN SHEET:")
print("-" * 80)
director_plan = workbook.get_director_plan()
print(f"Total director_plan entries: {len(director_plan)}")

segment_counts = {}
for plan in director_plan:
    segment_id = plan.get('segment_id', 0)
    if segment_id not in segment_counts:
        segment_counts[segment_id] = 0
    segment_counts[segment_id] += 1

print("Director plan by segment:")
for seg_id in sorted(segment_counts.keys()):
    count = segment_counts[seg_id]
    print(f"  Segment {seg_id}: {count} entries")
print()

# Check scenes sheet
print("SCENES SHEET:")
print("-" * 80)
scenes = workbook.get_scenes()
print(f"Total scenes: {len(scenes)}")

scene_segment_counts = {}
for scene in scenes:
    segment_id = getattr(scene, 'segment_id', 0)
    if segment_id not in scene_segment_counts:
        scene_segment_counts[segment_id] = []
    scene_segment_counts[segment_id].append(scene.scene_id)

print("Scenes by segment:")
for seg_id in sorted(scene_segment_counts.keys()):
    scene_ids = scene_segment_counts[seg_id]
    print(f"  Segment {seg_id}: {len(scene_ids)} scenes")
    print(f"    Scene IDs: {scene_ids[:10]}{'...' if len(scene_ids) > 10 else ''}")
print()

# Check video_note distribution by actual segment
print("VIDEO_NOTE BY SEGMENT:")
print("-" * 80)
for seg_id in sorted(scene_segment_counts.keys()):
    scene_ids = scene_segment_counts[seg_id]
    scenes_in_segment = [s for s in scenes if s.scene_id in scene_ids]

    video_note_counts = {}
    for scene in scenes_in_segment:
        note = scene.video_note if scene.video_note else ''
        if note not in video_note_counts:
            video_note_counts[note] = 0
        video_note_counts[note] += 1

    print(f"\nSegment {seg_id}:")
    for note, count in sorted(video_note_counts.items()):
        note_display = repr(note) if note else "''"
        print(f"  video_note={note_display}: {count} scenes")

print()
print("=" * 80)

# Check if segment_id is actually stored in Excel
print("\nCHECK RAW EXCEL DATA:")
print("-" * 80)

from openpyxl import load_workbook
wb = load_workbook(str(excel_file), read_only=True)
if "scenes" in wb.sheetnames:
    ws = wb["scenes"]
    header = [cell.value for cell in ws[1]]
    print(f"Scenes sheet headers ({len(header)} columns):")
    for i, h in enumerate(header, 1):
        if h and 'segment' in h.lower():
            print(f"  Column {i}: {h} ⭐")
        else:
            print(f"  Column {i}: {h}")

    # Check segment_id column position
    if 'segment_id' in header:
        segment_col_idx = header.index('segment_id') + 1
        print(f"\nsegment_id is at column {segment_col_idx}")

        # Check first 10 rows
        print("\nFirst 10 scenes - segment_id values:")
        for row_idx in range(2, min(12, ws.max_row + 1)):
            scene_id = ws.cell(row=row_idx, column=1).value
            segment_id = ws.cell(row=row_idx, column=segment_col_idx).value
            print(f"  Scene {scene_id}: segment_id={segment_id}")
    else:
        print("\n⚠️ WARNING: segment_id column NOT FOUND in scenes sheet!")

wb.close()

print()
print("=" * 80)
print("CHECK COMPLETE")
print("=" * 80)
