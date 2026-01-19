#!/usr/bin/env python3
"""
VE3 Tool - Worker VIDEO Mode (Video Generation ONLY)
Watch for images in Excel, create videos from them.

Usage:
    python run_worker_video.py                     (quét và xử lý tự động)
    python run_worker_video.py AR47-0028           (chạy 1 project cụ thể)
    python run_worker_video.py --video-count 5     (giới hạn số video)
"""

import sys
import os

# Fix Windows encoding issues
if sys.platform == "win32":
    if sys.stdout:
        sys.stdout.reconfigure(encoding='utf-8', errors='replace')
    if sys.stderr:
        sys.stderr.reconfigure(encoding='utf-8', errors='replace')
    os.environ['PYTHONIOENCODING'] = 'utf-8'
import time
import shutil
from pathlib import Path

# Add current directory to path
TOOL_DIR = Path(__file__).parent
sys.path.insert(0, str(TOOL_DIR))

# Import từ run_worker (dùng chung logic)
from run_worker import (
    detect_auto_path,
    get_channel_from_folder,
    matches_channel,
    is_project_complete_on_master,
    has_excel_with_prompts,
    SCAN_INTERVAL,
)

# Detect paths
AUTO_PATH = detect_auto_path()
if AUTO_PATH:
    MASTER_PROJECTS = AUTO_PATH / "ve3-tool-simple" / "PROJECTS"
    MASTER_VISUAL = AUTO_PATH / "VISUAL"
else:
    MASTER_PROJECTS = Path(r"\\tsclient\D\AUTO\ve3-tool-simple\PROJECTS")
    MASTER_VISUAL = Path(r"\\tsclient\D\AUTO\VISUAL")

LOCAL_PROJECTS = TOOL_DIR / "PROJECTS"
WORKER_CHANNEL = get_channel_from_folder()

# Global settings
WORKER_ID = 0
TOTAL_WORKERS = 1
VIDEO_COUNT = -1  # -1 = full (all images)
VIDEO_SCAN_INTERVAL = 5  # Scan nhanh hơn để bắt kịp ảnh mới


def get_scenes_with_images(project_dir: Path, name: str) -> list:
    """Get list of scenes that have images but no videos yet."""
    img_dir = project_dir / "img"
    if not img_dir.exists():
        return []

    scenes_need_video = []

    try:
        from modules.excel_manager import PromptWorkbook
        excel_path = project_dir / f"{name}_prompts.xlsx"

        if not excel_path.exists():
            return []

        wb = PromptWorkbook(str(excel_path))
        scenes = wb.get_scenes()
        scene_media_ids = wb.get_scene_media_ids()

        for scene in scenes:
            scene_id = str(scene.scene_id)

            # Check có media_id không (đã upload ảnh)
            media_id = scene_media_ids.get(scene_id, '')
            if not media_id:
                continue

            # Check đã có video chưa
            video_path = img_dir / f"{scene_id}.mp4"
            if video_path.exists():
                continue

            # Có media_id + chưa có video → cần tạo
            scenes_need_video.append({
                'scene_id': scene_id,
                'media_id': media_id,
                'video_prompt': scene.video_prompt or "Subtle cinematic motion"
            })

    except Exception as e:
        print(f"    Error reading Excel: {e}")

    return scenes_need_video


def count_videos(project_dir: Path) -> int:
    """Count existing videos in project."""
    img_dir = project_dir / "img"
    if not img_dir.exists():
        return 0
    return len(list(img_dir.glob("*.mp4")))


def process_project_video(code: str, video_count: int = -1, callback=None) -> bool:
    """Process a single project - VIDEO ONLY."""
    global WORKER_ID

    def log(msg, level="INFO"):
        if callback:
            callback(msg, level)
        else:
            print(msg)

    local_dir = LOCAL_PROJECTS / code

    if not local_dir.exists():
        log(f"  [FAIL] Project not found locally: {code}")
        return False

    excel_path = local_dir / f"{code}_prompts.xlsx"
    if not excel_path.exists():
        log(f"  [FAIL] Excel not found: {code}")
        return False

    # Get scenes that need videos
    scenes = get_scenes_with_images(local_dir, code)

    if not scenes:
        log(f"  [VIDEO] No scenes need video for: {code}")
        return True

    # Apply video_count limit
    if video_count > 0:
        scenes = scenes[:video_count]

    log(f"\n{'='*60}")
    log(f"[VIDEO] Processing: {code} ({len(scenes)} videos to create)")
    log(f"{'='*60}")

    # Setup DrissionPage API for video
    try:
        import yaml
        from modules.drission_flow_api import DrissionFlowAPI

        # Load config
        config = {}
        config_path = TOOL_DIR / "config" / "settings.yaml"
        if config_path.exists():
            with open(config_path, 'r', encoding='utf-8') as f:
                config = yaml.safe_load(f) or {}

        # Use separate Chrome profile for video
        chrome_portable = config.get('chrome_portable', '')
        chrome_portable_2 = config.get('chrome_portable_2', chrome_portable)

        # Get project URL from Excel metadata or cache
        project_url = None

        # Method 1: Read from Excel sheet 'config' (same as BrowserFlowGenerator)
        try:
            import openpyxl
            wb_xl = openpyxl.load_workbook(excel_path, data_only=True)

            # Try sheet 'config' first
            if 'config' in wb_xl.sheetnames:
                ws = wb_xl['config']
                log(f"  [LIST] Reading from sheet 'config'...")
            else:
                ws = wb_xl.active
                log(f"  [LIST] No 'config' sheet, using active sheet...")

            for row in ws.iter_rows(min_row=1, max_row=30, values_only=True):
                if not row:
                    continue

                # Method 1a: Find URL directly in any cell
                for cell_val in row:
                    if cell_val and isinstance(cell_val, str):
                        cell_str = str(cell_val).strip()
                        if '/project/' in cell_str and cell_str.startswith('http'):
                            project_url = cell_str
                            log(f"  [LIST] Found project URL in Excel!")
                            break

                if project_url:
                    break

                # Method 1b: Key-value format (column A = key, column B = value)
                if len(row) >= 2 and row[0]:
                    key = str(row[0]).strip().lower()
                    val = str(row[1] or '').strip()
                    if key == 'flow_project_url' and '/project/' in val:
                        project_url = val
                        log(f"  [LIST] Found project URL from key 'flow_project_url'!")
                        break
                    elif key == 'flow_project_id' and val:
                        project_url = f"https://labs.google/fx/vi/tools/flow/project/{val}"
                        log(f"  [LIST] Built project URL from project_id!")
                        break

            wb_xl.close()
        except Exception as e:
            log(f"  [WARN] Error reading Excel: {e}")

        # Method 2: Read from .media_cache.json (same as SmartEngine uses)
        # Try both flat and nested locations
        if not project_url:
            cache_locations = [
                local_dir / ".media_cache.json",  # Flat structure
                local_dir / "prompts" / ".media_cache.json",  # Nested structure
            ]
            for cache_file in cache_locations:
                if cache_file.exists():
                    try:
                        import json
                        with open(cache_file, 'r', encoding='utf-8') as f:
                            cache_data = json.load(f)
                        project_url = cache_data.get('_project_url', '')
                        if not project_url:
                            project_id = cache_data.get('_project_id', '')
                            if project_id:
                                project_url = f"https://labs.google/fx/vi/tools/flow/project/{project_id}"
                        if project_url:
                            log(f"  [PKG] Found project URL from cache: {cache_file.name}")
                            break
                    except Exception as e:
                        log(f"  [WARN] Error reading cache {cache_file}: {e}")

        if not project_url:
            log(f"  [FAIL] No project URL in Excel or cache!")
            log(f"  [TIP] Run run_worker_pic first to create images and save project URL")
            return False

        log(f"  [LIST] Project URL: {project_url[:50]}...")

        # Create API instance for video Chrome
        api = DrissionFlowAPI(
            worker_id=100 + WORKER_ID,  # Different worker_id for video
            headless=False,
            profile_dir="./chrome_profiles/video",
            chrome_portable=chrome_portable_2
        )

        # Setup Chrome
        if not api.setup(project_url=project_url, skip_mode_selection=True):
            log(f"  [FAIL] Failed to setup Chrome for video!")
            return False

        # Chuyển sang mode T2V ("Từ văn bản sang video")
        # Interceptor sẽ convert T2V request → I2V request
        log(f"  [VIDEO] Switching to T2V mode...")
        if api.switch_to_t2v_mode():
            log(f"  [v] Switched to T2V mode (Từ văn bản sang video)")
        else:
            log(f"  [WARN] Could not switch to T2V mode, trying anyway...", "WARN")
        time.sleep(1)

        # Create videos
        img_dir = local_dir / "img"
        video_created = 0

        consecutive_failures = 0  # Đếm số lần fail liên tiếp
        MAX_CONSECUTIVE_FAILURES = 5  # Nếu fail 5 lần liên tiếp, restart Chrome

        for scene_info in scenes:
            scene_id = scene_info['scene_id']
            media_id = scene_info['media_id']
            video_prompt = scene_info['video_prompt']

            mp4_path = img_dir / f"{scene_id}.mp4"

            log(f"\n  [VIDEO] Creating video: {scene_id}")
            log(f"     Media ID: {media_id[:40]}...")
            log(f"     Prompt: {video_prompt[:50]}...")

            # Kiểm tra Chrome còn sống không
            if not api._ready or api.driver is None:
                log(f"     [WARN] Chrome không sẵn sàng, restart...", "WARN")
                try:
                    if api.restart_chrome():
                        log(f"     [v] Chrome restarted")
                        consecutive_failures = 0
                    else:
                        log(f"     [x] Không restart được Chrome, skip video {scene_id}", "WARN")
                        continue
                except Exception as e:
                    log(f"     [x] Restart error: {e}, skip video {scene_id}", "WARN")
                    continue

            try:
                # Use T2V→I2V MODE:
                # - Chrome ở mode "Từ văn bản sang video" (T2V)
                # - Interceptor convert: T2V request → I2V request
                # - Đổi URL, thêm referenceImages, đổi model
                ok, result_path, error = api.generate_video_t2v_mode(
                    media_id=media_id,
                    prompt=video_prompt,
                    save_path=mp4_path
                )

                if ok:
                    video_created += 1
                    consecutive_failures = 0  # Reset counter on success
                    log(f"     [v] Video created: {scene_id}.mp4")

                    # Di chuyển ảnh gốc sang thư mục img_src để tránh nhầm lẫn khi edit
                    png_path = img_dir / f"{scene_id}.png"
                    if png_path.exists():
                        img_src_dir = local_dir / "img_src"
                        img_src_dir.mkdir(exist_ok=True)
                        dst_path = img_src_dir / f"{scene_id}.png"
                        try:
                            shutil.move(str(png_path), str(dst_path))
                            log(f"     [PKG] Moved image to img_src/")
                        except Exception as e:
                            log(f"     [WARN] Cannot move image: {e}", "WARN")
                else:
                    consecutive_failures += 1
                    log(f"     [x] Failed: {error} (fail {consecutive_failures}/{MAX_CONSECUTIVE_FAILURES})", "WARN")

                    # Nếu fail nhiều lần liên tiếp, restart Chrome
                    if consecutive_failures >= MAX_CONSECUTIVE_FAILURES:
                        log(f"     [WARN] {consecutive_failures} failures, restarting Chrome...", "WARN")
                        try:
                            if api.restart_chrome():
                                log(f"     [v] Chrome restarted")
                                consecutive_failures = 0
                        except:
                            pass

            except Exception as e:
                consecutive_failures += 1
                log(f"     [x] Exception: {e} (fail {consecutive_failures}/{MAX_CONSECUTIVE_FAILURES})", "ERROR")

                # Nếu fail nhiều lần liên tiếp, restart Chrome
                if consecutive_failures >= MAX_CONSECUTIVE_FAILURES:
                    log(f"     [WARN] {consecutive_failures} failures, restarting Chrome...", "WARN")
                    try:
                        if api.restart_chrome():
                            log(f"     [v] Chrome restarted")
                            consecutive_failures = 0
                    except:
                        pass

        # Cleanup
        try:
            api.close()
        except:
            pass

        log(f"\n  [OK] Created {video_created}/{len(scenes)} videos")
        return video_created > 0

    except Exception as e:
        log(f"  [FAIL] Exception: {e}", "ERROR")
        import traceback
        traceback.print_exc()
        return False


def scan_projects_need_video() -> list:
    """Scan local PROJECTS for projects that have images but need videos."""
    projects_need_video = []

    if not LOCAL_PROJECTS.exists():
        return projects_need_video

    for item in LOCAL_PROJECTS.iterdir():
        if not item.is_dir():
            continue

        code = item.name

        if not matches_channel(code):
            continue

        # Check có scenes cần video không
        scenes = get_scenes_with_images(item, code)
        if scenes:
            print(f"    - {code}: {len(scenes)} videos needed")
            projects_need_video.append((code, len(scenes)))

    # Sort by number of scenes (process smaller first)
    projects_need_video.sort(key=lambda x: x[1])

    return [p[0] for p in projects_need_video]


def run_scan_loop():
    """Run continuous scan loop for VIDEO generation."""
    global VIDEO_COUNT

    print(f"\n{'='*60}")
    print(f"  VE3 TOOL - WORKER VIDEO (Video Only)")
    print(f"{'='*60}")
    print(f"  Worker folder:   {TOOL_DIR.parent.name}")
    print(f"  Channel filter:  {WORKER_CHANNEL or 'ALL'}")
    print(f"  Video count:     {VIDEO_COUNT if VIDEO_COUNT > 0 else 'ALL'}")
    print(f"  Mode:            VIDEO ONLY (watch for images)")
    print(f"{'='*60}")

    cycle = 0

    while True:
        cycle += 1
        print(f"\n[VIDEO CYCLE {cycle}] Scanning for projects with images...")

        # Find projects that need videos
        pending = scan_projects_need_video()

        if not pending:
            print(f"  No projects need video")
            print(f"\n  Waiting {VIDEO_SCAN_INTERVAL}s... (Ctrl+C to stop)")
            try:
                time.sleep(VIDEO_SCAN_INTERVAL)
            except KeyboardInterrupt:
                print("\n\nStopped by user.")
                break
        else:
            print(f"  Found: {len(pending)} projects need video")

            for code in pending:
                try:
                    success = process_project_video(code, video_count=VIDEO_COUNT)
                    if not success:
                        print(f"  [SKIP] Skipping {code}, moving to next...")
                        continue
                except KeyboardInterrupt:
                    print("\n\nStopped by user.")
                    return
                except Exception as e:
                    print(f"  [FAIL] Error processing {code}: {e}")
                    continue

            print(f"\n  [OK] Processed all projects!")
            print(f"  Waiting {VIDEO_SCAN_INTERVAL}s... (Ctrl+C to stop)")
            try:
                time.sleep(VIDEO_SCAN_INTERVAL)
            except KeyboardInterrupt:
                print("\n\nStopped by user.")
                break


def main():
    global WORKER_ID, TOTAL_WORKERS, VIDEO_COUNT

    import argparse
    parser = argparse.ArgumentParser(description='VE3 Worker VIDEO - Video Generation Only')
    parser.add_argument('project', nargs='?', default=None, help='Project code')
    parser.add_argument('--worker-id', type=int, default=0, help='Worker ID')
    parser.add_argument('--total-workers', type=int, default=1, help='Total workers')
    parser.add_argument('--video-count', type=int, default=-1, help='Number of videos (-1=all)')
    args = parser.parse_args()

    WORKER_ID = args.worker_id
    TOTAL_WORKERS = args.total_workers
    VIDEO_COUNT = args.video_count

    if args.project:
        process_project_video(args.project, video_count=VIDEO_COUNT)
    else:
        run_scan_loop()


if __name__ == "__main__":
    main()
