"""
Force chạy lại Step 7 bằng cách xóa scenes sheet
"""
import sys
import io
from pathlib import Path
import yaml
import time

# Fix encoding
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
sys.path.insert(0, str(Path(__file__).parent))

from openpyxl import load_workbook

# Setup
project_code = "AR8-0003"
project_dir = Path(__file__).parent / "PROJECTS" / project_code
excel_file = project_dir / f"{project_code}_prompts.xlsx"

print("=" * 80)
print("FORCE CHẠY LẠI STEP 7")
print("=" * 80)
print()

if not excel_file.exists():
    print("[ERROR] Excel not found!")
    sys.exit(1)

# Load workbook và xóa scenes sheet để force Step 7 chạy lại
wb = load_workbook(str(excel_file))

if "scenes" in wb.sheetnames:
    print(f"Deleting scenes sheet to force Step 7 re-run...")
    wb.remove(wb["scenes"])
    wb.save(str(excel_file))
    print(f"[OK] Scenes sheet deleted")
else:
    print(f"[INFO] Scenes sheet already empty")

wb.close()
print()

# Run Step 7 via run_all_steps
from modules.progressive_prompts import ProgressivePromptsGenerator

config_file = Path(__file__).parent / "config" / "settings.yaml"
with open(config_file, 'r', encoding='utf-8') as f:
    config = yaml.safe_load(f)

print(f"Mode: {config.get('excel_mode', 'full')}")
print(f"Max parallel API: {config.get('max_parallel_api', 6)}")
print()
print("Starting run_all_steps (will skip Steps 1-6, run Step 7)...")
print("-" * 80)

generator = ProgressivePromptsGenerator(config=config)

start_time = time.time()

try:
    result = generator.run_all_steps(
        project_dir=project_dir,
        code=project_code
    )

    elapsed = time.time() - start_time

    print()
    print("-" * 80)
    print(f"Completed in {elapsed:.1f}s ({elapsed/60:.1f} minutes)")
    print()

    if result:
        print("[OK] SUCCESS!")

        # Check result
        from modules.excel_manager import PromptWorkbook
        workbook = PromptWorkbook(str(excel_file))
        scenes = workbook.get_scenes()

        print()
        print(f"Total scenes: {len(scenes)}")

        if len(scenes) > 0:
            create_count = sum(1 for s in scenes if getattr(s, 'video_note', '') == "")
            skip_count = sum(1 for s in scenes if getattr(s, 'video_note', '') == "SKIP")

            print(f"  video_note='': {create_count} scenes (CREATE video)")
            print(f"  video_note='SKIP': {skip_count} scenes (SKIP video)")
            print()

            # Show distribution by segment estimate
            seg1_count = 0
            seg2_count = 0

            for scene in scenes:
                srt_start = getattr(scene, 'srt_start', '')
                if srt_start < "00:01:00":
                    seg1_count += 1
                else:
                    seg2_count += 1

            print(f"Estimated distribution:")
            print(f"  Segment 1 (~first 60s): {seg1_count} scenes")
            print(f"  Segment 2+ (after 60s): {seg2_count} scenes")
            print()

            # Sample first 10 scenes
            print("Sample first 10 scenes:")
            for i, scene in enumerate(scenes[:10]):
                video_note = getattr(scene, 'video_note', '')
                srt_start = getattr(scene, 'srt_start', '')
                note_display = "CREATE" if video_note == "" else "SKIP"
                print(f"  Scene {scene.scene_id:3d} | {srt_start} | {note_display}")

    else:
        print("[ERROR] FAILED!")

except Exception as e:
    elapsed = time.time() - start_time
    print()
    print(f"[ERROR] Exception after {elapsed:.1f}s: {e}")
    import traceback
    traceback.print_exc()

print()
print("=" * 80)
