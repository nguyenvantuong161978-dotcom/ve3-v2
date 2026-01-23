"""
Full audit của Excel - check TẤT CẢ sheets
"""
import sys
import io
from pathlib import Path
from openpyxl import load_workbook

# Fix encoding
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

excel_file = Path(__file__).parent / "PROJECTS" / "AR8-0003" / "AR8-0003_prompts.xlsx"

print("=" * 80)
print("FULL EXCEL AUDIT - AR8-0003")
print("=" * 80)
print()

if not excel_file.exists():
    print("[ERROR] Excel not found!")
    sys.exit(1)

wb = load_workbook(str(excel_file), read_only=False)

print(f"File: {excel_file.name}")
print(f"Sheets: {wb.sheetnames}")
print()

# Check director_plan sheet
print("=" * 80)
print("DIRECTOR_PLAN SHEET - First 10 rows")
print("=" * 80)

if "director_plan" in wb.sheetnames:
    ws = wb["director_plan"]

    # Get header
    header = [cell.value for cell in ws[1]]
    print(f"Headers: {header}")
    print()

    # Get first 10 data rows
    for row_idx in range(2, min(12, ws.max_row + 1)):
        row_data = [cell.value for cell in ws[row_idx]]
        print(f"Row {row_idx}: {row_data[:10]}...")  # First 10 columns

        # Check for data issues
        plan_id = row_data[0] if len(row_data) > 0 else None
        segment_id = row_data[1] if len(row_data) > 1 else None
        characters_used = row_data[6] if len(row_data) > 6 else None
        location_used = row_data[7] if len(row_data) > 7 else None

        print(f"  plan_id={plan_id}, segment_id={segment_id}")
        print(f"  characters_used='{characters_used}'")
        print(f"  location_used='{location_used}'")
        print()

# Check scenes sheet
print("=" * 80)
print("SCENES SHEET - First 10 rows")
print("=" * 80)

if "scenes" in wb.sheetnames:
    ws = wb["scenes"]

    # Get header
    header = [cell.value for cell in ws[1]]
    print(f"Headers ({len(header)} columns):")
    for i, h in enumerate(header, 1):
        print(f"  Col {i:2d}: {h}")
    print()

    # Get first 10 data rows
    for row_idx in range(2, min(12, ws.max_row + 1)):
        row_data = [cell.value for cell in ws[row_idx]]

        scene_id = row_data[0] if len(row_data) > 0 else None
        srt_start = row_data[1] if len(row_data) > 1 else None
        characters_used = row_data[13] if len(row_data) > 13 else None  # Column 14
        location_used = row_data[14] if len(row_data) > 14 else None  # Column 15
        reference_files = row_data[15] if len(row_data) > 15 else None  # Column 16

        print(f"Scene {scene_id}:")
        print(f"  srt_start: {srt_start}")
        print(f"  characters_used: '{characters_used}'")
        print(f"  location_used: '{location_used}'")
        print(f"  reference_files: '{reference_files}'")

        # Check for swap
        if location_used and isinstance(location_used, str):
            if location_used.startswith('['):
                print(f"  ⚠️ SWAP! location_used looks like JSON array")

        if characters_used and isinstance(characters_used, str):
            if not characters_used.startswith('nv') and not characters_used.startswith('loc'):
                if len(characters_used) < 10:  # Short string, likely location ID
                    print(f"  ⚠️ SWAP! characters_used looks like location ID")

        print()

wb.close()

print("=" * 80)
print("AUDIT COMPLETE")
print("=" * 80)
