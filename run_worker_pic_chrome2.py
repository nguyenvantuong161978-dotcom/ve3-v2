#!/usr/bin/env python3
"""
VE3 Tool - Worker PIC Chrome 2 (Image Generation với Chrome 2)
==============================================================
Script riêng cho Chrome 2, dùng DrissionFlowAPI trực tiếp.
Được gọi từ run_worker_pic_basic_2.py như subprocess.

Usage:
    python run_worker_pic_chrome2.py --excel <path>
"""

import sys
import time
from pathlib import Path

# Add current directory to path
TOOL_DIR = Path(__file__).parent
sys.path.insert(0, str(TOOL_DIR))


def load_chrome2_path() -> str:
    """Load chrome_portable_2 from settings.yaml."""
    import yaml

    settings_path = TOOL_DIR / "config" / "settings.yaml"
    if not settings_path.exists():
        return None

    with open(settings_path, 'r', encoding='utf-8') as f:
        settings = yaml.safe_load(f) or {}

    chrome2 = settings.get('chrome_portable_2', '')

    # Auto-detect if not configured
    if not chrome2:
        copy_chrome = TOOL_DIR / "GoogleChromePortable - Copy" / "GoogleChromePortable.exe"
        if copy_chrome.exists():
            chrome2 = str(copy_chrome)

    return chrome2


def get_prompts_from_excel(excel_path: str) -> list:
    """Read prompts from Excel that need image generation."""
    from modules.excel_manager import PromptWorkbook

    prompts = []
    img_dir = Path(excel_path).parent / "img"

    try:
        wb = PromptWorkbook(excel_path)

        # Characters
        for char in wb.get_characters():
            char_id = char.id
            if not char.english_prompt:
                continue
            # Check if image exists
            img_path = img_dir / f"{char_id}.png"
            if img_path.exists():
                continue
            prompts.append({
                'id': char_id,
                'prompt': char.english_prompt,
                'type': 'character',
                'media_id': getattr(char, 'media_id', None)
            })

        # Locations
        for loc in wb.get_locations():
            loc_id = loc.id
            if not loc.english_prompt:
                continue
            img_path = img_dir / f"{loc_id}.png"
            if img_path.exists():
                continue
            prompts.append({
                'id': loc_id,
                'prompt': loc.english_prompt,
                'type': 'location',
                'media_id': getattr(loc, 'media_id', None)
            })

        # Scenes
        scene_media_ids = wb.get_scene_media_ids()
        for scene in wb.get_scenes():
            scene_id = str(scene.scene_id)
            if not scene.img_prompt:
                continue
            img_path = img_dir / f"{scene_id}.png"
            if img_path.exists():
                continue

            # Get reference files
            ref_files = []
            if hasattr(scene, 'reference_files') and scene.reference_files:
                try:
                    import json
                    ref_files = json.loads(scene.reference_files) if isinstance(scene.reference_files, str) else scene.reference_files
                except:
                    pass

            prompts.append({
                'id': scene_id,
                'prompt': scene.img_prompt,
                'type': 'scene',
                'reference_files': ref_files,
                'media_id': scene_media_ids.get(scene_id)
            })

    except Exception as e:
        print(f"[Chrome2] Error reading Excel: {e}", flush=True)

    return prompts


def run_chrome2_pic_worker(excel_path: str):
    """
    Chrome 2 worker for image generation.
    Uses DrissionFlowAPI directly like run_worker_video.py.
    """
    import sys
    sys.stdout.reconfigure(line_buffering=True)

    print(f"[Chrome2-PIC] Starting worker...", flush=True)

    chrome2 = load_chrome2_path()
    if not chrome2:
        print(f"[Chrome2-PIC] ERROR: chrome_portable_2 not configured!", flush=True)
        return

    print(f"[Chrome2-PIC] Chrome: {chrome2}", flush=True)
    print(f"[Chrome2-PIC] Excel: {excel_path}", flush=True)

    # Get prompts that need processing
    prompts = get_prompts_from_excel(excel_path)

    # Filter to only ODD indices (Chrome 2 handles odd, Chrome 1 handles even)
    # This way both Chrome work on different images
    prompts = [p for i, p in enumerate(prompts) if i % 2 == 1]

    print(f"[Chrome2-PIC] Found {len(prompts)} prompts to process (odd indices)", flush=True)

    if not prompts:
        print(f"[Chrome2-PIC] No prompts to process!", flush=True)
        return

    # Get project URL from Excel
    project_url = None
    try:
        import openpyxl
        wb_xl = openpyxl.load_workbook(excel_path, data_only=True)
        if 'config' in wb_xl.sheetnames:
            ws = wb_xl['config']
            for row in ws.iter_rows(min_row=1, max_row=30, values_only=True):
                if row and row[0] == 'flow_project_url':
                    project_url = row[1]
                    break
        wb_xl.close()
    except Exception as e:
        print(f"[Chrome2-PIC] Error reading project URL: {e}", flush=True)

    if not project_url:
        print(f"[Chrome2-PIC] ERROR: No project URL found in Excel!", flush=True)
        return

    print(f"[Chrome2-PIC] Project URL: {project_url[:60]}...", flush=True)

    # Safe print function to handle encoding errors on Windows
    def safe_print(msg):
        try:
            print(msg, flush=True)
        except UnicodeEncodeError:
            print(msg.encode('ascii', 'replace').decode('ascii'), flush=True)

    # Create DrissionFlowAPI with Chrome 2 settings
    from modules.drission_flow_api import DrissionFlowAPI

    api = DrissionFlowAPI(
        profile_dir="./chrome_profiles/chrome2",  # Profile RIÊNG cho Chrome 2
        verbose=True,
        log_callback=lambda msg, lvl="INFO": safe_print(f"[Chrome2-PIC] {msg}"),
        webshare_enabled=False,
        worker_id=1,  # Chrome 2 = bên phải
        total_workers=2,  # Chia đôi màn hình
        headless=False,
        machine_id=102,  # Khác với Chrome 1 (1)
        chrome_portable=chrome2
    )

    print(f"[Chrome2-PIC] Setting up Chrome...", flush=True)

    if not api.setup(project_url=project_url):
        print(f"[Chrome2-PIC] ERROR: Failed to setup Chrome!", flush=True)
        return

    print(f"[Chrome2-PIC] Chrome ready! Generating images...", flush=True)

    # Generate images
    img_dir = Path(excel_path).parent / "img"
    img_dir.mkdir(exist_ok=True)

    success_count = 0
    fail_count = 0

    for i, p in enumerate(prompts):
        prompt_id = p['id']
        prompt_text = p['prompt']
        ref_files = p.get('reference_files', [])

        print(f"[Chrome2-PIC] [{i+1}/{len(prompts)}] ID: {prompt_id}", flush=True)

        # Check if image already exists (might have been created by Chrome 1)
        img_path = img_dir / f"{prompt_id}.png"
        if img_path.exists():
            print(f"[Chrome2-PIC]   -> Already exists, skip", flush=True)
            continue

        try:
            # Generate image using DrissionFlowAPI
            result = api.generate_image(
                prompt=prompt_text,
                reference_media_ids=None,  # TODO: handle references
                output_path=str(img_path)
            )

            if result and img_path.exists():
                print(f"[Chrome2-PIC]   -> OK: {img_path.name}", flush=True)
                success_count += 1
            else:
                print(f"[Chrome2-PIC]   -> FAILED", flush=True)
                fail_count += 1

        except Exception as e:
            print(f"[Chrome2-PIC]   -> ERROR: {e}", flush=True)
            fail_count += 1

        # Small delay between requests
        time.sleep(0.5)

    print(f"[Chrome2-PIC] Done! Success: {success_count}, Failed: {fail_count}", flush=True)

    # Cleanup
    try:
        api.close()
    except:
        pass


def main():
    import argparse
    parser = argparse.ArgumentParser(description='VE3 Worker PIC Chrome 2')
    parser.add_argument('--excel', type=str, required=True, help='Excel path')
    args = parser.parse_args()

    run_chrome2_pic_worker(args.excel)


if __name__ == "__main__":
    main()
