#!/usr/bin/env python3
"""
VE3 Tool - Worker PIC BASIC Mode - CHROME 2 with Agent Protocol
================================================================
Chrome Worker 2 - Xử lý scenes lẻ (1, 3, 5, ...)

Tích hợp Agent Protocol để:
- Báo cáo trạng thái cho VM Manager
- Ghi log chi tiết
- Báo cáo kết quả thành công/thất bại

Usage:
    python _run_chrome2.py                     (quét và xử lý tự động)
    python _run_chrome2.py AR47-0028           (chạy 1 project cụ thể)
"""

import sys
import os

# Fix Windows encoding issues
if sys.platform == "win32":
    if sys.stdout:
        sys.stdout.reconfigure(encoding='utf-8', errors='replace')
    if sys.stderr:
        sys.stderr.reconfigure(encoding='utf-8', errors='replace')
    os.environ['PYTHONIOENCODING'] = 'utf-8'

import time
import shutil
from pathlib import Path
from datetime import datetime

# Add current directory to path
TOOL_DIR = Path(__file__).parent
sys.path.insert(0, str(TOOL_DIR))

# ================================================================================
# AGENT PROTOCOL
# ================================================================================

# Worker ID for this Chrome worker
WORKER_ID = "chrome_2"

# Agent Protocol - giao tiếp với VM Manager
try:
    from modules.agent_protocol import AgentWorker, ErrorType
    AGENT_ENABLED = True
except ImportError:
    AGENT_ENABLED = False
    AgentWorker = None

# Central Logger - để log hiển thị trong GUI
try:
    from modules.central_logger import get_logger
    _logger = get_logger(WORKER_ID)
except ImportError:
    class FakeLogger:
        def info(self, msg): print(f"[{WORKER_ID}] {msg}")
        def warn(self, msg): print(f"[{WORKER_ID}] WARN: {msg}")
        def error(self, msg): print(f"[{WORKER_ID}] ERROR: {msg}")
    _logger = FakeLogger()

# Global agent instance
_agent = None


# Override print CHÍNH XÁC - tránh recursion với central_logger
import builtins
_original_print = builtins.print

# Flag để tránh recursion khi central_logger gọi print
_in_logger = False

def _logger_print(*args, **kwargs):
    """Override print() to log to central_logger (tránh recursion)."""
    global _in_logger

    # Nếu đang trong logger, dùng print gốc
    if _in_logger:
        _original_print(*args, **kwargs)
        return

    try:
        _in_logger = True
        msg = ' '.join(str(arg) for arg in args)

        # Remove timestamp prefix if present (avoid duplication)
        if msg.startswith('[') and ']' in msg[:12]:
            msg = msg.split(']', 1)[-1].strip()

        if msg.strip():
            _logger.info(msg)
    finally:
        _in_logger = False

builtins.print = _logger_print


def log(msg: str, level: str = "INFO"):
    """Log to console + central logger + agent."""
    global _agent

    # Log to central logger (cho GUI)
    if level == "ERROR":
        _logger.error(msg)
    elif level == "WARN":
        _logger.warn(msg)
    else:
        _logger.info(msg)

    # Gửi đến Agent nếu có
    if _agent:
        if level == "ERROR":
            _agent.log_error(msg)
        else:
            _agent.log(msg, level)


def init_agent():
    """Khởi tạo Agent Protocol."""
    global _agent
    if AGENT_ENABLED and _agent is None:
        _agent = AgentWorker(WORKER_ID)
        _agent.start_status_updater(interval=5)
        _agent.update_status(state="idle")
        print(f"[{WORKER_ID}] Agent Protocol enabled")
    return _agent


def close_agent():
    """Đóng Agent Protocol."""
    global _agent
    if _agent:
        _agent.close()
        _agent = None


def safe_str(s) -> str:
    """Convert any value to a safe ASCII-friendly string."""
    try:
        text = str(s)
        # Replace non-ASCII characters with '?'
        return text.encode('ascii', errors='replace').decode('ascii')
    except:
        return "[encoding error]"


def agent_log(msg: str, level: str = "INFO"):
    """Log và gửi đến Agent + Central Logger (cho GUI)."""
    global _agent
    safe_msg = safe_str(msg)

    # Print to console
    print(f"[{datetime.now().strftime('%H:%M:%S')}] {safe_msg}")

    # Send to central logger for GUI display
    if CENTRAL_LOGGER and central_log:
        central_log(WORKER_ID, safe_msg, level)

    # Send to agent
    if _agent:
        if level == "ERROR":
            _agent.log_error(safe_msg)
        else:
            _agent.log(safe_msg, level)

# Import từ run_worker (dùng chung logic)
from run_worker import (
    detect_auto_path,
    POSSIBLE_AUTO_PATHS,
    get_channel_from_folder,
    matches_channel,
    is_project_complete_on_master,
    has_excel_with_prompts,
    needs_api_completion,
    copy_from_master,
    copy_to_visual,
    delete_local_project,
    SCAN_INTERVAL,
)


def safe_path_exists(path: Path) -> bool:
    """
    Safely check if a path exists, handling network disconnection errors.
    Returns False if path doesn't exist OR if network is disconnected.
    """
    try:
        return path.exists()
    except (OSError, PermissionError) as e:
        # WinError 1167: The device is not connected
        # WinError 53: The network path was not found
        # WinError 64: The specified network name is no longer available
        print(f"  [WARN] Network error checking path: {e}")
        return False


def safe_iterdir(path: Path) -> list:
    """
    Safely iterate over a directory, handling network disconnection errors.
    Returns empty list if path doesn't exist OR if network is disconnected.
    """
    try:
        if not path.exists():
            return []
        return list(path.iterdir())
    except (OSError, PermissionError) as e:
        print(f"  [WARN] Network error listing directory: {e}")
        return []

# Detect paths
AUTO_PATH = detect_auto_path()
if AUTO_PATH:
    MASTER_PROJECTS = AUTO_PATH / "ve3-tool-simple" / "PROJECTS"
    MASTER_VISUAL = AUTO_PATH / "VISUAL"
else:
    MASTER_PROJECTS = Path(r"\\tsclient\D\AUTO\ve3-tool-simple\PROJECTS")
    MASTER_VISUAL = Path(r"\\tsclient\D\AUTO\VISUAL")

LOCAL_PROJECTS = TOOL_DIR / "PROJECTS"
WORKER_CHANNEL = get_channel_from_folder()


def get_chrome2_path():
    """Get chrome_portable_2 from settings.yaml."""
    import yaml
    config_path = TOOL_DIR / "config" / "settings.yaml"
    if config_path.exists():
        with open(config_path, "r", encoding="utf-8") as f:
            config = yaml.safe_load(f) or {}
        chrome2 = config.get('chrome_portable_2', '')
        if chrome2:
            return chrome2

    # Auto-detect
    copy_chrome = TOOL_DIR / "GoogleChromePortable - Copy" / "GoogleChromePortable.exe"
    if copy_chrome.exists():
        return str(copy_chrome)
    return None


def is_local_pic_complete(project_dir: Path, name: str) -> bool:
    """Check if local project has images created."""
    img_dir = project_dir / "img"
    if not img_dir.exists():
        return False

    img_files = list(img_dir.glob("*.png")) + list(img_dir.glob("*.jpg"))

    if len(img_files) == 0:
        return False

    try:
        from modules.excel_manager import PromptWorkbook
        excel_path = project_dir / f"{name}_prompts.xlsx"
        if excel_path.exists():
            wb = PromptWorkbook(str(excel_path))
            scenes = wb.get_scenes()
            expected = len([s for s in scenes if s.img_prompt])

            if len(img_files) >= expected:
                print(f"    [{name}] Images: {len(img_files)}/{expected} - COMPLETE")
                return True
            else:
                print(f"    [{name}] Images: {len(img_files)}/{expected} - incomplete")
                return False
    except Exception as e:
        print(f"    [{name}] Warning: {e}")

    return len(img_files) > 0


def process_project_pic_basic_chrome2(code: str, callback=None) -> bool:
    """Process a single project - CHROME 2."""

    def log(msg, level="INFO"):
        if callback:
            callback(msg, level)
        else:
            print(f"[Chrome2] {msg}")

    log(f"\n{'='*60}")
    log(f"[PIC BASIC CHROME2] Processing: {code}")
    log(f"{'='*60}")

    # Get Chrome 2 path
    chrome2_path = get_chrome2_path()
    if not chrome2_path:
        log(f"  ERROR: chrome_portable_2 not found!", "ERROR")
        return False
    log(f"  Chrome2: {chrome2_path}")

    # Step 1: Check if already done on master
    if is_project_complete_on_master(code):
        log(f"  Already in VISUAL folder, skip!")
        return True

    # Step 2: Copy from master
    local_dir = copy_from_master(code)
    if not local_dir:
        return False

    # Step 3: Check Excel - Chrome 2 KHÔNG tạo Excel, đợi Chrome 1
    excel_path = local_dir / f"{code}_prompts.xlsx"

    if not excel_path.exists():
        log(f"  Waiting for Chrome 1 to create Excel...")
        # Đợi tối đa 120s
        for i in range(24):
            time.sleep(5)
            if excel_path.exists():
                log(f"  Excel found!")
                break
            log(f"  Waiting... ({(i+1)*5}s)")

        if not excel_path.exists():
            log(f"  No Excel after 120s, skip!")
            return False

    # Step 3.5: VALIDATOR - Kiểm tra tất cả NV references trước khi tạo scenes
    log(f"  ============================================================")
    log(f"  STEP 3.5: REFERENCE VALIDATOR")
    log(f"  ============================================================")

    try:
        from modules.excel_manager import PromptWorkbook

        # BƯỚC 1: LẬP KẾ HOẠCH - Đọc Excel để biết có bao nhiêu references cần validate
        workbook = PromptWorkbook(str(excel_path))
        all_chars = workbook.get_characters()

        # Đếm references cần validate (cả NV và LOC, bỏ qua skip=True hoặc đã verified)
        refs_to_validate = []
        for char in all_chars:
            # Kiểm tra cả NV và LOC
            if not char.id.lower().startswith(('nv', 'loc')):
                continue

            # Bỏ qua references có skip=True
            skip = getattr(char, 'skip', False)
            if skip:
                log(f"  [SKIP] {char.id} - marked as skip")
                continue

            # Bỏ qua references đã verified rồi
            status = getattr(char, 'status', '')
            if status in ['verified', 'verified_fixed']:
                log(f"  [SKIP] {char.id} - already verified (status={status})")
                continue

            refs_to_validate.append(char.id)

        if not refs_to_validate:
            log(f"  [v] All references already validated, skip validator mode")
        else:
            nv_count = len([r for r in refs_to_validate if r.lower().startswith('nv')])
            loc_count = len([r for r in refs_to_validate if r.lower().startswith('loc')])
            log(f"  [PLAN] Cần validate {len(refs_to_validate)} references: {nv_count} NV + {loc_count} LOC")
            log(f"  {refs_to_validate}")

            # Lấy project URL từ Excel (config sheet)
            project_url = None
            try:
                import openpyxl
                wb_openpyxl = openpyxl.load_workbook(str(excel_path), read_only=True)
                if 'config' in wb_openpyxl.sheetnames:
                    config_sheet = wb_openpyxl['config']
                    for row in config_sheet.iter_rows(min_row=2):  # Skip header
                        if row[0].value == 'flow_project_url':
                            project_url = row[1].value
                            break
                wb_openpyxl.close()

                if project_url:
                    log(f"  [v] Project URL from Excel: {project_url[:60]}...")
                else:
                    log(f"  [!] No project URL in Excel - will create new project")
            except Exception as e:
                log(f"  [WARN] Could not read project URL from Excel: {e}")

            # Track validated NVs
            validated_refs = set()

            # Setup Chrome API cho validator (tạo 1 lần, dùng chung)
            validator_api = None
            validator = None

            try:
                from modules.reference_validator import ReferenceValidator
                from modules.drission_flow_api import DrissionFlowAPI
                import yaml

                # Load config
                config_file = Path("config/settings.yaml")
                with open(config_file, 'r', encoding='utf-8') as f:
                    config = yaml.safe_load(f)

                # Setup Chrome 2 API
                chrome_path = chrome2_path
                validator_api = DrissionFlowAPI(
                    chrome_portable=chrome_path,
                    worker_id=2,
                    total_workers=2,
                    headless=False,
                    webshare_enabled=False
                )

                # VÀO ĐÚNG PROJECT (dùng URL từ Excel)
                if project_url:
                    log(f"  [VALIDATOR] Starting Chrome with project URL...")
                    if not validator_api.setup(project_url=project_url):
                        log(f"  [ERROR] Failed to start Chrome!", "error")
                        raise Exception("Chrome setup failed")
                else:
                    log(f"  [ERROR] No project URL - cannot validate!", "error")
                    raise Exception("No project URL")

                # Create validator
                validator = ReferenceValidator(
                    drission_api=validator_api,
                    workbook=workbook,
                    config=config,
                    project_code=code
                )

                log(f"  [v] Validator ready!")

            except Exception as e:
                log(f"  [ERROR] Failed to setup validator: {e}", "error")
                import traceback
                traceback.print_exc()
                # Skip validation
                validated_refs = set(refs_to_validate)

            # BƯỚC 2: CHỜ + VALIDATE TỪNG NV CÓ MEDIA_ID
            log(f"  [VALIDATOR] Waiting for Chrome 1 to create references...")
            log(f"  [VALIDATOR] Will validate each NV as soon as it has media_id...")

            wait_interval = 10
            waited = 0

            while len(validated_refs) < len(refs_to_validate) and validator:
                # Re-load Excel để check media_id mới
                workbook = PromptWorkbook(str(excel_path))
                all_chars = workbook.get_characters()

                # Tìm NV nào có media_id mà chưa validate
                for char in all_chars:
                    if char.id not in refs_to_validate:
                        continue
                    if char.id in validated_refs:
                        continue

                    # Check status
                    status = getattr(char, 'status', '')
                    if status in ['verified', 'verified_fixed']:
                        # Đã validate rồi
                        validated_refs.add(char.id)
                        log(f"  [v] {char.id} already validated (status={status})")
                        continue

                    # Check media_id
                    if not char.media_id:
                        continue  # Chưa có media_id, chờ tiếp

                    # CÓ MEDIA_ID → VALIDATE NGAY!
                    log(f"  ")
                    log(f"  [VALIDATOR] Found {char.id} with media_id!")
                    log(f"  [VALIDATOR] Validating {char.id}... ({len(validated_refs)+1}/{len(refs_to_validate)})")

                    # Validate
                    try:
                        result = validator.validate_and_fix(char.id)
                        log(f"  [VALIDATOR] {char.id} result: {result}")
                    except Exception as e:
                        log(f"  [ERROR] Validation failed for {char.id}: {e}", "error")
                        import traceback
                        traceback.print_exc()

                    # Mark validated
                    validated_refs.add(char.id)

                # Check progress
                if len(validated_refs) >= len(refs_to_validate):
                    log(f"  ")
                    log(f"  [v] VALIDATOR COMPLETED!")
                    log(f"  [v] Validated {len(validated_refs)}/{len(refs_to_validate)} references")
                    break

                # Log progress
                if waited % 30 == 0:
                    log(f"  [WAIT] Validated: {len(validated_refs)}/{len(refs_to_validate)} - Waiting... ({waited}s)")

                time.sleep(wait_interval)
                waited += wait_interval

            # Close validator API
            if validator_api:
                try:
                    validator_api.close()
                    log(f"  [v] Validator Chrome closed")
                except:
                    pass

    except Exception as e:
        log(f"  [ERROR] Validator error: {e}", "error")
        import traceback
        traceback.print_exc()
        # Cleanup
        if validator_api:
            try:
                validator_api.close()
            except:
                pass

    # Step 3.6: Đợi Chrome 1 bắt đầu tạo SCENE images
    img_dir = local_dir / "img"
    log(f"  [WAIT] Waiting for Chrome 1 to START creating scene images...")

    wait_interval = 10
    waited = 0

    while True:
        if img_dir.exists():
            # Tìm ảnh scene (số hoặc scene_X)
            scene_files = []
            for f in img_dir.glob("*.png"):
                name = f.stem
                if name.isdigit() or name.startswith("scene_"):
                    scene_files.append(f)
            for f in img_dir.glob("*.jpg"):
                name = f.stem
                if name.isdigit() or name.startswith("scene_"):
                    scene_files.append(f)

            if scene_files:
                log(f"  [v] Chrome 1 đã bắt đầu tạo scenes! Found {len(scene_files)} scene images")
                log(f"     → Chrome 2 bắt đầu tạo scenes lẻ...")
                time.sleep(3)
                break

        # Log mỗi 30 giây
        if waited % 30 == 0:
            log(f"  [WAIT] Waiting... ({waited}s)")

        time.sleep(wait_interval)
        waited += wait_interval

    # Step 4: Create images using SmartEngine
    try:
        from modules.smart_engine import SmartEngine

        # Chrome 2: worker_id=1, total_workers=2, dùng chrome_portable_2
        engine = SmartEngine(
            worker_id=1,
            total_workers=2,
            chrome_portable=chrome2_path
        )

        log(f"  Excel: {excel_path.name}")
        log(f"  Mode: CHROME 2 (scenes lẻ: 1,3,5,...)")

        # Run engine - images only, skip video, skip references (Chrome 1 tạo)
        result = engine.run(
            str(excel_path),
            callback=callback,
            skip_compose=True,
            skip_video=True,
            skip_references=True
        )

        if result.get('error'):
            log(f"  Error: {result.get('error')}", "ERROR")
            return False

    except Exception as e:
        log(f"  Exception: {e}", "ERROR")
        import traceback
        traceback.print_exc()
        return False

    # Step 5: Check completion
    if is_local_pic_complete(local_dir, code):
        log(f"  Images complete!")
        return True
    else:
        log(f"  Images incomplete", "WARN")
        return False


def scan_incomplete_local_projects() -> list:
    """Scan local PROJECTS for incomplete projects."""
    incomplete = []

    if not LOCAL_PROJECTS.exists():
        return incomplete

    for item in LOCAL_PROJECTS.iterdir():
        if not item.is_dir():
            continue

        code = item.name

        if not matches_channel(code):
            continue

        if is_project_complete_on_master(code):
            continue

        if is_local_pic_complete(item, code):
            continue

        srt_path = item / f"{code}.srt"
        if has_excel_with_prompts(item, code):
            print(f"    - {code}: incomplete (has Excel, no images)")
            incomplete.append(code)
        elif srt_path.exists():
            print(f"    - {code}: has SRT, no Excel")
            incomplete.append(code)

    return sorted(incomplete)


def scan_master_projects() -> list:
    """Scan master PROJECTS folder for pending projects."""
    pending = []

    if not safe_path_exists(MASTER_PROJECTS):
        return pending

    for item in safe_iterdir(MASTER_PROJECTS):
        try:
            if not item.is_dir():
                continue

            code = item.name

            if not matches_channel(code):
                continue

            if is_project_complete_on_master(code):
                continue

            excel_path = item / f"{code}_prompts.xlsx"
            srt_path = item / f"{code}.srt"

            # Wrap network path checks in try-except
            try:
                if has_excel_with_prompts(item, code):
                    print(f"    - {code}: ready (has prompts)")
                    pending.append(code)
                elif srt_path.exists():
                    print(f"    - {code}: has SRT")
                    pending.append(code)
            except (OSError, PermissionError) as e:
                print(f"  [WARN] Network error checking {code}: {e}")
                continue

        except (OSError, PermissionError) as e:
            # Network disconnected while iterating
            print(f"  [WARN] Network error scanning: {e}")
            break

    return sorted(pending)


def run_scan_loop():
    """Run continuous scan loop for IMAGE generation - CHROME 2."""
    chrome2_path = get_chrome2_path()

    print(f"\n{'='*60}")
    print(f"  VE3 TOOL - WORKER PIC BASIC - CHROME 2")
    print(f"{'='*60}")
    print(f"  Worker folder:   {TOOL_DIR.parent.name}")
    print(f"  Channel filter:  {WORKER_CHANNEL or 'ALL'}")
    print(f"  Chrome2:         {chrome2_path or 'NOT FOUND'}")
    print(f"  Mode:            Scenes lẻ (1,3,5,...)")
    print(f"{'='*60}")

    if not chrome2_path:
        print("ERROR: chrome_portable_2 not configured!")
        return

    # Đợi Chrome 1 khởi động trước
    print("\n[Chrome2] Waiting 10s for Chrome 1 to start...")
    time.sleep(10)

    cycle = 0
    current_project = None  # Track project đang làm

    while True:
        cycle += 1
        print(f"\n[Chrome2 CYCLE {cycle}] Scanning...")

        incomplete_local = scan_incomplete_local_projects()
        pending_master = scan_master_projects()
        pending = list(dict.fromkeys(incomplete_local + pending_master))

        if not pending:
            print(f"  No pending projects")
            current_project = None  # Reset khi không còn project
            print(f"\n  Waiting {SCAN_INTERVAL}s... (Ctrl+C to stop)")
            try:
                time.sleep(SCAN_INTERVAL)
            except KeyboardInterrupt:
                print("\n\nStopped by user.")
                break
        else:
            print(f"  Found: {len(pending)} pending projects")

            # CHỈ XỬ LÝ 1 PROJECT - ưu tiên project đang làm dở
            if current_project and current_project in pending:
                target = current_project
                print(f"  [Chrome2] Continuing: {target}")
            else:
                target = pending[0]
                current_project = target
                print(f"  [Chrome2] Starting: {target}")

            try:
                success = process_project_pic_basic_chrome2(target)
                if not success:
                    print(f"  [Chrome2] Project {target} incomplete, will retry...")
                else:
                    print(f"  [Chrome2] Project {target} completed!")
                    current_project = None  # Move to next project
            except KeyboardInterrupt:
                print("\n\nStopped by user.")
                return
            except Exception as e:
                print(f"  [Chrome2] Error processing {target}: {e}")

            print(f"\n  Waiting {SCAN_INTERVAL}s... (Ctrl+C to stop)")
            try:
                time.sleep(SCAN_INTERVAL)
            except KeyboardInterrupt:
                print("\n\nStopped by user.")
                break


def validate_references_if_needed(code: str, callback=None) -> bool:
    """
    VALIDATOR MODE - Chrome 2 validate references trước khi làm scenes.

    Kiểm tra xem project có references chưa validate không.
    Nếu có → Validate & fix bằng DeepSeek
    Nếu không → Return False (sẽ chạy normal mode)

    Returns:
        True nếu đã validate xong (hoặc không cần)
        False nếu có lỗi
    """
    def log(msg, level="INFO"):
        if callback:
            callback(msg, level)
        else:
            print(f"[Validator] {msg}")

    try:
        from modules.reference_validator import ReferenceValidator
        from modules.excel_manager import PromptWorkbook
        from modules.drission_flow_api import DrissionFlowAPI
        import yaml

        local_dir = LOCAL_PROJECTS / code
        excel_path = local_dir / f"{code}_prompts.xlsx"

        if not excel_path.exists():
            return True  # No Excel, skip validation

        # Load workbook
        workbook = PromptWorkbook(str(excel_path))

        # Get all references
        all_chars = workbook.get_characters()
        refs_to_validate = []

        for char in all_chars:
            # CHỈ VALIDATE NHI VẬT (nv) - BỎ QUA LOCATIONS (loc)
            if not char.id.lower().startswith('nv'):
                continue

            # Check if has media_id and not verified yet
            if char.media_id and char.status not in ['verified', 'verified_fixed']:
                refs_to_validate.append(char.id)

        if not refs_to_validate:
            log(f"No references need validation, skip validator mode")
            return True

        log(f"\n{'='*60}")
        log(f"VALIDATOR MODE - {len(refs_to_validate)} references to validate")
        log(f"{'='*60}")
        log(f"References: {refs_to_validate}")

        # Setup DrissionFlowAPI
        chrome2_path = get_chrome2_path()
        if not chrome2_path:
            log("ERROR: Chrome 2 path not found!", "ERROR")
            return False

        # Load config
        config_path = TOOL_DIR / "config" / "settings.yaml"
        with open(config_path, "r", encoding="utf-8") as f:
            config = yaml.safe_load(f) or {}

        api = DrissionFlowAPI(
            chrome_portable=chrome2_path,
            worker_id=1,
            total_workers=2,
            headless=False,
            webshare_enabled=False
        )

        # Setup Chrome
        log("Starting Chrome for validation...")
        if not api.setup(project_url="https://labs.google/fx/vi/tools/flow"):
            log("ERROR: Failed to start Chrome!", "ERROR")
            return False

        log("Chrome started, creating validator...")

        # Create validator
        validator = ReferenceValidator(
            drission_api=api,
            workbook=workbook,
            config=config
        )

        # Run validation
        stats = validator.validate_all_references(refs_to_validate)

        # Cleanup
        api.close()

        log(f"\n{'='*60}")
        log(f"VALIDATION COMPLETED")
        log(f"{'='*60}")
        log(f"Tested: {stats['tested']}")
        log(f"Verified: {stats['verified']}")
        log(f"Violated: {stats['violated']}")
        log(f"Fixed: {stats['fixed']}")
        log(f"Failed: {stats['failed']}")

        return True

    except Exception as e:
        log(f"ERROR in validator mode: {e}", "ERROR")
        import traceback
        traceback.print_exc()
        return False


def main():
    import argparse
    parser = argparse.ArgumentParser(description='VE3 Worker PIC BASIC - Chrome 2')
    parser.add_argument('project', nargs='?', default=None, help='Project code')
    parser.add_argument('--validate-only', action='store_true', help='Only run validator mode')
    args = parser.parse_args()

    # Khởi tạo Agent Protocol
    init_agent()

    try:
        if args.project:
            # Single project mode

            # VALIDATOR MODE: Validate references trước
            if not args.validate_only:
                print(f"\n[Chrome2] Checking if validation needed for {args.project}...")
                validate_references_if_needed(args.project)
                print(f"[Chrome2] Validation check completed, proceeding to normal mode...")

            # Normal mode hoặc validate-only
            if args.validate_only:
                success = validate_references_if_needed(args.project)
            else:
                success = process_project_with_agent(args.project)

            sys.exit(0 if success else 1)
        else:
            # Loop mode
            run_scan_loop_with_agent()
    finally:
        # Cleanup
        close_agent()


def process_project_with_agent(code: str) -> bool:
    """Process project với Agent Protocol."""
    global _agent

    task_id = f"image_{code}_{datetime.now().strftime('%H%M%S')}"
    start_time = time.time()

    # Get total scenes for progress tracking
    total_scenes = 0
    current_scene = [0]  # Use list to allow modification in callback

    try:
        from modules.excel_manager import PromptWorkbook
        local_dir = LOCAL_PROJECTS / code
        excel_path = local_dir / f"{code}_prompts.xlsx"
        if excel_path.exists():
            wb = PromptWorkbook(str(excel_path))
            total_scenes = len(wb.get_scenes())
    except:
        pass

    # Update agent status
    if _agent:
        _agent.update_status(
            state="working",
            current_project=code,
            current_task=task_id,
            current_scene=0,
            total_scenes=total_scenes,
            progress=0
        )

    # Callback to track progress
    def progress_callback(msg, level="INFO"):
        """Callback that also updates agent with scene progress."""
        print(msg)

        # Try to extract scene number from message
        if _agent and "Scene" in msg:
            import re
            match = re.search(r'Scene\s*(\d+)', msg, re.IGNORECASE)
            if match:
                scene_num = int(match.group(1))
                current_scene[0] = scene_num
                progress = int((scene_num / total_scenes * 100) if total_scenes > 0 else 0)
                _agent.update_status(
                    current_scene=scene_num,
                    progress=progress
                )

    # Process
    try:
        success = process_project_pic_basic_chrome2(code, callback=progress_callback)
        duration = time.time() - start_time

        # Report result
        if _agent:
            if success:
                _agent.report_success(
                    task_id=task_id,
                    project_code=code,
                    task_type="image",
                    duration=duration
                )
            else:
                _agent.report_failure(
                    task_id=task_id,
                    project_code=code,
                    task_type="image",
                    error="Processing failed",
                    duration=duration
                )
            _agent.update_status(
                state="idle",
                current_project="",
                current_task="",
                current_scene=0,
                total_scenes=0,
                progress=0
            )

        return success

    except Exception as e:
        duration = time.time() - start_time
        if _agent:
            _agent.report_failure(
                task_id=task_id,
                project_code=code,
                task_type="image",
                error=str(e),
                duration=duration
            )
            _agent.update_status(
                state="error",
                current_project="",
                current_task="",
                current_scene=0,
                total_scenes=0,
                progress=0
            )
        raise


def run_scan_loop_with_agent():
    """Run scan loop với Agent Protocol."""
    global _agent

    print(f"\n{'='*60}")
    print(f"  CHROME WORKER 2 - PIC BASIC MODE")
    print(f"  Agent: {'Enabled' if _agent else 'Disabled'}")
    print(f"{'='*60}\n")

    cycle = 0
    current_project = None  # Track project đang làm

    try:
        while True:
            cycle += 1
            print(f"\n[CYCLE {cycle}] Scanning...")

            # Tìm projects cần xử lý (từ local và master)
            projects = scan_incomplete_local_projects()
            if not projects:
                projects = scan_master_projects()

            if projects:
                # CHỈ XỬ LÝ 1 PROJECT - ưu tiên project đang làm dở
                if current_project and current_project in projects:
                    target = current_project
                    print(f"  [Chrome2] Continuing: {target}")
                else:
                    target = projects[0]
                    current_project = target
                    print(f"  [Chrome2] Starting: {target}")

                try:
                    success = process_project_with_agent(target)
                    if success:
                        print(f"  [Chrome2] Project {target} completed!")
                        current_project = None  # Move to next project
                    else:
                        print(f"  [Chrome2] Project {target} incomplete, will retry...")
                except KeyboardInterrupt:
                    raise
                except Exception as e:
                    agent_log(f"Error processing {target}: {e}", "ERROR")
            else:
                current_project = None  # Reset khi không còn project

            print(f"\nWaiting {SCAN_INTERVAL}s... (Ctrl+C to stop)")
            time.sleep(SCAN_INTERVAL)

    except KeyboardInterrupt:
        print("\n\nStopped by user.")


if __name__ == "__main__":
    main()
